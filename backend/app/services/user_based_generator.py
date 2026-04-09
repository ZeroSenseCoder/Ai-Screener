"""
User-Based Valuation for SaaS / Platform companies.
LTV/CAC analysis + 5-year user projection + DCF.
"""
import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.services.excel_utils import (
    font, fill, thin_border, align, safe, cr,
    section_header, build_cover, build_inputs, build_sensitivity_table,
)

THEME = {
    "cover_bg":      "001A0D",
    "primary":       "003319",
    "sub":           "005526",
    "accent":        "00C853",
    "input_color":   "1B5E20",
    "positive_fill": "E8F5E9",
    "positive_text": "1B5E20",
    "subtotal_fill": "C8E6C9",
    "key_fill":      "FDFEFE",
}


def _col(c: int) -> str:
    return get_column_letter(c)


def _lbl(ws, row, text, indent=0, bold=False):
    c = ws.cell(row=row, column=2, value=("    " * indent) + text)
    c.font = font(bold=bold, size=9)
    c.border = thin_border()
    c.alignment = align()
    return c


def _val(ws, row, col, val, fmt="#,##0.00", bold=False, bg=None, color="000000"):
    c = ws.cell(row=row, column=col, value=val)
    c.font = font(bold=bold, size=9, color=color)
    c.number_format = fmt
    c.alignment = align(h="right")
    c.border = thin_border()
    if bg:
        c.fill = fill(bg)
    return c


def _hdr(ws, row, cols, theme):
    for i, txt in enumerate(cols):
        c = ws.cell(row=row, column=2 + i, value=txt)
        c.font = font(bold=True, color="FFFFFF", size=9)
        c.fill = fill(theme["primary"])
        c.alignment = align(h="center")
        c.border = thin_border()


def _project_5y(users_0, arpu_0, arpu_growth, user_growth, churn,
                gm, ebitda_start_margin, ebitda_step, ke):
    """Returns list of dicts for each of 5 years."""
    results = []
    users_t = users_0
    arpu_t  = arpu_0
    em_t    = ebitda_start_margin
    for t in range(1, 6):
        new_acq   = users_t * user_growth
        churned   = users_t * churn
        end_users = users_t + new_acq - churned
        arpu_t    = arpu_t * (1 + arpu_growth)
        rev_t     = end_users * arpu_t / 1e7    # in Crores
        gp_t      = rev_t * gm
        ebitda_t  = rev_t * em_t
        fcf_t     = rev_t * 0.20
        results.append(dict(
            t=t, start=users_t, new_acq=new_acq, churned=churned,
            end=end_users, arpu=arpu_t, rev=rev_t,
            gp=gp_t, ebitda=ebitda_t, fcf=fcf_t,
        ))
        users_t = end_users
        em_t    = min(em_t + ebitda_step, 0.40)
    return results


def _compute(fin: dict):
    price   = safe(fin.get("price"), 100.0)
    shares  = safe(fin.get("shares"), 1e8)
    beta    = safe(fin.get("beta"), 1.0)
    revenue = safe(fin.get("revenue"), 1.0)
    ebitda  = safe(fin.get("ebitda"), 0.0)
    debt_cr = cr(fin.get("total_debt"))
    cash_cr = cr(fin.get("cash"))

    shares_cr = shares / 1e7
    mktcap_cr = price * shares / 1e7
    ke        = 0.071 + beta * 0.055
    g         = 0.05

    # User economics
    arpu       = revenue * 1e-6   # proxy: revenue / 1M users
    arpu       = max(arpu, 1.0)
    users_0    = revenue / (arpu * 1e7) if arpu else 1e6
    churn      = 0.10
    gm         = 0.70
    ltv        = arpu * gm / churn if churn else 0.0
    cac        = arpu * 0.20
    ltv_cac    = ltv / cac if cac else 0.0
    lifetime   = 1.0 / churn if churn else 10.0

    # Current EBITDA margin
    ebitda_cr  = ebitda / 1e7
    rev_cr     = revenue / 1e7
    em_base    = ebitda_cr / rev_cr if rev_cr else 0.05

    user_growth = 0.20
    arpu_growth = 0.03
    proj = _project_5y(users_0, arpu, arpu_growth, user_growth, churn,
                       gm, em_base, 0.05, ke)

    # DCF value build
    pv_fcf = sum(p["fcf"] / (1 + ke) ** p["t"] for p in proj)
    fcf_5  = proj[-1]["fcf"]
    tv     = fcf_5 * (1 + g) / (ke - g) if ke > g else 0.0
    pv_tv  = tv / (1 + ke) ** 5
    ev     = pv_fcf + pv_tv
    eq_val = ev - debt_cr + cash_cr
    implied_price = eq_val / shares_cr if shares_cr else 0.0
    upside = (implied_price - price) / price if price else 0.0

    # Platform metrics
    monthly_rev_pu = arpu / 12
    payback        = cac / monthly_rev_pu if monthly_rev_pu else 0.0
    nrr            = 1.05

    return dict(
        price=price, shares_cr=shares_cr, mktcap_cr=mktcap_cr,
        beta=beta, ke=ke, g=g,
        users_0=users_0, arpu=arpu, churn=churn, gm=gm,
        ltv=ltv, cac=cac, ltv_cac=ltv_cac, lifetime=lifetime,
        user_growth=user_growth, arpu_growth=arpu_growth,
        proj=proj, pv_fcf=pv_fcf, tv=tv, pv_tv=pv_tv,
        ev=ev, eq_val=eq_val, implied_price=implied_price, upside=upside,
        debt_cr=debt_cr, cash_cr=cash_cr,
        nrr=nrr, payback=payback, monthly_rev_pu=monthly_rev_pu,
    )


def _build_analysis(wb, theme, fin, d):
    ws = wb.create_sheet("User-Based Valuation")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 16
    for i in range(4, 10):
        ws.column_dimensions[_col(i)].width = 14

    row = 2
    t = ws.cell(row=row, column=2,
                value="User-Based Valuation  —  LTV/CAC  ×  5Y User Projection  ×  DCF")
    t.font = font(bold=True, size=13)
    ws.merge_cells(f"B{row}:{_col(9)}{row}")
    row += 2

    # ── USER ECONOMICS ──
    section_header(ws, row, 2, "▌ USER ECONOMICS", theme, span=3)
    row += 1
    econ_rows = [
        ("Total Users  (estimated)",                 d["users_0"],    "#,##0"),
        ("ARPU  (₹ / year)",                         d["arpu"],       "#,##0.00"),
        ("Annual Churn Rate",                         d["churn"],      "0.0%"),
        ("User Lifetime  =  1 / Churn  (years)",      d["lifetime"],   "0.0"),
        ("Gross Margin",                              d["gm"],         "0.0%"),
        ("LTV  =  ARPU × GM / Churn  (₹)",           d["ltv"],        "#,##0.00"),
        ("CAC  =  ARPU × 20%  (₹)",                  d["cac"],        "#,##0.00"),
        ("★ LTV / CAC  (>3× = good)",                d["ltv_cac"],    "0.00x"),
    ]
    for lbl_txt, val, fmt_ in econ_rows:
        is_key = lbl_txt.startswith("★")
        _lbl(ws, row, lbl_txt, bold=is_key)
        _val(ws, row, 3, val, fmt_, bold=is_key,
             bg=theme["key_fill"] if is_key else None,
             color=theme["accent"] if is_key else "000000")
        row += 1
    row += 1

    # ── 5-YEAR PROJECTIONS ──
    YEAR_COLS = [f"Year {p['t']}" for p in d["proj"]]
    section_header(ws, row, 2, "▌ 5-YEAR USER PROJECTIONS", theme, span=7)
    row += 1
    _hdr(ws, row, ["Metric"] + YEAR_COLS, theme)
    row += 1

    proj_rows = [
        ("Starting Users",            [p["start"]   for p in d["proj"]], "#,##0"),
        ("New Acquired  (20% growth)",[p["new_acq"] for p in d["proj"]], "#,##0"),
        ("Churned  (opening × churn)",[p["churned"] for p in d["proj"]], "#,##0"),
        ("Ending Users",              [p["end"]     for p in d["proj"]], "#,##0"),
        ("ARPU  (₹)",                 [p["arpu"]    for p in d["proj"]], "#,##0.00"),
        ("Revenue  (₹ Cr)",           [p["rev"]     for p in d["proj"]], "#,##0"),
        ("Gross Profit  (₹ Cr)",      [p["gp"]      for p in d["proj"]], "#,##0"),
        ("EBITDA  (₹ Cr)",            [p["ebitda"]  for p in d["proj"]], "#,##0"),
        ("FCF  (₹ Cr)",               [p["fcf"]     for p in d["proj"]], "#,##0"),
    ]
    for lbl_txt, vals, fmt_ in proj_rows:
        _lbl(ws, row, lbl_txt, indent=1)
        for j, v in enumerate(vals):
            _val(ws, row, 3 + j, v, fmt_)
        row += 1
    row += 1

    # ── DCF VALUE BUILD ──
    section_header(ws, row, 2, "▌ DCF VALUE BUILD", theme, span=3)
    row += 1
    dcf_rows = [
        ("PV of FCF  Years 1-5  (₹ Cr)",              d["pv_fcf"],        "#,##0",       False),
        ("Terminal Value  =  FCF5×(1+g)/(ke−g)  (₹ Cr)", d["tv"],         "#,##0",       False),
        ("PV of Terminal Value  (₹ Cr)",               d["pv_tv"],         "#,##0",       False),
        ("Enterprise Value  (₹ Cr)",                   d["ev"],            "#,##0",       False),
        ("Less: Total Debt  (₹ Cr)",                   d["debt_cr"],       "#,##0",       False),
        ("Add: Cash  (₹ Cr)",                          d["cash_cr"],       "#,##0",       False),
        ("Equity Value  (₹ Cr)",                       d["eq_val"],        "#,##0",       False),
        ("Shares Outstanding  (Cr)",                   d["shares_cr"],     "#,##0.00",    False),
        ("★ Implied Share Price  (₹)",                 d["implied_price"], "#,##0.00",    True),
        ("Current Market Price  (₹)",                  d["price"],         "#,##0.00",    False),
        ("★ Upside / (Downside)",                      d["upside"],        "+0.0%;-0.0%", True),
    ]
    for lbl_txt, val, fmt_, is_key in dcf_rows:
        _lbl(ws, row, lbl_txt, bold=is_key)
        _val(ws, row, 3, val, fmt_, bold=is_key,
             bg=theme["key_fill"] if is_key else None,
             color=theme["accent"] if is_key else "000000")
        row += 1
    row += 1

    # ── PLATFORM METRICS ──
    section_header(ws, row, 2, "▌ PLATFORM METRICS", theme, span=3)
    row += 1
    plat = [
        ("LTV / CAC  (>3× = healthy)",         d["ltv_cac"],         "0.00x"),
        ("Net Revenue Retention  (est.)",      d["nrr"],              "0.0%"),
        ("Payback Period  (months)",           d["payback"],          "0.0"),
        ("Monthly Revenue per User  (₹)",     d["monthly_rev_pu"],   "#,##0.00"),
    ]
    for lbl_txt, val, fmt_ in plat:
        _lbl(ws, row, lbl_txt, indent=1)
        _val(ws, row, 3, val, fmt_)
        row += 1

    ws.freeze_panes = "C4"


def _build_results_sheet(wb, theme, company_name, d):
    ws = wb.create_sheet("Results & Sensitivity")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 16

    row = 2
    section_header(ws, row, 2, "▌ VALUATION RESULTS SUMMARY", theme, span=4)
    row += 1
    results = [
        ("Total Users (est.)",               d["users_0"],       "#,##0",       False),
        ("LTV / CAC",                         d["ltv_cac"],       "0.00x",       False),
        ("Enterprise Value  (₹ Cr)",          d["ev"],            "#,##0",       False),
        ("★ Implied Share Price  (₹)",        d["implied_price"], "#,##0.00",    True),
        ("Current Price  (₹)",                d["price"],         "#,##0.00",    False),
        ("★ Upside / (Downside)",             d["upside"],        "+0.0%;-0.0%", True),
        ("ke  (Cost of Equity)",              d["ke"],            "0.0%",        False),
        ("Terminal Growth  g",               d["g"],             "0.0%",        False),
    ]
    for lbl_txt, val, fmt_, is_key in results:
        lc = ws.cell(row=row, column=2, value=lbl_txt)
        lc.font = font(bold=is_key, size=9)
        lc.border = thin_border()
        if is_key:
            lc.fill = fill(theme["key_fill"])
        vc = ws.cell(row=row, column=3, value=val)
        vc.font = font(bold=is_key, size=9,
                       color=theme["accent"] if is_key else "000000")
        vc.number_format = fmt_
        vc.alignment = align(h="right")
        vc.border = thin_border()
        if is_key:
            vc.fill = fill(theme["key_fill"])
        row += 1

    # Sensitivity: User Growth Rate × Annual Churn Rate
    ugrowth_vals = [0.05, 0.10, 0.15, 0.20, 0.25, 0.30, 0.35, 0.40]
    churn_vals   = [0.03, 0.05, 0.08, 0.10, 0.13, 0.15, 0.20]

    users_0  = d["users_0"]
    arpu     = d["arpu"]
    arpu_g   = d["arpu_growth"]
    gm       = d["gm"]
    ke       = d["ke"]
    g        = d["g"]
    debt_cr  = d["debt_cr"]
    cash_cr  = d["cash_cr"]
    sc       = d["shares_cr"]
    price    = d["price"]
    em_base  = 0.10

    base_ui = min(range(len(ugrowth_vals)), key=lambda i: abs(ugrowth_vals[i] - d["user_growth"]))
    base_ci = min(range(len(churn_vals)),   key=lambda i: abs(churn_vals[i]   - d["churn"]))

    matrix = []
    for ug in ugrowth_vals:
        row_data = []
        for ch in churn_vals:
            proj_ = _project_5y(users_0, arpu, arpu_g, ug, ch, gm, em_base, 0.05, ke)
            pv_   = sum(p["fcf"] / (1 + ke) ** p["t"] for p in proj_)
            fcf5_ = proj_[-1]["fcf"]
            tv_   = fcf5_ * (1 + g) / (ke - g) if ke > g else 0.0
            pv_tv_= tv_ / (1 + ke) ** 5
            ev_   = pv_ + pv_tv_
            eq_   = ev_ - debt_cr + cash_cr
            ip_   = eq_ / sc if sc else 0.0
            row_data.append(round(ip_, 2))
        matrix.append(row_data)

    build_sensitivity_table(
        ws, theme,
        row_label="User Growth Rate",
        col_label="Annual Churn Rate",
        row_vals=ugrowth_vals,
        col_vals=churn_vals,
        matrix=matrix,
        current_price=price,
        base_row_idx=base_ui,
        base_col_idx=base_ci,
        start_row=row + 2,
        start_col=2,
        row_fmt="0.0%",
        col_fmt="0.0%",
    )


def generate_excel(fin: dict, symbol: str, company_name: str) -> bytes:
    d  = _compute(fin)
    wb = Workbook()

    build_cover(
        wb, THEME, symbol, company_name,
        model_label="User-Based Valuation",
        model_desc="LTV/CAC analysis + 5-year user projections + DCF. "
                   "Designed for SaaS, platform, and subscription businesses.",
        sheets_index=[
            (1, "Cover",                "Model overview & index"),
            (2, "Inputs & Assumptions", "Financial data & model parameters"),
            (3, "User-Based Valuation", "User economics, projections & DCF build"),
            (4, "Results & Sensitivity","Summary + user growth × churn sensitivity"),
        ],
        meta_extra={
            "price":     d["price"],
            "mktcap_cr": d["mktcap_cr"],
            "sector":    safe(fin.get("sector"), "—"),
        },
    )

    params_def = [
        ("Annual Churn Rate",         d["churn"],      "%",  "Default 10%",            "0.0%", True),
        ("User Growth Rate",          d["user_growth"],"%",  "Default 20% YoY",        "0.0%", True),
        ("ARPU  (₹ / year)",          d["arpu"],       "₹",  "Revenue per user proxy",  "#,##0.00", True),
        ("Gross Margin",              d["gm"],         "%",  "Default 70%",             "0.0%", False),
        ("ARPU Growth Rate",          d["arpu_growth"],"%",  "3% pa default",           "0.0%", False),
        ("FCF Margin Assumption",     0.20,            "%",  "20% of revenue",          "0.0%", False),
        ("ke (Cost of Equity)",       d["ke"],         "%",  "rf + β × ERP",            "0.0%", True),
        ("Terminal Growth g",         d["g"],          "%",  "Long-run growth",         "0.0%", False),
        ("Risk-Free Rate (rf)",       0.071,           "%",  "India 10Y GSec",          "0.0%", False),
        ("Equity Risk Premium",       0.055,           "%",  "Damodaran India ERP",      "0.0%", False),
        ("Beta",                      d["beta"],       "x",  "Market beta",             "0.00", False),
    ]

    build_inputs(wb, THEME, company_name, fin, params_def)
    _build_analysis(wb, THEME, fin, d)
    _build_results_sheet(wb, THEME, company_name, d)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
