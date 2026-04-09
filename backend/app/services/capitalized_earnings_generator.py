"""
Capitalized Earnings Excel Generator
Value = Normalized EPS / Required Return Rate
Simple, intuitive for stable, mature earners.
"""
import io
from openpyxl import Workbook

from app.services.excel_utils import (
    font, fill, thin_border, align, safe, cr,
    section_header, sub_header, label, value_cell, pct_cell,
    build_cover, build_inputs, build_results, build_sensitivity_table,
)

THEME = {
    "cover_bg":      "001A1A",
    "primary":       "003333",
    "sub":           "005555",
    "accent":        "00CED1",
    "input_color":   "0E6655",
    "positive_fill": "D0F0F0",
    "positive_text": "0E6655",
    "subtotal_fill": "A8D8D8",
    "key_fill":      "FFF9C4",
}

RF_RATE  = 0.072   # 10Y G-Sec
ERP      = 0.055   # Equity Risk Premium
DEFAULT_EPS_GROWTH = 0.05


# ─────────────────────────────────────────────────────────────────────────────
def _capm_ke(beta):
    return RF_RATE + beta * ERP


def _cap_value(eps, ke, eps_growth):
    cap_rate = ke - eps_growth
    if cap_rate <= 0:
        return 0.0
    return eps / cap_rate


# ─────────────────────────────────────────────────────────────────────────────
def _build_ce_analysis(wb, theme, fin, ke, eps_growth):
    ws = wb.create_sheet("Capitalized Earnings")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16

    eps     = safe(fin.get("eps"), 0.0)
    price   = safe(fin.get("price"), 0.0)
    beta    = safe(fin.get("beta"), 1.0)
    rev_g   = safe(fin.get("revenue_growth"), 0.10)

    # 3Y/5Y average EPS estimate: back-calculate using revenue growth as proxy
    eps_3y  = eps / ((1 + rev_g) ** 1.5) if rev_g > -1 else eps
    eps_5y  = eps / ((1 + rev_g) ** 2.5) if rev_g > -1 else eps
    norm_eps = (eps + eps_3y + eps_5y) / 3

    cap_rate    = ke - eps_growth
    if cap_rate <= 0:
        cap_rate = 0.01
    intrinsic   = norm_eps / cap_rate
    upside      = (intrinsic - price) / price if price else 0
    imp_pe_intr = intrinsic / norm_eps if norm_eps else 0
    imp_pe_curr = price     / norm_eps if norm_eps else 0
    ey_curr     = norm_eps  / price    if price    else 0
    ey_intr     = norm_eps  / intrinsic if intrinsic else 0
    erp_spread  = ey_curr - RF_RATE

    row = 2
    t = ws.cell(row=row, column=2, value="Capitalized Earnings Valuation")
    t.font = font(bold=True, size=13)
    ws.merge_cells(f"B{row}:F{row}")

    # ── Section 1: Earnings Normalization ─────────────────────────────────────
    row += 2
    section_header(ws, row, 2, "▌ EARNINGS NORMALIZATION", theme, span=5)
    row += 1

    for lbl_txt, val, fmt, is_key in [
        ("Reported EPS (₹)",                   eps,      "#,##0.00", False),
        ("3Y Avg EPS (est., rev growth proxy)", eps_3y,  "#,##0.00", False),
        ("5Y Avg EPS (est., rev growth proxy)", eps_5y,  "#,##0.00", False),
        ("Normalized EPS (avg of above)",       norm_eps, "#,##0.00", False),
        ("Adj for extraordinary items",         0.0,     "#,##0.00", False),
        ("★ Normalized EPS Used (₹)",           norm_eps, "#,##0.00", True),
    ]:
        c = ws.cell(row=row, column=2, value=lbl_txt)
        c.font = font(bold=is_key, size=9)
        c.border = thin_border()
        if is_key:
            c.fill = fill(theme["key_fill"])
        v = ws.cell(row=row, column=3, value=val)
        v.number_format = fmt
        v.font = font(bold=is_key, size=9,
                      color=theme["accent"] if is_key else "000000")
        v.alignment = align(h="right")
        v.border = thin_border()
        if is_key:
            v.fill = fill(theme["key_fill"])
        row += 1

    # EPS forward projections
    row += 1
    sub_header(ws, row, 2, "EPS Growth Projections", theme, span=5)
    row += 1

    hdr = ["Year", "EPS (₹)", "Growth Rate", "vs Current EPS"]
    for j, h in enumerate(hdr):
        c = ws.cell(row=row, column=2 + j, value=h)
        c.font = font(bold=True, color="FFFFFF", size=9)
        c.fill = fill(theme["sub"])
        c.border = thin_border()
        c.alignment = align()
    row += 1

    for yr in range(1, 4):
        proj_eps = norm_eps * (1 + rev_g) ** yr
        vs_curr  = (proj_eps - eps) / eps if eps else 0
        ws.cell(row=row, column=2, value=f"Year {yr}").font = font(size=9)
        ws.cell(row=row, column=2).border = thin_border()
        value_cell(ws, row, 3, proj_eps, "#,##0.00")
        pct_cell(ws,   row, 4, rev_g)
        pct_cell(ws,   row, 5, vs_curr)
        row += 1

    # ── Section 2: Capitalization Calculation ─────────────────────────────────
    row += 1
    section_header(ws, row, 2, "▌ CAPITALIZATION CALCULATION", theme, span=5)
    row += 1

    for lbl_txt, val, fmt, is_key in [
        ("Normalized EPS (₹)",                  norm_eps,   "#,##0.00", False),
        ("Required Return / ke (CAPM)",          ke,         "0.0%",    False),
        ("EPS Growth Rate (g)",                  eps_growth, "0.0%",    False),
        ("Capitalization Rate = ke − g",         cap_rate,   "0.0%",    True),
        ("★ Intrinsic Value = EPS / Cap Rate (₹)", intrinsic,"#,##0.00", True),
        ("Current Market Price (₹)",             price,      "#,##0.00", False),
        ("★ Upside / (Downside)",                upside,     "0.0%",    True),
        ("Implied P/E at Intrinsic Value",       imp_pe_intr,"0.0x",    False),
        ("Implied P/E at Current Price",         imp_pe_curr,"0.0x",    False),
    ]:
        c = ws.cell(row=row, column=2, value=lbl_txt)
        c.font = font(bold=is_key, size=9)
        c.border = thin_border()
        if is_key:
            c.fill = fill(theme["key_fill"])
        v = ws.cell(row=row, column=3, value=val)
        v.number_format = fmt
        v.font = font(bold=is_key, size=9,
                      color=theme["accent"] if is_key else "000000")
        v.alignment = align(h="right")
        v.border = thin_border()
        if is_key:
            v.fill = fill(theme["key_fill"])
        row += 1

    # ── Section 3: Earnings Yield Analysis ────────────────────────────────────
    row += 1
    section_header(ws, row, 2, "▌ EARNINGS YIELD ANALYSIS", theme, span=5)
    row += 1

    for lbl_txt, val, fmt, is_key in [
        ("Earnings Yield at Current Price",      ey_curr,   "0.0%", False),
        ("Earnings Yield at Intrinsic Value",    ey_intr,   "0.0%", False),
        ("Risk-Free Rate (10Y G-Sec)",           RF_RATE,   "0.0%", False),
        ("Equity Risk Premium (ERP)",            ERP,       "0.0%", False),
        ("Earnings Yield Spread over Risk-Free", erp_spread,"0.0%", True),
    ]:
        c = ws.cell(row=row, column=2, value=lbl_txt)
        c.font = font(bold=is_key, size=9)
        c.border = thin_border()
        if is_key:
            c.fill = fill(theme["key_fill"])
        v = ws.cell(row=row, column=3, value=val)
        v.number_format = fmt
        v.font = font(bold=is_key, size=9,
                      color=theme["accent"] if is_key else "000000")
        v.alignment = align(h="right")
        v.border = thin_border()
        if is_key:
            v.fill = fill(theme["key_fill"])
        row += 1

    # ── Section 4: Capitalization Schedule ────────────────────────────────────
    row += 1
    section_header(ws, row, 2, "▌ EARNINGS CAPITALIZATION SCHEDULE (5 YEARS)", theme, span=5)
    row += 1

    hdr2 = ["Year", "EPS (₹)", "Required Return", "Cap Rate", "Capitalized Value (₹)", "vs Current Price"]
    for j, h in enumerate(hdr2):
        c = ws.cell(row=row, column=2 + j, value=h)
        c.font = font(bold=True, color="FFFFFF", size=9)
        c.fill = fill(theme["sub"])
        c.border = thin_border()
        c.alignment = align()
    row += 1

    for yr in range(1, 6):
        yr_eps  = norm_eps * (1 + rev_g) ** yr
        yr_ke   = ke
        yr_cr   = yr_ke - eps_growth
        yr_cv   = yr_eps / yr_cr if yr_cr > 0 else 0
        vs_pr   = (yr_cv - price) / price if price else 0

        ws.cell(row=row, column=2, value=f"Year {yr}").font = font(size=9)
        ws.cell(row=row, column=2).border = thin_border()
        value_cell(ws, row, 3, yr_eps, "#,##0.00")
        pct_cell(ws,   row, 4, yr_ke)
        pct_cell(ws,   row, 5, yr_cr)
        value_cell(ws, row, 6, yr_cv, "#,##0.00",
                   bg=theme["positive_fill"] if yr == 1 else None)
        pct_cell(ws,   row, 7, vs_pr)
        row += 1

    ws.freeze_panes = "C4"
    return ws


def _build_ce_sensitivity(wb, theme, fin, eps):
    ws = wb.create_sheet("Sensitivity")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 22

    price = safe(fin.get("price"), 0.0)

    row_vals = [0.10, 0.11, 0.12, 0.13, 0.14, 0.15, 0.16, 0.17, 0.18]
    col_vals = [0.02, 0.03, 0.04, 0.05, 0.06, 0.08, 0.10, 0.12]

    base_ke_idx  = 3   # 13%
    base_g_idx   = 3   # 5%

    matrix = []
    for ke in row_vals:
        row_data = []
        for g in col_vals:
            cv = _cap_value(eps, ke, g)
            row_data.append(round(cv, 2))
        matrix.append(row_data)

    row = 2
    t = ws.cell(row=row, column=2, value="Capitalized Earnings Sensitivity Analysis")
    t.font = font(bold=True, size=13)
    ws.merge_cells(f"B{row}:J{row}")
    row += 2

    build_sensitivity_table(
        ws, theme,
        row_label="Required Return",
        col_label="EPS Growth (g)",
        row_vals=row_vals,
        col_vals=col_vals,
        matrix=matrix,
        current_price=price,
        base_row_idx=base_ke_idx,
        base_col_idx=base_g_idx,
        start_row=row,
    )
    return ws


# ─────────────────────────────────────────────────────────────────────────────
def generate_excel(fin: dict, symbol: str, company_name: str) -> bytes:
    wb = Workbook()

    price     = safe(fin.get("price"), 0.0)
    shares    = safe(fin.get("shares"), 1e8)
    mktcap_cr = price * shares / 1e7
    beta      = safe(fin.get("beta"), 1.0)
    ke        = _capm_ke(beta)
    eps       = safe(fin.get("eps"), 0.0)
    rev_g     = safe(fin.get("revenue_growth"), 0.10)

    # Normalized EPS for sensitivity
    eps_3y    = eps / ((1 + rev_g) ** 1.5) if rev_g > -1 else eps
    eps_5y    = eps / ((1 + rev_g) ** 2.5) if rev_g > -1 else eps
    norm_eps  = (eps + eps_3y + eps_5y) / 3

    build_cover(
        wb, THEME, symbol, company_name,
        model_label="Capitalized Earnings Model",
        model_desc="Intrinsic Value = Normalized EPS ÷ (Required Return − Growth Rate)",
        sheets_index=[
            (1, "Cover",               "Company overview & model index"),
            (2, "Inputs & Assumptions","Financial data & model parameters"),
            (3, "Capitalized Earnings","EPS normalization, cap rate, earnings yield"),
            (4, "Sensitivity",         "Required Return × EPS growth — implied value"),
        ],
        meta_extra={"price": price, "mktcap_cr": mktcap_cr, "sector": safe(fin.get("sector"), "—")},
    )

    params_def = [
        ("Risk-Free Rate (10Y G-Sec)",    RF_RATE,            "%",  "RBI 10Y bond yield",                  "0.0%", True),
        ("Equity Risk Premium",           ERP,                "%",  "Historical India ERP",                "0.0%", False),
        ("Beta",                          beta,               "x",  "5Y monthly regression",               "0.00", False),
        ("Cost of Equity / ke (CAPM)",    ke,                 "%",  "Rf + β × ERP",                        "0.0%", True),
        ("EPS Growth Rate (g)",           DEFAULT_EPS_GROWTH, "%",  "Sustainable long-run EPS growth",     "0.0%", True),
        ("Capitalization Rate = ke − g",  max(0.01, ke - DEFAULT_EPS_GROWTH), "%", "Denominator of cap model", "0.0%", True),
        ("Reported EPS (₹)",              eps,                "₹",  "Trailing twelve months",              "#,##0.00", False),
        ("Normalized EPS (₹)",            norm_eps,           "₹",  "3Y/5Y average smoothing",             "#,##0.00", True),
    ]
    build_inputs(wb, THEME, company_name, fin, params_def)

    _build_ce_analysis(wb, THEME, fin, ke, DEFAULT_EPS_GROWTH)
    _build_ce_sensitivity(wb, THEME, fin, norm_eps)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
