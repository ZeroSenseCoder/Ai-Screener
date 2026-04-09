"""
EV/Revenue Multiple Valuation Generator
==========================================
Enterprise Value = Revenue × EV/Revenue Multiple

Formula (EV/Revenue is an ENTERPRISE VALUE multiple — NOT P/S):
  Implied EV     = Revenue × EV/Revenue_Multiple
  Equity Value   = Implied EV  −  Net Debt  (= Total Debt − Cash)
  Implied Price  = Equity Value / Shares Outstanding

Key principle:
  EV/Revenue is debt-adjusted (unlike P/S which is a pure price/equity multiple).
  Debt & cash must always be brought through the equity bridge.

Both LTM and NTM (forward) revenue are shown — NTM is the IB standard since
investors price future performance.
"""

import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.services.excel_utils import (
    font, fill, thin_border, align, safe, cr,
    section_header, build_cover, build_inputs, build_sensitivity_table,
)

THEME = {
    "cover_bg":      "1A001A",
    "primary":       "3B003B",
    "sub":           "660066",
    "accent":        "E74C3C",
    "input_color":   "7B241C",
    "positive_fill": "FDEDEC",
    "positive_text": "7B241C",
    "subtotal_fill": "F5B7B1",
    "key_fill":      "FFF2CC",
}

# ── Sector EV/Revenue benchmark multiples (Indian market) ────────────────────
# (sector_key: (low, mid, high, default_multiple, description))
SECTOR_MULTIPLES = {
    "Information Technology": (3.5, 5.0, 7.0,  5.0,  "Large-cap IT services — TCS, Infosys range"),
    "Technology":             (4.0, 6.0, 10.0, 6.0,  "Tech / software companies"),
    "Software":               (5.0, 8.0, 14.0, 8.0,  "Pure-play software / SaaS"),
    "Consumer Defensive":     (2.5, 4.0, 6.0,  4.0,  "FMCG — stable revenues, brand premium"),
    "Consumer Staples":       (2.5, 4.0, 6.0,  4.0,  "Consumer staples — defensive names"),
    "FMCG":                   (2.5, 4.0, 6.0,  4.0,  "FMCG sector"),
    "Healthcare":             (2.5, 4.5, 7.0,  4.0,  "Hospitals, diagnostics, pharma"),
    "Pharmaceuticals":        (2.5, 4.0, 6.5,  4.0,  "Generics & specialty pharma"),
    "Financial Services":     (1.0, 2.0, 3.5,  2.0,  "Banks / NBFCs — interest income as revenue"),
    "Banking":                (1.0, 2.0, 3.5,  2.0,  "Banking sector"),
    "Energy":                 (0.5, 1.0, 1.8,  1.0,  "Oil & gas, petrochemicals — capital intensive"),
    "Oil & Gas":              (0.5, 1.0, 1.8,  1.0,  "Upstream / downstream energy"),
    "Consumer Cyclical":      (1.0, 2.0, 3.5,  2.0,  "Autos, retail, durables"),
    "Industrials":            (1.0, 2.0, 3.0,  1.8,  "Infra, engineering, capital goods"),
    "Communication Services": (1.5, 2.5, 4.0,  2.5,  "Telecom — Bharti, Jio range"),
    "Real Estate":            (3.0, 5.0, 8.0,  5.0,  "Residential + commercial developers"),
    "Basic Materials":        (0.8, 1.5, 2.5,  1.5,  "Metals, mining, chemicals"),
    "Utilities":              (1.0, 2.0, 3.0,  1.8,  "Power generation & distribution"),
}
DEFAULT_MULTIPLE = (1.0, 3.0, 6.0, 3.0, "General market benchmark")


def _get_sector_info(sector: str):
    """Return (low, mid, high, default_mult, description) for sector."""
    if not sector:
        return DEFAULT_MULTIPLE
    if sector in SECTOR_MULTIPLES:
        return SECTOR_MULTIPLES[sector]
    low = sector.lower()
    for k, v in SECTOR_MULTIPLES.items():
        if k.lower() == low or k.lower() in low or low in k.lower():
            return v
    return DEFAULT_MULTIPLE


def _col(c: int) -> str:
    return get_column_letter(c)


def _lbl(ws, row, text, bold=False, bg=None):
    c = ws.cell(row=row, column=2, value=text)
    c.font      = font(bold=bold, size=9)
    c.border    = thin_border()
    c.alignment = align()
    if bg:
        c.fill = fill(bg)
    return c


def _val(ws, row, col, val, fmt="#,##0.00", bold=False, bg=None, color="000000"):
    c = ws.cell(row=row, column=col, value=val if val is not None else "N/M")
    c.font          = font(bold=bold, size=9, color=color)
    c.number_format = fmt if val is not None else "@"
    c.alignment     = align(h="right")
    c.border        = thin_border()
    if bg:
        c.fill = fill(bg)
    return c


def _hdr(ws, row, cols, theme):
    for i, txt in enumerate(cols):
        c = ws.cell(row=row, column=2 + i, value=txt)
        c.font      = font(bold=True, color="FFFFFF", size=9)
        c.fill      = fill(theme["sub"])
        c.border    = thin_border()
        c.alignment = align(h="center" if i > 0 else "left")


# ── Core computation ──────────────────────────────────────────────────────────

def _compute(fin: dict):
    price      = safe(fin.get("price"), 0.0)
    shares_raw = fin.get("shares") or 0            # from balance sheet filing
    shares_cr  = shares_raw / 1e7
    revenue    = safe(fin.get("revenue"), 0.0)
    ebitda     = safe(fin.get("ebitda"), 0.0)
    net_income = safe(fin.get("net_income"), 0.0)
    total_debt = safe(fin.get("total_debt"), 0.0)
    cash_val   = safe(fin.get("cash"), 0.0)
    rev_growth = safe(fin.get("revenue_growth"), 0.10)
    sector     = fin.get("sector", "")

    rev_cr     = cr(revenue)
    ebitda_cr  = cr(ebitda)
    ni_cr      = cr(net_income)
    debt_cr    = cr(total_debt)
    cash_cr    = cr(cash_val)
    net_debt   = debt_cr - cash_cr
    mktcap_cr  = price * shares_cr
    ev_mkt     = cr(fin.get("enterprise_value") or 0) or (mktcap_cr + net_debt)

    ebitda_margin   = ebitda_cr / rev_cr  if rev_cr > 0 else 0.0
    net_margin      = ni_cr    / rev_cr  if rev_cr > 0 else 0.0
    # Rule of 40: growth% + EBITDA margin% (expressed as percentage points)
    rule_of_40      = (rev_growth + ebitda_margin) * 100

    # Sector-derived multiple
    s_low, s_mid, s_high, s_default, s_desc = _get_sector_info(sector)
    multiple = s_default

    # LTM (current revenue)
    ltm_ev    = rev_cr * multiple
    ltm_eq    = ltm_ev - net_debt
    ltm_price = ltm_eq / shares_cr if shares_cr > 0 else 0.0
    ltm_up    = (ltm_price / price - 1) if price > 0 else 0.0

    # NTM (next twelve months — forward revenue, standard IB approach)
    ntm_rev   = rev_cr * (1 + rev_growth)
    ntm_ev    = ntm_rev * multiple
    ntm_eq    = ntm_ev - net_debt
    ntm_price = ntm_eq / shares_cr if shares_cr > 0 else 0.0
    ntm_up    = (ntm_price / price - 1) if price > 0 else 0.0

    # Market implied EV/Revenue (current)
    mkt_ev_rev = ev_mkt / rev_cr  if rev_cr > 0 else 0.0
    mkt_ev_ntm = ev_mkt / ntm_rev if ntm_rev > 0 else 0.0

    # 5-year forward revenue build (at constant growth)
    fwd = []
    rev_t = rev_cr
    for t in range(1, 6):
        rev_t = rev_t * (1 + rev_growth)
        ev_t  = rev_t * multiple
        ip_t  = (ev_t - net_debt) / shares_cr if shares_cr > 0 else 0.0
        fwd.append((t, rev_t, rev_growth, ev_t, ip_t))

    return dict(
        price=price, shares_cr=shares_cr,
        rev_cr=rev_cr, ntm_rev=ntm_rev, rev_growth=rev_growth,
        ebitda_cr=ebitda_cr, ni_cr=ni_cr,
        debt_cr=debt_cr, cash_cr=cash_cr, net_debt=net_debt,
        mktcap_cr=mktcap_cr, ev_mkt=ev_mkt,
        ebitda_margin=ebitda_margin, net_margin=net_margin,
        rule_of_40=rule_of_40,
        multiple=multiple,
        s_low=s_low, s_mid=s_mid, s_high=s_high, s_desc=s_desc,
        mkt_ev_rev=mkt_ev_rev, mkt_ev_ntm=mkt_ev_ntm,
        ltm_ev=ltm_ev, ltm_eq=ltm_eq, ltm_price=ltm_price, ltm_up=ltm_up,
        ntm_ev=ntm_ev, ntm_eq=ntm_eq, ntm_price=ntm_price, ntm_up=ntm_up,
        fwd=fwd,
    )


# ── Sheet 3: Revenue Multiple Analysis ───────────────────────────────────────

def _build_analysis(wb, theme, fin, d):
    ws = wb.create_sheet("Revenue Multiple Analysis")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16
    for i in range(5, 10):
        ws.column_dimensions[_col(i)].width = 14

    row = 2
    t = ws.cell(row=row, column=2,
                value="EV/Revenue Multiple Valuation  —  EV = Revenue × Multiple")
    t.font = font(bold=True, size=13)
    ws.merge_cells(f"B{row}:{_col(9)}{row}")
    row += 2

    # ── SECTION A: Company Revenue Profile ───────────────────────────────────
    section_header(ws, row, 2, "▌ A.  COMPANY REVENUE PROFILE", theme, span=3)
    row += 1
    _hdr(ws, row, ["Metric", "LTM Value", "Notes"], theme)
    row += 1
    for lbl_txt, val, fmt, note in [
        ("Revenue — LTM  (₹ Cr)",         d["rev_cr"],       "#,##0",
         "Last twelve months"),
        ("Revenue — NTM  (₹ Cr)",         d["ntm_rev"],      "#,##0",
         f"= LTM × (1 + {d['rev_growth']:.1%}) — NTM is IB standard"),
        ("YoY Revenue Growth",            d["rev_growth"],   "0.0%",
         "From yfinance / filings"),
        ("EBITDA  (₹ Cr)",                d["ebitda_cr"],    "#,##0",  ""),
        ("EBITDA Margin",                 d["ebitda_margin"],"0.0%",   ""),
        ("Net Margin",                    d["net_margin"],   "0.0%",   ""),
        ("Rule of 40  (Growth% + EBITDA Margin%)", d["rule_of_40"], "0.0",
         "> 40 = healthy SaaS/growth profile"),
        ("Market Cap  (₹ Cr)",            d["mktcap_cr"],    "#,##0",  ""),
        ("Enterprise Value — Market  (₹ Cr)", d["ev_mkt"],   "#,##0",  ""),
        ("Market EV/Revenue  (LTM)",      d["mkt_ev_rev"],   "0.0x",
         "Current market-implied multiple on LTM"),
        ("Market EV/Revenue  (NTM)",      d["mkt_ev_ntm"],   "0.0x",
         "Current market-implied multiple on NTM"),
    ]:
        _lbl(ws, row, "    " + lbl_txt)
        _val(ws, row, 3, val, fmt)
        nc = ws.cell(row=row, column=4, value=note)
        nc.font = font(size=8, color="595959"); nc.border = thin_border()
        row += 1
    row += 1

    # ── SECTION B: Sector Multiple Benchmarks ────────────────────────────────
    sector = fin.get("sector", "—")
    section_header(ws, row, 2,
                   f"▌ B.  SECTOR EV/REVENUE BENCHMARKS  —  {sector}", theme, span=4)
    row += 1
    _hdr(ws, row, ["Segment", "Low", "Mid", "High", "Note"], theme)
    row += 1

    all_benchmarks = [
        ("IT Services (large-cap)",  "3.5x", "5.0x", "7.0x",  "TCS, Infosys, Wipro"),
        ("Software / SaaS",          "5.0x", "8.0x", "14.0x", "Product / subscription model"),
        ("FMCG / Consumer Staples",  "2.5x", "4.0x", "6.0x",  "Stable revenues, brand premium"),
        ("Healthcare / Pharma",      "2.5x", "4.5x", "7.0x",  "Mix of generics & specialty"),
        ("Banking / Financial Svc",  "1.0x", "2.0x", "3.5x",  "NII as revenue proxy"),
        ("Energy / Oil & Gas",       "0.5x", "1.0x", "1.8x",  "Capital intensive, cyclical"),
        ("Autos / Consumer Cyclical","1.0x", "2.0x", "3.5x",  "Volume & ASP driven"),
        ("Telecom",                  "1.5x", "2.5x", "4.0x",  "ARPU growth key"),
        ("Infrastructure",           "1.0x", "2.0x", "3.0x",  "Long-cycle, regulated assets"),
    ]
    for seg, lo, mi, hi, note in all_benchmarks:
        is_active = (
            seg.split(" ")[0].lower() in sector.lower() or
            sector.lower() in seg.lower()
        )
        bg_c = theme["positive_fill"] if is_active else None
        for col_off, val in enumerate([seg, lo, mi, hi, note]):
            c = ws.cell(row=row, column=2 + col_off, value=val)
            c.font      = font(bold=is_active, size=9)
            c.border    = thin_border()
            c.alignment = align(h="right" if col_off in [1,2,3] else "left")
            if bg_c:
                c.fill = fill(bg_c)
        row += 1

    # Highlight current sector default
    note_c = ws.cell(row=row, column=2,
                     value=f"  ★  Using {d['multiple']:.1f}x for {sector}  —  {d['s_desc']}")
    note_c.font   = font(bold=True, size=9, color=theme["accent"])
    note_c.border = thin_border()
    ws.merge_cells(f"B{row}:F{row}")
    row += 2

    # ── SECTION C: LTM vs NTM Valuation ──────────────────────────────────────
    section_header(ws, row, 2, "▌ C.  LTM vs NTM VALUATION  (IB Standard: use NTM)", theme, span=4)
    row += 1
    _hdr(ws, row, ["Step", "LTM  (Trailing)", "NTM  (Forward)"], theme)
    row += 1

    steps = [
        ("Revenue  (₹ Cr)",               d["rev_cr"],    d["ntm_rev"],    "#,##0"),
        ("× EV/Revenue Multiple",          d["multiple"],  d["multiple"],   "0.0x"),
        ("= Implied EV  (₹ Cr)",           d["ltm_ev"],    d["ntm_ev"],     "#,##0"),
        ("− Net Debt  (₹ Cr)",             d["net_debt"],  d["net_debt"],   "#,##0"),
        ("= Equity Value  (₹ Cr)",         d["ltm_eq"],    d["ntm_eq"],     "#,##0"),
        ("÷ Shares  (Cr)  [balance sheet]",d["shares_cr"], d["shares_cr"],  "#,##0.00"),
        ("★  Implied Price  (₹)",          d["ltm_price"], d["ntm_price"],  "#,##0.00"),
        ("Current Price  (₹)",             d["price"],     d["price"],      "#,##0.00"),
        ("★  Upside / (Downside)",         d["ltm_up"],    d["ntm_up"],     "+0.0%;-0.0%;0.0%"),
    ]
    for lbl_txt, ltm_v, ntm_v, fmt in steps:
        is_key = lbl_txt.startswith("★")
        _lbl(ws, row, "    " + lbl_txt, bold=is_key,
             bg=theme["key_fill"] if is_key else None)
        for col_off, val in [(1, ltm_v), (2, ntm_v)]:
            _val(ws, row, 2 + col_off, val, fmt, bold=is_key,
                 bg=theme["key_fill"] if is_key else None,
                 color=theme["accent"] if is_key else "000000")
        row += 1
    row += 1

    # ── SECTION D: 5-Year Forward Revenue Build ──────────────────────────────
    section_header(ws, row, 2, "▌ D.  5-YEAR FORWARD REVENUE BUILD", theme, span=7)
    row += 1
    _hdr(ws, row, ["Metric", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"], theme)
    row += 1

    revs    = [f[1] for f in d["fwd"]]
    growths = [f[2] for f in d["fwd"]]
    evs     = [f[3] for f in d["fwd"]]
    ips     = [f[4] for f in d["fwd"]]

    for lbl_txt, vals, fmt in [
        ("Revenue  (₹ Cr)",             revs,    "#,##0"),
        ("Growth %",                    growths, "0.0%"),
        (f"Implied EV  ({d['multiple']:.1f}x)  (₹ Cr)", evs, "#,##0"),
        ("Implied Price  (₹)",          ips,     "#,##0.00"),
    ]:
        _lbl(ws, row, "    " + lbl_txt)
        for j, v in enumerate(vals):
            _val(ws, row, 3 + j, v, fmt)
        row += 1

    ws.freeze_panes = "C4"


# ── Sheet 4: Results & Sensitivity ───────────────────────────────────────────

def _build_sensitivity(wb, theme, d):
    ws = wb.create_sheet("Results & Sensitivity")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 28

    t = ws.cell(row=2, column=2,
                value="Results & Sensitivity  —  EV/Revenue Multiple  ×  NTM Revenue Growth")
    t.font = font(bold=True, size=13, color=theme["primary"])
    ws.merge_cells("B2:M2")

    # NTM-based sensitivity: rows = multiple, cols = NTM growth
    mult_vals   = [1.0, 1.5, 2.0, 2.5, 3.0, 4.0, 5.0, 6.0, 8.0, 10.0]
    growth_vals = [-0.10, -0.05, 0.0, 0.05, 0.10, 0.15, 0.20, 0.25, 0.30]

    base_row_idx = min(range(len(mult_vals)),  key=lambda i: abs(mult_vals[i]  - d["multiple"]))
    base_col_idx = min(range(len(growth_vals)),key=lambda i: abs(growth_vals[i]- d["rev_growth"]))

    rev     = d["rev_cr"]
    net_d   = d["net_debt"]
    sc      = d["shares_cr"]

    matrix = []
    for mult in mult_vals:
        r_row = []
        for g in growth_vals:
            ntm_r  = rev * (1 + g)
            ev_    = mult * ntm_r
            eq_    = ev_ - net_d
            ip_    = eq_ / sc if sc > 0 else 0.0
            r_row.append(round(ip_, 2))
        matrix.append(r_row)

    build_sensitivity_table(
        ws, theme,
        row_label="EV/Revenue Multiple",
        col_label="NTM Revenue Growth",
        row_vals=mult_vals,
        col_vals=growth_vals,
        matrix=matrix,
        current_price=d["price"],
        base_row_idx=base_row_idx,
        base_col_idx=base_col_idx,
        start_row=4,
        start_col=2,
        row_fmt="0.0x",
        col_fmt="0.0%",
    )
    ws.freeze_panes = "C4"


# ── Entry point ───────────────────────────────────────────────────────────────

def generate_excel(fin: dict, symbol: str, company_name: str) -> bytes:
    d  = _compute(fin)
    wb = Workbook()

    build_cover(
        wb, THEME, symbol, company_name,
        model_label="EV/Revenue Multiple",
        model_desc=(
            "Enterprise Value = Revenue × Multiple  →  Equity = EV − Net Debt  →  Price = Equity / Shares. "
            "EV/Revenue is an enterprise multiple (NOT P/S) — debt is always brought through the bridge. "
            "NTM (forward) revenue is the investment banking standard."
        ),
        sheets_index=[
            (1, "Cover",                    "Model overview & index"),
            (2, "Inputs & Assumptions",     "Financial data & model parameters"),
            (3, "Revenue Multiple Analysis","Profile, benchmarks, LTM vs NTM, forward build"),
            (4, "Results & Sensitivity",    "EV/Revenue multiple × NTM growth sensitivity"),
        ],
        meta_extra={
            "price":     d["price"],
            "mktcap_cr": d["mktcap_cr"],
            "sector":    fin.get("sector", "—"),
        },
    )

    params_def = [
        ("EV/Revenue Multiple  (sector default)", d["multiple"], "x",
         d["s_desc"],                                              "0.0x",    True),
        ("Sector Multiple Range",
         f"{d['s_low']:.1f}x – {d['s_high']:.1f}x", "",
         "Low / High range for this sector",                       "@",       False),
        ("NTM Revenue Growth",                    d["rev_growth"], "%",
         "Used for forward (NTM) revenue estimate",                "0.0%",    True),
        ("Revenue — LTM  (₹ Cr)",                 d["rev_cr"],    "₹ Cr",
         "Last twelve months — from filings",                      "#,##0",   False),
        ("Revenue — NTM  (₹ Cr)",                 d["ntm_rev"],   "₹ Cr",
         "= LTM × (1 + growth) — IB standard",                    "#,##0",   False),
        ("Net Debt  (₹ Cr)",                       d["net_debt"],  "₹ Cr",
         "Total Debt − Cash",                                       "#,##0",   False),
        ("Shares Outstanding  (Cr)",               d["shares_cr"], "Cr",
         "From balance sheet (filing)",                             "#,##0.00",False),
    ]

    build_inputs(wb, THEME, company_name, fin, params_def)
    _build_analysis(wb, THEME, fin, d)
    _build_sensitivity(wb, THEME, d)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
