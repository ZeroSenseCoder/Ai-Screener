"""
VC Method Valuation
Terminal value discounted at target IRR = pre-money valuation.
"""
import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.services.excel_utils import (
    font, fill, thin_border, align, safe, cr,
    section_header, build_cover, build_inputs, build_sensitivity_table,
)

THEME = {
    "cover_bg":      "110022",
    "primary":       "220044",
    "sub":           "380066",
    "accent":        "E91E8C",
    "input_color":   "880E4F",
    "positive_fill": "FCE4EC",
    "positive_text": "880E4F",
    "subtotal_fill": "F8BBD9",
    "key_fill":      "FFF2CC",
}


def _col(c: int) -> str:
    return get_column_letter(c)


def _lbl(ws, row, text, indent=0, bold=False):
    c = ws.cell(row=row, column=2, value=("    " * indent) + text)
    c.font = font(bold=bold, size=9)
    c.border = thin_border()
    c.alignment = align()
    return c


def _val(ws, row, col, val, fmt="#,##0.00", bold=False, bg=None, color="000000"):
    c = ws.cell(row=row, column=col, value=val)
    c.font = font(bold=bold, size=9, color=color)
    c.number_format = fmt
    c.alignment = align(h="right")
    c.border = thin_border()
    if bg:
        c.fill = fill(bg)
    return c


def _hdr(ws, row, cols, theme):
    for i, txt in enumerate(cols):
        c = ws.cell(row=row, column=2 + i, value=txt)
        c.font = font(bold=True, color="FFFFFF", size=9)
        c.fill = fill(theme["primary"])
        c.alignment = align(h="center")
        c.border = thin_border()


def _compute(fin: dict):
    price      = safe(fin.get("price"), 100.0)
    shares     = safe(fin.get("shares"), 1e8)
    beta       = safe(fin.get("beta"), 1.0)
    rev_cr     = cr(fin.get("revenue"))
    rev_growth = safe(fin.get("revenue_growth"), 0.20)
    debt_cr    = cr(fin.get("total_debt"))
    cash_cr    = cr(fin.get("cash"))

    shares_cr  = shares / 1e7
    mktcap_cr  = price * shares / 1e7

    # Business overview
    rev_exit    = rev_cr * (1 + rev_growth) ** 5
    exit_margin = 0.20
    ebitda_exit = rev_exit * exit_margin
    exit_mult   = 3.0
    exit_ev     = rev_exit * exit_mult

    # VC return parameters
    irr        = 0.25
    hold       = 5
    disc_f     = 1.0 / (1 + irr) ** hold
    pre_money  = exit_ev * disc_f
    investment = pre_money * 0.20
    post_money = pre_money + investment
    ownership  = investment / post_money if post_money else 0.0
    exit_equity = exit_ev * ownership
    moic       = exit_equity / investment if investment else 0.0

    implied_price = (pre_money - debt_cr + cash_cr) / shares_cr if shares_cr else 0.0
    upside = (implied_price - price) / price if price else 0.0

    # Revenue bridge
    bridge = []
    rev_t  = rev_cr
    for t in range(1, 6):
        rev_t   = rev_t * (1 + rev_growth)
        ebitda_t = rev_t * exit_margin
        fcf_t   = rev_t * 0.10
        bridge.append((t, rev_t, rev_growth, ebitda_t, fcf_t))
    cum_rev = sum(b[1] for b in bridge)

    # Return scenarios
    scenarios = []
    for label_s, mult_s, irr_s in [("Bear", 2.0, 0.20), ("Base", 3.0, 0.25), ("Bull", 5.0, 0.35)]:
        ev_s   = rev_exit * mult_s
        pre_s  = ev_s / (1 + irr_s) ** 5
        inv_s  = pre_s * 0.20
        post_s = pre_s + inv_s
        own_s  = inv_s / post_s if post_s else 0.0
        exit_s = ev_s * own_s
        moic_s = exit_s / inv_s if inv_s else 0.0
        ret_s  = (exit_s - inv_s) / inv_s if inv_s else 0.0
        scenarios.append((label_s, mult_s, irr_s, pre_s, moic_s, ret_s))

    # IRR compounding table: 100 invested × 6 IRR levels × 5 years
    irr_levels = [0.15, 0.20, 0.25, 0.30, 0.35, 0.40]
    year_levels = [3, 4, 5, 6, 7]
    irr_table = []
    for irr_ in irr_levels:
        row_d = []
        for yr_ in year_levels:
            row_d.append(100 * (1 + irr_) ** yr_)
        irr_table.append(row_d)

    # Dilution table
    dilution = [
        ("Round A", 1.0, 1.2, 1.2 / (pre_money if pre_money else 1)),
        ("Round B", 2.0, 2.5, 2.5 / (pre_money * 2 if pre_money else 1)),
        ("Round C", 5.0, 6.5, 6.5 / (pre_money * 4 if pre_money else 1)),
    ]

    return dict(
        price=price, shares_cr=shares_cr, mktcap_cr=mktcap_cr,
        beta=beta, rev_cr=rev_cr, rev_growth=rev_growth,
        debt_cr=debt_cr, cash_cr=cash_cr,
        rev_exit=rev_exit, ebitda_exit=ebitda_exit,
        exit_mult=exit_mult, exit_ev=exit_ev,
        irr=irr, hold=hold, disc_f=disc_f,
        pre_money=pre_money, investment=investment,
        post_money=post_money, ownership=ownership,
        exit_equity=exit_equity, moic=moic,
        implied_price=implied_price, upside=upside,
        bridge=bridge, cum_rev=cum_rev,
        scenarios=scenarios,
        irr_levels=irr_levels, year_levels=year_levels, irr_table=irr_table,
        dilution=dilution,
    )


def _build_analysis(wb, theme, fin, d):
    ws = wb.create_sheet("VC PE Method")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 16
    for i in range(4, 12):
        ws.column_dimensions[_col(i)].width = 14

    row = 2
    t = ws.cell(row=row, column=2,
                value="VC / PE Method  —  Exit EV × Disc Factor = Pre-Money Valuation")
    t.font = font(bold=True, size=13)
    ws.merge_cells(f"B{row}:{_col(11)}{row}")
    row += 2

    # ── BUSINESS OVERVIEW ──
    section_header(ws, row, 2, "▌ BUSINESS OVERVIEW", theme, span=3)
    row += 1
    biz_rows = [
        ("Current Revenue  (₹ Cr)",                          d["rev_cr"],      "#,##0"),
        ("Revenue CAGR  (5Y)",                                d["rev_growth"],  "0.0%"),
        ("Revenue at Exit  (Year 5)  (₹ Cr)",                d["rev_exit"],    "#,##0"),
        ("EBITDA Margin at Exit  (target 20%)",               0.20,             "0.0%"),
        ("EBITDA at Exit  (₹ Cr)",                            d["ebitda_exit"], "#,##0"),
        ("Exit Revenue Multiple",                             d["exit_mult"],   "0.0x"),
        ("★ Exit Enterprise Value  (₹ Cr)",                  d["exit_ev"],     "#,##0"),
    ]
    for lbl_txt, val, fmt_ in biz_rows:
        is_key = lbl_txt.startswith("★")
        _lbl(ws, row, lbl_txt, bold=is_key)
        _val(ws, row, 3, val, fmt_, bold=is_key,
             bg=theme["key_fill"] if is_key else None,
             color=theme["accent"] if is_key else "000000")
        row += 1
    row += 1

    # ── VC RETURN PARAMETERS ──
    section_header(ws, row, 2, "▌ VC RETURN PARAMETERS", theme, span=3)
    row += 1
    vc_rows = [
        ("Target IRR  (default 25%)",                         d["irr"],        "0.0%",       False),
        ("Hold Period  (years)",                              d["hold"],       "0",          False),
        ("Discount Factor  =  1/(1+IRR)^5",                  d["disc_f"],     "0.0000",     False),
        ("★ Pre-Money Valuation  =  ExitEV × DiscFactor",    d["pre_money"],  "#,##0",      True),
        ("Investment  =  PreMoney × 20%  (₹ Cr)",            d["investment"], "#,##0",      False),
        ("Post-Money  =  PreMoney + Investment  (₹ Cr)",     d["post_money"], "#,##0",      False),
        ("Ownership %  =  Investment / PostMoney",           d["ownership"],  "0.0%",       False),
        ("Exit Equity  =  ExitEV × Ownership  (₹ Cr)",      d["exit_equity"],"#,##0",      False),
        ("MOIC  =  ExitEquity / Investment",                 d["moic"],       "0.00x",      False),
        ("★ Implied Share Price  (₹)",                       d["implied_price"],"#,##0.00", True),
        ("Current Market Price  (₹)",                        d["price"],      "#,##0.00",   False),
        ("★ Upside / (Downside)",                            d["upside"],     "+0.0%;-0.0%",True),
    ]
    for lbl_txt, val, fmt_, is_key in vc_rows:
        _lbl(ws, row, lbl_txt, bold=is_key)
        _val(ws, row, 3, val, fmt_, bold=is_key,
             bg=theme["key_fill"] if is_key else None,
             color=theme["accent"] if is_key else "000000")
        row += 1
    row += 1

    # ── REVENUE BRIDGE ──
    YEAR_COLS = [f"Year {b[0]}" for b in d["bridge"]]
    section_header(ws, row, 2, "▌ REVENUE BRIDGE  (5 YEARS)", theme, span=7)
    row += 1
    _hdr(ws, row, ["Metric"] + YEAR_COLS, theme)
    row += 1
    bridge_rows = [
        ("Revenue  (₹ Cr)",          [b[1] for b in d["bridge"]], "#,##0"),
        ("YoY Growth",               [b[2] for b in d["bridge"]], "0.0%"),
        ("EBITDA  (₹ Cr)",           [b[3] for b in d["bridge"]], "#,##0"),
        ("FCF  (est 10% of rev)  (₹ Cr)", [b[4] for b in d["bridge"]], "#,##0"),
    ]
    for lbl_txt, vals, fmt_ in bridge_rows:
        _lbl(ws, row, lbl_txt, indent=1)
        for j, v in enumerate(vals):
            _val(ws, row, 3 + j, v, fmt_)
        row += 1
    _lbl(ws, row, "Cumulative Revenue  (₹ Cr)", indent=1)
    _val(ws, row, 3, d["cum_rev"], "#,##0")
    row += 2

    # ── RETURN SCENARIOS ──
    section_header(ws, row, 2, "▌ RETURN SCENARIOS", theme, span=5)
    row += 1
    _hdr(ws, row, ["Metric", "Bear  2x / 20%", "Base  3x / 25%", "Bull  5x / 35%"], theme)
    row += 1
    for lbl_txt, idx, fmt_ in [
        ("Exit Revenue Multiple",   1, "0.0x"),
        ("Target IRR",              2, "0.0%"),
        ("Pre-Money  (₹ Cr)",       3, "#,##0"),
        ("MOIC",                    4, "0.00x"),
        ("Equity Return",           5, "+0.0%;-0.0%"),
    ]:
        _lbl(ws, row, lbl_txt, indent=1)
        for si, s in enumerate(d["scenarios"]):
            _val(ws, row, 3 + si, s[idx], fmt_)
        row += 1
    row += 1

    # ── IRR COMPOUNDING TABLE ──
    section_header(ws, row, 2, "▌ IRR COMPOUNDING TABLE  (₹100 invested)", theme, span=7)
    row += 1
    _hdr(ws, row, ["IRR"] + [f"Year {y}" for y in d["year_levels"]], theme)
    row += 1
    for ii, irr_ in enumerate(d["irr_levels"]):
        _lbl(ws, row, f"{irr_:.0%}", indent=1)
        for ji, val_ in enumerate(d["irr_table"][ii]):
            _val(ws, row, 3 + ji, val_, "#,##0.00")
        row += 1
    row += 1

    # ── DILUTION TABLE ──
    section_header(ws, row, 2, "▌ DILUTION TABLE", theme, span=5)
    row += 1
    _hdr(ws, row, ["Round", "Pre-Money  (₹ Cr)", "Post-Money  (₹ Cr)", "Ownership %"], theme)
    row += 1
    for rnd, pre_s, post_s, own_s in d["dilution"]:
        _lbl(ws, row, rnd, indent=1)
        _val(ws, row, 3, pre_s,  "#,##0")
        _val(ws, row, 4, post_s, "#,##0")
        _val(ws, row, 5, own_s,  "0.0%")
        row += 1

    ws.freeze_panes = "C4"


def _build_results_sheet(wb, theme, company_name, d):
    ws = wb.create_sheet("Results & Sensitivity")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 16

    row = 2
    section_header(ws, row, 2, "▌ VALUATION RESULTS SUMMARY", theme, span=4)
    row += 1
    results = [
        ("Revenue at Exit  (₹ Cr)",      d["rev_exit"],     "#,##0",       False),
        ("Exit Multiple",                 d["exit_mult"],    "0.0x",        False),
        ("Exit Enterprise Value  (₹ Cr)", d["exit_ev"],      "#,##0",       False),
        ("Target IRR",                    d["irr"],          "0.0%",        False),
        ("★ Pre-Money Valuation  (₹ Cr)", d["pre_money"],   "#,##0",       True),
        ("MOIC",                          d["moic"],         "0.00x",       False),
        ("★ Implied Share Price  (₹)",    d["implied_price"],"#,##0.00",    True),
        ("Current Price  (₹)",            d["price"],        "#,##0.00",    False),
        ("★ Upside / (Downside)",         d["upside"],       "+0.0%;-0.0%", True),
    ]
    for lbl_txt, val, fmt_, is_key in results:
        lc = ws.cell(row=row, column=2, value=lbl_txt)
        lc.font = font(bold=is_key, size=9)
        lc.border = thin_border()
        if is_key:
            lc.fill = fill(theme["key_fill"])
        vc = ws.cell(row=row, column=3, value=val)
        vc.font = font(bold=is_key, size=9,
                       color=theme["accent"] if is_key else "000000")
        vc.number_format = fmt_
        vc.alignment = align(h="right")
        vc.border = thin_border()
        if is_key:
            vc.fill = fill(theme["key_fill"])
        row += 1

    # Sensitivity: Target IRR × Exit Revenue Multiple
    irr_vals  = [0.15, 0.18, 0.20, 0.25, 0.30, 0.35, 0.40]
    mult_vals = [1.5, 2.0, 2.5, 3.0, 4.0, 5.0, 6.0]

    rev_exit = d["rev_exit"]
    debt_cr  = d["debt_cr"]
    cash_cr  = d["cash_cr"]
    sc       = d["shares_cr"]

    base_ii = min(range(len(irr_vals)),  key=lambda i: abs(irr_vals[i]  - d["irr"]))
    base_mi = min(range(len(mult_vals)), key=lambda i: abs(mult_vals[i] - d["exit_mult"]))

    matrix = []
    for irr_ in irr_vals:
        row_data = []
        for mult_ in mult_vals:
            ev_      = rev_exit * mult_
            pre_     = ev_ / (1 + irr_) ** 5
            ip_      = (pre_ - debt_cr + cash_cr) / sc if sc else 0.0
            row_data.append(round(ip_, 2))
        matrix.append(row_data)

    build_sensitivity_table(
        ws, theme,
        row_label="Target IRR",
        col_label="Exit Revenue Multiple",
        row_vals=irr_vals,
        col_vals=mult_vals,
        matrix=matrix,
        current_price=d["price"],
        base_row_idx=base_ii,
        base_col_idx=base_mi,
        start_row=row + 2,
        start_col=2,
        row_fmt="0.0%",
        col_fmt="0.0x",
    )


def generate_excel(fin: dict, symbol: str, company_name: str) -> bytes:
    d  = _compute(fin)
    wb = Workbook()

    build_cover(
        wb, THEME, symbol, company_name,
        model_label="VC / PE Method Valuation",
        model_desc="Exit EV discounted at target IRR gives pre-money valuation. "
                   "Used by venture capital and private equity to price early-stage investments.",
        sheets_index=[
            (1, "Cover",                "Model overview & index"),
            (2, "Inputs & Assumptions", "Financial data & model parameters"),
            (3, "VC PE Method",       "Business overview, VC returns & scenarios"),
            (4, "Results & Sensitivity","Summary + IRR × exit multiple sensitivity"),
        ],
        meta_extra={
            "price":     d["price"],
            "mktcap_cr": d["mktcap_cr"],
            "sector":    safe(fin.get("sector"), "—"),
        },
    )

    params_def = [
        ("Target IRR",               d["irr"],       "%",  "Required return",           "0.0%", True),
        ("Hold Period  (years)",     d["hold"],      "yr", "Time to exit",              "0",    True),
        ("Exit Revenue Multiple",    d["exit_mult"], "x",  "EV/Revenue at exit",        "0.0x", True),
        ("EBITDA Margin at Exit",    0.20,           "%",  "Target 20%",                "0.0%", False),
        ("Revenue CAGR  (5Y)",       d["rev_growth"],"%",  "Growth assumption",         "0.0%", True),
        ("Investment %",             0.20,           "%",  "20% of pre-money",          "0.0%", False),
        ("Risk-Free Rate (rf)",      0.071,          "%",  "India 10Y GSec",            "0.0%", False),
        ("Equity Risk Premium",      0.055,          "%",  "Damodaran India ERP",        "0.0%", False),
        ("Beta",                     d["beta"],      "x",  "Market beta",               "0.00", False),
    ]

    build_inputs(wb, THEME, company_name, fin, params_def)
    _build_analysis(wb, THEME, fin, d)
    _build_results_sheet(wb, THEME, company_name, d)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
