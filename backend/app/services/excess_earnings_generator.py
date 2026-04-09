"""
Excess Earnings Valuation Model
Value = Net Tangible Assets + PV(earnings above fair return on assets).
"""
import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.services.excel_utils import (
    font, fill, thin_border, align, safe, cr,
    section_header, build_cover, build_inputs, build_sensitivity_table,
)

THEME = {
    "cover_bg":      "1A1500",
    "primary":       "3B3100",
    "sub":           "645300",
    "accent":        "F1C40F",
    "input_color":   "7D6608",
    "positive_fill": "FEF9E7",
    "positive_text": "7D6608",
    "subtotal_fill": "FCF3CF",
    "key_fill":      "FFF2CC",
}


def _col(c: int) -> str:
    return get_column_letter(c)


def _lbl(ws, row, text, indent=0, bold=False):
    c = ws.cell(row=row, column=2, value=("    " * indent) + text)
    c.font = font(bold=bold, size=9)
    c.border = thin_border()
    c.alignment = align()
    return c


def _val(ws, row, col, val, fmt="#,##0.00", bold=False, bg=None, color="000000"):
    c = ws.cell(row=row, column=col, value=val)
    c.font = font(bold=bold, size=9, color=color)
    c.number_format = fmt
    c.alignment = align(h="right")
    c.border = thin_border()
    if bg:
        c.fill = fill(bg)
    return c


def _hdr(ws, row, cols, theme):
    for i, txt in enumerate(cols):
        c = ws.cell(row=row, column=2 + i, value=txt)
        c.font = font(bold=True, color="FFFFFF", size=9)
        c.fill = fill(theme["primary"])
        c.alignment = align(h="center")
        c.border = thin_border()


def _compute(fin: dict):
    price      = safe(fin.get("price"), 100.0)
    shares     = safe(fin.get("shares"), 1e8)
    beta       = safe(fin.get("beta"), 1.0)
    assets     = cr(fin.get("total_assets"))
    liab       = cr(fin.get("total_liabilities"))
    ni_cr      = cr(fin.get("net_income"))
    debt_cr    = cr(fin.get("total_debt"))
    cash_cr    = cr(fin.get("cash"))

    shares_cr  = shares / 1e7
    nta        = assets - liab           # net tangible assets
    actual_roa = ni_cr / assets if assets else 0.0
    fair_rate  = 0.08
    fair_earn  = assets * fair_rate
    excess_earn = ni_cr - fair_earn

    ke         = 0.071 + beta * 0.055
    g_excess   = 0.035
    cap_rate   = ke - g_excess if ke > g_excess else 0.01
    goodwill   = excess_earn / cap_rate if cap_rate else 0.0

    # 5-year schedule
    schedule = []
    excess_t = excess_earn
    cum_pv   = 0.0
    for t in range(1, 6):
        excess_t = excess_t * 1.05
        df       = 1 / (1 + ke) ** t
        pv_ex    = excess_t * df
        cum_pv  += pv_ex
        schedule.append((t, excess_t, df, pv_ex, cum_pv))

    # Terminal value at year 5
    excess_5    = schedule[-1][1]
    df_5        = schedule[-1][2]
    tv          = excess_5 * (1 + g_excess) / cap_rate * df_5 if cap_rate else 0.0
    sum_pv_5    = schedule[-1][4]

    ev          = nta + sum_pv_5 + tv
    net_debt    = debt_cr - cash_cr
    equity_val  = ev - net_debt
    implied_price = equity_val / shares_cr if shares_cr else 0.0
    upside      = (implied_price - price) / price if price else 0.0
    mktcap_cr   = price * shares / 1e7

    return dict(
        price=price, shares_cr=shares_cr, assets=assets, liab=liab,
        ni_cr=ni_cr, debt_cr=debt_cr, cash_cr=cash_cr, net_debt=net_debt,
        nta=nta, actual_roa=actual_roa, fair_rate=fair_rate,
        fair_earn=fair_earn, excess_earn=excess_earn, ke=ke,
        g_excess=g_excess, cap_rate=cap_rate, goodwill=goodwill,
        schedule=schedule, sum_pv_5=sum_pv_5, tv=tv,
        ev=ev, equity_val=equity_val, implied_price=implied_price,
        upside=upside, mktcap_cr=mktcap_cr, beta=beta,
    )


def _build_analysis(wb, theme, fin, d):
    ws = wb.create_sheet("Excess Earnings Model")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 16
    for i in range(4, 10):
        ws.column_dimensions[_col(i)].width = 14

    row = 2
    t = ws.cell(row=row, column=2,
                value="Excess Earnings Valuation  —  NTA + PV(Excess Earnings)")
    t.font = font(bold=True, size=13)
    ws.merge_cells(f"B{row}:{_col(9)}{row}")
    row += 2

    # ── ASSET BASE & EARNINGS ──
    section_header(ws, row, 2, "▌ ASSET BASE & EARNINGS", theme, span=3)
    row += 1
    base_rows = [
        ("Total Assets (₹ Cr)",                   d["assets"],       "#,##0"),
        ("Total Liabilities (₹ Cr)",               d["liab"],         "#,##0"),
        ("Net Tangible Assets (Assets − Liab)",    d["nta"],          "#,##0"),
        ("Net Income (₹ Cr)",                      d["ni_cr"],        "#,##0"),
        ("Actual ROA = NI / Assets",               d["actual_roa"],   "0.0%"),
        ("Fair Return Rate",                       d["fair_rate"],    "0.0%"),
        ("Fair Earnings = Assets × Fair Rate",     d["fair_earn"],    "#,##0"),
        ("★ Excess Earnings = NI − Fair Earnings", d["excess_earn"], "#,##0"),
    ]
    for lbl_txt, val, fmt_ in base_rows:
        is_key = lbl_txt.startswith("★")
        _lbl(ws, row, lbl_txt, bold=is_key)
        _val(ws, row, 3, val, fmt_, bold=is_key,
             bg=theme["key_fill"] if is_key else None,
             color=theme["accent"] if is_key else "000000")
        row += 1
    row += 1

    # ── EXCESS EARNINGS CAPITALISATION ──
    section_header(ws, row, 2, "▌ EXCESS EARNINGS CAPITALISATION", theme, span=3)
    row += 1
    cap_rows = [
        ("Excess Earnings (₹ Cr)",                                   d["excess_earn"], "#,##0",    False),
        ("Discount Rate ke (CAPM)",                                   d["ke"],          "0.0%",     False),
        ("Terminal Growth for Excess Earnings",                       d["g_excess"],    "0.0%",     False),
        ("Cap Rate = ke − g_excess",                                  d["cap_rate"],    "0.0%",     False),
        ("★ Value of Excess Earnings (Goodwill) = Excess / Cap Rate", d["goodwill"],   "#,##0",    True),
    ]
    for lbl_txt, val, fmt_, is_key in cap_rows:
        _lbl(ws, row, lbl_txt, bold=is_key)
        _val(ws, row, 3, val, fmt_, bold=is_key,
             bg=theme["key_fill"] if is_key else None,
             color=theme["accent"] if is_key else "000000")
        row += 1
    row += 1

    # ── 5-YEAR SCHEDULE ──
    YEAR_COLS = [f"Year {y}" for y in range(1, 6)]
    section_header(ws, row, 2, "▌ 5-YEAR EXCESS EARNINGS SCHEDULE", theme, span=7)
    row += 1
    _hdr(ws, row, ["Metric"] + YEAR_COLS, theme)
    row += 1

    excesses  = [s[1] for s in d["schedule"]]
    dfs       = [s[2] for s in d["schedule"]]
    pvs       = [s[3] for s in d["schedule"]]
    cum_pvs   = [s[4] for s in d["schedule"]]

    for lbl_txt, vals, fmt_ in [
        ("Excess Earnings (₹ Cr)",       excesses, "#,##0"),
        ("Discount Factor",              dfs,      "0.0000"),
        ("PV of Excess Earnings (₹ Cr)", pvs,      "#,##0"),
        ("Cumulative PV (₹ Cr)",         cum_pvs,  "#,##0"),
    ]:
        _lbl(ws, row, lbl_txt, indent=1)
        for j, v in enumerate(vals):
            _val(ws, row, 3 + j, v, fmt_)
        row += 1
    row += 1

    # ── VALUE BRIDGE ──
    section_header(ws, row, 2, "▌ VALUE BRIDGE", theme, span=3)
    row += 1
    bridge = [
        ("Net Tangible Assets (₹ Cr)",                      d["nta"],          "#,##0",       False),
        ("Add: Sum PV Excess Earnings — 5Y (₹ Cr)",         d["sum_pv_5"],     "#,##0",       False),
        ("Add: Terminal Value (₹ Cr)",                      d["tv"],           "#,##0",       False),
        ("Enterprise Value (₹ Cr)",                         d["ev"],           "#,##0",       False),
        ("Less: Net Debt (₹ Cr)",                           d["net_debt"],     "#,##0",       False),
        ("Equity Value (₹ Cr)",                             d["equity_val"],   "#,##0",       False),
        ("Shares Outstanding (Cr)",                         d["shares_cr"],    "#,##0.00",    False),
        ("★ Implied Share Price (₹)",                       d["implied_price"],"#,##0.00",    True),
        ("Current Price (₹)",                               d["price"],        "#,##0.00",    False),
        ("★ Upside / (Downside)",                           d["upside"],       "+0.0%;-0.0%", True),
    ]
    for lbl_txt, val, fmt_, is_key in bridge:
        _lbl(ws, row, lbl_txt, bold=is_key)
        _val(ws, row, 3, val, fmt_, bold=is_key,
             bg=theme["key_fill"] if is_key else None,
             color=theme["accent"] if is_key else "000000")
        row += 1

    ws.freeze_panes = "C4"


def _build_results_sheet(wb, theme, company_name, d):
    ws = wb.create_sheet("Results & Sensitivity")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 16

    row = 2
    section_header(ws, row, 2, "▌ VALUATION RESULTS SUMMARY", theme, span=4)
    row += 1
    results = [
        ("Net Tangible Assets (₹ Cr)",  d["nta"],          "#,##0",       False),
        ("Excess Earnings (₹ Cr)",      d["excess_earn"],  "#,##0",       False),
        ("PV Excess Earnings 5Y",       d["sum_pv_5"],     "#,##0",       False),
        ("Terminal Value (₹ Cr)",       d["tv"],           "#,##0",       False),
        ("Enterprise Value (₹ Cr)",     d["ev"],           "#,##0",       False),
        ("★ Implied Share Price (₹)",   d["implied_price"],"#,##0.00",    True),
        ("Current Price (₹)",           d["price"],        "#,##0.00",    False),
        ("★ Upside / (Downside)",       d["upside"],       "+0.0%;-0.0%", True),
    ]
    for lbl_txt, val, fmt_, is_key in results:
        lc = ws.cell(row=row, column=2, value=lbl_txt)
        lc.font = font(bold=is_key, size=9)
        lc.border = thin_border()
        if is_key:
            lc.fill = fill(theme["key_fill"])
        vc = ws.cell(row=row, column=3, value=val)
        vc.font = font(bold=is_key, size=9,
                       color=theme["accent"] if is_key else "000000")
        vc.number_format = fmt_
        vc.alignment = align(h="right")
        vc.border = thin_border()
        if is_key:
            vc.fill = fill(theme["key_fill"])
        row += 1

    # Sensitivity: Fair Return Rate × Discount Rate ke
    fair_vals = [0.05, 0.06, 0.07, 0.08, 0.09, 0.10, 0.11, 0.12]
    ke_vals   = [0.10, 0.11, 0.12, 0.13, 0.14, 0.15, 0.16, 0.17, 0.18]

    assets  = d["assets"]
    liab    = d["liab"]
    ni_cr   = d["ni_cr"]
    debt_cr = d["debt_cr"]
    cash_cr = d["cash_cr"]
    sc      = d["shares_cr"]
    g_ex    = d["g_excess"]

    base_fi = min(range(len(fair_vals)), key=lambda i: abs(fair_vals[i] - d["fair_rate"]))
    base_ki = min(range(len(ke_vals)),   key=lambda i: abs(ke_vals[i]   - d["ke"]))

    matrix = []
    for fr in fair_vals:
        row_data = []
        for ke_ in ke_vals:
            excess_ = ni_cr - assets * fr
            cap_    = ke_ - g_ex if ke_ > g_ex else 0.01
            gw      = excess_ / cap_
            ev_     = assets - liab + gw
            net_d   = debt_cr - cash_cr
            eq_     = ev_ - net_d
            ip_     = eq_ / sc if sc else 0.0
            row_data.append(round(ip_, 2))
        matrix.append(row_data)

    build_sensitivity_table(
        ws, theme,
        row_label="Fair Return Rate",
        col_label="Discount Rate ke",
        row_vals=fair_vals,
        col_vals=ke_vals,
        matrix=matrix,
        current_price=d["price"],
        base_row_idx=base_fi,
        base_col_idx=base_ki,
        start_row=row + 2,
        start_col=2,
        row_fmt="0.0%",
        col_fmt="0.0%",
    )


def generate_excel(fin: dict, symbol: str, company_name: str) -> bytes:
    d  = _compute(fin)
    wb = Workbook()

    build_cover(
        wb, THEME, symbol, company_name,
        model_label="Excess Earnings Valuation",
        model_desc="Value = Net Tangible Assets + PV(Excess Earnings above fair return). "
                   "Decomposes value into tangible assets and intangible/franchise goodwill.",
        sheets_index=[
            (1, "Cover",                 "Model overview & index"),
            (2, "Inputs & Assumptions",  "Financial data & model parameters"),
            (3, "Excess Earnings Model", "Asset base, capitalisation & 5-year schedule"),
            (4, "Results & Sensitivity", "Summary + fair rate × ke sensitivity"),
        ],
        meta_extra={
            "price":     d["price"],
            "mktcap_cr": d["mktcap_cr"],
            "sector":    safe(fin.get("sector"), "—"),
        },
    )

    params_def = [
        ("Fair Return Rate on Assets",  d["fair_rate"],  "%",  "Opportunity cost benchmark",    "0.0%", True),
        ("Discount Rate ke (CAPM)",     d["ke"],         "%",  "rf + β × ERP",                  "0.0%", True),
        ("Terminal Growth — Excess",    d["g_excess"],   "%",  "Long-run excess earnings growth","0.0%", False),
        ("Cap Rate = ke − g_excess",    d["cap_rate"],   "%",  "Capitalisation rate",            "0.0%", False),
        ("Risk-Free Rate (rf)",         0.071,           "%",  "India 10Y GSec",                 "0.0%", False),
        ("Equity Risk Premium",         0.055,           "%",  "Damodaran India ERP",             "0.0%", False),
        ("Beta",                        d["beta"],       "x",  "Market beta",                    "0.00", False),
    ]

    build_inputs(wb, THEME, company_name, fin, params_def)
    _build_analysis(wb, THEME, fin, d)
    _build_results_sheet(wb, THEME, company_name, d)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
