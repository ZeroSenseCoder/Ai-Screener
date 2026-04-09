"""
Shared Excel styling utilities for all valuation model generators.
Each generator imports from here to maintain consistency.
"""
import io
import math
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def font(bold=False, color="000000", size=10, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)


def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def thin_border(color="BFBFBF"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def bottom_border(color="BFBFBF"):
    s = Side(style="thin", color=color)
    return Border(bottom=s)


def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def safe(v, default=0.0):
    if v is None or (isinstance(v, float) and (math.isnan(v) or math.isinf(v))):
        return default
    return v


def cr(v):
    """Convert INR raw value to Crores."""
    return safe(v, 0.0) / 1e7


def section_header(ws, row, col, text, theme, span=8):
    """Dark filled section header spanning `span` columns."""
    c = ws.cell(row=row, column=col, value=text)
    c.font = font(bold=True, color="FFFFFF", size=10)
    c.fill = fill(theme["primary"])
    c.alignment = align()
    ws.merge_cells(
        start_row=row, start_column=col,
        end_row=row, end_column=col + span - 1
    )


def sub_header(ws, row, col, text, theme, span=8):
    c = ws.cell(row=row, column=col, value=text)
    c.font = font(bold=True, color="FFFFFF", size=9)
    c.fill = fill(theme["sub"])
    c.alignment = align()
    ws.merge_cells(
        start_row=row, start_column=col,
        end_row=row, end_column=col + span - 1
    )


def label(ws, row, col, text, indent=0, bold=False, theme=None):
    prefix = "    " * indent
    c = ws.cell(row=row, column=col, value=prefix + text)
    c.font = font(bold=bold, size=9)
    c.border = thin_border()
    c.alignment = align()
    return c


def value_cell(ws, row, col, val, fmt="#,##0", color="000000",
               bg=None, bold=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font = font(bold=bold, size=9, color=color)
    c.number_format = fmt
    c.alignment = align(h="right")
    c.border = thin_border()
    if bg:
        c.fill = fill(bg)
    return c


def pct_cell(ws, row, col, val, color="000000", bg=None, bold=False):
    return value_cell(ws, row, col, val, fmt="0.0%",
                      color=color, bg=bg, bold=bold)


def build_cover(wb: Workbook, theme: dict, symbol: str, company_name: str,
                model_label: str, model_desc: str, sheets_index: list,
                meta_extra: dict = None):
    """
    Builds the Cover sheet on wb.active.
    theme keys: cover_bg, primary, accent, text_light
    sheets_index: list of (num, sheet_name, description)
    meta_extra: additional metadata rows {label: value}
    """
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 3
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 22

    COVER  = theme["cover_bg"]
    ACCENT = theme["accent"]

    # Paint background
    for r in range(1, 42):
        for c in range(1, 12):
            ws.cell(row=r, column=c).fill = fill(COVER)

    # Company + Model title
    t = ws.cell(row=5, column=2,
        value=f"{company_name}  ({symbol})")
    t.font = Font(name="Calibri", bold=True, size=22, color="FFFFFF")
    t.alignment = align()
    ws.merge_cells("B5:J5")
    ws.row_dimensions[5].height = 30

    sub = ws.cell(row=8, column=2,
        value=f"{model_label}  |  Equity Research  |  "
              f"{datetime.date.today().strftime('%B %Y')}")
    sub.font = Font(name="Calibri", size=13, color=ACCENT, italic=True)
    sub.alignment = align()
    ws.merge_cells("B8:J8")

    model_sub = ws.cell(row=9, column=2, value=model_desc)
    model_sub.font = Font(name="Calibri", size=10, color="A0A0A0", italic=True)
    model_sub.alignment = align()
    ws.merge_cells("B9:J9")

    # Accent separator
    for col in range(2, 11):
        ws.cell(row=11, column=col).fill = fill(ACCENT)

    # Metadata
    base_meta = [
        ("Analyst",            "Equity Research  |  Indian Markets"),
        ("Model",              model_label),
        ("Base Currency",      "INR (₹ Crores)"),
        ("Share Price (ref)",  f"₹{(meta_extra or {}).get('price', 0):.2f}"),
        ("Market Cap (ref)",   f"₹{(meta_extra or {}).get('mktcap_cr', 0):,.0f} Cr"),
        ("Sector",             (meta_extra or {}).get('sector', '—')),
        ("Model Date",         datetime.date.today().strftime("%B %d, %Y").replace(" 0", " ")),
    ]
    for i, (lbl, val) in enumerate(base_meta):
        r = 13 + i
        lc = ws.cell(row=r, column=2, value=lbl)
        lc.font = Font(name="Calibri", size=10, color="A0A0A0")
        vc = ws.cell(row=r, column=4, value=val)
        vc.font = Font(name="Calibri", size=10, color="FFFFFF", bold=True)

    for col in range(2, 11):
        ws.cell(row=22, column=col).fill = fill(ACCENT)

    # Index
    idx = ws.cell(row=24, column=2, value="MODEL INDEX")
    idx.font = Font(name="Calibri", bold=True, size=11, color=ACCENT)

    for i, (num, sname, desc) in enumerate(sheets_index):
        r = 26 + i
        ws.cell(row=r, column=2, value=num).font = \
            Font(name="Calibri", size=10, color=ACCENT, bold=True)
        ws.cell(row=r, column=3, value=sname).font = \
            Font(name="Calibri", size=10, color="FFFFFF", bold=True)
        ws.cell(row=r, column=5, value=desc).font = \
            Font(name="Calibri", size=10, color="808080")

    # Disclaimer
    disc = ws.cell(row=38, column=2,
        value="DISCLAIMER: For educational purposes only. Not investment advice. "
              "Data from NSE/BSE filings, screener.in, and yfinance. "
              "Analyst estimates — not guaranteed. Past performance ≠ future results.")
    disc.font = Font(name="Calibri", size=8, color="606060", italic=True)
    disc.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("B38:J40")


def build_inputs(wb: Workbook, theme: dict, company_name: str,
                 fin: dict, params_def: list):
    """
    Builds the 'Inputs & Assumptions' sheet.
    params_def: list of (label, key_in_params_or_fin, value, unit, description)
    """
    ws = wb.create_sheet("Inputs & Assumptions")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 46

    # Title
    t = ws.cell(row=2, column=2, value=f"{company_name} — Model Inputs & Assumptions")
    t.font = font(bold=True, size=13)
    ws.merge_cells("B2:E2")

    note = ws.cell(row=3, column=2,
        value="Blue = hardcoded input  |  Yellow fill = key driver  |  Source / notes in column E")
    note.font = font(size=8, color="808080", italic=True)
    ws.merge_cells("B3:E3")

    # Header
    for col, txt in enumerate(["Parameter", "Value", "Unit", "Source / Note"], 2):
        c = ws.cell(row=5, column=col, value=txt)
        c.font = font(bold=True, color="FFFFFF", size=9)
        c.fill = fill(theme["primary"])
        c.alignment = align()
        c.border = thin_border()

    row = 7
    section_header(ws, row, 2, "▌ FINANCIAL DATA (FROM FILINGS / YFINANCE)", theme, span=4)
    row += 1

    FIN_ROWS = [
        ("Share Price (₹)",          safe(fin.get("price"), 0),        "₹",    "Market reference price",          "#,##0.00"),
        ("Revenue (₹ Cr)",           cr(fin.get("revenue")),            "₹ Cr", "Latest annual / TTM",             "#,##0"),
        ("EBITDA (₹ Cr)",            cr(fin.get("ebitda")),             "₹ Cr", "Latest annual / TTM",             "#,##0"),
        ("Net Income (₹ Cr)",        cr(fin.get("net_income")),         "₹ Cr", "Latest annual / TTM",             "#,##0"),
        ("Free Cash Flow (₹ Cr)",    cr(fin.get("fcf") or fin.get("operating_cf")), "₹ Cr", "FCF or Op CF proxy", "#,##0"),
        ("Total Debt (₹ Cr)",        cr(fin.get("total_debt")),         "₹ Cr", "From balance sheet",              "#,##0"),
        ("Cash & Equiv (₹ Cr)",      cr(fin.get("cash")),               "₹ Cr", "From balance sheet",              "#,##0"),
        ("Total Assets (₹ Cr)",      cr(fin.get("total_assets")),       "₹ Cr", "From balance sheet",              "#,##0"),
        ("Total Liabilities (₹ Cr)", cr(fin.get("total_liabilities")),  "₹ Cr", "From balance sheet",              "#,##0"),
        ("Book Value/Share (₹)",     safe(fin.get("book_value_per_share")), "₹", "From balance sheet",             "#,##0.00"),
        ("EPS (₹)",                  safe(fin.get("eps")),               "₹",   "Trailing twelve months",          "#,##0.00"),
        ("DPS (₹)",                  safe(fin.get("dps")),               "₹",   "Last declared dividend",          "#,##0.00"),
        ("ROE",                      safe(fin.get("roe")),               "%",   "Return on equity (annual)",       "0.0%"),
        ("Beta (vs Nifty 50)",       safe(fin.get("beta"), 1.0),         "x",   "5Y monthly regression",           "0.00"),
        ("Revenue Growth (YoY)",     safe(fin.get("revenue_growth")),    "%",   "Most recent YoY",                 "0.0%"),
        ("Annualised Volatility",    safe(fin.get("volatility_annual"), 0.30), "%", "1Y daily returns",            "0.0%"),
        ("Shares Outstanding (Cr)",  safe(fin.get("shares"), 1) / 1e7,  "Cr",  "From NSE/BSE filing",             "#,##0.00"),
        ("Invested Capital (₹ Cr)",  cr(fin.get("invested_capital")),   "₹ Cr","Book equity + net debt",           "#,##0"),
        ("NOPAT (₹ Cr)",             cr(fin.get("nopat")),              "₹ Cr", "Net operating profit after tax",  "#,##0"),
        ("Profit Margin",            safe(fin.get("profit_margins")),   "%",    "Net income / revenue",            "0.0%"),
    ]

    for lbl, val, unit, src, fmt in FIN_ROWS:
        lc = ws.cell(row=row, column=2, value=lbl)
        lc.font = font(size=9)
        lc.border = thin_border()

        vc = ws.cell(row=row, column=3, value=val)
        vc.font = font(size=9, color=theme["input_color"])
        vc.number_format = fmt
        vc.alignment = align(h="right")
        vc.border = thin_border()

        ws.cell(row=row, column=4, value=unit).font = font(size=9)
        ws.cell(row=row, column=4).border = thin_border()

        sc = ws.cell(row=row, column=5, value=src)
        sc.font = font(size=8, color="606060", italic=True)
        sc.alignment = align(wrap=True)
        sc.border = thin_border()

        row += 1

    row += 1
    section_header(ws, row, 2, "▌ MODEL PARAMETERS", theme, span=4)
    row += 1

    for lbl, val, unit, src, fmt, is_key in params_def:
        lc = ws.cell(row=row, column=2, value=lbl)
        lc.font = font(size=9)
        lc.border = thin_border()

        vc = ws.cell(row=row, column=3, value=val)
        vc.font = font(size=9, color=theme["input_color"])
        vc.number_format = fmt
        vc.alignment = align(h="right")
        vc.border = thin_border()
        if is_key:
            vc.fill = fill("FFF2CC")

        ws.cell(row=row, column=4, value=unit).font = font(size=9)
        ws.cell(row=row, column=4).border = thin_border()

        sc = ws.cell(row=row, column=5, value=src)
        sc.font = font(size=8, color="606060", italic=True)
        sc.alignment = align(wrap=True)
        sc.border = thin_border()
        row += 1

    ws.freeze_panes = "C6"


def build_results(ws, theme: dict, company_name: str, model_label: str,
                  results_rows: list, start_row: int = 2) -> int:
    """
    Writes a Results Summary block.
    results_rows: list of (label, value, fmt, is_key)
    Returns next available row.
    """
    r = start_row
    section_header(ws, r, 2, "▌ VALUATION RESULTS SUMMARY", theme, span=4)
    r += 1

    for lbl, val, fmt, is_key in results_rows:
        lc = ws.cell(row=r, column=2, value=lbl)
        lc.font = font(bold=is_key, size=9)
        lc.border = thin_border()
        if is_key:
            lc.fill = fill("FFF2CC")

        vc = ws.cell(row=r, column=3, value=val)
        vc.font = font(bold=is_key, size=9,
                       color=theme["accent"] if is_key else "000000")
        vc.number_format = fmt
        vc.alignment = align(h="right")
        vc.border = thin_border()
        if is_key:
            vc.fill = fill("FFF2CC")
        r += 1

    return r + 1


def build_sensitivity_table(ws, theme: dict,
                             row_label: str, col_label: str,
                             row_vals: list, col_vals: list,
                             matrix: list,
                             current_price: float,
                             base_row_idx: int, base_col_idx: int,
                             start_row: int, start_col: int = 2,
                             row_fmt: str = "0.0%",
                             col_fmt: str = "0.0%"):
    """Generic sensitivity table writer with heat-map colouring."""
    # Title
    t = ws.cell(row=start_row, column=start_col,
        value=f"Sensitivity: {row_label} (rows) × {col_label} (columns)  — Implied Share Price (₹)")
    t.font = font(bold=True, size=10, color=theme["input_color"])
    ws.merge_cells(
        start_row=start_row, start_column=start_col,
        end_row=start_row, end_column=start_col + len(col_vals)
    )
    start_row += 1

    # Corner
    ws.cell(row=start_row, column=start_col,
        value=f"{row_label[:12]} \\ {col_label[:12]}").font = font(bold=True, size=9)

    # Column headers
    for j, cv in enumerate(col_vals):
        c = ws.cell(row=start_row, column=start_col + 1 + j, value=cv)
        c.font = font(bold=True, color="FFFFFF", size=9)
        c.fill = fill(theme["primary"])
        c.number_format = col_fmt
        c.alignment = align(h="center")
        c.border = thin_border()
        ws.column_dimensions[get_column_letter(start_col + 1 + j)].width = 13

    # Data rows
    for i, rv in enumerate(row_vals):
        r = start_row + 1 + i
        rh = ws.cell(row=r, column=start_col, value=rv)
        rh.font = font(bold=True, color="FFFFFF", size=9)
        rh.fill = fill(theme["primary"])
        rh.number_format = row_fmt
        rh.alignment = align(h="center")
        rh.border = thin_border()

        for j, cv in enumerate(col_vals):
            price = matrix[i][j]
            c = ws.cell(row=r, column=start_col + 1 + j, value=price)
            c.number_format = "#,##0.00"
            c.alignment = align(h="right")
            c.border = thin_border()

            is_base = (i == base_row_idx and j == base_col_idx)
            if is_base:
                c.fill = fill("FFD700")
                c.font = font(bold=True, size=9)
            elif price is not None and current_price:
                if price >= current_price * 1.20:
                    c.fill = fill("C6EFCE"); c.font = font(size=9, color="375623")
                elif price <= current_price * 0.80:
                    c.fill = fill("FFC7CE"); c.font = font(size=9, color="9C0006")
                else:
                    c.font = font(size=9)

    note_row = start_row + 1 + len(row_vals) + 1
    n = ws.cell(row=note_row, column=start_col,
        value=f"★ Gold = base case. Green ≥ +20% vs ₹{current_price:.2f}. Red ≤ −20%. All values in ₹.")
    n.font = font(size=8, italic=True, color="606060")
    ws.merge_cells(
        start_row=note_row, start_column=start_col,
        end_row=note_row, end_column=start_col + len(col_vals)
    )
    return note_row + 2
