"""
PEG Ratio Valuation Excel Generator
Fair Value = EPS × (EPS Growth % × Target PEG). Growth-adjusted P/E.
Sheets: Cover | Inputs & Assumptions | PEG Ratio Analysis | Results & Sensitivity
"""

import io
from openpyxl import Workbook

from app.services.excel_utils import (
    font, fill, thin_border, align, safe, cr,
    section_header, sub_header, label, value_cell, pct_cell,
    build_cover, build_inputs, build_results, build_sensitivity_table,
)

THEME = {
    "cover_bg":      "001220",
    "primary":       "002D4A",
    "sub":           "004A7A",
    "accent":        "5DADE2",
    "input_color":   "1A5276",
    "positive_fill": "D6EAF8",
    "positive_text": "1A5276",
    "subtotal_fill": "AED6F1",
    "key_fill":      "FFF2CC",
}

# Sector-adjusted PEG benchmarks
SECTOR_PEG_BENCHMARKS = {
    "Technology":      {"fair_peg": 1.0, "sector_avg": 1.1, "comment": "High-growth IT; PEG 1.0–1.3 typical"},
    "FMCG":            {"fair_peg": 1.5, "sector_avg": 1.6, "comment": "Stable growers; premium PEG acceptable"},
    "Banking":         {"fair_peg": 0.8, "sector_avg": 0.9, "comment": "Cyclical; lower PEG tolerance"},
    "Pharmaceuticals": {"fair_peg": 1.2, "sector_avg": 1.3, "comment": "Patent-driven growth; mid PEG range"},
    "Energy":          {"fair_peg": 0.7, "sector_avg": 0.8, "comment": "Commodity cyclical; conservative PEG"},
    "Automobile":      {"fair_peg": 0.9, "sector_avg": 1.0, "comment": "Cyclical + EV transition premium"},
    "Real Estate":     {"fair_peg": 1.0, "sector_avg": 1.1, "comment": "Cycle-adjusted; RERA visibility"},
    "Telecom":         {"fair_peg": 1.1, "sector_avg": 1.2, "comment": "Data growth premium"},
}

DEFAULT_SECTOR_BENCH = {"fair_peg": 1.0, "sector_avg": 1.2, "comment": "Indian market average (Nifty 50 reference)"}


def _build_analysis(wb, fin, theme, target_peg, eps_growth):
    ws = wb.create_sheet("PEG Ratio Analysis")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 18

    price    = safe(fin.get("price"), 0.0)
    eps      = safe(fin.get("eps"), 0.0)
    roe      = safe(fin.get("roe"), 0.0)
    dps      = safe(fin.get("dps"), 0.0)
    g        = eps_growth
    sector   = fin.get("sector", "")

    trailing_pe  = (price / eps) if eps > 0 else 0.0
    forward_eps  = eps * (1 + g)
    current_peg  = (trailing_pe / (g * 100)) if g > 0 else 0.0
    fair_pe      = target_peg * (g * 100)
    fair_value   = fair_pe * eps
    upside       = (fair_value / price - 1) if price > 0 else 0.0

    # Sustainable growth = ROE × retention ratio
    payout_ratio      = (dps / eps) if eps > 0 else 0.0
    retention_ratio   = 1 - payout_ratio
    sustainable_growth = roe * retention_ratio

    bench = SECTOR_PEG_BENCHMARKS.get(sector, DEFAULT_SECTOR_BENCH)
    is_expensive = current_peg > bench["fair_peg"]

    row = 2

    # ── Section 1: Earnings & Growth Data ────────────────────────────────────
    section_header(ws, row, 2, "▌ EARNINGS & GROWTH DATA", theme, span=3)
    row += 1

    eg_rows = [
        ("EPS — Trailing 12 Months (₹)",     eps,              "#,##0.00"),
        ("Trailing P/E (×)",                 trailing_pe,      "0.0x"),
        ("Forward EPS Estimate (₹)",         forward_eps,      "#,##0.00"),
        ("EPS Growth Rate (YoY %)",          g,                "0.0%"),
        ("Long-Run EPS CAGR (5Y estimate)",  g,                "0.0%"),
        ("ROE",                              roe,              "0.0%"),
        ("Dividend Payout Ratio",            payout_ratio,     "0.0%"),
        ("Retention Ratio",                  retention_ratio,  "0.0%"),
        ("Sustainable Growth Rate (ROE × b)", sustainable_growth, "0.0%"),
    ]
    for lbl_txt, val, fmt in eg_rows:
        label(ws, row, 2, lbl_txt)
        value_cell(ws, row, 3, val, fmt=fmt)
        row += 1

    row += 1

    # ── Section 2: PEG Ratio Calculation ─────────────────────────────────────
    section_header(ws, row, 2, "▌ PEG RATIO CALCULATION", theme, span=3)
    row += 1

    peg_rows = [
        ("Current P/E (×)",               trailing_pe,  "0.0x",  False),
        ("EPS Growth Rate (%)",            g,            "0.0%",  False),
        ("Current PEG Ratio",             current_peg,  "0.00x", False),
        ("Target PEG (input)",            target_peg,   "0.00x", True),
        ("Fair P/E = Target PEG × Growth %", fair_pe,  "0.0x",  False),
        ("★ Fair Value (₹)",             fair_value,   "#,##0.00", True),
        ("Current Market Price (₹)",      price,        "#,##0.00", False),
        ("★ Upside / (Downside)",         upside,       "+0.0%;-0.0%;0.0%", True),
        ("Premium / (Discount) to Fair PEG", (current_peg / bench["fair_peg"] - 1) if bench["fair_peg"] > 0 else 0.0, "+0.0%;-0.0%;0.0%", False),
    ]
    for lbl_txt, val, fmt, is_key in peg_rows:
        lc = ws.cell(row=row, column=2, value=lbl_txt)
        lc.font = font(bold=is_key, size=9)
        lc.border = thin_border()
        if is_key:
            lc.fill = fill(theme["key_fill"])
        vc = value_cell(ws, row, 3, val, fmt=fmt,
                        bold=is_key,
                        bg=theme["key_fill"] if is_key else None,
                        color=theme["accent"] if is_key else "000000")
        row += 1

    # Cheap/expensive label
    label_text = "⚠ Stock appears EXPENSIVE vs fair PEG" if is_expensive else "✔ Stock appears UNDERVALUED vs fair PEG"
    lc = ws.cell(row=row, column=2, value=label_text)
    lc.font = font(bold=True, size=9,
                   color="9C0006" if is_expensive else "375623")
    lc.fill = fill("FFC7CE" if is_expensive else "C6EFCE")
    lc.border = thin_border()
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    row += 2

    # ── Section 3: EPS Growth Projections (5 Years) ───────────────────────────
    section_header(ws, row, 2, "▌ EPS GROWTH PROJECTIONS (5-YEAR FORWARD)", theme, span=7)
    row += 1

    # Year headers
    ws.cell(row=row, column=2, value="Metric").font = font(bold=True, color="FFFFFF", size=9)
    ws.cell(row=row, column=2).fill = fill(theme["sub"])
    ws.cell(row=row, column=2).border = thin_border()
    for yr in range(1, 6):
        c = ws.cell(row=row, column=2 + yr, value=f"Year {yr}")
        c.font = font(bold=True, color="FFFFFF", size=9)
        c.fill = fill(theme["sub"])
        c.border = thin_border()
        c.alignment = align(h="center")
    row += 1

    proj_metrics = ["EPS (₹)", "YoY Growth", "P/E at Fair PEG", "Fair Value (₹)", "Cumul. EPS CAGR"]
    proj_data = {m: [] for m in proj_metrics}

    eps_t = eps
    for yr in range(1, 6):
        eps_t = eps_t * (1 + g)
        fair_pe_yr = target_peg * (g * 100)
        fv_yr      = fair_pe_yr * eps_t
        cagr_yr    = ((eps_t / eps) ** (1 / yr) - 1) if eps > 0 else 0.0
        proj_data["EPS (₹)"].append(eps_t)
        proj_data["YoY Growth"].append(g)
        proj_data["P/E at Fair PEG"].append(fair_pe_yr)
        proj_data["Fair Value (₹)"].append(fv_yr)
        proj_data["Cumul. EPS CAGR"].append(cagr_yr)

    fmts = {"EPS (₹)": "#,##0.00", "YoY Growth": "0.0%", "P/E at Fair PEG": "0.0x",
            "Fair Value (₹)": "#,##0.00", "Cumul. EPS CAGR": "0.0%"}

    for m_name in proj_metrics:
        lc = ws.cell(row=row, column=2, value=m_name)
        lc.font = font(size=9)
        lc.border = thin_border()
        for yr_idx, val in enumerate(proj_data[m_name]):
            vc = ws.cell(row=row, column=3 + yr_idx, value=val)
            vc.number_format = fmts[m_name]
            vc.alignment = align(h="right")
            vc.border = thin_border()
            vc.font = font(size=9)
            vc.fill = fill(theme["positive_fill"])
        row += 1

    row += 1

    # ── Section 4: PEG Context ────────────────────────────────────────────────
    section_header(ws, row, 2, "▌ PEG CONTEXT & BENCHMARKS", theme, span=3)
    row += 1

    context_rows = [
        ("Peter Lynch Rule: PEG < 1",    "Undervalued — Growth not fully priced in"),
        ("Peter Lynch Rule: PEG 1–2",    "Fairly valued — Growth reasonably priced"),
        ("Peter Lynch Rule: PEG > 2",    "Overvalued — Market pricing in above-trend growth"),
        ("Nifty 50 Average PEG (ref)",   "~1.2x – 1.5x (historical range)"),
        ("Sector",                       sector if sector else "—"),
        ("Sector Fair PEG Benchmark",    f"{bench['fair_peg']:.1f}x"),
        ("Sector Avg PEG (ref)",         f"{bench['sector_avg']:.1f}x"),
        ("Sector Note",                  bench["comment"]),
        ("This Company's Current PEG",   f"{current_peg:.2f}x"),
        ("Assessment",                   "EXPENSIVE vs benchmark" if is_expensive else "FAIR / CHEAP vs benchmark"),
    ]
    for lbl_txt, val_txt in context_rows:
        lc = ws.cell(row=row, column=2, value=lbl_txt)
        lc.font = font(size=9)
        lc.border = thin_border()
        vc = ws.cell(row=row, column=3, value=val_txt)
        vc.font = font(size=9, italic=True)
        vc.border = thin_border()
        vc.alignment = align(wrap=True)
        row += 1

    ws.freeze_panes = "C4"
    return fair_value, upside


def _build_sensitivity(wb, fin, theme, target_peg):
    ws = wb.create_sheet("Results & Sensitivity")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 20

    price = safe(fin.get("price"), 0.0)
    eps   = safe(fin.get("eps"), 0.0)

    t = ws.cell(row=2, column=2,
        value="Results & Sensitivity Analysis — PEG Ratio Valuation")
    t.font = font(bold=True, size=13, color=theme["primary"])
    ws.merge_cells("B2:L2")

    # EPS Growth % (rows: 8%–30% in 2pp steps) × Target PEG (cols: 0.5–2.0 in 0.25 steps)
    growth_rows = [g / 100 for g in range(8, 32, 2)]   # 0.08, 0.10 … 0.30
    peg_cols    = [0.5, 0.75, 1.0, 1.25, 1.5, 1.75, 2.0]

    base_g      = safe(fin.get("earnings_growth"), 0.15)
    base_row_idx = min(range(len(growth_rows)), key=lambda i: abs(growth_rows[i] - base_g))
    base_col_idx = peg_cols.index(1.0) if 1.0 in peg_cols else 2

    matrix = []
    for g in growth_rows:
        r_row = []
        for tpeg in peg_cols:
            fair_pe_val = tpeg * (g * 100)
            implied     = fair_pe_val * eps
            r_row.append(round(implied, 2))
        matrix.append(r_row)

    build_sensitivity_table(
        ws, theme,
        row_label="EPS Growth %",
        col_label="Target PEG",
        row_vals=growth_rows,
        col_vals=peg_cols,
        matrix=matrix,
        current_price=price,
        base_row_idx=base_row_idx,
        base_col_idx=base_col_idx,
        start_row=4,
        start_col=2,
        row_fmt="0.0%",
        col_fmt="0.00x",
    )

    ws.freeze_panes = "C4"


def generate_excel(fin: dict, symbol: str, company_name: str) -> bytes:
    wb = Workbook()

    price     = safe(fin.get("price"), 0.0)
    shares    = safe(fin.get("shares"), 1.0)
    mktcap_cr = price * shares / 1e7
    sector    = fin.get("sector", "—")

    eps_growth = safe(fin.get("earnings_growth"), 0.15)
    target_peg = 1.0

    build_cover(
        wb, THEME, symbol, company_name,
        model_label="PEG Ratio Valuation",
        model_desc="Fair Value = EPS × (Growth % × Target PEG) — growth-adjusted P/E methodology",
        sheets_index=[
            (1, "Cover",                "Model overview & metadata"),
            (2, "Inputs & Assumptions", "Financial data & PEG parameters"),
            (3, "PEG Ratio Analysis",   "PEG calculation, projections & context"),
            (4, "Results & Sensitivity","Sensitivity: EPS Growth % × Target PEG"),
        ],
        meta_extra={"price": price, "mktcap_cr": mktcap_cr, "sector": sector},
    )

    eps        = safe(fin.get("eps"), 0.0)
    trailing_pe = (price / eps) if eps > 0 else 0.0
    current_peg = (trailing_pe / (eps_growth * 100)) if eps_growth > 0 else 0.0
    fair_pe     = target_peg * (eps_growth * 100)

    params_def = [
        ("EPS (₹)",                eps,        "₹",  "Trailing 12 months",          "#,##0.00", True),
        ("EPS Growth Rate",        eps_growth, "%",  "YoY earnings growth",          "0.0%",    True),
        ("Trailing P/E (×)",       trailing_pe,"x",  "Price / EPS",                  "0.0x",    False),
        ("Current PEG Ratio",      current_peg,"x",  "P/E ÷ Growth%",                "0.00x",   False),
        ("Target PEG (valuation)", target_peg, "x",  "Peter Lynch: 1.0x = fair",     "0.00x",   True),
        ("Fair P/E = PEG × Growth", fair_pe,   "x",  "Derived fair multiple",        "0.0x",    False),
        ("★ Fair Value (₹)",        fair_pe * eps, "₹", "Fair P/E × EPS",            "#,##0.00", True),
        ("Sector",                 sector,     "—",  "Yahoo Finance classification",  "@",       False),
    ]

    build_inputs(wb, THEME, company_name, fin, params_def)
    _build_analysis(wb, fin, THEME, target_peg, eps_growth)
    _build_sensitivity(wb, fin, THEME, target_peg)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
