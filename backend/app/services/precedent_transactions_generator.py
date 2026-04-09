"""
Precedent Transactions Excel Generator
=========================================
M&A-implied valuation using the acquisition premium approach.

Why acquisition-premium (not EV/EBITDA applied to EBITDA directly):
  - Indian filings often provide *standalone* EBITDA while yfinance/market
    prices reflect *consolidated* operations. Applying a deal multiple to
    standalone EBITDA against a consolidated EV produces a >60x trading
    multiple and massively understates value.
  - The acquisition premium approach is immune to this: it starts from the
    market price (already consolidated), adds the premium, and uses EBITDA
    only for sizing synergies (which are a small incremental term).

Correct M&A Offer Price Formula:
  Acquisition Value  = Market Cap × (1 + deal_premium)
                     = current_price × shares × (1 + deal_premium)
  Synergy EV         = EBITDA × synergies_pct × synergy_multiple
  Synergy per Share  = Synergy EV / Shares
  Implied Offer Price = current_price × (1 + deal_premium)
                      + Synergy EV / Shares

  Alternatively (equivalent):
  Total Offer Equity = Acquisition Value (equity) + Synergy EV
  Implied Price      = Total Offer Equity / Shares

  The EV/EBITDA deal-table multiples are shown as reference benchmarks.
  The deal_premium parameter drives the valuation (30-40% typical in India).
"""

import io
import math
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.services.excel_utils import (
    font, fill, thin_border, align, safe, cr,
    section_header, build_cover, build_inputs, build_sensitivity_table,
)

THEME = {
    "cover_bg":      "1A0500",
    "primary":       "3E1200",
    "sub":           "6B2000",
    "accent":        "E67E22",
    "input_color":   "784212",
    "positive_fill": "FEF0E7",
    "positive_text": "784212",
    "subtotal_fill": "FDEBD0",
    "key_fill":      "FFF2CC",
}

# ── Comparable M&A Deal Tables (Indian sector reference) ─────────────────────
# Columns: (Target, Acquirer, Year, Deal_EV_Cr, EV/EBITDA, EV/Revenue, Acq_Premium)
# Acq_Premium = premium paid over unaffected share price (1-day prior)

_TECH_DEALS = [
    ("InfoEdge Target A",   "Strategic Buyer X", 2022, 12500, 22.0, 5.2, 0.32),
    ("SaaS Target B",       "PE Sponsor Y",      2021, 8500,  26.0, 6.1, 0.38),
    ("IT Services C",       "Global Corp Z",     2023, 45000, 19.5, 4.8, 0.28),
    ("Platform D",          "Conglomerate W",    2020, 6200,  24.5, 5.8, 0.35),
    ("FinTech E",           "Bank V",            2022, 18000, 28.0, 7.0, 0.42),
    ("Digital Media F",     "Media Corp U",      2021, 3800,  20.0, 4.5, 0.30),
]

_FMCG_DEALS = [
    ("FMCG Target A",       "Unilever Group",    2022, 28000, 32.0, 6.0, 0.35),
    ("Foods B",             "Global FMCG Y",     2021, 15000, 28.5, 5.5, 0.30),
    ("Beverages C",         "Beverage Corp Z",   2023, 42000, 35.0, 6.8, 0.40),
    ("Personal Care D",     "Beauty Group W",    2020, 12000, 25.0, 5.0, 0.28),
    ("Dairy E",             "Dairy Giant V",     2022, 9500,  22.0, 4.2, 0.25),
    ("Snacks F",            "Food Congl U",      2021, 7200,  27.0, 5.2, 0.32),
]

_BANKING_DEALS = [
    ("NBFC Target A",       "Large Bank X",      2022, 22000,  8.5, 2.8, 0.22),
    ("Small Finance B",     "Private Bank Y",    2021, 8500,   7.0, 2.2, 0.18),
    ("Microfinance C",      "Bank Corp Z",       2023, 5500,   9.0, 3.0, 0.25),
    ("Insurance D",         "Financial Group W", 2020, 35000, 10.5, 3.5, 0.28),
    ("AMC E",               "Wealth Mgmt V",     2022, 12000, 12.0, 4.0, 0.30),
    ("Broking F",           "Capital Markets U", 2021, 4500,   8.0, 2.5, 0.20),
]

_PHARMA_DEALS = [
    ("Pharma Target A",     "Global Pharma X",   2022, 18000, 18.0, 4.2, 0.28),
    ("Generics B",          "MNC Pharma Y",      2021, 12000, 15.5, 3.8, 0.24),
    ("API Maker C",         "Chemical Corp Z",   2023, 8500,  14.0, 3.5, 0.22),
    ("Hospital Chain D",    "Healthcare W",      2020, 45000, 22.0, 5.0, 0.32),
    ("Diagnostics E",       "Diagnostics Corp V",2022, 15000, 25.0, 5.5, 0.35),
    ("MedTech F",           "Device Corp U",     2021, 9200,  20.0, 4.8, 0.30),
]

_ENERGY_DEALS = [
    ("Oil Refiner A",       "Energy Corp X",     2022, 85000, 10.0, 1.2, 0.20),
    ("Gas Distributor B",   "Utility Y",         2021, 32000,  9.5, 1.5, 0.18),
    ("Renewable C",         "Power Corp Z",      2023, 18000, 15.0, 3.2, 0.25),
    ("Pipelines D",         "Infrastructure W",  2020, 55000, 11.0, 1.8, 0.22),
    ("Petrochemicals E",    "Chemical Group V",  2022, 42000,  8.5, 1.0, 0.15),
    ("City Gas F",          "Gas Utility U",     2021, 12000, 12.0, 2.5, 0.20),
]

_INFRA_DEALS = [
    ("Roads Target A",      "Infrastructure X",  2022, 25000, 14.0, 3.5, 0.25),
    ("Ports B",             "Port Corp Y",       2021, 38000, 16.0, 4.0, 0.28),
    ("Airports C",          "Aviation Group Z",  2023, 85000, 20.0, 5.0, 0.35),
    ("Construction D",      "EPC Corp W",        2020, 8500,   8.5, 1.5, 0.18),
    ("Power Infra E",       "Power Corp V",      2022, 15000, 12.0, 2.8, 0.22),
    ("Telecom Tower F",     "Tower Corp U",      2021, 42000, 18.0, 4.5, 0.30),
]

DEFAULT_DEALS = [
    ("Target A",  "Acquirer X", 2022, 50000, 12.0, 3.5, 0.25),
    ("Target B",  "Acquirer Y", 2021, 30000, 10.0, 2.8, 0.20),
    ("Target C",  "Acquirer Z", 2023, 75000, 14.0, 4.0, 0.30),
    ("Target D",  "Acquirer W", 2020, 25000,  9.0, 2.5, 0.18),
    ("Target E",  "Acquirer V", 2022, 60000, 13.0, 3.8, 0.28),
    ("Target F",  "Acquirer U", 2021, 40000, 11.0, 3.2, 0.22),
]

SECTOR_DEALS = {
    "Technology":             _TECH_DEALS,
    "Information Technology": _TECH_DEALS,
    "Software":               _TECH_DEALS,
    "FMCG":                   _FMCG_DEALS,
    "Consumer Defensive":     _FMCG_DEALS,
    "Consumer Staples":       _FMCG_DEALS,
    "Banking":                _BANKING_DEALS,
    "Financial Services":     _BANKING_DEALS,
    "Finance":                _BANKING_DEALS,
    "Pharmaceuticals":        _PHARMA_DEALS,
    "Healthcare":             _PHARMA_DEALS,
    "Pharma":                 _PHARMA_DEALS,
    "Energy":                 _ENERGY_DEALS,
    "Oil & Gas":              _ENERGY_DEALS,
    "Industrials":            _INFRA_DEALS,
    "Infrastructure":         _INFRA_DEALS,
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _median(vals):
    s = sorted(v for v in vals if v is not None and v > 0)
    if not s:
        return 0.0
    n = len(s)
    return s[n // 2] if n % 2 else (s[n // 2 - 1] + s[n // 2]) / 2


def _mean(vals):
    v = [x for x in vals if x is not None and x > 0]
    return sum(v) / len(v) if v else 0.0


def _lbl(ws, row, text, bold=False, bg=None):
    c = ws.cell(row=row, column=2, value=text)
    c.font      = font(bold=bold, size=9)
    c.border    = thin_border()
    c.alignment = align()
    if bg:
        c.fill = fill(bg)
    return c


def _val(ws, row, col, val, fmt="#,##0.00", bold=False, bg=None, color="000000"):
    c = ws.cell(row=row, column=col, value=val if val is not None else "N/M")
    c.font          = font(bold=bold, size=9, color=color)
    c.number_format = fmt if val is not None else "@"
    c.alignment     = align(h="right")
    c.border        = thin_border()
    if bg:
        c.fill = fill(bg)
    return c


def _hdr(ws, row, cols, theme):
    for i, txt in enumerate(cols):
        c = ws.cell(row=row, column=2 + i, value=txt)
        c.font      = font(bold=True, color="FFFFFF", size=9)
        c.fill      = fill(theme["sub"])
        c.border    = thin_border()
        c.alignment = align(h="center" if i > 0 else "left")


def _stat_row(ws, row, lbl_txt, vals, theme, fmts=None):
    c = ws.cell(row=row, column=2, value=lbl_txt)
    c.font      = font(bold=True, size=9)
    c.fill      = fill(theme["subtotal_fill"])
    c.border    = thin_border()
    c.alignment = align()
    fmts = fmts or ["0.0x"] * len(vals)
    for j, (v, fmt) in enumerate(zip(vals, fmts)):
        vc = ws.cell(row=row, column=3 + j, value=round(v, 2) if v else "—")
        vc.font         = font(bold=True, size=9)
        vc.number_format = fmt
        vc.alignment    = align(h="right")
        vc.border       = thin_border()
        vc.fill         = fill(theme["subtotal_fill"])


def _resolve_deals(sector: str):
    if not sector:
        return DEFAULT_DEALS, "General Market"
    if sector in SECTOR_DEALS:
        return SECTOR_DEALS[sector], sector
    low = sector.lower()
    for k, v in SECTOR_DEALS.items():
        if k.lower() == low or k.lower() in low or low in k.lower():
            return v, k
    return DEFAULT_DEALS, "General Market"


# ── Core valuation ────────────────────────────────────────────────────────────

def _compute(fin, deal_premium, synergies_pct, synergy_multiple=10.0):
    """
    Acquisition-premium approach — robust to standalone vs consolidated data.

    Formula:
        Acquisition Value (equity) = Market Cap × (1 + deal_premium)
        Synergy EV                 = EBITDA × synergies_pct × synergy_multiple
        Total Offer Equity         = Acquisition Value + Synergy EV
        Implied Price              = Total Offer Equity / Shares

    Note: Synergy EV is added directly to equity (assumes synergies create
    equity value for target shareholders; acquirer pays for them via the offer).
    """
    price      = safe(fin.get("price"), 0.0)
    shares_raw = fin.get("shares") or 0          # from balance sheet filing
    shares_cr  = shares_raw / 1e7
    ebitda_cr  = cr(fin.get("ebitda"))
    mktcap_cr  = price * shares_cr

    acq_equity    = mktcap_cr * (1.0 + deal_premium)   # offer equity to shareholders
    synergy_ev    = ebitda_cr * synergies_pct * synergy_multiple
    total_equity  = acq_equity + synergy_ev
    implied_price = total_equity / shares_cr if shares_cr > 0 else 0.0
    upside        = (implied_price / price - 1) if price > 0 else 0.0

    return implied_price, upside, acq_equity, synergy_ev, total_equity


# ── Sheet 3: M&A Transaction Analysis ────────────────────────────────────────

def _build_analysis(wb, fin, theme, deal_premium, synergies_pct, synergy_multiple):
    ws = wb.create_sheet("M&A Transaction Analysis")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14

    price      = safe(fin.get("price"), 0.0)
    shares_raw = fin.get("shares") or 0
    shares_cr  = shares_raw / 1e7
    rev_cr     = cr(fin.get("revenue"))
    ebitda_cr  = cr(fin.get("ebitda"))
    ni_cr      = cr(fin.get("net_income"))
    debt_cr    = cr(fin.get("total_debt"))
    cash_cr    = cr(fin.get("cash"))
    net_debt   = debt_cr - cash_cr
    mktcap_cr  = price * shares_cr
    ev_trading = cr(fin.get("enterprise_value")) or (mktcap_cr + net_debt)

    ebitda_margin     = ebitda_cr / rev_cr        if rev_cr > 0     else 0.0
    trading_ev_ebitda = ev_trading / ebitda_cr    if ebitda_cr > 0  else 0.0
    ev_rev_trading    = ev_trading / rev_cr       if rev_cr > 0     else 0.0

    ip, up, acq_eq, syn_ev, total_eq = _compute(fin, deal_premium, synergies_pct, synergy_multiple)

    acq_premium_check = (ip / price - 1) if price > 0 else 0.0

    row = 2
    t = ws.cell(row=row, column=2,
                value="Precedent Transactions  —  M&A Acquisition Premium Valuation")
    t.font = font(bold=True, size=13)
    ws.merge_cells(f"B{row}:H{row}")
    row += 2

    # ── SECTION A: Subject Company Snapshot ──────────────────────────────────
    section_header(ws, row, 2, "▌ A.  SUBJECT COMPANY SNAPSHOT", theme, span=3)
    row += 1
    _hdr(ws, row, ["Metric", "Value", "Unit"], theme)
    row += 1
    for lbl_txt, val, unit, fmt in [
        ("Current Share Price  (₹)",        price,             "₹",    "#,##0.00"),
        ("Shares Outstanding  (Cr)  [BS]",  shares_cr,         "Cr",   "#,##0.00"),
        ("Market Cap  (₹ Cr)",              mktcap_cr,         "₹ Cr", "#,##0"),
        ("Enterprise Value — Market  (₹ Cr)", ev_trading,      "₹ Cr", "#,##0"),
        ("Revenue  (₹ Cr)",                 rev_cr,            "₹ Cr", "#,##0"),
        ("EBITDA  (₹ Cr)",                  ebitda_cr,         "₹ Cr", "#,##0"),
        ("EBITDA Margin",                   ebitda_margin,     "%",    "0.0%"),
        ("Net Income  (₹ Cr)",              ni_cr,             "₹ Cr", "#,##0"),
        ("Total Debt  (₹ Cr)",              debt_cr,           "₹ Cr", "#,##0"),
        ("Cash & Equivalents  (₹ Cr)",      cash_cr,           "₹ Cr", "#,##0"),
        ("Net Debt / (Cash)  (₹ Cr)",       net_debt,          "₹ Cr", "#,##0"),
        ("Trading EV / EBITDA  (mkt)",      trading_ev_ebitda, "x",    "0.0x"),
        ("Trading EV / Revenue  (mkt)",     ev_rev_trading,    "x",    "0.0x"),
    ]:
        _lbl(ws, row, "    " + lbl_txt)
        _val(ws, row, 3, val, fmt)
        c = ws.cell(row=row, column=4, value=unit)
        c.font = font(size=9); c.border = thin_border()
        row += 1
    row += 1

    # ── SECTION B: Acquisition Assumptions ───────────────────────────────────
    section_header(ws, row, 2, "▌ B.  ACQUISITION ASSUMPTIONS", theme, span=3)
    row += 1
    _hdr(ws, row, ["Parameter", "Value", "Note"], theme)
    row += 1
    for lbl_txt, val, fmt, note in [
        ("Deal Premium over Market Price",
         deal_premium, "0.0%",
         "Premium paid over unaffected share price  |  Typical India M&A: 25-40%"),
        ("Synergies (% of Target EBITDA)",
         synergies_pct, "0.0%",
         "Annual synergy EBITDA as % of target EBITDA"),
        ("Synergy Capitalisation Multiple",
         synergy_multiple, "0.0x",
         "EV/EBITDA multiple used to capitalise synergy value"),
        ("Synergy EV  (₹ Cr)",
         syn_ev, "#,##0",
         f"= EBITDA {ebitda_cr:,.0f} x {synergies_pct:.0%} x {synergy_multiple:.1f}x"),
        ("Acquisition Value — Equity  (₹ Cr)",
         acq_eq, "#,##0",
         f"= Market Cap {mktcap_cr:,.0f} x (1 + {deal_premium:.0%})"),
        ("Total Offer Equity  (₹ Cr)",
         total_eq, "#,##0",
         "= Acquisition Value + Synergy EV"),
    ]:
        is_key = lbl_txt.startswith("Total Offer")
        _lbl(ws, row, "    " + lbl_txt, bold=is_key,
             bg=theme["key_fill"] if is_key else None)
        _val(ws, row, 3, val, fmt, bold=is_key,
             bg=theme["key_fill"] if is_key else None)
        nc = ws.cell(row=row, column=4, value=note)
        nc.font = font(size=8, color="595959")
        nc.border = thin_border()
        ws.merge_cells(f"D{row}:H{row}")
        row += 1
    row += 1

    # ── SECTION C: Offer Price Bridge ────────────────────────────────────────
    section_header(ws, row, 2, "▌ C.  OFFER PRICE BRIDGE", theme, span=3)
    row += 1
    bridge = [
        ("Unaffected Market Price  (₹)",        price,       "#,##0.00",           False),
        ("× (1 + Deal Premium)",                deal_premium,"0.0%",               False),
        ("= Acquisition Price  (₹)",            price*(1+deal_premium), "#,##0.00",False),
        ("+ Synergy Value per Share  (₹)",      syn_ev/shares_cr if shares_cr else 0,
                                                "#,##0.00",                         False),
        ("★  Implied Offer Price  (₹)",         ip,          "#,##0.00",           True),
        ("Current Market Price  (₹)",           price,       "#,##0.00",           False),
        ("★  Upside to Offer Price",            up,          "+0.0%;-0.0%;0.0%",   True),
        ("÷ Shares Outstanding  (Cr)  [BS]",    shares_cr,   "#,##0.00",           False),
        ("Total Offer Equity  (₹ Cr)",          total_eq,    "#,##0",              False),
    ]
    for lbl_txt, val, fmt, is_key in bridge:
        _lbl(ws, row, "    " + lbl_txt, bold=is_key,
             bg=theme["key_fill"] if is_key else None)
        _val(ws, row, 3, val, fmt, bold=is_key,
             bg=theme["key_fill"] if is_key else None,
             color=theme["accent"] if is_key else "000000")
        row += 1
    row += 1

    # ── SECTION D: Comparable Deal Table ─────────────────────────────────────
    sector = fin.get("sector", "")
    deals, sector_label = _resolve_deals(sector)

    section_header(ws, row, 2,
                   f"▌ D.  COMPARABLE DEAL TABLE  —  Sector: {sector_label}", theme, span=7)
    row += 1
    _hdr(ws, row,
         ["Target (Anonymous)", "Acquirer", "Year",
          "Deal EV  (₹ Cr)", "EV/EBITDA  (deal)", "EV/Revenue", "Acq. Premium"],
         theme)
    row += 1

    ev_mult_list, ev_rev_list, prem_list = [], [], []
    for (target, acquirer, year, ev_deal, ev_eb, ev_rev, prem) in deals:
        ev_mult_list.append(ev_eb)
        ev_rev_list.append(ev_rev)
        prem_list.append(prem)
        _lbl(ws, row, "    " + target)
        for col_off, val, fmt in [
            (1, acquirer, "@"),
            (2, year,     "0"),
            (3, ev_deal,  "#,##0"),
            (4, ev_eb,    "0.0x"),
            (5, ev_rev,   "0.0x"),
            (6, prem,     "0.0%"),
        ]:
            c = ws.cell(row=row, column=2 + col_off, value=val)
            c.number_format = fmt
            c.alignment = align(h="right" if col_off > 1 else "left")
            c.border    = thin_border()
            c.font      = font(size=9)
        row += 1

    for stat_lbl, fn in [("Median", _median), ("Mean", _mean)]:
        _stat_row(ws, row, stat_lbl,
                  [fn(ev_mult_list), fn(ev_rev_list), fn(prem_list)],
                  theme, fmts=["0.0x", "0.0x", "0.0%"])
        for col_off in [1, 2, 3]:
            c = ws.cell(row=row, column=2 + col_off)
            c.border = thin_border()
            c.fill   = fill(theme["subtotal_fill"])
        row += 1
    row += 1

    # ── SECTION E: Scenario Analysis ─────────────────────────────────────────
    med_premium = _median(prem_list)
    section_header(ws, row, 2, "▌ E.  SCENARIO ANALYSIS", theme, span=6)
    row += 1
    _hdr(ws, row,
         ["Scenario", "Deal Premium", "Synergies %",
          "Offer Equity  (₹ Cr)", "Implied Price  (₹)", "vs Current"],
         theme)
    row += 1

    scenarios = [
        ("Bear  — Low Premium",   max(med_premium * 0.60, 0.10), 0.00,           "FFC7CE"),
        ("Base  — Median Deal",   med_premium,                    synergies_pct,  theme["key_fill"]),
        ("Bull  — Strategic Fit", min(med_premium * 1.40, 0.60),  synergies_pct * 2, "C6EFCE"),
    ]
    for scen_name, dp, syn_p, bg_c in scenarios:
        s_iv, s_up, _, s_syn, s_eq = _compute(fin, dp, syn_p, synergy_multiple)
        for col_off, (val, fmt) in enumerate([
            (scen_name, "@"),
            (dp,        "0.0%"),
            (syn_p,     "0.0%"),
            (s_eq,      "#,##0"),
            (s_iv,      "#,##0.00"),
            (s_up,      "+0.0%;-0.0%;0.0%"),
        ]):
            c = ws.cell(row=row, column=2 + col_off, value=val)
            c.number_format = fmt
            c.alignment = align(h="left" if col_off == 0 else "right")
            c.border    = thin_border()
            c.fill      = fill(bg_c)
            c.font      = font(bold=(scen_name.startswith("Base")), size=9)
        row += 1

    ws.freeze_panes = "C4"
    return ip, up


# ── Sheet 4: Results & Sensitivity ───────────────────────────────────────────

def _build_sensitivity(wb, fin, theme, deal_premium, synergies_pct, synergy_multiple):
    ws = wb.create_sheet("Results & Sensitivity")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 24

    price = safe(fin.get("price"), 0.0)

    t = ws.cell(row=2, column=2,
                value="Results & Sensitivity  —  Deal Premium  x  Synergies %")
    t.font = font(bold=True, size=13, color=theme["primary"])
    ws.merge_cells("B2:L2")

    # Rows = deal premium, cols = synergies %
    prem_vals = [0.10, 0.15, 0.20, 0.25, 0.30, 0.35, 0.40, 0.45, 0.50]
    syn_vals  = [0.00, 0.025, 0.05, 0.075, 0.10, 0.125, 0.15]

    base_row_idx = min(range(len(prem_vals)), key=lambda i: abs(prem_vals[i] - deal_premium))
    base_col_idx = min(range(len(syn_vals)),  key=lambda i: abs(syn_vals[i]  - synergies_pct))

    matrix = []
    for dp in prem_vals:
        r_row = []
        for syn_p in syn_vals:
            ip, _, _, _, _ = _compute(fin, dp, syn_p, synergy_multiple)
            r_row.append(round(ip, 2))
        matrix.append(r_row)

    build_sensitivity_table(
        ws, theme,
        row_label="Deal Premium",
        col_label="Synergies % of EBITDA",
        row_vals=prem_vals,
        col_vals=syn_vals,
        matrix=matrix,
        current_price=price,
        base_row_idx=base_row_idx,
        base_col_idx=base_col_idx,
        start_row=4,
        start_col=2,
        row_fmt="0.0%",
        col_fmt="0.0%",
    )
    ws.freeze_panes = "C4"


# ── Entry point ───────────────────────────────────────────────────────────────

def generate_excel(fin: dict, symbol: str, company_name: str) -> bytes:
    wb = Workbook()

    price      = safe(fin.get("price"), 0.0)
    shares_raw = fin.get("shares") or 0
    mktcap_cr  = price * shares_raw / 1e7
    sector     = fin.get("sector", "—")

    # Default parameters
    deal_premium     = 0.30   # 30% acquisition premium over market (typical India M&A)
    synergies_pct    = 0.05   # 5% synergy EBITDA
    synergy_multiple = 10.0   # capitalise synergies at 10x EV/EBITDA

    build_cover(
        wb, THEME, symbol, company_name,
        model_label="Precedent Transactions",
        model_desc=(
            "M&A-implied value via acquisition premium + capitalized synergies. "
            "Uses market price as base to avoid standalone/consolidated data mismatch."
        ),
        sheets_index=[
            (1, "Cover",                    "Model overview & metadata"),
            (2, "Inputs & Assumptions",     "Financial data & deal parameters"),
            (3, "M&A Transaction Analysis", "Assumptions, offer bridge & deal table"),
            (4, "Results & Sensitivity",    "Sensitivity: deal premium × synergies %"),
        ],
        meta_extra={"price": price, "mktcap_cr": mktcap_cr, "sector": sector},
    )

    ebitda_cr = cr(fin.get("ebitda"))
    net_debt  = cr(fin.get("total_debt")) - cr(fin.get("cash"))

    params_def = [
        ("Deal Premium over Market Price", deal_premium,     "%",
         "Acquisition premium over unaffected price  |  Typical: 25-40%", "0.0%",  True),
        ("Synergies (% of Target EBITDA)", synergies_pct,    "%",
         "Annual synergy EBITDA uplift as % of target EBITDA",            "0.0%",  True),
        ("Synergy Capitalisation Multiple",synergy_multiple, "x",
         "EV/EBITDA multiple to capitalise synergy EBITDA into EV",       "0.0x",  True),
        ("EBITDA  (₹ Cr)",                 ebitda_cr,        "₹ Cr",
         "From filings (used for synergy sizing only)",                    "#,##0", False),
        ("Net Debt / (Cash)  (₹ Cr)",      net_debt,         "₹ Cr",
         "Total Debt − Cash  (reference)",                                 "#,##0", False),
        ("Shares Outstanding  (Cr)",       shares_raw / 1e7, "Cr",
         "From balance sheet (filing)",                                    "#,##0.00", False),
    ]

    build_inputs(wb, THEME, company_name, fin, params_def)
    _build_analysis(wb, fin, THEME, deal_premium, synergies_pct, synergy_multiple)
    _build_sensitivity(wb, fin, THEME, deal_premium, synergies_pct, synergy_multiple)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
