"""
Replacement Cost Valuation Excel Generator
Values the firm based on what it would cost to recreate its asset base today —
reflects economic replacement value vs book.
"""
import io
from openpyxl import Workbook

from app.services.excel_utils import (
    font, fill, thin_border, align, safe, cr,
    section_header, sub_header, label, value_cell, pct_cell,
    build_cover, build_inputs, build_results, build_sensitivity_table,
)

THEME = {
    "cover_bg":      "0A1500",
    "primary":       "1E3A00",
    "sub":           "345E00",
    "accent":        "82960A",
    "input_color":   "4A5E00",
    "positive_fill": "EAFAF1",
    "positive_text": "1E8449",
    "subtotal_fill": "A9DFBF",
    "key_fill":      "FFF2CC",
}

DEFAULT_REBUILD_MULT = 1.2
DEFAULT_DEP_ADJ      = 0.30


# ─────────────────────────────────────────────────────────────────────────────
def _replacement_per_share(total_assets_cr, total_liab_cr, shares_cr,
                            rebuild_mult=DEFAULT_REBUILD_MULT,
                            dep_adj=DEFAULT_DEP_ADJ):
    ppe_bv      = total_assets_cr * 0.40
    bldg_bv     = total_assets_cr * 0.20
    tech_bv     = total_assets_cr * 0.08
    inv_wc_bv   = total_assets_cr * 0.15

    ppe_rep     = ppe_bv    * rebuild_mult * (1 - dep_adj)
    bldg_rep    = bldg_bv   * 1.3         * (1 - dep_adj)
    tech_rep    = tech_bv   * 1.5         * (1 - 0.5 * dep_adj)
    inv_rep     = inv_wc_bv * 1.0

    total_rep   = ppe_rep + bldg_rep + tech_rep + inv_rep
    eq_val      = total_rep - total_liab_cr
    if shares_cr <= 0:
        return 0.0
    return eq_val / shares_cr * 1e7   # ₹ per share


# ─────────────────────────────────────────────────────────────────────────────
def _build_rc_analysis(wb, theme, fin, params):
    ws = wb.create_sheet("Replacement Cost Analysis")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 18

    total_assets_cr = cr(fin.get("total_assets"))
    total_liab_cr   = cr(fin.get("total_liabilities"))
    shares          = safe(fin.get("shares"), 1e8)
    shares_cr       = shares / 1e7
    price           = safe(fin.get("price"), 0.0)
    total_debt_cr   = cr(fin.get("total_debt"))
    ev_cr           = cr(fin.get("enterprise_value"))
    if ev_cr <= 0:
        ev_cr = price * shares / 1e7 + total_debt_cr - cr(fin.get("cash"))

    rebuild_mult = params.get("rebuild_mult", DEFAULT_REBUILD_MULT)
    dep_adj      = params.get("dep_adj", DEFAULT_DEP_ADJ)

    ppe_bv      = total_assets_cr * 0.40
    bldg_bv     = total_assets_cr * 0.20
    tech_bv     = total_assets_cr * 0.08
    inv_wc_bv   = total_assets_cr * 0.15
    intang_bv   = total_assets_cr * 0.05
    fixed_bv    = ppe_bv + bldg_bv
    curr_bv     = inv_wc_bv + tech_bv
    book_net    = total_assets_cr - total_liab_cr

    # Replacement rows
    RC_ROWS = [
        ("PP&E / Machinery",           ppe_bv,    rebuild_mult, dep_adj,       ppe_bv * rebuild_mult * (1 - dep_adj)),
        ("Buildings / Civil Works",    bldg_bv,   1.30,         dep_adj,       bldg_bv * 1.30 * (1 - dep_adj)),
        ("Technology / IT Infra",      tech_bv,   1.50,         0.5 * dep_adj, tech_bv * 1.50 * (1 - 0.5 * dep_adj)),
        ("Inventory / Working Capital",inv_wc_bv, 1.00,         0.00,          inv_wc_bv),
        ("Intangibles / Brand (excl.)",0.0,       0.00,         0.00,          0.0),
    ]
    total_rep = sum(r[4] for r in RC_ROWS)
    rep_eq    = total_rep - total_liab_cr
    imp_ps    = rep_eq / shares_cr * 1e7 if shares_cr > 0 else 0
    upside    = (imp_ps - price) / price if price else 0
    tobins_q  = ev_cr / total_rep if total_rep else 0

    row = 2
    t = ws.cell(row=row, column=2, value="Replacement Cost Analysis")
    t.font = font(bold=True, size=13)
    ws.merge_cells(f"B{row}:G{row}")

    # ── Section 1: Balance Sheet ───────────────────────────────────────────────
    row += 2
    section_header(ws, row, 2, "▌ BALANCE SHEET (CURRENT)", theme, span=6)
    row += 1

    for lbl_txt, val, fmt, is_key in [
        ("Total Assets (₹ Cr)",              total_assets_cr, "#,##0.0", False),
        ("Fixed / PP&E Assets (~40% total)", fixed_bv,        "#,##0.0", False),
        ("Current Assets (~15% total)",      curr_bv,         "#,##0.0", False),
        ("Intangibles / Goodwill (~5%)",     intang_bv,       "#,##0.0", False),
        ("Total Liabilities (₹ Cr)",         total_liab_cr,   "#,##0.0", False),
        ("Book Net Worth (₹ Cr)",            book_net,        "#,##0.0", True),
    ]:
        c = ws.cell(row=row, column=2, value=lbl_txt)
        c.font = font(bold=is_key, size=9)
        c.border = thin_border()
        if is_key:
            c.fill = fill(theme["positive_fill"])
        v = ws.cell(row=row, column=3, value=val)
        v.number_format = fmt
        v.font = font(bold=is_key, size=9)
        v.alignment = align(h="right")
        v.border = thin_border()
        if is_key:
            v.fill = fill(theme["positive_fill"])
        row += 1

    # ── Section 2: Replacement Cost Adjustment ────────────────────────────────
    row += 1
    section_header(ws, row, 2, "▌ REPLACEMENT COST ADJUSTMENT", theme, span=6)
    row += 1

    hdr = ["Asset Category", "Book Value (₹ Cr)", "Rebuild Mult",
           "Gross Repl. Cost", "Depr. Adj %", "Net Repl. Value (₹ Cr)"]
    for j, h in enumerate(hdr):
        c = ws.cell(row=row, column=2 + j, value=h)
        c.font = font(bold=True, color="FFFFFF", size=9)
        c.fill = fill(theme["sub"])
        c.border = thin_border()
        c.alignment = align()
    row += 1

    for name, bv, mult, dep, nrv in RC_ROWS:
        ws.cell(row=row, column=2, value=name).font = font(size=9)
        ws.cell(row=row, column=2).border = thin_border()
        value_cell(ws, row, 3, bv,          "#,##0.0")
        value_cell(ws, row, 4, mult,        "0.00x")
        value_cell(ws, row, 5, bv * mult,   "#,##0.0")
        pct_cell(ws,   row, 6, dep)
        value_cell(ws, row, 7, nrv,         "#,##0.0")
        row += 1

    for col, val in [(3, sum(r[1] for r in RC_ROWS)), (7, total_rep)]:
        value_cell(ws, row, col, val, "#,##0.0", bold=True, bg=theme["subtotal_fill"])
    ws.cell(row=row, column=2, value="TOTAL Replacement Value of Assets").font = font(bold=True, size=9)
    ws.cell(row=row, column=2).border = thin_border()
    ws.cell(row=row, column=2).fill = fill(theme["subtotal_fill"])
    row += 1

    # ── Section 3: Tobin's Q Analysis ─────────────────────────────────────────
    row += 1
    section_header(ws, row, 2, "▌ TOBIN'S Q ANALYSIS", theme, span=6)
    row += 1

    mkt_eq = price * shares / 1e7
    tobins_q_label = "Q > 1 → overvalued vs assets" if tobins_q > 1 else "Q < 1 → potential value opportunity"

    for lbl_txt, val, fmt, is_key in [
        ("Market Value of Equity (₹ Cr)",    mkt_eq,      "#,##0.0", False),
        ("Market Value of Debt (≈ Book)",    total_debt_cr,"#,##0.0", False),
        ("Total Market Value / EV (₹ Cr)",   ev_cr,       "#,##0.0", False),
        ("Total Replacement Value (₹ Cr)",   total_rep,   "#,##0.0", False),
        ("★ Tobin's Q (Market / Replacement)", tobins_q,  "0.00x",   True),
        (tobins_q_label,                     "",          "",        False),
    ]:
        c = ws.cell(row=row, column=2, value=lbl_txt)
        c.font = font(bold=is_key, size=9,
                      color=theme["accent"] if is_key else "000000")
        c.border = thin_border()
        if is_key:
            c.fill = fill(theme["key_fill"])
        if val != "":
            v = ws.cell(row=row, column=3, value=val)
            v.number_format = fmt
            v.font = font(bold=is_key, size=9,
                          color=theme["accent"] if is_key else "000000")
            v.alignment = align(h="right")
            v.border = thin_border()
            if is_key:
                v.fill = fill(theme["key_fill"])
        row += 1

    # ── Section 4: Equity Value Bridge ────────────────────────────────────────
    row += 1
    section_header(ws, row, 2, "▌ EQUITY VALUE BRIDGE", theme, span=6)
    row += 1

    for lbl_txt, val, fmt, is_key in [
        ("Total Net Replacement Value (₹ Cr)",   total_rep,  "#,##0.0",  False),
        ("Less: Total Liabilities (₹ Cr)",       total_liab_cr, "#,##0.0", False),
        ("Replacement Equity Value (₹ Cr)",      rep_eq,     "#,##0.0",  True),
        ("Shares Outstanding (Cr)",              shares_cr,  "#,##0.00", False),
        ("★ Implied Share Price (₹)",            imp_ps,     "#,##0.00", True),
        ("Current Market Price (₹)",             price,      "#,##0.00", False),
        ("Upside / (Downside)",                  upside,     "0.0%",     True),
    ]:
        c = ws.cell(row=row, column=2, value=lbl_txt)
        c.font = font(bold=is_key, size=9)
        c.border = thin_border()
        if is_key:
            c.fill = fill(theme["key_fill"])
        v = ws.cell(row=row, column=3, value=val)
        v.number_format = fmt
        v.font = font(bold=is_key, size=9,
                      color=theme["accent"] if is_key else "000000")
        v.alignment = align(h="right")
        v.border = thin_border()
        if is_key:
            v.fill = fill(theme["key_fill"])
        row += 1

    ws.freeze_panes = "C4"
    return ws


def _build_rc_sensitivity(wb, theme, fin):
    ws = wb.create_sheet("Sensitivity")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 22

    total_assets_cr = cr(fin.get("total_assets"))
    total_liab_cr   = cr(fin.get("total_liabilities"))
    shares          = safe(fin.get("shares"), 1e8)
    shares_cr       = shares / 1e7
    price           = safe(fin.get("price"), 0.0)

    row_vals = [0.80, 0.90, 1.00, 1.10, 1.20, 1.40, 1.60, 1.80]
    col_vals = [0.10, 0.20, 0.30, 0.40, 0.50, 0.60]

    base_row_idx = 4   # 1.20
    base_col_idx = 2   # 0.30

    matrix = []
    for mult in row_vals:
        row_data = []
        for dep in col_vals:
            ps = _replacement_per_share(total_assets_cr, total_liab_cr, shares_cr, mult, dep)
            row_data.append(round(ps, 2))
        matrix.append(row_data)

    row = 2
    t = ws.cell(row=row, column=2, value="Replacement Cost Sensitivity Analysis")
    t.font = font(bold=True, size=13)
    ws.merge_cells(f"B{row}:I{row}")
    row += 2

    build_sensitivity_table(
        ws, theme,
        row_label="Rebuild Multiplier",
        col_label="Depreciation Adj %",
        row_vals=row_vals,
        col_vals=col_vals,
        matrix=matrix,
        current_price=price,
        base_row_idx=base_row_idx,
        base_col_idx=base_col_idx,
        start_row=row,
        row_fmt="0.00x",
        col_fmt="0.0%",
    )
    return ws


# ─────────────────────────────────────────────────────────────────────────────
def generate_excel(fin: dict, symbol: str, company_name: str) -> bytes:
    wb = Workbook()

    price     = safe(fin.get("price"), 0.0)
    shares    = safe(fin.get("shares"), 1e8)
    mktcap_cr = price * shares / 1e7

    params = {
        "rebuild_mult": DEFAULT_REBUILD_MULT,
        "dep_adj":      DEFAULT_DEP_ADJ,
    }

    build_cover(
        wb, THEME, symbol, company_name,
        model_label="Replacement Cost Valuation",
        model_desc="Equity value = Net Replacement Value of Assets − Total Liabilities",
        sheets_index=[
            (1, "Cover",                  "Company overview & model index"),
            (2, "Inputs & Assumptions",   "Financial data & model parameters"),
            (3, "Replacement Cost Analysis", "Asset rebuild cost, Tobin's Q, equity bridge"),
            (4, "Sensitivity",            "Rebuild mult × depreciation adj — implied price"),
        ],
        meta_extra={"price": price, "mktcap_cr": mktcap_cr, "sector": safe(fin.get("sector"), "—")},
    )

    beta = safe(fin.get("beta"), 1.0)
    ke   = 0.072 + beta * 0.055

    params_def = [
        ("Rebuild Multiplier (PP&E)",      DEFAULT_REBUILD_MULT, "x",  "Cost to rebuild vs book; 1.2 = 20% above book", "0.00x", True),
        ("Buildings Rebuild Multiplier",   1.30,                 "x",  "Construction cost premium",                     "0.00x", False),
        ("Tech/IT Rebuild Multiplier",     1.50,                 "x",  "Modern equipment premium",                      "0.00x", False),
        ("Depreciation Adjustment %",      DEFAULT_DEP_ADJ,      "%",  "Reduction for accumulated depreciation",        "0.0%",  True),
        ("Cost of Equity (CAPM)",          ke,                   "%",  "Rf + β × ERP (ref only)",                       "0.0%",  False),
        ("PP&E % of Total Assets",         0.40,                 "%",  "Used if PP&E not separately reported",          "0.0%",  False),
        ("Buildings % of Total Assets",    0.20,                 "%",  "Civil works allocation",                        "0.0%",  False),
        ("Tech/IT % of Total Assets",      0.08,                 "%",  "IT infrastructure allocation",                  "0.0%",  False),
        ("Inventory/WC % of Total Assets", 0.15,                 "%",  "Working capital proxy",                         "0.0%",  False),
    ]
    build_inputs(wb, THEME, company_name, fin, params_def)

    _build_rc_analysis(wb, THEME, fin, params)
    _build_rc_sensitivity(wb, THEME, fin)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
