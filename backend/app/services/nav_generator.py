"""
NAV (Net Asset Value) Excel Generator
Equity value = Adjusted Total Assets − Total Liabilities
Used for asset-heavy companies, holding companies, REITs, infrastructure firms.
"""
import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.services.excel_utils import (
    font, fill, thin_border, align, safe, cr,
    section_header, sub_header, label, value_cell, pct_cell,
    build_cover, build_inputs, build_results, build_sensitivity_table,
)

THEME = {
    "cover_bg":      "1A0A00",
    "primary":       "4A1900",
    "sub":           "7A2E00",
    "accent":        "D35400",
    "input_color":   "784212",
    "positive_fill": "FAE5D3",
    "positive_text": "6E2C00",
    "subtotal_fill": "E59866",
    "key_fill":      "FFF2CC",
}


# ─────────────────────────────────────────────────────────────────────────────
def _capm_ke(beta: float, rf: float = 0.072, erp: float = 0.055) -> float:
    return rf + beta * erp


def _nav_per_share(total_assets_cr, total_liabilities_cr, shares_cr,
                   adj_pct=0.0, goodwill_pct=0.05):
    """
    adj_pct: uniform additional adjustment % on top of the weighted fair-value
    goodwill_pct: goodwill as % of total_assets
    """
    cash_bv       = total_assets_cr * 0.05
    recv_bv       = total_assets_cr * 0.15
    inv_bv        = total_assets_cr * 0.10
    ppe_bv        = total_assets_cr * 0.40
    invest_bv     = total_assets_cr * 0.15
    other_bv      = total_assets_cr - cash_bv - recv_bv - inv_bv - ppe_bv - invest_bv

    adjs = {
        "cash":    1.00,
        "recv":    0.95,
        "inv":     0.85,
        "ppe":     1.10,
        "invest":  1.05,
        "other":   0.70,
    }
    fv = (cash_bv * adjs["cash"] +
          recv_bv * adjs["recv"] +
          inv_bv  * adjs["inv"] +
          ppe_bv  * adjs["ppe"] +
          invest_bv * adjs["invest"] +
          other_bv * adjs["other"])
    fv *= (1 + adj_pct)
    goodwill = total_assets_cr * goodwill_pct
    nav = fv - total_liabilities_cr - goodwill
    if shares_cr <= 0:
        return 0.0
    return nav / shares_cr   # per share in ₹  (Crore ₹ / Crore shares = ₹/share)


# ─────────────────────────────────────────────────────────────────────────────
def _build_nav_analysis(wb, theme, fin, params):
    ws = wb.create_sheet("NAV Analysis")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 22

    total_assets_cr     = cr(fin.get("total_assets"))
    total_liab_cr       = cr(fin.get("total_liabilities"))
    cash_cr             = cr(fin.get("cash"))
    shares              = fin.get("shares") or 0
    shares_cr           = shares / 1e7
    price               = safe(fin.get("price"), 0.0)
    bvps                = safe(fin.get("book_value_per_share"), 0.0)
    bv_total            = cr(fin.get("book_value_total"))

    # Asset estimates
    recv_cr    = total_assets_cr * 0.15
    inv_cr     = total_assets_cr * 0.10
    ppe_cr     = total_assets_cr * 0.40
    invest_cr  = total_assets_cr * 0.15
    other_cr   = total_assets_cr - cash_cr - recv_cr - inv_cr - ppe_cr - invest_cr
    if other_cr < 0:
        other_cr = total_assets_cr - recv_cr - inv_cr - ppe_cr - invest_cr

    book_equity = total_assets_cr - total_liab_cr
    goodwill_pct = params.get("goodwill_pct", 0.05)
    goodwill_cr  = total_assets_cr * goodwill_pct

    # Fair value adjustments
    ADJS = [
        ("Cash & Equivalents",     cash_cr,   1.00, "Full value assumed"),
        ("Net Receivables",        recv_cr,   0.95, "5% haircut for collection risk"),
        ("Inventory",              inv_cr,    0.85, "15% haircut for liquidation discount"),
        ("PP&E / Fixed Assets",    ppe_cr,    1.10, "10% replacement premium over book"),
        ("Investments / Other",    invest_cr, 1.05, "5% market premium"),
        ("Other Assets",           other_cr,  0.70, "30% haircut – misc/intangibles"),
    ]
    fv_rows = [(name, bv, adj, bv * adj, note) for name, bv, adj, note in ADJS]
    total_fv = sum(r[3] for r in fv_rows)
    nav_adj  = total_fv - total_liab_cr
    nav_ex_gw = nav_adj - goodwill_cr
    nav_ps   = nav_adj / shares_cr if shares_cr > 0 else 0
    nav_ex_gw_ps = nav_ex_gw / shares_cr if shares_cr > 0 else 0
    pnav     = price / nav_ps if nav_ps else 0
    premium  = (price - nav_ps) / nav_ps if nav_ps else 0

    row = 2
    t = ws.cell(row=row, column=2, value="NAV Analysis — Net Asset Value Model")
    t.font = font(bold=True, size=13)
    ws.merge_cells(f"B{row}:F{row}")

    # ── Section 1: Balance Sheet Summary ──────────────────────────────────────
    row += 2
    section_header(ws, row, 2, "▌ BALANCE SHEET SUMMARY", theme, span=5)
    row += 1

    bs_rows = [
        ("Total Assets (₹ Cr)",             total_assets_cr, "#,##0.0", False),
        ("  Cash & Equivalents (₹ Cr)",     cash_cr,         "#,##0.0", False),
        ("  Net Receivables (~15% assets)", recv_cr,         "#,##0.0", False),
        ("  Inventory (~10% assets)",       inv_cr,          "#,##0.0", False),
        ("  PP&E / Fixed Assets (~40%)",    ppe_cr,          "#,##0.0", False),
        ("  Investments / Other (~15%)",    invest_cr,       "#,##0.0", False),
        ("Total Liabilities (₹ Cr)",        total_liab_cr,   "#,##0.0", False),
        ("Book Equity (Assets − Liab) (₹ Cr)", book_equity, "#,##0.0", True),
    ]
    for lbl_txt, val, fmt, bold_ in bs_rows:
        c = ws.cell(row=row, column=2, value=lbl_txt)
        c.font = font(bold=bold_, size=9)
        c.border = thin_border()
        v = ws.cell(row=row, column=3, value=val)
        v.font = font(bold=bold_, size=9)
        v.number_format = fmt
        v.alignment = align(h="right")
        v.border = thin_border()
        if bold_:
            c.fill = fill(theme["positive_fill"])
            v.fill = fill(theme["positive_fill"])
        row += 1

    # ── Section 2: Fair Value Adjustments ─────────────────────────────────────
    row += 1
    section_header(ws, row, 2, "▌ FAIR VALUE ADJUSTMENTS", theme, span=5)
    row += 1

    hdr_cols = ["Asset Class", "Book Value (₹ Cr)", "Adjustment %", "Fair Value (₹ Cr)", "Notes"]
    for j, h in enumerate(hdr_cols):
        c = ws.cell(row=row, column=2 + j, value=h)
        c.font = font(bold=True, color="FFFFFF", size=9)
        c.fill = fill(theme["sub"])
        c.border = thin_border()
        c.alignment = align()
    row += 1

    for name, bv, adj, fv, note in fv_rows:
        ws.cell(row=row, column=2, value=name).font = font(size=9)
        ws.cell(row=row, column=2).border = thin_border()
        value_cell(ws, row, 3, bv,  "#,##0.0")
        pct_cell(ws,   row, 4, adj)
        value_cell(ws, row, 5, fv,  "#,##0.0")
        ws.cell(row=row, column=6, value=note).font = font(size=8, color="606060", italic=True)
        ws.cell(row=row, column=6).border = thin_border()
        row += 1

    # Total Fair Value of Assets (sum of fair values, not book values)
    ws.cell(row=row, column=2, value="Total Fair Value Assets").font = font(bold=True, size=9)
    ws.cell(row=row, column=2).border = thin_border()
    ws.cell(row=row, column=2).fill = fill(theme["subtotal_fill"])
    value_cell(ws, row, 3, total_fv, "#,##0.0", bold=True, bg=theme["subtotal_fill"])
    row += 1

    # Less: Total Liabilities — separate row per NAV formula (NAV = Assets − Liabilities)
    ws.cell(row=row, column=2, value="  Less: Total Liabilities (₹ Cr)").font = font(size=9)
    ws.cell(row=row, column=2).border = thin_border()
    value_cell(ws, row, 3, -total_liab_cr, "#,##0.0")
    row += 1

    # Net Asset Value = Fair Value Assets − Total Liabilities
    ws.cell(row=row, column=2, value="Net Asset Value (Adjusted) (₹ Cr)").font = font(bold=True, size=9)
    ws.cell(row=row, column=2).border = thin_border()
    ws.cell(row=row, column=2).fill = fill(theme["positive_fill"])
    value_cell(ws, row, 3, nav_adj, "#,##0.0", bold=True, bg=theme["positive_fill"])
    row += 1

    ws.cell(row=row, column=2, value=f"  Less: Goodwill & Intangibles (~{goodwill_pct*100:.0f}% of assets)").font = font(size=9)
    ws.cell(row=row, column=2).border = thin_border()
    value_cell(ws, row, 3, goodwill_cr, "#,##0.0")
    row += 1

    ws.cell(row=row, column=2, value="NAV ex-Goodwill (₹ Cr)").font = font(bold=True, size=9)
    ws.cell(row=row, column=2).border = thin_border()
    ws.cell(row=row, column=2).fill = fill(theme["key_fill"])
    value_cell(ws, row, 3, nav_ex_gw, "#,##0.0", bold=True, bg=theme["key_fill"])
    row += 1

    # ── Section 3: NAV Per Share ───────────────────────────────────────────────
    row += 1
    section_header(ws, row, 2, "▌ NAV PER SHARE CALCULATION", theme, span=5)
    row += 1

    ps_rows = [
        ("Net Asset Value (₹ Cr)",        nav_adj,     "#,##0.0",  False),
        ("Shares Outstanding (Cr)",        shares_cr,   "#,##0.00", False),
        ("★ NAV Per Share (₹)",            nav_ps,      "#,##0.00", True),
        ("Current Market Price (₹)",       price,       "#,##0.00", False),
        ("Premium/(Discount) to NAV",      premium,     "0.0%",     True),
        ("P/NAV Ratio",                    pnav,        "0.00x",    False),
    ]
    for lbl_txt, val, fmt, is_key in ps_rows:
        c = ws.cell(row=row, column=2, value=lbl_txt)
        c.font = font(bold=is_key, size=9)
        c.border = thin_border()
        if is_key:
            c.fill = fill(theme["key_fill"])
        v = ws.cell(row=row, column=3, value=val)
        v.number_format = fmt
        v.font = font(bold=is_key, size=9,
                      color=theme["accent"] if is_key else "000000")
        v.alignment = align(h="right")
        v.border = thin_border()
        if is_key:
            v.fill = fill(theme["key_fill"])
        row += 1

    # ── Section 4: NAV Bridge ─────────────────────────────────────────────────
    row += 1
    section_header(ws, row, 2, "▌ NAV BRIDGE (WATERFALL SUMMARY)", theme, span=5)
    row += 1

    fv_adj_total = total_fv - total_assets_cr   # total fair-value uplift/haircut
    gw_deduct    = -goodwill_cr

    bridge = [
        ("Book Value per Share (₹)",    bvps,                     "#,##0.00", False),
        ("  (+) Fair Value Adjustment (₹/share)",
            fv_adj_total / shares_cr if shares_cr > 0 else 0, "#,##0.00", False),
        ("  (−) Goodwill Deduction (₹/share)",
            gw_deduct / shares_cr if shares_cr > 0 else 0,    "#,##0.00", False),
        ("★ NAV per Share (₹)",          nav_ps,                   "#,##0.00", True),
        ("  NAV ex-Goodwill per Share (₹)", nav_ex_gw_ps,          "#,##0.00", False),
    ]
    for lbl_txt, val, fmt, is_key in bridge:
        c = ws.cell(row=row, column=2, value=lbl_txt)
        c.font = font(bold=is_key, size=9)
        c.border = thin_border()
        if is_key:
            c.fill = fill(theme["key_fill"])
        v = ws.cell(row=row, column=3, value=val)
        v.number_format = fmt
        v.font = font(bold=is_key, size=9,
                      color=theme["accent"] if is_key else "000000")
        v.alignment = align(h="right")
        v.border = thin_border()
        if is_key:
            v.fill = fill(theme["key_fill"])
        row += 1

    ws.freeze_panes = "C4"
    return ws


def _build_nav_sensitivity(wb, theme, fin):
    ws = wb.create_sheet("Sensitivity")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 22

    total_assets_cr  = cr(fin.get("total_assets"))
    total_liab_cr    = cr(fin.get("total_liabilities"))
    shares           = fin.get("shares") or 0
    shares_cr        = shares / 1e7
    price            = safe(fin.get("price"), 0.0)

    row_vals = [-0.15, -0.10, -0.05, 0.00, 0.05, 0.10, 0.15, 0.20]
    col_vals = [0.00, 0.03, 0.05, 0.07, 0.10, 0.12, 0.15]

    base_row_idx = 3   # 0.00 adjustment
    base_col_idx = 2   # 5% goodwill

    matrix = []
    for adj in row_vals:
        row_data = []
        for gw_pct in col_vals:
            ps = _nav_per_share(total_assets_cr, total_liab_cr, shares_cr, adj, gw_pct)
            row_data.append(round(ps, 2))
        matrix.append(row_data)

    row = 2
    t = ws.cell(row=row, column=2, value="NAV Sensitivity Analysis")
    t.font = font(bold=True, size=13)
    ws.merge_cells(f"B{row}:I{row}")
    row += 2

    build_sensitivity_table(
        ws, theme,
        row_label="Asset FV Adj %",
        col_label="Goodwill % of Assets",
        row_vals=row_vals,
        col_vals=col_vals,
        matrix=matrix,
        current_price=price,
        base_row_idx=base_row_idx,
        base_col_idx=base_col_idx,
        start_row=row,
    )
    return ws


# ─────────────────────────────────────────────────────────────────────────────
def generate_excel(fin: dict, symbol: str, company_name: str) -> bytes:
    wb = Workbook()

    total_assets_cr  = cr(fin.get("total_assets"))
    total_liab_cr    = cr(fin.get("total_liabilities"))
    shares           = fin.get("shares") or 0
    shares_cr        = shares / 1e7
    price            = safe(fin.get("price"), 0.0)
    mktcap_cr        = price * shares / 1e7

    goodwill_pct = 0.05
    beta         = safe(fin.get("beta"), 1.0)
    ke           = _capm_ke(beta)

    params = {"goodwill_pct": goodwill_pct}

    build_cover(
        wb, THEME, symbol, company_name,
        model_label="Net Asset Value (NAV) Model",
        model_desc="Equity value = Adjusted Fair Value of Assets − Total Liabilities",
        sheets_index=[
            (1, "Cover",               "Company overview & model index"),
            (2, "Inputs & Assumptions","Financial data & model parameters"),
            (3, "NAV Analysis",        "Balance sheet, fair value adj, NAV/share"),
            (4, "Sensitivity",         "Asset adj % × Goodwill % implied NAV/share"),
        ],
        meta_extra={"price": price, "mktcap_cr": mktcap_cr, "sector": safe(fin.get("sector"), "—")},
    )

    params_def = [
        ("Risk-Free Rate (10Y G-Sec)", 0.072,         "%",   "RBI 10Y bond yield",           "0.0%", True),
        ("Equity Risk Premium",        0.055,         "%",   "Historical India ERP",         "0.0%", False),
        ("Beta",                       beta,          "x",   "5Y monthly regression",        "0.00", False),
        ("Cost of Equity (CAPM)",      ke,            "%",   "Rf + β × ERP",                 "0.0%", True),
        ("Goodwill % of Assets",       goodwill_pct,  "%",   "Intangibles est. of total assets", "0.0%", True),
        ("Cash Adj Rate",              1.00,          "x",   "Book = Fair for cash",         "0.00", False),
        ("Receivables Adj Rate",       0.95,          "x",   "5% collection haircut",        "0.00", False),
        ("Inventory Adj Rate",         0.85,          "x",   "15% distressed sale haircut",  "0.00", False),
        ("PP&E Adj Rate",              1.10,          "x",   "Replacement cost premium",     "0.00", True),
        ("Investments Adj Rate",       1.05,          "x",   "Market premium",               "0.00", False),
        ("Other Assets Adj Rate",      0.70,          "x",   "Intangibles largely unrecov.", "0.00", False),
    ]
    build_inputs(wb, THEME, company_name, fin, params_def)

    _build_nav_analysis(wb, THEME, fin, params)
    _build_nav_sensitivity(wb, THEME, fin)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
