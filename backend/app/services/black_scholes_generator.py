"""
Black-Scholes Option Pricing Model
Prices European call and put options and computes the Greeks.
norm_cdf implemented via math.erf — no scipy dependency.
"""
import io
import math
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.services.excel_utils import (
    font, fill, thin_border, align, safe, cr,
    section_header, build_cover, build_inputs, build_sensitivity_table,
)

THEME = {
    "cover_bg":      "150025",
    "primary":       "2D0050",
    "sub":           "4A0080",
    "accent":        "9B59B6",
    "input_color":   "6C3483",
    "positive_fill": "F5EEF8",
    "positive_text": "6C3483",
    "subtotal_fill": "E8DAEF",
    "key_fill":      "FFF9C4",
}


def _col(c: int) -> str:
    return get_column_letter(c)


def _lbl(ws, row, text, indent=0, bold=False):
    c = ws.cell(row=row, column=2, value=("    " * indent) + text)
    c.font = font(bold=bold, size=9)
    c.border = thin_border()
    c.alignment = align()
    return c


def _val(ws, row, col, val, fmt="#,##0.00", bold=False, bg=None, color="000000"):
    c = ws.cell(row=row, column=col, value=val)
    c.font = font(bold=bold, size=9, color=color)
    c.number_format = fmt
    c.alignment = align(h="right")
    c.border = thin_border()
    if bg:
        c.fill = fill(bg)
    return c


def _hdr(ws, row, cols, theme):
    for i, txt in enumerate(cols):
        c = ws.cell(row=row, column=2 + i, value=txt)
        c.font = font(bold=True, color="FFFFFF", size=9)
        c.fill = fill(theme["primary"])
        c.alignment = align(h="center")
        c.border = thin_border()


def _norm_cdf(x: float) -> float:
    return 0.5 * (1.0 + math.erf(x / math.sqrt(2)))


def _norm_pdf(x: float) -> float:
    return math.exp(-x * x / 2.0) / math.sqrt(2 * math.pi)




def _bs_merton_call(S, K, T, r, sigma, q=0.0):
    """BSM call with continuous dividend yield q."""
    if T <= 0 or sigma <= 0:
        return max(S * math.exp(-q * T) - K * math.exp(-r * T), 0.0)
    d1 = (math.log(S / K) + (r - q + sigma ** 2 / 2) * T) / (sigma * math.sqrt(T))
    d2 = d1 - sigma * math.sqrt(T)
    return S * math.exp(-q * T) * _norm_cdf(d1) - K * math.exp(-r * T) * _norm_cdf(d2)


def _compute(fin: dict):
    price  = safe(fin.get("price"), 100.0)
    sigma  = safe(fin.get("volatility_annual"), 0.30)
    beta   = safe(fin.get("beta"), 1.0)
    shares = safe(fin.get("shares"), 1e8)
    q      = safe(fin.get("dividend_yield"), 0.0)   # continuous dividend yield

    S     = price
    K     = price          # ATM strike (more intuitive default than 1.05×)
    T     = 1.0
    r     = 0.071
    sigma = max(sigma, 0.01)

    sqrt_T  = math.sqrt(T)
    disc_q  = math.exp(-q * T)
    disc_r  = math.exp(-r * T)

    # Merton d1: includes (r − q + σ²/2)
    d1 = (math.log(S / K) + (r - q + sigma ** 2 / 2) * T) / (sigma * sqrt_T)
    d2 = d1 - sigma * sqrt_T

    Nd1  = _norm_cdf(d1);  Nd2  = _norm_cdf(d2)
    Nnd1 = _norm_cdf(-d1); Nnd2 = _norm_cdf(-d2)
    npd1 = _norm_pdf(d1)

    # Merton call and put prices
    call_price = S * disc_q * Nd1 - K * disc_r * Nd2
    put_price  = K * disc_r * Nnd2 - S * disc_q * Nnd1
    pcp        = S * disc_q - K * disc_r  # put-call parity: C − P = S·e^(−qT) − K·e^(−rT)

    # Greeks (Merton-adjusted)
    delta_call = disc_q * Nd1
    delta_put  = disc_q * (Nd1 - 1)
    gamma      = disc_q * npd1 / (S * sigma * sqrt_T) if S * sigma * sqrt_T else 0.0
    theta_call = (
        -S * disc_q * npd1 * sigma / (2 * sqrt_T)
        - r * K * disc_r * Nd2
        + q * S * disc_q * Nd1
    ) / 365
    theta_put  = (
        -S * disc_q * npd1 * sigma / (2 * sqrt_T)
        + r * K * disc_r * Nnd2
        - q * S * disc_q * Nnd1
    ) / 365
    vega       = S * disc_q * npd1 * sqrt_T / 100
    rho_call   = K * T * disc_r * Nd2 / 100

    # Intrinsic and time value
    intrinsic  = max(S - K, 0.0)
    time_val   = call_price - intrinsic

    # Scenario grid: 5 spot × 3 vol
    spot_levels = [S * f for f in (0.85, 0.92, 1.0, 1.08, 1.15)]
    vol_levels  = [sigma * f for f in (0.75, 1.0, 1.25)]
    grid = []
    for sv in spot_levels:
        row_g = [_bs_merton_call(sv, K, T, r, vv, q) for vv in vol_levels]
        grid.append(row_g)

    mktcap_cr = price * shares / 1e7

    return dict(
        price=price, S=S, K=K, T=T, r=r, sigma=sigma, q=q,
        d1=d1, d2=d2, Nd1=Nd1, Nd2=Nd2, Nnd1=Nnd1, Nnd2=Nnd2, npd1=npd1,
        call_price=call_price, put_price=put_price, pcp=pcp,
        delta_call=delta_call, delta_put=delta_put, gamma=gamma,
        theta_call=theta_call, theta_put=theta_put,
        vega=vega, rho_call=rho_call,
        intrinsic=intrinsic, time_val=time_val,
        spot_levels=spot_levels, vol_levels=vol_levels, grid=grid,
        beta=beta, mktcap_cr=mktcap_cr,
    )


def _build_analysis(wb, theme, fin, d):
    ws = wb.create_sheet("Black-Scholes Model")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 16
    for i in range(4, 10):
        ws.column_dimensions[_col(i)].width = 14

    row = 2
    t = ws.cell(row=row, column=2,
                value="Black-Scholes Option Pricing  —  European Call & Put")
    t.font = font(bold=True, size=13)
    ws.merge_cells(f"B{row}:{_col(9)}{row}")
    row += 2

    # ── OPTION PARAMETERS ──
    section_header(ws, row, 2, "▌ OPTION PARAMETERS", theme, span=3)
    row += 1
    params = [
        ("Spot Price  S  (₹)",          d["S"],     "#,##0.00"),
        ("Strike Price  K  (₹)",        d["K"],     "#,##0.00"),
        ("Time to Expiry  T  (years)",  d["T"],     "0.00"),
        ("Risk-Free Rate  r",           d["r"],     "0.0%"),
        ("Dividend Yield  q (annual)",  d["q"],     "0.00%"),
        ("Volatility  σ (annual)",      d["sigma"], "0.0%"),
        ("Option Type",                 "Call & Put (European)", "@"),
    ]
    for lbl_txt, val, fmt_ in params:
        _lbl(ws, row, lbl_txt, indent=1)
        _val(ws, row, 3, val, fmt_)
        row += 1
    row += 1

    # ── BLACK-SCHOLES CALCULATION ──
    section_header(ws, row, 2, "▌ BLACK-SCHOLES CALCULATION", theme, span=3)
    row += 1
    calc_rows = [
        ("d₁  =  [ln(S/K) + (r − q + σ²/2)·T] / (σ·√T)", d["d1"],         "0.0000",   False),
        ("d₂  =  d₁ − σ·√T",                               d["d2"],         "0.0000",   False),
        ("N(d₁)",                                           d["Nd1"],        "0.0000",   False),
        ("N(d₂)",                                           d["Nd2"],        "0.0000",   False),
        ("N(−d₁)",                                          d["Nnd1"],       "0.0000",   False),
        ("N(−d₂)",                                          d["Nnd2"],       "0.0000",   False),
        ("★ Call  =  S·e^(−qT)·N(d₁) − K·e^(−rT)·N(d₂)",  d["call_price"], "#,##0.00", True),
        ("★ Put   =  K·e^(−rT)·N(−d₂) − S·e^(−qT)·N(−d₁)",d["put_price"],  "#,##0.00", True),
        ("Put-Call Parity  =  S·e^(−qT) − K·e^(−rT)",      d["pcp"],        "#,##0.00", False),
    ]
    for lbl_txt, val, fmt_, is_key in calc_rows:
        _lbl(ws, row, lbl_txt, bold=is_key)
        _val(ws, row, 3, val, fmt_, bold=is_key,
             bg=theme["key_fill"] if is_key else None,
             color=theme["accent"] if is_key else "000000")
        row += 1
    row += 1

    # ── GREEKS ──
    section_header(ws, row, 2, "▌ OPTION GREEKS", theme, span=3)
    row += 1
    greeks = [
        ("Delta  (Call)  =  e^(−qT)·N(d₁)",                   d["delta_call"],  "0.0000"),
        ("Delta  (Put)   =  e^(−qT)·(N(d₁) − 1)",             d["delta_put"],   "0.0000"),
        ("Gamma  =  e^(−qT)·N'(d₁) / (S·σ·√T)",              d["gamma"],       "0.000000"),
        ("Theta  (Call)  per calendar day  (₹)",               d["theta_call"],  "#,##0.0000"),
        ("Theta  (Put)   per calendar day  (₹)",               d["theta_put"],   "#,##0.0000"),
        ("Vega   =  S·e^(−qT)·N'(d₁)·√T / 100",              d["vega"],        "#,##0.00"),
        ("Rho    (Call)  =  K·T·e^(−rT)·N(d₂)/100",          d["rho_call"],    "#,##0.0000"),
    ]
    for lbl_txt, val, fmt_ in greeks:
        _lbl(ws, row, lbl_txt, indent=1)
        _val(ws, row, 3, val, fmt_)
        row += 1
    row += 1

    # ── INTRINSIC vs TIME VALUE ──
    section_header(ws, row, 2, "▌ INTRINSIC vs TIME VALUE", theme, span=3)
    row += 1
    for lbl_txt, val, fmt_ in [
        ("Intrinsic Value = max(S−K, 0)  (₹)", d["intrinsic"], "#,##0.00"),
        ("Time Value = Call − Intrinsic  (₹)",  d["time_val"],  "#,##0.00"),
        ("Total Call Price  (₹)",               d["call_price"],"#,##0.00"),
    ]:
        _lbl(ws, row, lbl_txt, indent=1)
        _val(ws, row, 3, val, fmt_)
        row += 1
    row += 1

    # ── SCENARIO GRID ──
    vol_cols = [f"σ × {f:.2f}" for f in (0.75, 1.0, 1.25)]
    section_header(ws, row, 2, "▌ OPTION PRICE SCENARIO GRID  (Call Prices)", theme, span=5)
    row += 1
    _hdr(ws, row, ["Spot Level"] + vol_cols, theme)
    row += 1
    spot_labels = [f"S × {f:.2f}" for f in (0.85, 0.92, 1.00, 1.08, 1.15)]
    for si, sl in enumerate(spot_labels):
        _lbl(ws, row, sl, indent=1)
        for vi in range(3):
            _val(ws, row, 3 + vi, d["grid"][si][vi], "#,##0.00")
        row += 1

    ws.freeze_panes = "C4"


def _build_results_sheet(wb, theme, company_name, d):
    ws = wb.create_sheet("Results & Sensitivity")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 16

    row = 2
    section_header(ws, row, 2, "▌ RESULTS SUMMARY", theme, span=4)
    row += 1
    results = [
        ("Spot  S  (₹)",           d["S"],          "#,##0.00",  False),
        ("Strike  K  (₹)",         d["K"],          "#,##0.00",  False),
        ("Volatility  σ",          d["sigma"],      "0.0%",      False),
        ("★ Call Price  (₹)",      d["call_price"], "#,##0.00",  True),
        ("★ Put Price   (₹)",      d["put_price"],  "#,##0.00",  True),
        ("Delta (Call)",           d["delta_call"], "0.0000",    False),
        ("Gamma",                  d["gamma"],      "0.000000",  False),
        ("Vega",                   d["vega"],       "#,##0.00",  False),
    ]
    for lbl_txt, val, fmt_, is_key in results:
        lc = ws.cell(row=row, column=2, value=lbl_txt)
        lc.font = font(bold=is_key, size=9)
        lc.border = thin_border()
        if is_key:
            lc.fill = fill(theme["key_fill"])
        vc = ws.cell(row=row, column=3, value=val)
        vc.font = font(bold=is_key, size=9,
                       color=theme["accent"] if is_key else "000000")
        vc.number_format = fmt_
        vc.alignment = align(h="right")
        vc.border = thin_border()
        if is_key:
            vc.fill = fill(theme["key_fill"])
        row += 1

    # Sensitivity Table 1: Spot × Volatility → Call Price
    S      = d["S"]
    K      = d["K"]
    T      = d["T"]
    r      = d["r"]
    sigma  = d["sigma"]

    # 7 spot levels from 0.7*S to 1.3*S
    spot_vals = [round(S * (0.7 + 0.1 * i), 2) for i in range(7)]
    # 6 vol levels from 0.5*sigma to 1.5*sigma
    vol_vals  = [round(sigma * (0.5 + 0.2 * i), 4) for i in range(6)]

    base_si = 3   # S*1.0 is index 3
    base_vi = 2   # sigma*0.9 ≈ index 2; nearest to base sigma

    q = d["q"]
    matrix = []
    for sv in spot_vals:
        row_data = []
        for vv in vol_vals:
            cp = _bs_merton_call(sv, K, T, r, vv, q)
            row_data.append(round(cp, 2))
        matrix.append(row_data)

    build_sensitivity_table(
        ws, theme,
        row_label="Spot Price (₹)",
        col_label="Volatility σ",
        row_vals=spot_vals,
        col_vals=vol_vals,
        matrix=matrix,
        current_price=d["call_price"],   # use call price as colour proxy
        base_row_idx=base_si,
        base_col_idx=base_vi,
        start_row=row + 2,
        start_col=2,
        row_fmt="#,##0.00",
        col_fmt="0.0%",
    )


def generate_excel(fin: dict, symbol: str, company_name: str) -> bytes:
    d  = _compute(fin)
    wb = Workbook()

    build_cover(
        wb, THEME, symbol, company_name,
        model_label="Black-Scholes Option Pricing",
        model_desc="European call & put pricing with full Greeks. "
                   "norm_cdf via math.erf — no scipy dependency.",
        sheets_index=[
            (1, "Cover",                "Model overview & index"),
            (2, "Inputs & Assumptions", "Financial data & model parameters"),
            (3, "Black-Scholes Model",  "Pricing, Greeks & scenario grid"),
            (4, "Results & Sensitivity","Summary + Spot × Volatility sensitivity"),
        ],
        meta_extra={
            "price":     d["price"],
            "mktcap_cr": d["mktcap_cr"],
            "sector":    safe(fin.get("sector"), "—"),
        },
    )

    params_def = [
        ("Spot Price S (₹)",          d["S"],     "₹",  "Current market price",       "#,##0.00", True),
        ("Strike Price K (₹)",        d["K"],     "₹",  "ATM (= Spot) default",       "#,##0.00", True),
        ("Time to Expiry T (years)",  d["T"],     "yr", "1 year default",              "0.00",     False),
        ("Risk-Free Rate r",          d["r"],     "%",  "India 10Y GSec",              "0.0%",     False),
        ("Dividend Yield q (annual)", d["q"],     "%",  "From yfinance — Merton adj.", "0.00%",    False),
        ("Volatility σ (annual)",     d["sigma"], "%",  "Historical annualised vol",   "0.0%",     True),
        ("Option Type",               "European Call & Put", "", "Merton BSM",         "@",        False),
    ]

    build_inputs(wb, THEME, company_name, fin, params_def)
    _build_analysis(wb, THEME, fin, d)
    _build_results_sheet(wb, THEME, company_name, d)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
