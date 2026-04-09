"""
CFROI (Cash Flow Return on Investment) Excel Generator
Compares CFROI (OpCF / Asset Base) vs required return.
Positive spread → value creating.
"""

import io
import math
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.services.excel_utils import (
    font, fill, thin_border, align, safe, cr,
    section_header, sub_header, label, value_cell, pct_cell,
    build_cover, build_inputs, build_results, build_sensitivity_table,
)

THEME = {
    "cover_bg":      "001A26",
    "primary":       "003344",
    "sub":           "005566",
    "accent":        "1ABC9C",
    "input_color":   "0E6655",
    "positive_fill": "D1F2EB",
    "positive_text": "0E6655",
    "subtotal_fill": "A2D9CE",
    "key_fill":      "FFF9C4",
}

SHEETS_INDEX = [
    (1, "Cover",              "Model overview & metadata"),
    (2, "Inputs & Assumptions","Financial data & model parameters"),
    (3, "CFROI Analysis",     "CFROI calculation, projections & firm value bridge"),
    (4, "Sensitivity",        "Required Return × Asset Life sensitivity table"),
]


def _lv(ws, row, col, text, bold=False, indent=0):
    return label(ws, row, col, text, indent=indent, bold=bold, theme=THEME)


def _vc(ws, row, col, val, fmt="#,##0", bold=False, bg=None, color="000000"):
    return value_cell(ws, row, col, val, fmt=fmt, bold=bold, bg=bg, color=color)


def _pc(ws, row, col, val, bold=False, bg=None):
    return pct_cell(ws, row, col, val, bold=bold, bg=bg)


def _sh(ws, row, text, span=9):
    section_header(ws, row, 2, text, THEME, span=span)


def _key_row(ws, row, lbl_txt, val, fmt="#,##0"):
    _lv(ws, row, 2, lbl_txt, bold=True)
    ws.cell(row=row, column=2).fill = fill(THEME["key_fill"])
    _vc(ws, row, 3, val, fmt=fmt, bold=True,
        bg=THEME["key_fill"], color=THEME["input_color"])


def generate_excel(fin: dict, symbol: str, company_name: str) -> bytes:
    wb = Workbook()

    price      = safe(fin.get("price"), 0.0)
    shares_raw = safe(fin.get("shares"), 1e7)
    shares_cr  = shares_raw / 1e7
    mktcap_cr  = price * shares_cr
    sector     = fin.get("sector", "—")

    op_cf_cr   = cr(fin.get("operating_cf"))
    total_assets_cr = cr(fin.get("total_assets"))
    net_income_cr   = cr(fin.get("net_income"))
    ebitda_cr  = cr(fin.get("ebitda"))
    debt_cr    = cr(fin.get("total_debt"))
    cash_cr    = cr(fin.get("cash"))
    net_debt_cr = debt_cr - cash_cr
    rev_growth = safe(fin.get("revenue_growth"), 0.10)
    beta       = safe(fin.get("beta"), 1.0)
    roe        = safe(fin.get("roe"), 0.0)
    invested_capital_cr = cr(fin.get("invested_capital"))
    nopat_cr   = cr(fin.get("nopat"))
    equity_cr   = cr(fin.get("book_value_total"))  # shareholders' equity

    # Required return via CAPM
    rf  = 0.07
    erp = 0.055
    ke  = rf + beta * erp

    # CFROI
    # Numerator  : Operating Cash Flow (already post-tax, post-WC)
    # Denominator: Capital Employed = Total Assets − Current Liabilities
    #              ≈ Invested Capital (Equity + Financial Debt) when current liabilities unavailable
    gcf_cr = op_cf_cr
    # Capital Employed proxy: use invested_capital (equity + fin. debt) if available,
    # otherwise approximate as Total Assets − Non-financial liabilities
    #   = Total Assets − (Total Liabilities − Total Debt) = Equity + Debt = invested_capital
    gci_cr = invested_capital_cr if invested_capital_cr > 0 else total_assets_cr
    asset_life  = 10  # default years
    annual_dep_cr  = gci_cr / asset_life if asset_life > 0 else 0
    cash_conv_rate = op_cf_cr / net_income_cr if net_income_cr > 0 else 0

    cfroi           = gcf_cr / gci_cr if gci_cr > 0 else 0
    cfroi_spread    = cfroi - ke
    value_flag      = "Value Creating ✓" if cfroi > ke else "Value Destroying ✗"

    # ROA / ROIC
    roa  = net_income_cr / total_assets_cr if total_assets_cr > 0 else 0
    roic = nopat_cr / invested_capital_cr if invested_capital_cr > 0 else 0

    # ── Cover ──────────────────────────────────────────────────────────────────
    build_cover(
        wb, THEME, symbol, company_name,
        "Cash Flow Return on Investment (CFROI) Model",
        "Compares operating cash return on asset base vs cost of capital. "
        "Positive spread drives firm value creation.",
        SHEETS_INDEX,
        {"price": price, "mktcap_cr": mktcap_cr, "sector": sector},
    )

    # ── Inputs ─────────────────────────────────────────────────────────────────
    params_def = [
        ("Risk-Free Rate (rf)",         rf,         "%",    "India 10Y G-Sec proxy",                "0.0%", False),
        ("Equity Risk Premium (ERP)",   erp,        "%",    "Damodaran India ERP estimate",         "0.0%", False),
        ("Required Return (ke)",        ke,         "%",    "CAPM: rf + β × ERP",                   "0.0%", True),
        ("Asset Life (years)",          asset_life, "yrs",  "Average useful life of assets",        "#,##0", True),
        ("Terminal Growth Rate",        0.05,       "%",    "Long-run nominal GDP growth",          "0.0%", True),
        ("Near-term CF Growth",         min(rev_growth, 0.15), "%", "min(rev_growth, 15%)",         "0.0%", True),
    ]
    build_inputs(wb, THEME, company_name, fin, params_def)

    # ── CFROI Analysis sheet ───────────────────────────────────────────────────
    ws = wb.create_sheet("CFROI Analysis")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 16
    for ci in range(4, 10):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    t = ws.cell(row=2, column=2, value=f"{company_name} — CFROI Analysis")
    t.font = font(bold=True, size=13)
    ws.merge_cells("B2:J2")

    row = 4

    # ── SECTION 1: CFROI CALCULATION ──────────────────────────────────────────
    _sh(ws, row, "▌ CFROI CALCULATION")
    row += 1

    s1_rows = [
        ("Operating Cash Flow (₹ Cr)",                          op_cf_cr,   "#,##0.0", False),
        ("Capital Employed = Equity + Financial Debt (₹ Cr)",   gci_cr,     "#,##0.0", False),
        ("EBITDA (₹ Cr)",                                        ebitda_cr,  "#,##0.0", False),
        ("CFROI = Operating Cash Flow / Capital Employed",       cfroi,      "0.00%",   False),
        ("Required Return (ke from CAPM)",                       ke,         "0.00%",   False),
        ("★ CFROI Spread = CFROI − Required Return",             cfroi_spread, "0.00%", True),
    ]
    for lbl_txt, val, fmt, is_key in s1_rows:
        if is_key:
            _key_row(ws, row, lbl_txt, val, fmt)
        else:
            _lv(ws, row, 2, lbl_txt)
            _vc(ws, row, 3, val, fmt=fmt)
        row += 1

    # Value creation flag
    _lv(ws, row, 2, "Value Creation Assessment")
    c = ws.cell(row=row, column=3, value=value_flag)
    c.font = font(size=9, bold=True,
                  color=THEME["positive_text"] if cfroi > ke else "9C0006")
    c.fill = fill(THEME["positive_fill"] if cfroi > ke else "FFC7CE")
    c.border = thin_border()
    c.alignment = align(h="center")
    row += 2

    # ── SECTION 2: ASSET LIFE & DEPRECIATION ──────────────────────────────────
    _sh(ws, row, "▌ ASSET LIFE & DEPRECIATION ANALYSIS")
    row += 1

    s2_rows = [
        ("Asset Life (years)",                               asset_life,     "#,##0",    False),
        ("Annual Depreciation = Capital Employed / Life (₹ Cr)", annual_dep_cr, "#,##0.0", False),
        ("EBITDA (₹ Cr)",                                    ebitda_cr,      "#,##0.0",  False),
        ("Net Income (₹ Cr)",                                net_income_cr,  "#,##0.0",  False),
        ("Cash Conversion Rate = OpCF / Net Income",         cash_conv_rate, "#,##0.00", False),
    ]
    for lbl_txt, val, fmt, is_key in s2_rows:
        _lv(ws, row, 2, lbl_txt)
        _vc(ws, row, 3, val, fmt=fmt)
        row += 1

    row += 1

    # ── SECTION 3: CFROI TREND PROJECTION (5 years) ───────────────────────────
    _sh(ws, row, "▌ CFROI TREND PROJECTION (5 YEARS)", span=9)
    row += 1

    yr_cols   = list(range(4, 9))
    yr_labels = [f"Year {i}" for i in range(1, 6)]
    _lv(ws, row, 2, "Metric")
    for ci, lbl_txt in zip(yr_cols, yr_labels):
        c = ws.cell(row=row, column=ci, value=lbl_txt)
        c.font = font(bold=True, color="FFFFFF", size=9)
        c.fill = fill(THEME["sub"])
        c.alignment = align(h="center")
        c.border = thin_border()
    row += 1

    g_cf_initial  = min(rev_growth, 0.15)
    g_terminal    = 0.05
    g_asset_growth = 0.06

    proj_cf   = []
    proj_ab   = []
    proj_cfroi= []
    proj_req  = []
    proj_spr  = []
    proj_cer  = []
    proj_pv   = []

    for t_idx in range(5):
        g_cf = g_cf_initial + (g_terminal - g_cf_initial) * t_idx / 4
        cf_prev = proj_cf[-1]  if proj_cf  else gcf_cr
        ab_prev = proj_ab[-1]  if proj_ab  else gci_cr

        cf = cf_prev * (1 + g_cf)
        ab = ab_prev * (1 + g_asset_growth)
        cfr = cf / ab if ab > 0 else 0
        spr = cfr - ke
        cer = cf - ke * ab
        df  = 1 / (1 + ke) ** (t_idx + 1)
        pv  = cer * df

        proj_cf.append(cf)
        proj_ab.append(ab)
        proj_cfroi.append(cfr)
        proj_req.append(ke)
        proj_spr.append(spr)
        proj_cer.append(cer)
        proj_pv.append(pv)

    proj_metric_rows = [
        ("Gross Cash Flow (₹ Cr)",         proj_cf,    "#,##0.0", False),
        ("Gross Cash Investment (₹ Cr)",   proj_ab,    "#,##0.0", False),
        ("CFROI",                          proj_cfroi, "0.00%",   False),
        ("Required Return",                proj_req,   "0.00%",   False),
        ("Spread (CFROI − Required)",      proj_spr,   "0.00%",   False),
        ("Cash Economic Return (₹ Cr)",    proj_cer,   "#,##0.0", False),
        ("PV of Cash Economic Return (₹ Cr)", proj_pv, "#,##0.0", True),
    ]

    for lbl_txt, vals, fmt, is_key in proj_metric_rows:
        _lv(ws, row, 2, lbl_txt, bold=is_key)
        if is_key:
            ws.cell(row=row, column=2).fill = fill(THEME["positive_fill"])
        for ci, v in zip(yr_cols, vals):
            _vc(ws, row, ci, v, fmt=fmt,
                bg=THEME["positive_fill"] if is_key else None,
                color=THEME["positive_text"] if is_key else "000000")
        row += 1

    row += 1

    # ── SECTION 4: FIRM VALUE BRIDGE ──────────────────────────────────────────
    _sh(ws, row, "▌ FIRM VALUE BRIDGE")
    row += 1

    pv_cer_sum  = sum(proj_pv)
    last_cer    = proj_cer[-1]
    last_ab     = proj_ab[-1]
    term_cer    = last_cer * (1 + g_terminal)
    pv_term_cer = term_cer / (ke - g_terminal) * (1 / (1 + ke) ** 5) if ke > g_terminal else 0

    firm_ev_cr  = gci_cr + pv_cer_sum + pv_term_cer
    eq_val_cr   = firm_ev_cr - net_debt_cr
    implied_px  = eq_val_cr * 1e7 / shares_raw if shares_raw > 0 else 0
    upside      = (implied_px / price - 1) if price > 0 else 0

    bridge_rows = [
        ("Gross Cash Investment / Asset Base (₹ Cr)", gci_cr,  "#,##0.0", False),
        ("Add: PV(Economic Returns, 5Y) (₹ Cr)", pv_cer_sum,    "#,##0.0", False),
        ("Add: Terminal Value of Economic Returns (₹ Cr)", pv_term_cer, "#,##0.0", False),
        ("Enterprise Value (₹ Cr)",              firm_ev_cr,    "#,##0.0", False),
        ("Less: Net Debt (₹ Cr)",                net_debt_cr,   "#,##0.0", False),
        ("Equity Value (₹ Cr)",                  eq_val_cr,     "#,##0.0", False),
        ("Shares Outstanding (Cr)",              shares_cr,     "#,##0.00",False),
        ("★ Implied Share Price (₹)",            implied_px,    "#,##0.00",True),
        ("Current Market Price (₹)",             price,         "#,##0.00",False),
        ("★ Upside/(Downside)",                  upside,        "0.0%",    True),
    ]
    for lbl_txt, val, fmt, is_key in bridge_rows:
        if is_key:
            _key_row(ws, row, lbl_txt, val, fmt)
        else:
            _lv(ws, row, 2, lbl_txt)
            _vc(ws, row, 3, val, fmt=fmt)
        row += 1

    row += 1

    # ── SECTION 5: RETURN DECOMPOSITION ───────────────────────────────────────
    _sh(ws, row, "▌ RETURN DECOMPOSITION")
    row += 1

    decomp_rows = [
        ("CFROI = Operating CF / Capital Employed",    cfroi,   "0.00%"),
        ("ROIC (NOPAT / Invested Capital)",             roic,    "0.00%"),
        ("ROE (Net Income / Equity)",                   roe,     "0.00%"),
        ("ROA (Net Income / Total Assets)",             roa,     "0.00%"),
        ("Required Return (CAPM ke)",                   ke,      "0.00%"),
        ("CFROI Spread (CFROI − ke)",                   cfroi_spread, "0.00%"),
        ("Industry Benchmark CFROI (est.)",             0.12,    "0.00%"),
        ("Industry Benchmark ROIC (est.)",              0.14,    "0.00%"),
    ]
    for lbl_txt, val, fmt in decomp_rows:
        _lv(ws, row, 2, lbl_txt)
        bg = THEME["positive_fill"] if (fmt == "0.00%" and val > ke) else None
        _vc(ws, row, 3, val, fmt=fmt, bg=bg)
        row += 1

    ws.freeze_panes = "C5"

    # ── Sensitivity: Required Return × Asset Life ──────────────────────────────
    ws_s = wb.create_sheet("Sensitivity")
    ws_s.sheet_view.showGridLines = False
    ws_s.column_dimensions["A"].width = 2
    ws_s.column_dimensions["B"].width = 18

    t2 = ws_s.cell(row=2, column=2, value=f"{company_name} — CFROI Sensitivity Analysis")
    t2.font = font(bold=True, size=13)
    ws_s.merge_cells("B2:L2")

    req_vals   = [r/100 for r in range(9, 16)]   # 9%–15%
    life_vals  = [5, 7, 10, 15, 20]

    base_r_idx = min(range(len(req_vals)), key=lambda i: abs(req_vals[i] - ke))
    base_l_idx = min(range(len(life_vals)), key=lambda i: abs(life_vals[i] - asset_life))

    def _implied_price_cfroi(req, life):
        dep = gci_cr / life if life > 0 else 0
        pv_sum = 0
        cf = gcf_cr
        ab = gci_cr
        for t_idx in range(5):
            g_cf = g_cf_initial + (g_terminal - g_cf_initial) * t_idx / 4
            cf = cf * (1 + g_cf)
            ab = ab + cf - dep * (1 + g_cf * 0.5)
            ab = max(ab, gci_cr * 0.5)
            cer = cf - req * ab
            df  = 1 / (1 + req) ** (t_idx + 1)
            pv_sum += cer * df
        term_cer = proj_cer[-1] * (1 + g_terminal)
        pv_t = term_cer / (req - g_terminal) * (1 / (1 + req) ** 5) if req > g_terminal else 0
        fv   = gci_cr + pv_sum + pv_t
        eq   = fv - net_debt_cr
        return eq * 1e7 / shares_raw if shares_raw > 0 else 0

    matrix = [[_implied_price_cfroi(r, life) for life in life_vals] for r in req_vals]

    build_sensitivity_table(
        ws_s, THEME,
        "Required Return", "Asset Life (yrs)",
        req_vals, life_vals, matrix,
        price, base_r_idx, base_l_idx,
        start_row=4,
        row_fmt="0.0%", col_fmt="#,##0",
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
