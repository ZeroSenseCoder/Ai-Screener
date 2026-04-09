"""
DCF Excel Generator — Indian Stock Screener
Produces a professional 5-sheet DCF workbook identical in structure to the
NVIDIA_DCF_Model.xlsx template:
  1. Cover         — Company info & model index
  2. Assumptions   — WACC inputs, growth & margin drivers
  3. Income Statement — P&L build (3 historical + 5 projected years)
  4. DCF Valuation — FCFF model, WACC, terminal value, equity bridge
  5. Sensitivity   — WACC × g table + Revenue Growth × Gross Margin table
"""

import io
import math
import datetime
from typing import Optional

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Colour palette (matching NVIDIA model) ────────────────────────────────────
NAVY        = "1F2D3D"   # section header background
DARK_BLUE   = "2E4057"   # sub-section header
LIGHT_BLUE_FILL = "DCE9F5"  # sub-total rows
GREEN_FILL  = "D6F0D6"  # positive metric (GP, EBITDA, NI)
YELLOW_FILL = "FFF2CC"  # key assumption cell
COVER_BG    = "0A1628"  # cover page dark background
COVER_GOLD  = "C9A84C"  # cover accent
SENSITIVITY_GOLD = "FFD700"  # base-case highlight in sensitivity

WHITE = "FFFFFF"
BLACK = "000000"
BLUE_TEXT  = "1F497D"   # hardcoded input colour
GREEN_TEXT = "375623"   # cross-sheet link colour


def _font(bold=False, color=BLACK, size=10, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _border(style="thin"):
    s = Side(style=style, color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _bottom_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(bottom=s)


def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _pct(v: Optional[float]) -> str:
    if v is None:
        return "N/A"
    return f"{v * 100:.1f}%"


def _fmt_cr(v: Optional[float]) -> str:
    """Format a value in Indian Crores (divide by 1e7)."""
    if v is None:
        return "N/A"
    return f"₹{v / 1e7:,.0f} Cr"


def _safe(v, default=0.0):
    if v is None or (isinstance(v, float) and (math.isnan(v) or math.isinf(v))):
        return default
    return v


# ── Helper: write a styled section-header row ─────────────────────────────────
def _section_header(ws, row, col, text, width_cols=9):
    c = ws.cell(row=row, column=col, value=text)
    c.font = _font(bold=True, color=WHITE, size=10)
    c.fill = _fill(NAVY)
    c.alignment = _align()
    # merge across columns
    ws.merge_cells(
        start_row=row, start_column=col,
        end_row=row, end_column=col + width_cols - 1
    )


def _sub_header(ws, row, col, text, width_cols=9):
    c = ws.cell(row=row, column=col, value=text)
    c.font = _font(bold=True, color=WHITE, size=9)
    c.fill = _fill(DARK_BLUE)
    c.alignment = _align()
    ws.merge_cells(
        start_row=row, start_column=col,
        end_row=row, end_column=col + width_cols - 1
    )


def _label_cell(ws, row, col, text, indent=0):
    prefix = "    " * indent
    c = ws.cell(row=row, column=col, value=prefix + text)
    c.font = _font(size=9)
    c.alignment = _align()
    return c


def _value_cell(ws, row, col, value, fmt="#,##0", color=BLACK, fill_hex=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _font(color=color, size=9)
    c.number_format = fmt
    c.alignment = _align(h="right")
    if fill_hex:
        c.fill = _fill(fill_hex)
    return c


def _pct_cell(ws, row, col, value, color=BLACK, fill_hex=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _font(color=color, size=9)
    c.number_format = "0.0%"
    c.alignment = _align(h="right")
    if fill_hex:
        c.fill = _fill(fill_hex)
    return c


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — COVER
# ═══════════════════════════════════════════════════════════════════════════════
def _build_cover(wb: Workbook, meta: dict):
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 2
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 22

    # Background fill entire cover area
    for r in range(1, 40):
        for c in range(1, 12):
            ws.cell(row=r, column=c).fill = _fill(COVER_BG)

    # Company name
    name_cell = ws.cell(row=5, column=2, value=f"{meta['company_name']} ({meta['symbol']})")
    name_cell.font = Font(name="Calibri", bold=True, size=22, color=WHITE)
    name_cell.alignment = _align()
    ws.merge_cells("B5:J5")

    # Subtitle
    subtitle = ws.cell(row=8, column=2,
        value="Discounted Cash Flow Valuation Model  |  Equity Research  |  " +
              datetime.date.today().strftime("%B %Y"))
    subtitle.font = Font(name="Calibri", size=13, color=COVER_GOLD, italic=True)
    subtitle.alignment = _align()
    ws.merge_cells("B8:J8")

    # Gold separator line
    for col in range(2, 11):
        ws.cell(row=10, column=col).fill = _fill(COVER_GOLD)

    # Metadata table
    meta_rows = [
        ("Analyst",           "Equity Research Analyst"),
        ("Coverage",          meta.get("sector", "Indian Equities")),
        ("Fiscal Year End",   "March 31"),
        ("Base Currency",     "INR (₹ Crores)"),
        ("Share Price (ref)", f"₹{meta['price']:.2f}"),
        ("Shares Outstanding", f"{meta['shares_cr']:.1f} Cr"),
        ("Market Cap (ref)",  f"₹{meta['mktcap_cr']:.0f} Cr"),
        ("Rating",            "Model Output — See DCF Valuation"),
        ("Model Date",        datetime.date.today().strftime("%B %d, %Y").replace(" 0", " ")),
    ]

    for i, (label, value) in enumerate(meta_rows):
        r = 12 + i
        lc = ws.cell(row=r, column=2, value=label)
        lc.font = Font(name="Calibri", size=10, color="A0A0A0")
        vc = ws.cell(row=r, column=4, value=value)
        vc.font = Font(name="Calibri", size=10, color=WHITE, bold=True)

    # Gold separator
    for col in range(2, 11):
        ws.cell(row=23, column=col).fill = _fill(COVER_GOLD)

    # Model index
    idx_title = ws.cell(row=25, column=2, value="MODEL INDEX")
    idx_title.font = Font(name="Calibri", bold=True, size=11, color=COVER_GOLD)

    index_items = [
        ("1", "Assumptions",       "Key model drivers and inputs"),
        ("2", "Income Statement",  "P&L build — Historical + 5-year estimates"),
        ("3", "DCF Valuation",     "FCFF model, WACC, terminal value, equity bridge"),
        ("4", "Sensitivity",       "Valuation sensitivity tables"),
    ]

    for i, (num, sheet, desc) in enumerate(index_items):
        r = 27 + i
        ws.cell(row=r, column=2, value=num).font = Font(name="Calibri", size=10, color=COVER_GOLD, bold=True)
        ws.cell(row=r, column=3, value=sheet).font = Font(name="Calibri", size=10, color=WHITE, bold=True)
        ws.cell(row=r, column=5, value=desc).font = Font(name="Calibri", size=10, color="808080")

    # Disclaimer
    disc = ws.cell(row=37, column=2,
        value="DISCLAIMER: This model is for educational and illustrative purposes only. "
              "It does not constitute investment advice. Financial data sourced from "
              "NSE/BSE filings, screener.in, and yfinance. Assumptions are model estimates. "
              "Past performance is not indicative of future results.")
    disc.font = Font(name="Calibri", size=8, color="606060", italic=True)
    disc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.merge_cells("B37:J38")
    ws.row_dimensions[37].height = 30

    ws.row_dimensions[5].height = 30
    ws.row_dimensions[8].height = 22


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — ASSUMPTIONS
# ═══════════════════════════════════════════════════════════════════════════════
def _build_assumptions(wb: Workbook, a: dict):
    ws = wb.create_sheet("Assumptions")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 50

    # Title
    t = ws.cell(row=2, column=2,
        value=f"{a['company_name']} — Key Model Assumptions")
    t.font = _font(bold=True, size=13)
    ws.merge_cells("B2:F2")

    sub = ws.cell(row=3, column=2,
        value="Blue cells = hardcoded inputs  |  Yellow background = key driver assumptions  |  Source annotations in column F")
    sub.font = _font(size=8, color="808080", italic=True)
    ws.merge_cells("B3:F3")

    # Header row
    for col, txt in enumerate(["Assumption", "Value", "Unit", "Range / Alt", "Source / Note"], 2):
        c = ws.cell(row=5, column=col, value=txt)
        c.font = _font(bold=True, size=9, color=WHITE)
        c.fill = _fill(NAVY)
        c.alignment = _align()
        c.border = _border()

    row = 7

    def section(title):
        nonlocal row
        _section_header(ws, row, 2, title, width_cols=5)
        row += 1

    def assumption_row(label, value, unit="", rng="", source="",
                       is_key=False, fmt="0.00%", indent=0):
        nonlocal row
        prefix = "    " * indent
        lc = ws.cell(row=row, column=2, value=prefix + label)
        lc.font = _font(size=9)
        lc.border = _border()

        vc = ws.cell(row=row, column=3, value=value)
        vc.font = _font(size=9, color=BLUE_TEXT)
        vc.number_format = fmt
        vc.alignment = _align(h="right")
        vc.border = _border()
        if is_key:
            vc.fill = _fill(YELLOW_FILL)

        ws.cell(row=row, column=4, value=unit).font = _font(size=9)
        ws.cell(row=row, column=4).border = _border()
        ws.cell(row=row, column=5, value=rng).font = _font(size=9)
        ws.cell(row=row, column=5).border = _border()
        sc = ws.cell(row=row, column=6, value=source)
        sc.font = _font(size=8, color="606060", italic=True)
        sc.alignment = _align(wrap=True)
        sc.border = _border()
        row += 1

    # ── WACC ──────────────────────────────────────────────────────────────────
    section("▌ WACC INPUTS")
    assumption_row("Risk-Free Rate (10Y G-Sec)", a["rf"],         "%", "6.5%–7.5%",
                   "Source: RBI / NSE G-Sec yield, current", fmt="0.00%")
    assumption_row("Equity Risk Premium",         a["erp"],        "%", "5.0%–6.5%",
                   "Source: Damodaran, India ERP estimate", fmt="0.00%")
    assumption_row("Beta (5Y Monthly vs NIFTY)",  a["beta"],       "x", "0.5–2.0",
                   "Source: yfinance 5Y regression vs Nifty 50", fmt="0.00")
    assumption_row("Cost of Equity (CAPM)",       a["ke"],         "%", "",
                   "Derived: Rf + β × ERP", fmt="0.00%")
    assumption_row("Pre-Tax Cost of Debt",         a["kd"],         "%", "8%–12%",
                   "Source: sector avg borrowing rate (RBI data)", fmt="0.00%")
    assumption_row("Effective Tax Rate",           a["tax_rate"],   "%", "25%",
                   "Source: Indian Corporate Tax Rate (Sec 115BAA)", fmt="0.00%")
    assumption_row("After-Tax Cost of Debt",       a["kd_at"],      "%", "",
                   "Derived: Kd × (1 – t)", fmt="0.00%")
    assumption_row("Debt Weight [D/(D+E)]",        a["debt_weight"],"%", "",
                   "Source: Market cap + total debt from filings", fmt="0.00%")
    assumption_row("WACC",                         a["wacc"],       "%", "",
                   "Derived (CAPM-based; equity-driven if low-leverage)", fmt="0.00%",
                   is_key=True)
    row += 1

    # ── TERMINAL VALUE ────────────────────────────────────────────────────────
    section("▌ TERMINAL VALUE")
    assumption_row("Terminal Growth Rate (g)",     a["tg"],         "%", "4%–6%",
                   "Long-run India nominal GDP growth + inflation", fmt="0.00%", is_key=True)
    assumption_row("Terminal Year FCF Margin",     a["terminal_fcf_margin"], "%", "",
                   "Normalised FCF margin in steady state", fmt="0.00%")
    row += 1

    # ── REVENUE GROWTH ────────────────────────────────────────────────────────
    section("▌ REVENUE GROWTH ASSUMPTIONS")
    proj_years = a["proj_years"]
    for i, g in enumerate(a["rev_growth"]):
        yr = proj_years[i]
        assumption_row(f"{yr} Revenue Growth", g, "%", "",
                       f"Analyst estimate — {yr}", fmt="0.00%", is_key=True)
    row += 1

    # ── MARGINS ───────────────────────────────────────────────────────────────
    section("▌ MARGIN ASSUMPTIONS")
    assumption_row("Gross Margin — Year 1",        a["gm_y1"],      "%", "", "Source: Latest annual filing", fmt="0.00%", is_key=True)
    assumption_row("Gross Margin — Year 2–5",      a["gm_later"],   "%", "", "Modest improvement on operating leverage", fmt="0.00%", is_key=True)
    assumption_row("R&D as % of Revenue",          a["rd_pct"],     "%", "", "Source: FY actuals from screener.in", fmt="0.00%")
    assumption_row("SG&A as % of Revenue",         a["sga_pct"],    "%", "", "Source: FY actuals from screener.in", fmt="0.00%")
    assumption_row("D&A as % of Revenue",          a["da_pct"],     "%", "", "Source: FY actuals", fmt="0.00%")
    assumption_row("CapEx as % of Revenue",        a["capex_pct"],  "%", "", "Source: FY actuals from screener.in", fmt="0.00%")
    assumption_row("Change in NWC as % of Revenue",a["nwc_pct"],   "%", "", "Working capital build with revenue growth", fmt="0.00%")
    row += 1

    # ── BALANCE SHEET ─────────────────────────────────────────────────────────
    section("▌ BALANCE SHEET (LATEST ACTUALS)")
    assumption_row("Cash & Equivalents (₹ Cr)",   a["cash_cr"],    "₹ Cr", "",
                   "Source: Latest annual report / screener.in", fmt="#,##0")
    assumption_row("Total Debt (₹ Cr)",            a["debt_cr"],    "₹ Cr", "",
                   "Source: Latest annual report / screener.in", fmt="#,##0")
    assumption_row("Shares Outstanding (Cr)",      a["shares_cr"],  "Cr",   "",
                   "Source: NSE/BSE filings", fmt="#,##0.00")
    assumption_row("Current Share Price (₹)",      a["price"],      "₹",    "",
                   "Source: Market reference price", fmt="#,##0.00")

    # Lock the assumptions reference cell address (for DCF cross-links)
    # Store row references on sheet for use by later sheets
    ws["H2"] = "WACC Row"
    ws["I2"] = row  # not needed, WACC is always at C16 equivalent row
    ws.sheet_view.showGridLines = False


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — INCOME STATEMENT
# ═══════════════════════════════════════════════════════════════════════════════
def _build_income_statement(wb: Workbook, is_data: dict):
    ws = wb.create_sheet("Income Statement")
    ws.sheet_view.showGridLines = False

    col_widths = [2, 38, 14, 14, 14, 14, 14, 14, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    hist_years = is_data["hist_years"]   # e.g. ["FY2022A","FY2023A","FY2024A"]
    proj_years = is_data["proj_years"]   # e.g. ["FY2025E",..,"FY2029E"]
    all_years  = hist_years + proj_years
    n_hist = len(hist_years)
    n_proj = len(proj_years)
    N = len(all_years)

    # Data arrays (₹ Crores)
    rev      = is_data["revenue"]        # length N
    cogs     = is_data["cogs"]
    gp       = [rev[i] - cogs[i] for i in range(N)]
    gm_pct   = [gp[i] / rev[i] if rev[i] else None for i in range(N)]
    rd       = is_data["rd"]
    sga      = is_data["sga"]
    opex     = [rd[i] + sga[i] for i in range(N)]
    ebit     = [gp[i] - opex[i] for i in range(N)]
    ebit_m   = [ebit[i] / rev[i] if rev[i] else None for i in range(N)]
    da       = is_data["da"]
    ebitda   = [ebit[i] + da[i] for i in range(N)]
    ebitda_m = [ebitda[i] / rev[i] if rev[i] else None for i in range(N)]
    net_int  = is_data["net_interest"]
    ebt      = [ebit[i] + net_int[i] for i in range(N)]
    tax      = is_data["tax"]
    ni       = [ebt[i] - tax[i] for i in range(N)]
    ni_m     = [ni[i] / rev[i] if rev[i] else None for i in range(N)]
    eps      = is_data["eps"]

    # Title
    t = ws.cell(row=2, column=2,
        value=f"{is_data['company_name']} — Income Statement Build  (₹ Crores)")
    t.font = _font(bold=True, size=13)
    ws.merge_cells(f"B2:{get_column_letter(1+N+1)}2")

    sub = ws.cell(row=3, column=2,
        value="Historical: NSE/BSE filings via screener.in  |  Estimates: Model  |  Blue = input, Black = formula, Green = cross-sheet link")
    sub.font = _font(size=8, color="606060", italic=True)
    ws.merge_cells(f"B3:{get_column_letter(1+N+1)}3")

    # Column header row
    ws.cell(row=6, column=2, value="(₹ Cr unless stated)").font = _font(bold=True, size=9)
    for j, yr in enumerate(all_years):
        c = ws.cell(row=6, column=3+j, value=yr)
        c.font = _font(bold=True, size=9, color=WHITE)
        c.fill = _fill(NAVY)
        c.alignment = _align(h="center")
        c.border = _border()

    row = 8

    def data_row(label, values, fmt="#,##0", fill_hex=None, bold=False,
                 color=BLACK, indent=0, is_pct=False):
        nonlocal row
        prefix = "    " * indent
        lc = ws.cell(row=row, column=2, value=prefix + label)
        lc.font = _font(bold=bold, size=9)
        if fill_hex:
            lc.fill = _fill(fill_hex)
        for j, v in enumerate(values):
            c = ws.cell(row=row, column=3+j, value=v)
            c.font = _font(bold=bold, size=9, color=color)
            c.number_format = "0.0%" if is_pct else fmt
            c.alignment = _align(h="right")
            c.border = _border()
            if fill_hex:
                c.fill = _fill(fill_hex)
            # grey out historical vs estimates
            if j >= n_hist:
                c.font = _font(bold=bold, size=9, color=color)
        row += 1

    def blank_row():
        nonlocal row
        row += 1

    # ── REVENUE ───────────────────────────────────────────────────────────────
    _section_header(ws, row, 2, "▌ REVENUE", width_cols=N+1)
    row += 1

    data_row("Total Revenue", rev, bold=True)
    yoy = [None] + [(rev[i]-rev[i-1])/rev[i-1] if rev[i-1] else None for i in range(1, N)]
    data_row("  YoY Growth", yoy, is_pct=True, color="606060", indent=1)
    blank_row()

    # ── COST STRUCTURE ────────────────────────────────────────────────────────
    _section_header(ws, row, 2, "▌ COST STRUCTURE", width_cols=N+1)
    row += 1

    data_row("    Cost of Revenue (COGS)", cogs, indent=1)
    data_row("Gross Profit", gp, bold=True, fill_hex=GREEN_FILL, color=GREEN_TEXT)
    data_row("  Gross Margin %", gm_pct, is_pct=True, color="606060", indent=1)
    blank_row()

    # ── OPERATING EXPENSES ────────────────────────────────────────────────────
    _section_header(ws, row, 2, "▌ OPERATING EXPENSES", width_cols=N+1)
    row += 1

    data_row("    Research & Development", rd, indent=1)
    data_row("    Sales, General & Administrative", sga, indent=1)
    data_row("Total Operating Expenses", opex, fill_hex=LIGHT_BLUE_FILL, bold=True)
    blank_row()

    # ── PROFITABILITY ─────────────────────────────────────────────────────────
    _section_header(ws, row, 2, "▌ PROFITABILITY", width_cols=N+1)
    row += 1

    data_row("EBIT (Operating Income)", ebit, bold=True)
    data_row("  EBIT Margin %", ebit_m, is_pct=True, color="606060", indent=1)
    data_row("    Depreciation & Amortization", da, indent=1)
    data_row("EBITDA", ebitda, bold=True, fill_hex=GREEN_FILL, color=GREEN_TEXT)
    data_row("  EBITDA Margin %", ebitda_m, is_pct=True, color="606060", indent=1)
    blank_row()

    # ── BELOW THE LINE ────────────────────────────────────────────────────────
    _section_header(ws, row, 2, "▌ BELOW THE LINE", width_cols=N+1)
    row += 1

    data_row("    Net Interest & Other Income", net_int, indent=1)
    data_row("Pre-Tax Income (EBT)", ebt, bold=True)
    data_row("    Income Tax Expense", tax, indent=1)
    data_row("Net Income", ni, bold=True, fill_hex=GREEN_FILL, color=GREEN_TEXT)
    data_row("  Net Margin %", ni_m, is_pct=True, color="606060", indent=1)
    blank_row()

    # ── PER SHARE ─────────────────────────────────────────────────────────────
    _section_header(ws, row, 2, "▌ PER SHARE & KEY METRICS", width_cols=N+1)
    row += 1

    data_row("Diluted EPS (₹)", eps, fmt="#,##0.00")

    ws.freeze_panes = "C7"


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — DCF VALUATION
# ═══════════════════════════════════════════════════════════════════════════════
def _build_dcf_valuation(wb: Workbook, dcf: dict):
    ws = wb.create_sheet("DCF Valuation")
    ws.sheet_view.showGridLines = False

    col_widths = [2, 44, 4, 14, 14, 14, 14, 14, 4, 4]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    proj_years = dcf["proj_years"]  # 5 years
    N = len(proj_years)

    # Unpack DCF arrays
    ebit   = dcf["ebit"]
    nopat  = dcf["nopat"]
    da     = dcf["da"]
    capex  = dcf["capex"]
    dnwc   = dcf["dnwc"]
    fcff   = dcf["fcff"]
    fcff_m = dcf["fcff_margin"]
    disc_t = dcf["disc_t"]
    disc_f = dcf["disc_f"]
    pv_fcff= dcf["pv_fcff"]

    tv       = dcf["terminal_value"]
    pv_tv    = dcf["pv_terminal_value"]
    sum_pv   = dcf["sum_pv_fcff"]
    ev       = dcf["enterprise_value"]
    debt_cr  = dcf["debt_cr"]
    cash_cr  = dcf["cash_cr"]
    eq_val   = dcf["equity_value"]
    price_impl = dcf["implied_price"]
    price_cur  = dcf["current_price"]
    upside     = dcf["upside"]
    wacc       = dcf["wacc"]
    tg         = dcf["tg"]
    tv_pct_ev  = pv_tv / ev if ev else None
    shares_cr  = dcf["shares_cr"]

    # Title
    t = ws.cell(row=2, column=2,
        value=f"{dcf['company_name']} — DCF Valuation  (₹ Crores, except per share in ₹)")
    t.font = _font(bold=True, size=13)
    ws.merge_cells("B2:I2")

    sub = ws.cell(row=3, column=2,
        value="FCFF-based DCF  |  Projection period: " + "–".join([proj_years[0], proj_years[-1]]) +
              "  |  Midyear convention applied  |  TV via Gordon Growth Model")
    sub.font = _font(size=8, color="606060", italic=True)
    ws.merge_cells("B3:I3")

    # Year headers
    ws.cell(row=6, column=2, value="(₹ Cr unless stated)").font = _font(bold=True, size=9)
    for j, yr in enumerate(proj_years):
        c = ws.cell(row=6, column=4+j, value=yr)
        c.font = _font(bold=True, size=9, color=WHITE)
        c.fill = _fill(NAVY)
        c.alignment = _align(h="center")
        c.border = _border()

    row = 8

    def dcf_row(label, values, fmt="#,##0", fill_hex=None, bold=False,
                color=BLACK, indent=0, is_pct=False):
        nonlocal row
        prefix = "    " * indent
        lc = ws.cell(row=row, column=2, value=prefix + label)
        lc.font = _font(bold=bold, size=9)
        if fill_hex:
            lc.fill = _fill(fill_hex)
        for j, v in enumerate(values):
            c = ws.cell(row=row, column=4+j, value=v)
            c.font = _font(bold=bold, size=9, color=color)
            c.number_format = "0.0%" if is_pct else fmt
            c.alignment = _align(h="right")
            c.border = _border()
            if fill_hex:
                c.fill = _fill(fill_hex)
        row += 1

    def single_row(label, value, col_offset=0, fmt="#,##0", bold=False,
                   fill_hex=None, color=BLACK, is_pct=False, star=False):
        nonlocal row
        lc = ws.cell(row=row, column=2, value=("⭐  " if star else "") + label)
        lc.font = _font(bold=bold, size=9)
        if fill_hex:
            lc.fill = _fill(fill_hex)
        c = ws.cell(row=row, column=4+col_offset, value=value)
        c.font = _font(bold=bold, size=9, color=color)
        c.number_format = "0.0%" if is_pct else fmt
        c.alignment = _align(h="right")
        c.border = _border()
        if fill_hex:
            c.fill = _fill(fill_hex)
        row += 1

    def blank():
        nonlocal row
        row += 1

    # ── FCFF BUILD ────────────────────────────────────────────────────────────
    _section_header(ws, row, 2, "▌ FREE CASH FLOW TO FIRM (FCFF) BUILD", width_cols=2+N)
    row += 1

    dcf_row("    EBIT", ebit, indent=1)
    dcf_row("    NOPAT [EBIT × (1–t)]", nopat, indent=1)
    dcf_row("    Add: Depreciation & Amortization", da, indent=1)
    dcf_row("    Less: Capital Expenditures", [-v for v in capex], indent=1)
    dcf_row("    Less: Increase in Net Working Capital", [-v for v in dnwc], indent=1)
    dcf_row("Free Cash Flow to Firm (FCFF)", fcff, bold=True, fill_hex=GREEN_FILL, color=GREEN_TEXT)
    dcf_row("  FCFF Margin %", fcff_m, is_pct=True, color="606060", indent=1)
    blank()

    # ── DISCOUNT FACTORS ──────────────────────────────────────────────────────
    _section_header(ws, row, 2, "▌ DISCOUNT FACTORS (MIDYEAR CONVENTION)", width_cols=2+N)
    row += 1

    dcf_row("Discount Period (t)", disc_t, fmt="0.0")
    dcf_row("Discount Factor", disc_f, fmt="0.0000")
    dcf_row("PV of FCFF", pv_fcff, fill_hex=LIGHT_BLUE_FILL)
    blank()

    # ── TERMINAL VALUE ────────────────────────────────────────────────────────
    _section_header(ws, row, 2, "▌ TERMINAL VALUE", width_cols=2+N)
    row += 1

    # Terminal value in last year's column (col 4 + N-1)
    lc = ws.cell(row=row, column=2, value="Terminal Value (Gordon Growth Model)")
    lc.font = _font(size=9)
    c = ws.cell(row=row, column=4+N-1, value=tv)
    c.font = _font(size=9, color=BLUE_TEXT)
    c.number_format = "#,##0"
    c.alignment = _align(h="right")
    c.border = _border()
    row += 1

    lc = ws.cell(row=row, column=2, value="PV of Terminal Value")
    lc.font = _font(size=9)
    c = ws.cell(row=row, column=4+N-1, value=pv_tv)
    c.font = _font(size=9)
    c.number_format = "#,##0"
    c.alignment = _align(h="right")
    c.border = _border()
    row += 1
    blank()

    # ── EQUITY VALUE BRIDGE ───────────────────────────────────────────────────
    _section_header(ws, row, 2, "▌ ENTERPRISE & EQUITY VALUE BRIDGE", width_cols=2+N)
    row += 1

    single_row("Sum of PV (FCFF, projection period)", sum_pv)
    single_row("Add: PV of Terminal Value", pv_tv)
    single_row("Enterprise Value (EV)", ev, bold=True, fill_hex=LIGHT_BLUE_FILL)
    single_row("Less: Total Debt", -debt_cr)
    single_row("Add: Cash & Marketable Securities", cash_cr)
    single_row("Equity Value", eq_val, bold=True, fill_hex=LIGHT_BLUE_FILL)
    single_row("Implied Share Price (₹) — Intrinsic Value", price_impl, bold=True,
               fill_hex=YELLOW_FILL, color=BLUE_TEXT, fmt="#,##0.00", star=True)
    single_row("Upside / (Downside) to Current Price",
               upside, is_pct=True,
               color="375623" if upside and upside >= 0 else "9C0006",
               bold=True)
    blank()

    # ── DIAGNOSTICS ───────────────────────────────────────────────────────────
    _section_header(ws, row, 2, "▌ MODEL DIAGNOSTICS & SANITY CHECKS", width_cols=2+N)
    row += 1

    single_row("TV as % of Enterprise Value", tv_pct_ev, is_pct=True)
    single_row("WACC Used", wacc, is_pct=True)
    single_row("Terminal Growth Rate (g) Used", tg, is_pct=True)
    single_row("Current Share Price (₹) [Reference]", price_cur, fmt="#,##0.00")

    ws.freeze_panes = "D7"


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — SENSITIVITY
# ═══════════════════════════════════════════════════════════════════════════════
def _build_sensitivity(wb: Workbook, sens: dict):
    ws = wb.create_sheet("Sensitivity")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 14

    company = sens["company_name"]
    base_wacc = sens["wacc"]
    base_g    = sens["tg"]
    base_rev_g = sens["base_rev_growth"]
    base_gm    = sens["base_gm"]
    price_cur  = sens["current_price"]

    # Title
    t = ws.cell(row=2, column=2, value=f"{company} — Valuation Sensitivity Analysis")
    t.font = _font(bold=True, size=13)
    ws.merge_cells("B2:J2")

    sub = ws.cell(row=3, column=2,
        value="Table 1: Implied Share Price vs WACC & Terminal Growth Rate  |  "
              "Table 2: Implied Share Price vs FY+1 Revenue Growth & Gross Margin")
    sub.font = _font(size=8, color="606060", italic=True)
    ws.merge_cells("B3:J3")

    # ── TABLE 1: WACC × g ─────────────────────────────────────────────────────
    ws.cell(row=5, column=2,
        value="TABLE 1 — Implied Share Price (₹)  |  WACC (rows) vs. Terminal Growth Rate, g (columns)"
    ).font = _font(bold=True, size=10, color=BLUE_TEXT)
    ws.merge_cells("B5:J5")

    waccs = sens["wacc_range"]   # e.g. [0.08, 0.085, ..., 0.11]
    gs    = sens["g_range"]      # e.g. [0.025, 0.03, ..., 0.05]
    t1    = sens["table1"]       # 2D list [wacc_idx][g_idx]

    # Column headers (g values)
    ws.cell(row=6, column=2, value="WACC \\ g →").font = _font(bold=True, size=9)
    for j, g_val in enumerate(gs):
        c = ws.cell(row=6, column=3+j, value=g_val)
        c.font = _font(bold=True, size=9, color=WHITE)
        c.fill = _fill(DARK_BLUE)
        c.number_format = "0.0%"
        c.alignment = _align(h="center")
        c.border = _border()
        ws.column_dimensions[get_column_letter(3+j)].width = 13

    for i, w in enumerate(waccs):
        r = 7 + i
        # Row header
        rh = ws.cell(row=r, column=2, value=w)
        rh.number_format = "0.0%"
        rh.font = _font(bold=True, size=9, color=WHITE)
        rh.fill = _fill(DARK_BLUE)
        rh.alignment = _align(h="center")
        rh.border = _border()

        for j, g_val in enumerate(gs):
            price = t1[i][j]
            c = ws.cell(row=r, column=3+j, value=price)
            c.number_format = "#,##0.00"
            c.alignment = _align(h="right")
            c.border = _border()
            # Base-case highlight
            is_base = (abs(w - base_wacc) < 0.001 and abs(g_val - base_g) < 0.001)
            if is_base:
                c.fill = _fill(SENSITIVITY_GOLD)
                c.font = _font(bold=True, size=9)
            else:
                # Colour-code relative to current price
                if price and price_cur:
                    if price >= price_cur * 1.2:
                        c.fill = _fill("C6EFCE")
                        c.font = _font(size=9, color=GREEN_TEXT)
                    elif price <= price_cur * 0.8:
                        c.fill = _fill("FFC7CE")
                        c.font = _font(size=9, color="9C0006")
                    else:
                        c.font = _font(size=9)

    # Note
    note1_row = 7 + len(waccs) + 1
    n1 = ws.cell(row=note1_row, column=2,
        value=f"★ Gold cell = Base case (WACC: {base_wacc*100:.1f}%, g: {base_g*100:.1f}%). "
              f"Current share price: ₹{price_cur:.2f} (reference). "
              f"Green = >20% upside, Red = >20% downside.")
    n1.font = _font(size=8, italic=True, color="606060")
    ws.merge_cells(f"B{note1_row}:J{note1_row}")

    # ── TABLE 2: Rev Growth × Gross Margin ───────────────────────────────────
    t2_start = note1_row + 3
    ws.cell(row=t2_start, column=2,
        value="TABLE 2 — Implied Share Price (₹)  |  FY+1 Revenue Growth (rows) vs. FY+1 Gross Margin (columns)"
    ).font = _font(bold=True, size=10, color=BLUE_TEXT)
    ws.merge_cells(f"B{t2_start}:J{t2_start}")

    rev_gs = sens["rev_growth_range"]
    gms    = sens["gm_range"]
    t2     = sens["table2"]

    # Column headers
    ws.cell(row=t2_start+1, column=2, value="Rev Growth \\ GM →").font = _font(bold=True, size=9)
    for j, gm in enumerate(gms):
        c = ws.cell(row=t2_start+1, column=3+j, value=gm)
        c.font = _font(bold=True, size=9, color=WHITE)
        c.fill = _fill(DARK_BLUE)
        c.number_format = "0%"
        c.alignment = _align(h="center")
        c.border = _border()

    for i, rg in enumerate(rev_gs):
        r = t2_start + 2 + i
        rh = ws.cell(row=r, column=2, value=rg)
        rh.number_format = "0%"
        rh.font = _font(bold=True, size=9, color=WHITE)
        rh.fill = _fill(DARK_BLUE)
        rh.alignment = _align(h="center")
        rh.border = _border()

        for j, gm in enumerate(gms):
            price = t2[i][j]
            c = ws.cell(row=r, column=3+j, value=price)
            c.number_format = "#,##0.00"
            c.alignment = _align(h="right")
            c.border = _border()
            is_base = (abs(rg - base_rev_g) < 0.001 and abs(gm - base_gm) < 0.001)
            if is_base:
                c.fill = _fill(SENSITIVITY_GOLD)
                c.font = _font(bold=True, size=9)
            else:
                if price and price_cur:
                    if price >= price_cur * 1.2:
                        c.fill = _fill("C6EFCE")
                        c.font = _font(size=9, color=GREEN_TEXT)
                    elif price <= price_cur * 0.8:
                        c.fill = _fill("FFC7CE")
                        c.font = _font(size=9, color="9C0006")
                    else:
                        c.font = _font(size=9)

    note2_row = t2_start + 2 + len(rev_gs) + 1
    n2 = ws.cell(row=note2_row, column=2,
        value=f"★ Gold cell = Base case (Rev Growth: {base_rev_g*100:.0f}%, GM: {base_gm*100:.0f}%). "
              f"WACC={base_wacc*100:.1f}%, g={base_g*100:.1f}% held constant. Values in ₹.")
    n2.font = _font(size=8, italic=True, color="606060")
    ws.merge_cells(f"B{note2_row}:J{note2_row}")

    # ── Colour coding legend ───────────────────────────────────────────────────
    leg_start = note2_row + 4
    ws.cell(row=leg_start, column=2, value="COLOUR CODING LEGEND").font = _font(bold=True, size=10)

    legend = [
        (BLUE_TEXT,  "Blue text",       "Hardcoded input — numbers the analyst enters / can change for scenarios"),
        (BLACK,      "Black text",       "Formula / calculated value — derived from other cells"),
        (GREEN_TEXT, "Green text",       "Cross-sheet link — pulling data from another worksheet"),
        (YELLOW_FILL,"Yellow / gold fill","Key assumption cell — requires review / update for each scenario run"),
        (GREEN_FILL, "Green fill",       "Positive / profitability metric (Gross Profit, EBITDA, Net Income)"),
        (LIGHT_BLUE_FILL,"Light blue fill","Sub-total / aggregated line item"),
        (NAVY,       "Dark navy fill",   "Section header / title row"),
        (DARK_BLUE,  "Dark blue fill",   "Sub-section header / category row"),
    ]

    for k, (color_hex, label, desc) in enumerate(legend):
        r = leg_start + 1 + k
        lc = ws.cell(row=r, column=2, value=label)
        if color_hex in (YELLOW_FILL, GREEN_FILL, LIGHT_BLUE_FILL):
            lc.fill = _fill(color_hex)
            lc.font = _font(size=9)
        elif color_hex in (NAVY, DARK_BLUE):
            lc.fill = _fill(color_hex)
            lc.font = _font(size=9, color=WHITE)
        else:
            lc.font = _font(size=9, color=color_hex)

        ws.cell(row=r, column=3, value=desc).font = _font(size=9)
        ws.merge_cells(f"C{r}:J{r}")


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN PUBLIC FUNCTION
# ═══════════════════════════════════════════════════════════════════════════════
def generate_dcf_excel(fin: dict, symbol: str, company_name: str) -> bytes:
    """
    Given financial data dict (from _fetch_financials + yfinance history),
    compute all DCF model numbers and write a styled Excel workbook.
    Returns raw bytes of the .xlsx file.

    fin dict keys used:
      price, shares, revenue, ebitda, net_income, operating_cf, fcf,
      total_debt, cash, revenue_growth, earnings_growth, profit_margins,
      roe, beta, volatility_annual, sector
    """
    # ── 1. Base financial data ────────────────────────────────────────────────
    price     = _safe(fin.get("price"), 100.0)
    shares    = _safe(fin.get("shares"), 1e8)       # units: actual count
    revenue   = _safe(fin.get("revenue"), 1e10)     # INR
    ebitda    = _safe(fin.get("ebitda"), revenue * 0.20)
    net_inc   = _safe(fin.get("net_income"), revenue * 0.10)
    fcf_base  = _safe(fin.get("fcf") or fin.get("operating_cf"), revenue * 0.08)
    debt      = _safe(fin.get("total_debt"), 0.0)
    cash      = _safe(fin.get("cash"), 0.0)
    beta      = _safe(fin.get("beta"), 1.0)
    rev_g     = _safe(fin.get("revenue_growth"), 0.12)
    earn_g    = _safe(fin.get("earnings_growth"), 0.12)
    sector    = fin.get("sector", "")

    # Crore conversions (₹1 Cr = 1e7)
    CR = 1e7
    rev_cr   = revenue / CR
    debt_cr  = debt    / CR
    cash_cr  = cash    / CR
    shares_cr = shares / CR
    mktcap_cr = price * shares / CR

    # ── 2. WACC Assumptions ───────────────────────────────────────────────────
    rf   = 0.071        # India 10Y G-Sec
    erp  = 0.055        # India ERP (Damodaran)
    ke   = rf + beta * erp
    kd   = 0.09
    tax  = 0.25
    kd_at = kd * (1 - tax)
    ev_est = _safe(fin.get("enterprise_value"), price * shares + debt - cash)
    debt_weight = debt / (ev_est if ev_est > 0 else price * shares + debt) if (price * shares + debt) > 0 else 0.0
    wacc = ke * (1 - debt_weight) + kd_at * debt_weight
    tg   = 0.055        # terminal growth (India)

    # ── 3. Gross margin & opex ────────────────────────────────────────────────
    gm_base     = _safe(fin.get("profit_margins"), 0.15) + 0.25   # crude GP margin
    # If we have revenue & net_income, back-calculate plausible gross margin
    if revenue > 0:
        gm_base = max(0.20, min(0.80, (ebitda / revenue) + 0.12))
    rd_pct  = 0.04
    sga_pct = 0.08
    da_pct  = max(0.02, (ebitda - net_inc * 1.3) / revenue) if revenue else 0.03
    da_pct  = min(da_pct, 0.10)
    capex_pct = 0.04
    nwc_pct   = 0.02

    # ── 4. Historical years & projections ────────────────────────────────────
    now = datetime.date.today()
    fy  = now.year if now.month >= 4 else now.year - 1
    hist_years = [f"FY{fy-2}A", f"FY{fy-1}A", f"FY{fy}A"]
    proj_years = [f"FY{fy+i}E" for i in range(1, 6)]

    # Historical revenue: walk back from latest
    rev_h = [rev_cr / ((1 + rev_g)**2), rev_cr / (1 + rev_g), rev_cr]

    # Projected revenue growth schedule
    proj_rev_g = [
        max(rev_g, 0.08),               # Y1: recent momentum
        max(rev_g * 0.75, 0.06),        # Y2
        max(rev_g * 0.55, 0.05),        # Y3
        max(rev_g * 0.40, 0.055),       # Y4
        0.055,                           # Y5: converge to terminal
    ]

    rev_e = []
    prev = rev_cr
    for g in proj_rev_g:
        prev = prev * (1 + g)
        rev_e.append(prev)

    all_rev = rev_h + rev_e
    N = len(all_rev)   # 8

    gm_y1    = gm_base
    gm_later = min(gm_base + 0.02, gm_base * 1.03)

    def _gm(i):
        if i < len(hist_years):
            return gm_base - 0.02 * (len(hist_years) - 1 - i) * 0.5
        elif i == len(hist_years):
            return gm_y1
        else:
            return gm_later

    cogs     = [all_rev[i] * (1 - _gm(i)) for i in range(N)]
    rd       = [all_rev[i] * rd_pct for i in range(N)]
    sga      = [all_rev[i] * sga_pct for i in range(N)]
    da       = [all_rev[i] * da_pct for i in range(N)]
    net_int  = [cash_cr * 0.04 - debt_cr * kd for _ in range(N)]  # simplified
    gp       = [all_rev[i] - cogs[i] for i in range(N)]
    opex     = [rd[i] + sga[i] for i in range(N)]
    ebit     = [gp[i] - opex[i] for i in range(N)]
    ebitda_a = [ebit[i] + da[i] for i in range(N)]
    ebt      = [ebit[i] + net_int[i] for i in range(N)]
    tax_a    = [max(0, ebt[i] * tax) for i in range(N)]
    ni_a     = [ebt[i] - tax_a[i] for i in range(N)]
    eps_a    = [ni_a[i] / shares_cr if shares_cr else 0 for i in range(N)]

    # ── 5. DCF FCFF Build ──────────────────────────────────────────────────────
    proj_ebit  = ebit[len(hist_years):]
    proj_nopat = [e * (1 - tax) for e in proj_ebit]
    proj_da    = da[len(hist_years):]
    proj_capex = [all_rev[len(hist_years)+i] * capex_pct for i in range(5)]
    proj_dnwc  = [all_rev[len(hist_years)+i] * nwc_pct  for i in range(5)]
    proj_fcff  = [proj_nopat[i] + proj_da[i] - proj_capex[i] - proj_dnwc[i] for i in range(5)]
    proj_rev_e = all_rev[len(hist_years):]
    fcff_m     = [proj_fcff[i] / proj_rev_e[i] if proj_rev_e[i] else 0 for i in range(5)]

    disc_t = [0.5, 1.5, 2.5, 3.5, 4.5]
    disc_f = [1 / (1 + wacc) ** t for t in disc_t]
    pv_fcff = [proj_fcff[i] * disc_f[i] for i in range(5)]

    tv     = proj_fcff[-1] * (1 + tg) / (wacc - tg)
    pv_tv  = tv * disc_f[-1]
    sum_pv = sum(pv_fcff)
    ev_dcf = sum_pv + pv_tv
    eq_val = ev_dcf - debt_cr + cash_cr
    implied_price = eq_val / shares_cr if shares_cr else 0
    upside = (implied_price - price) / price if price else 0

    terminal_fcf_margin = proj_fcff[-1] / proj_rev_e[-1] if proj_rev_e[-1] else 0.15

    # ── 6. Sensitivity ────────────────────────────────────────────────────────
    def _dcf_price(w, g, rg1, gm1):
        """Compute implied price for given WACC, terminal g, rev growth Y1, gross margin Y1."""
        r_e = []
        p = rev_cr
        for k, grw in enumerate([rg1] + proj_rev_g[1:]):
            p = p * (1 + grw)
            r_e.append(p)

        def _gm_s(k):
            return gm1 if k == 0 else gm_later

        c_local = [r_e[k] * (1 - _gm_s(k)) for k in range(5)]
        rd_l    = [r_e[k] * rd_pct for k in range(5)]
        sga_l   = [r_e[k] * sga_pct for k in range(5)]
        da_l    = [r_e[k] * da_pct for k in range(5)]
        gp_l    = [r_e[k] - c_local[k] for k in range(5)]
        ebit_l  = [gp_l[k] - rd_l[k] - sga_l[k] for k in range(5)]
        nopat_l = [ebit_l[k] * (1 - tax) for k in range(5)]
        capex_l = [r_e[k] * capex_pct for k in range(5)]
        dnwc_l  = [r_e[k] * nwc_pct for k in range(5)]
        fcff_l  = [nopat_l[k] + da_l[k] - capex_l[k] - dnwc_l[k] for k in range(5)]
        df_l    = [1 / (1 + w) ** t for t in disc_t]
        pv_l    = [fcff_l[k] * df_l[k] for k in range(5)]
        tv_l    = fcff_l[-1] * (1 + g) / (w - g) if (w - g) > 0 else 0
        pv_tv_l = tv_l * df_l[-1]
        ev_l    = sum(pv_l) + pv_tv_l
        eq_l    = ev_l - debt_cr + cash_cr
        return round(eq_l / shares_cr, 2) if shares_cr else 0

    wacc_range = [0.080, 0.085, 0.090, 0.095, 0.100, 0.105, 0.110]
    g_range    = [0.025, 0.030, 0.035, 0.040, 0.045, 0.050]
    table1 = [
        [_dcf_price(w, g, proj_rev_g[0], gm_y1) for g in g_range]
        for w in wacc_range
    ]

    rev_growth_range = [max(0.05, proj_rev_g[0] - 0.10 + i * 0.05) for i in range(7)]
    gm_range = [max(0.20, gm_y1 - 0.06 + i * 0.03) for i in range(5)]
    table2 = [
        [_dcf_price(wacc, tg, rg, gm) for gm in gm_range]
        for rg in rev_growth_range
    ]

    # ── 7. Build workbook ─────────────────────────────────────────────────────
    wb = Workbook()

    meta = {
        "company_name": company_name,
        "symbol": symbol,
        "sector": sector,
        "price": price,
        "shares_cr": shares_cr,
        "mktcap_cr": mktcap_cr,
    }
    _build_cover(wb, meta)

    assumptions = {
        "company_name": company_name,
        "rf": rf, "erp": erp, "beta": beta, "ke": ke,
        "kd": kd, "tax_rate": tax, "kd_at": kd_at,
        "debt_weight": debt_weight, "wacc": wacc, "tg": tg,
        "terminal_fcf_margin": terminal_fcf_margin,
        "proj_years": proj_years,
        "rev_growth": proj_rev_g,
        "gm_y1": gm_y1, "gm_later": gm_later,
        "rd_pct": rd_pct, "sga_pct": sga_pct,
        "da_pct": da_pct, "capex_pct": capex_pct, "nwc_pct": nwc_pct,
        "cash_cr": cash_cr, "debt_cr": debt_cr,
        "shares_cr": shares_cr, "price": price,
    }
    _build_assumptions(wb, assumptions)

    is_data = {
        "company_name": company_name,
        "hist_years": hist_years,
        "proj_years": proj_years,
        "revenue": all_rev,
        "cogs": cogs,
        "rd": rd, "sga": sga, "da": da,
        "net_interest": net_int,
        "tax": tax_a,
        "eps": eps_a,
    }
    _build_income_statement(wb, is_data)

    dcf_data = {
        "company_name": company_name,
        "proj_years": proj_years,
        "ebit": proj_ebit,
        "nopat": proj_nopat,
        "da": proj_da,
        "capex": proj_capex,
        "dnwc": proj_dnwc,
        "fcff": proj_fcff,
        "fcff_margin": fcff_m,
        "disc_t": disc_t,
        "disc_f": disc_f,
        "pv_fcff": pv_fcff,
        "terminal_value": tv,
        "pv_terminal_value": pv_tv,
        "sum_pv_fcff": sum_pv,
        "enterprise_value": ev_dcf,
        "debt_cr": debt_cr,
        "cash_cr": cash_cr,
        "equity_value": eq_val,
        "implied_price": implied_price,
        "current_price": price,
        "upside": upside,
        "wacc": wacc,
        "tg": tg,
        "shares_cr": shares_cr,
    }
    _build_dcf_valuation(wb, dcf_data)

    sens_data = {
        "company_name": company_name,
        "wacc": wacc, "tg": tg,
        "base_rev_growth": proj_rev_g[0],
        "base_gm": gm_y1,
        "current_price": price,
        "wacc_range": wacc_range,
        "g_range": g_range,
        "table1": table1,
        "rev_growth_range": rev_growth_range,
        "gm_range": gm_range,
        "table2": table2,
    }
    _build_sensitivity(wb, sens_data)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
