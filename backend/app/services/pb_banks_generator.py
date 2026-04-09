"""
P/B Model for Banks & NBFCs
Fair P/B = (ROE - g) / (ke - g).  Gordon Growth applied to book value.
"""
import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.services.excel_utils import (
    font, fill, thin_border, align, safe, cr,
    section_header, build_cover, build_inputs, build_sensitivity_table,
)

THEME = {
    "cover_bg":      "001033",
    "primary":       "002366",
    "sub":           "003D99",
    "accent":        "2E86AB",
    "input_color":   "1A5276",
    "positive_fill": "D6EAF8",
    "positive_text": "1A5276",
    "subtotal_fill": "AED6F1",
    "key_fill":      "FFF2CC",
}


def _col(c: int) -> str:
    return get_column_letter(c)


def _lbl(ws, row, text, indent=0, bold=False):
    c = ws.cell(row=row, column=2, value=("    " * indent) + text)
    c.font = font(bold=bold, size=9)
    c.border = thin_border()
    c.alignment = align()
    return c


def _val(ws, row, col, val, fmt="#,##0.00", bold=False, bg=None, color="000000"):
    c = ws.cell(row=row, column=col, value=val)
    c.font = font(bold=bold, size=9, color=color)
    c.number_format = fmt
    c.alignment = align(h="right")
    c.border = thin_border()
    if bg:
        c.fill = fill(bg)
    return c


def _hdr(ws, row, cols, theme):
    for i, txt in enumerate(cols):
        c = ws.cell(row=row, column=2 + i, value=txt)
        c.font = font(bold=True, color="FFFFFF", size=9)
        c.fill = fill(theme["primary"])
        c.alignment = align(h="center")
        c.border = thin_border()


def _compute(fin: dict):
    price   = safe(fin.get("price"), 100.0)
    shares  = safe(fin.get("shares"), 1e8)
    beta    = safe(fin.get("beta"), 1.0)
    roe     = safe(fin.get("roe"), 0.15)
    roa     = safe(fin.get("roa"), 0.012)
    eps     = safe(fin.get("eps"), 1.0)
    dps     = safe(fin.get("dps"), 0.0)
    bvps    = safe(fin.get("book_value_per_share"), 100.0)
    assets  = cr(fin.get("total_assets"))
    revenue = cr(fin.get("revenue"))
    ni_cr   = cr(fin.get("net_income"))

    shares_cr  = shares / 1e7
    mktcap_cr  = price * shares / 1e7
    pb_current = price / bvps if bvps else 0.0
    pe_current = price / eps  if eps  else 0.0

    ke     = 0.071 + beta * 0.055
    g      = 0.06
    payout = dps / eps if eps > 0 and dps > 0 else max(1 - g / roe, 0.30) if roe else 0.30
    payout = min(payout, 0.95)

    justified_pb    = (roe - g) / (ke - g) if ke != g else 0.0
    intrinsic_value = justified_pb * bvps
    upside          = (intrinsic_value - price) / price if price else 0.0

    pb_premium = pb_current - justified_pb

    # DuPont
    npm = ni_cr / revenue if revenue else 0.0
    au  = revenue / assets if assets else 0.0
    bvps_total  = bvps * shares_cr
    leverage    = assets / bvps_total if bvps_total else 0.0
    dupont_roe  = npm * au * leverage

    # 5-year BVPS projection
    projection = []
    bvps_t = bvps
    for t in range(1, 6):
        bvps_t  = bvps_t * (1 + g)
        eps_t   = bvps_t * roe
        dps_t   = eps_t * payout
        ret_t   = eps_t - dps_t
        jpb_t   = justified_pb
        ip_t    = jpb_t * bvps_t
        projection.append((t, bvps_t, eps_t, dps_t, ret_t, jpb_t, ip_t))

    return dict(
        price=price, shares_cr=shares_cr, mktcap_cr=mktcap_cr,
        beta=beta, roe=roe, roa=roa, eps=eps, dps=dps, bvps=bvps,
        pb_current=pb_current, pe_current=pe_current,
        ke=ke, g=g, payout=payout,
        justified_pb=justified_pb, intrinsic_value=intrinsic_value,
        upside=upside, pb_premium=pb_premium,
        npm=npm, au=au, leverage=leverage, dupont_roe=dupont_roe,
        projection=projection,
    )


def _build_analysis(wb, theme, fin, d):
    ws = wb.create_sheet("PB Bank Valuation")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 16
    for i in range(4, 10):
        ws.column_dimensions[_col(i)].width = 14

    row = 2
    t = ws.cell(row=row, column=2,
                value="P/B Valuation for Banks & NBFCs  —  Justified P/B = (ROE − g) / (ke − g)")
    t.font = font(bold=True, size=13)
    ws.merge_cells(f"B{row}:{_col(9)}{row}")
    row += 2

    # ── BANK FINANCIAL METRICS ──
    section_header(ws, row, 2, "▌ BANK FINANCIAL METRICS", theme, span=3)
    row += 1
    metrics = [
        ("Book Value Per Share  BVPS  (₹)",        d["bvps"],       "#,##0.00"),
        ("Current P/B  (Price / BVPS)",             d["pb_current"], "0.00x"),
        ("Return on Equity  ROE",                   d["roe"],        "0.0%"),
        ("Return on Assets  ROA",                   d["roa"],        "0.0%"),
        ("NIM Estimate  (industry default 3.5%)",   0.035,           "0.0%"),
        ("Cost-to-Income  (default 45%)",           0.45,            "0.0%"),
        ("Current P/E  (Price / EPS)",              d["pe_current"], "0.00x"),
    ]
    for lbl_txt, val, fmt_ in metrics:
        _lbl(ws, row, lbl_txt, indent=1)
        _val(ws, row, 3, val, fmt_)
        row += 1
    row += 1

    # ── GORDON GROWTH P/B FORMULA ──
    section_header(ws, row, 2, "▌ GORDON GROWTH P/B FORMULA", theme, span=3)
    row += 1
    sust_g_check = d["roe"] * (1 - d["payout"])
    formula_rows = [
        ("ke  (Cost of Equity, CAPM)",              d["ke"],            "0.0%",        False),
        ("g   (Bank growth rate, default 6%)",      d["g"],             "0.0%",        False),
        ("ROE",                                     d["roe"],           "0.0%",        False),
        ("Payout Ratio",                            d["payout"],        "0.0%",        False),
        ("Sustainable Growth = ROE × (1 − Payout)", sust_g_check,      "0.0%",        False),
        ("★ Justified P/B = (ROE − g) / (ke − g)", d["justified_pb"], "0.00x",        True),
        ("★ Intrinsic Value = Justified P/B × BVPS",d["intrinsic_value"],"#,##0.00",  True),
        ("Current Market Price  (₹)",               d["price"],         "#,##0.00",    False),
        ("★ Upside / (Downside)",                   d["upside"],        "+0.0%;-0.0%", True),
    ]
    for lbl_txt, val, fmt_, is_key in formula_rows:
        _lbl(ws, row, lbl_txt, bold=is_key)
        _val(ws, row, 3, val, fmt_, bold=is_key,
             bg=theme["key_fill"] if is_key else None,
             color=theme["accent"] if is_key else "000000")
        row += 1
    row += 1

    # ── P/B PREMIUM ANALYSIS ──
    section_header(ws, row, 2, "▌ P/B PREMIUM ANALYSIS", theme, span=3)
    row += 1
    premium_rows = [
        ("Current Market P/B",             d["pb_current"],  "0.00x"),
        ("Justified P/B (Gordon Growth)",  d["justified_pb"],"0.00x"),
        ("Premium / (Discount) to Justified", d["pb_premium"],"0.00x"),
        ("Source of Premium",              "Franchise value, growth optionality, brand", "@"),
    ]
    for lbl_txt, val, fmt_ in premium_rows:
        _lbl(ws, row, lbl_txt, indent=1)
        _val(ws, row, 3, val, fmt_)
        row += 1
    row += 1

    # ── DUPONT ──
    section_header(ws, row, 2, "▌ DUPONT ROE DECOMPOSITION", theme, span=3)
    row += 1
    dupont = [
        ("Net Profit Margin  =  NI / Revenue",    d["npm"],       "0.0%"),
        ("Asset Utilisation  =  Revenue / Assets", d["au"],       "0.0x"),
        ("Financial Leverage  =  Assets / Equity", d["leverage"], "0.00x"),
        ("ROE  =  Margin × Utilisation × Leverage",d["dupont_roe"],"0.0%"),
    ]
    for lbl_txt, val, fmt_ in dupont:
        _lbl(ws, row, lbl_txt, indent=1)
        _val(ws, row, 3, val, fmt_)
        row += 1
    row += 1

    # ── 5-YEAR PROJECTION ──
    YEAR_COLS = [f"Year {y}" for y in range(1, 6)]
    section_header(ws, row, 2, "▌ BVPS GROWTH PROJECTION (5 YEARS)", theme, span=7)
    row += 1
    _hdr(ws, row, ["Metric"] + YEAR_COLS, theme)
    row += 1

    p_bvps  = [p[1] for p in d["projection"]]
    p_eps   = [p[2] for p in d["projection"]]
    p_dps   = [p[3] for p in d["projection"]]
    p_ret   = [p[4] for p in d["projection"]]
    p_jpb   = [p[5] for p in d["projection"]]
    p_ip    = [p[6] for p in d["projection"]]

    for lbl_txt, vals, fmt_ in [
        ("BVPS (₹)",                   p_bvps, "#,##0.00"),
        ("EPS  =  BVPS × ROE  (₹)",    p_eps,  "#,##0.00"),
        ("DPS  =  EPS × Payout  (₹)",  p_dps,  "#,##0.00"),
        ("Retained = EPS − DPS  (₹)",  p_ret,  "#,##0.00"),
        ("Justified P/B",              p_jpb,  "0.00x"),
        ("Implied Price  (₹)",         p_ip,   "#,##0.00"),
    ]:
        _lbl(ws, row, lbl_txt, indent=1)
        for j, v in enumerate(vals):
            _val(ws, row, 3 + j, v, fmt_)
        row += 1

    ws.freeze_panes = "C4"


def _build_results_sheet(wb, theme, company_name, d):
    ws = wb.create_sheet("Results & Sensitivity")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 16

    row = 2
    section_header(ws, row, 2, "▌ VALUATION RESULTS SUMMARY", theme, span=4)
    row += 1
    results = [
        ("BVPS  (₹)",                      d["bvps"],           "#,##0.00",    False),
        ("Justified P/B",                   d["justified_pb"],   "0.00x",       False),
        ("★ Intrinsic Value  (₹)",          d["intrinsic_value"],"#,##0.00",    True),
        ("Current Price  (₹)",              d["price"],          "#,##0.00",    False),
        ("★ Upside / (Downside)",           d["upside"],         "+0.0%;-0.0%", True),
        ("ke  (Cost of Equity)",            d["ke"],             "0.0%",        False),
        ("g   (Bank growth rate)",          d["g"],              "0.0%",        False),
        ("ROE",                             d["roe"],            "0.0%",        False),
    ]
    for lbl_txt, val, fmt_, is_key in results:
        lc = ws.cell(row=row, column=2, value=lbl_txt)
        lc.font = font(bold=is_key, size=9)
        lc.border = thin_border()
        if is_key:
            lc.fill = fill(theme["key_fill"])
        vc = ws.cell(row=row, column=3, value=val)
        vc.font = font(bold=is_key, size=9,
                       color=theme["accent"] if is_key else "000000")
        vc.number_format = fmt_
        vc.alignment = align(h="right")
        vc.border = thin_border()
        if is_key:
            vc.fill = fill(theme["key_fill"])
        row += 1

    # Sensitivity: ROE × ke
    roe_vals = [0.10, 0.12, 0.14, 0.16, 0.18, 0.20, 0.22, 0.24]
    ke_vals  = [0.11, 0.12, 0.13, 0.14, 0.15, 0.16, 0.17]

    bvps = d["bvps"]
    g    = d["g"]

    base_ri = min(range(len(roe_vals)), key=lambda i: abs(roe_vals[i] - d["roe"]))
    base_ki = min(range(len(ke_vals)),  key=lambda i: abs(ke_vals[i]  - d["ke"]))

    matrix = []
    for roe_ in roe_vals:
        row_data = []
        for ke_ in ke_vals:
            denom = ke_ - g
            if denom <= 0:
                row_data.append(0.0)
            else:
                jpb_ = (roe_ - g) / denom
                ip_  = jpb_ * bvps
                row_data.append(round(ip_, 2))
        matrix.append(row_data)

    build_sensitivity_table(
        ws, theme,
        row_label="ROE",
        col_label="ke (Cost of Equity)",
        row_vals=roe_vals,
        col_vals=ke_vals,
        matrix=matrix,
        current_price=d["price"],
        base_row_idx=base_ri,
        base_col_idx=base_ki,
        start_row=row + 2,
        start_col=2,
        row_fmt="0.0%",
        col_fmt="0.0%",
    )


def generate_excel(fin: dict, symbol: str, company_name: str) -> bytes:
    d  = _compute(fin)
    wb = Workbook()

    build_cover(
        wb, THEME, symbol, company_name,
        model_label="P/B Bank Valuation Model",
        model_desc="Justified P/B = (ROE − g) / (ke − g). "
                   "Gordon Growth applied to book value. Suitable for banks, NBFCs and financial institutions.",
        sheets_index=[
            (1, "Cover",                "Model overview & index"),
            (2, "Inputs & Assumptions", "Financial data & model parameters"),
            (3, "PB Bank Valuation",    "Formula, DuPont & 5-year BVPS projection"),
            (4, "Results & Sensitivity","Summary + ROE × ke sensitivity"),
        ],
        meta_extra={
            "price":     d["price"],
            "mktcap_cr": d["mktcap_cr"],
            "sector":    safe(fin.get("sector"), "—"),
        },
    )

    params_def = [
        ("ke (Cost of Equity)",     d["ke"],    "%",  "rf + β × ERP",         "0.0%", True),
        ("g  (Bank growth rate)",   d["g"],     "%",  "Sustainable bank g",    "0.0%", True),
        ("ROE",                     d["roe"],   "%",  "Return on equity",      "0.0%", True),
        ("Justified P/B",           d["justified_pb"], "x", "(ROE-g)/(ke-g)", "0.00x", True),
        ("Risk-Free Rate (rf)",     0.071,      "%",  "India 10Y GSec",        "0.0%", False),
        ("Equity Risk Premium",     0.055,      "%",  "Damodaran India ERP",    "0.0%", False),
        ("Beta",                    d["beta"],  "x",  "Market beta",           "0.00", False),
    ]

    build_inputs(wb, THEME, company_name, fin, params_def)
    _build_analysis(wb, THEME, fin, d)
    _build_results_sheet(wb, THEME, company_name, d)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
