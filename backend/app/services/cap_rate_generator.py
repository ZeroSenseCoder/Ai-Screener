"""
Cap Rate — Real Estate / REITs
Property Value = NOI / Cap Rate.
"""
import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.services.excel_utils import (
    font, fill, thin_border, align, safe, cr,
    section_header, build_cover, build_inputs, build_sensitivity_table,
)

THEME = {
    "cover_bg":      "001100",
    "primary":       "002200",
    "sub":           "003800",
    "accent":        "27AE60",
    "input_color":   "1E8449",
    "positive_fill": "D5F5E3",
    "positive_text": "1E8449",
    "subtotal_fill": "A9DFBF",
    "key_fill":      "FFF2CC",
}


def _col(c: int) -> str:
    return get_column_letter(c)


def _lbl(ws, row, text, indent=0, bold=False):
    c = ws.cell(row=row, column=2, value=("    " * indent) + text)
    c.font = font(bold=bold, size=9)
    c.border = thin_border()
    c.alignment = align()
    return c


def _val(ws, row, col, val, fmt="#,##0.00", bold=False, bg=None, color="000000"):
    c = ws.cell(row=row, column=col, value=val)
    c.font = font(bold=bold, size=9, color=color)
    c.number_format = fmt
    c.alignment = align(h="right")
    c.border = thin_border()
    if bg:
        c.fill = fill(bg)
    return c


def _hdr(ws, row, cols, theme):
    for i, txt in enumerate(cols):
        c = ws.cell(row=row, column=2 + i, value=txt)
        c.font = font(bold=True, color="FFFFFF", size=9)
        c.fill = fill(theme["primary"])
        c.alignment = align(h="center")
        c.border = thin_border()


def _compute(fin: dict):
    price      = safe(fin.get("price"), 100.0)
    shares     = safe(fin.get("shares"), 1e8)
    beta       = safe(fin.get("beta"), 1.0)
    ebitda     = cr(fin.get("ebitda"))
    revenue    = cr(fin.get("revenue"))
    net_income = cr(fin.get("net_income"))
    total_debt = cr(fin.get("total_debt"))
    cash       = cr(fin.get("cash"))
    total_assets = cr(fin.get("total_assets"))

    shares_cr  = shares / 1e7
    mktcap_cr  = price * shares / 1e7

    # NOI build
    gross_rev   = ebitda / 0.65 if ebitda else revenue
    vacancy     = gross_rev * 0.05
    egr         = gross_rev - vacancy
    opex        = egr * 0.35
    prop_tax    = total_assets * 0.02
    noi         = egr - opex - prop_tax
    noi_margin  = noi / gross_rev if gross_rev else 0.0

    cap_rate    = 0.07
    prop_val    = noi / cap_rate if cap_rate else 0.0
    net_debt    = total_debt - cash
    equity_val  = prop_val - net_debt
    implied_price = equity_val / shares_cr if shares_cr else 0.0
    upside      = (implied_price - price) / price if price else 0.0

    # REIT metrics
    da          = revenue * 0.05
    ffo         = net_income + da
    capex       = revenue * 0.03
    affo        = ffo - capex
    ffo_ps      = ffo / shares_cr if shares_cr else 0.0
    affo_ps     = affo / shares_cr if shares_cr else 0.0
    p_ffo       = price / ffo_ps  if ffo_ps else 0.0
    dps         = safe(fin.get("dps"), 0.0)
    div_yield   = dps / price if price else 0.0
    ffo_payout  = dps / ffo_ps if ffo_ps else 0.0

    # Scenarios (Bear/Base/Bull — 5Y NPV of NOI + terminal)
    ke = 0.071 + beta * 0.055
    scenarios = []
    for label_s, g_s in [("Bear", 0.03), ("Base", 0.05), ("Bull", 0.08)]:
        pv_noi = 0.0
        noi_t  = noi
        for t in range(1, 6):
            noi_t   = noi_t * (1 + g_s)
            pv_noi += noi_t / (1 + ke) ** t
        tv_s   = noi_t * (1 + g_s) / (cap_rate - g_s) * (1 / (1 + ke) ** 5) if cap_rate > g_s else 0.0
        npv_s  = pv_noi + tv_s - net_debt
        ip_s   = npv_s / shares_cr if shares_cr else 0.0
        scenarios.append((label_s, g_s, pv_noi, tv_s, npv_s, ip_s))

    return dict(
        price=price, shares_cr=shares_cr, mktcap_cr=mktcap_cr,
        beta=beta, ke=ke,
        gross_rev=gross_rev, vacancy=vacancy, egr=egr,
        opex=opex, prop_tax=prop_tax, noi=noi, noi_margin=noi_margin,
        cap_rate=cap_rate, prop_val=prop_val, net_debt=net_debt,
        equity_val=equity_val, implied_price=implied_price, upside=upside,
        ffo=ffo, affo=affo, ffo_ps=ffo_ps, affo_ps=affo_ps,
        p_ffo=p_ffo, div_yield=div_yield, ffo_payout=ffo_payout,
        total_debt=total_debt, cash=cash,
        scenarios=scenarios,
    )


def _build_analysis(wb, theme, fin, d):
    ws = wb.create_sheet("Cap Rate REIT Valuation")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 16
    for i in range(4, 8):
        ws.column_dimensions[_col(i)].width = 14

    row = 2
    t = ws.cell(row=row, column=2,
                value="Cap Rate / REIT Valuation  —  Property Value = NOI / Cap Rate")
    t.font = font(bold=True, size=13)
    ws.merge_cells(f"B{row}:{_col(8)}{row}")
    row += 2

    # ── NOI BUILD ──
    section_header(ws, row, 2, "▌ NOI BUILD", theme, span=3)
    row += 1
    noi_rows = [
        ("Gross Revenue  (₹ Cr)",                      d["gross_rev"],  "#,##0",  False),
        ("Less: Vacancy  (5%)  (₹ Cr)",                d["vacancy"],    "#,##0",  False),
        ("Effective Gross Revenue  EGR  (₹ Cr)",       d["egr"],        "#,##0",  False),
        ("Less: Operating Expenses  (35%)  (₹ Cr)",    d["opex"],       "#,##0",  False),
        ("Less: Property Tax  (2% of Assets)  (₹ Cr)", d["prop_tax"],   "#,##0",  False),
        ("★ NOI  =  EGR − Opex − PropTax  (₹ Cr)",    d["noi"],        "#,##0",  True),
        ("NOI Margin  %",                              d["noi_margin"], "0.0%",   False),
    ]
    for lbl_txt, val, fmt_, is_key in noi_rows:
        _lbl(ws, row, lbl_txt, bold=is_key)
        _val(ws, row, 3, val, fmt_, bold=is_key,
             bg=theme["key_fill"] if is_key else None,
             color=theme["accent"] if is_key else "000000")
        row += 1
    row += 1

    # ── CAP RATE VALUATION ──
    section_header(ws, row, 2, "▌ CAP RATE VALUATION", theme, span=3)
    row += 1
    cap_rows = [
        ("NOI  (₹ Cr)",                             d["noi"],          "#,##0",       False),
        ("Cap Rate  (default 7%)",                  d["cap_rate"],     "0.0%",        False),
        ("★ Property Value = NOI / Cap Rate  (₹ Cr)", d["prop_val"],  "#,##0",       True),
        ("Less: Total Debt  (₹ Cr)",                d["total_debt"],   "#,##0",       False),
        ("Add: Cash  (₹ Cr)",                       d["cash"],         "#,##0",       False),
        ("Equity Value  (₹ Cr)",                    d["equity_val"],   "#,##0",       False),
        ("Shares Outstanding  (Cr)",                d["shares_cr"],    "#,##0.00",    False),
        ("★ Implied Share Price  (₹)",              d["implied_price"],"#,##0.00",    True),
        ("Current Market Price  (₹)",               d["price"],        "#,##0.00",    False),
        ("★ Upside / (Downside)",                   d["upside"],       "+0.0%;-0.0%", True),
    ]
    for lbl_txt, val, fmt_, is_key in cap_rows:
        _lbl(ws, row, lbl_txt, bold=is_key)
        _val(ws, row, 3, val, fmt_, bold=is_key,
             bg=theme["key_fill"] if is_key else None,
             color=theme["accent"] if is_key else "000000")
        row += 1
    row += 1

    # ── REIT METRICS ──
    section_header(ws, row, 2, "▌ REIT METRICS", theme, span=3)
    row += 1
    reit_rows = [
        ("FFO  =  Net Income + D&A  (₹ Cr)",  d["ffo"],       "#,##0"),
        ("AFFO  =  FFO − CapEx  (₹ Cr)",      d["affo"],      "#,##0"),
        ("FFO Per Share  (₹)",                 d["ffo_ps"],    "#,##0.00"),
        ("AFFO Per Share  (₹)",                d["affo_ps"],   "#,##0.00"),
        ("Price / FFO  (times)",               d["p_ffo"],     "0.0x"),
        ("Dividend Yield",                     d["div_yield"], "0.0%"),
        ("FFO Payout Ratio",                   d["ffo_payout"],"0.0%"),
    ]
    for lbl_txt, val, fmt_ in reit_rows:
        _lbl(ws, row, lbl_txt, indent=1)
        _val(ws, row, 3, val, fmt_)
        row += 1
    row += 1

    # ── BENCHMARKS ──
    section_header(ws, row, 2, "▌ CAP RATE BENCHMARKS", theme, span=4)
    row += 1
    _hdr(ws, row, ["Property Type", "Cap Rate Range", "Indian Market Notes"], theme)
    row += 1
    benchmarks = [
        ("Grade A Office — Mumbai",  "7–8.5%",  "Premium CBD assets"),
        ("Office — Other Cities",    "8–9.5%",  "Tier 1 non-Mumbai"),
        ("Retail",                   "8.5–10%", "Mall and high-street"),
        ("Industrial / Warehousing", "8–9%",    "Strong demand post-GST"),
        ("Residential",              "3–5%",    "Low yield, capital gain driven"),
    ]
    for prop, rng, note in benchmarks:
        ws.cell(row=row, column=2, value=prop).border = thin_border()
        ws.cell(row=row, column=3, value=rng).border = thin_border()
        ws.cell(row=row, column=4, value=note).border = thin_border()
        ws.cell(row=row, column=4).font = font(size=9, color="606060", italic=True)
        row += 1
    row += 1

    # ── SCENARIOS ──
    section_header(ws, row, 2, "▌ INCOME GROWTH SCENARIOS  (5-Year NPV)", theme, span=5)
    row += 1
    _hdr(ws, row, ["Metric", "Bear  (3%)", "Base  (5%)", "Bull  (8%)"], theme)
    row += 1
    sc = d["scenarios"]
    for lbl_txt, idx in [
        ("NOI Growth Rate",          2),
        ("PV of NOI — 5Y  (₹ Cr)",  3),
        ("Terminal Value  (₹ Cr)",   4),
        ("Equity Value  (₹ Cr)",     5),
        ("Implied Share Price  (₹)", 6),
    ]:
        _lbl(ws, row, lbl_txt, indent=1)
        for si, s in enumerate(sc):
            fmt_ = "0.0%" if idx == 2 else ("#,##0.00" if idx == 6 else "#,##0")
            _val(ws, row, 3 + si, s[idx - 1], fmt_)
        row += 1

    ws.freeze_panes = "C4"


def _build_results_sheet(wb, theme, company_name, d):
    ws = wb.create_sheet("Results & Sensitivity")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 16

    row = 2
    section_header(ws, row, 2, "▌ VALUATION RESULTS SUMMARY", theme, span=4)
    row += 1
    results = [
        ("NOI  (₹ Cr)",                   d["noi"],          "#,##0",       False),
        ("Cap Rate",                        d["cap_rate"],     "0.0%",        False),
        ("Property Value  (₹ Cr)",         d["prop_val"],     "#,##0",       False),
        ("Net Debt  (₹ Cr)",               d["net_debt"],     "#,##0",       False),
        ("★ Implied Share Price  (₹)",     d["implied_price"],"#,##0.00",    True),
        ("Current Price  (₹)",             d["price"],        "#,##0.00",    False),
        ("★ Upside / (Downside)",          d["upside"],       "+0.0%;-0.0%", True),
        ("Price / FFO",                    d["p_ffo"],        "0.0x",        False),
        ("Dividend Yield",                 d["div_yield"],    "0.0%",        False),
    ]
    for lbl_txt, val, fmt_, is_key in results:
        lc = ws.cell(row=row, column=2, value=lbl_txt)
        lc.font = font(bold=is_key, size=9)
        lc.border = thin_border()
        if is_key:
            lc.fill = fill(theme["key_fill"])
        vc = ws.cell(row=row, column=3, value=val)
        vc.font = font(bold=is_key, size=9,
                       color=theme["accent"] if is_key else "000000")
        vc.number_format = fmt_
        vc.alignment = align(h="right")
        vc.border = thin_border()
        if is_key:
            vc.fill = fill(theme["key_fill"])
        row += 1

    # Sensitivity: Cap Rate × NOI Growth
    cap_vals    = [0.05, 0.055, 0.06, 0.065, 0.07, 0.08, 0.09, 0.10]
    growth_vals = [0.02, 0.03, 0.04, 0.05, 0.06, 0.07, 0.08]

    noi     = d["noi"]
    debt    = d["total_debt"]
    cash    = d["cash"]
    sc      = d["shares_cr"]

    base_ci = min(range(len(cap_vals)),    key=lambda i: abs(cap_vals[i]    - d["cap_rate"]))
    base_gi = min(range(len(growth_vals)), key=lambda i: abs(growth_vals[i] - 0.05))

    matrix = []
    for cap_ in cap_vals:
        row_data = []
        for g_ in growth_vals:
            denom = cap_ - g_
            if denom <= 0:
                row_data.append(0.0)
            else:
                # Gordon Growth on NOI: growing perpetuity
                prop_v = noi * (1 + g_) / denom
                eq_    = prop_v - debt + cash
                ip_    = eq_ / sc if sc else 0.0
                row_data.append(round(ip_, 2))
        matrix.append(row_data)

    build_sensitivity_table(
        ws, theme,
        row_label="Cap Rate",
        col_label="NOI Growth Rate",
        row_vals=cap_vals,
        col_vals=growth_vals,
        matrix=matrix,
        current_price=d["price"],
        base_row_idx=base_ci,
        base_col_idx=base_gi,
        start_row=row + 2,
        start_col=2,
        row_fmt="0.0%",
        col_fmt="0.0%",
    )


def generate_excel(fin: dict, symbol: str, company_name: str) -> bytes:
    d  = _compute(fin)
    wb = Workbook()

    build_cover(
        wb, THEME, symbol, company_name,
        model_label="Cap Rate / REIT Valuation",
        model_desc="Property Value = NOI / Cap Rate. "
                   "Used for REITs and real estate companies. NOI divided by required yield gives intrinsic property value.",
        sheets_index=[
            (1, "Cover",                   "Model overview & index"),
            (2, "Inputs & Assumptions",    "Financial data & model parameters"),
            (3, "Cap Rate REIT Valuation",  "NOI build, metrics & scenarios"),
            (4, "Results & Sensitivity",   "Summary + cap rate × NOI growth sensitivity"),
        ],
        meta_extra={
            "price":     d["price"],
            "mktcap_cr": d["mktcap_cr"],
            "sector":    safe(fin.get("sector"), "—"),
        },
    )

    params_def = [
        ("Cap Rate",               d["cap_rate"], "%",  "Market cap rate for property type", "0.0%", True),
        ("Vacancy Rate",           0.05,          "%",  "Industry default",                  "0.0%", False),
        ("Operating Expense Ratio",0.35,          "%",  "As % of EGR",                       "0.0%", False),
        ("Property Tax",           0.02,          "%",  "% of total assets",                 "0.0%", False),
        ("D&A (FFO add-back)",     0.05,          "%",  "% of revenue",                      "0.0%", False),
        ("CapEx (AFFO deduct)",    0.03,          "%",  "% of revenue",                      "0.0%", False),
        ("Risk-Free Rate (rf)",    0.071,         "%",  "India 10Y GSec",                    "0.0%", False),
        ("Equity Risk Premium",    0.055,         "%",  "Damodaran India ERP",                "0.0%", False),
        ("Beta",                   d["beta"],     "x",  "Market beta",                       "0.00", False),
    ]

    build_inputs(wb, THEME, company_name, fin, params_def)
    _build_analysis(wb, THEME, fin, d)
    _build_results_sheet(wb, THEME, company_name, d)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
