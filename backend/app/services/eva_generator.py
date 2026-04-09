"""
EVA (Economic Value Added) Excel Generator
Firm Value = Invested Capital + PV(all future EVAs)
EVA = NOPAT - WACC x IC
"""

import io
import math
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.services.excel_utils import (
    font, fill, thin_border, align, safe, cr,
    section_header, sub_header, label, value_cell, pct_cell,
    build_cover, build_inputs, build_results, build_sensitivity_table,
)

THEME = {
    "cover_bg":      "0B1326",
    "primary":       "1A3A5C",
    "sub":           "2E6090",
    "accent":        "3498DB",
    "input_color":   "1A5276",
    "positive_fill": "D6EAF8",
    "positive_text": "1A5276",
    "subtotal_fill": "AED6F1",
    "key_fill":      "FFF2CC",
}

SHEETS_INDEX = [
    (1, "Cover",              "Model overview & metadata"),
    (2, "Inputs & Assumptions","Financial data & model parameters"),
    (3, "EVA Model",          "EVA projections & firm value bridge"),
    (4, "Sensitivity",        "WACC × NOPAT Growth sensitivity table"),
]


# ── helpers ────────────────────────────────────────────────────────────────────

def _lv(ws, row, col, text, bold=False, indent=0):
    c = label(ws, row, col, text, indent=indent, bold=bold, theme=THEME)
    return c


def _vc(ws, row, col, val, fmt="#,##0", bold=False, bg=None, color="000000"):
    return value_cell(ws, row, col, val, fmt=fmt, bold=bold, bg=bg, color=color)


def _pc(ws, row, col, val, bold=False, bg=None):
    return pct_cell(ws, row, col, val, bold=bold, bg=bg)


def _sh(ws, row, text, span=9):
    section_header(ws, row, 2, text, THEME, span=span)


def _key_row(ws, row, lbl, val, fmt="#,##0"):
    _lv(ws, row, 2, lbl, bold=True)
    ws.cell(row=row, column=2).fill = fill(THEME["key_fill"])
    _vc(ws, row, 3, val, fmt=fmt, bold=True,
        bg=THEME["key_fill"], color=THEME["input_color"])


# ── main ───────────────────────────────────────────────────────────────────────

def generate_excel(fin: dict, symbol: str, company_name: str) -> bytes:
    wb = Workbook()

    price       = safe(fin.get("price"), 0.0)
    shares_raw  = safe(fin.get("shares"), 1e7)
    shares_cr   = shares_raw / 1e7
    mktcap_cr   = price * shares_cr
    sector      = fin.get("sector", "—")

    nopat_cr    = cr(fin.get("nopat"))
    ic_cr       = cr(fin.get("invested_capital"))
    debt_cr     = cr(fin.get("total_debt"))
    cash_cr     = cr(fin.get("cash"))
    rev_growth  = safe(fin.get("revenue_growth"), 0.10)
    beta        = safe(fin.get("beta"), 1.0)
    ev_cr       = cr(fin.get("enterprise_value"))

    # WACC components
    rf          = 0.07           # India 10Y Gsec proxy
    erp         = 0.055          # equity risk premium India
    ke          = rf + beta * erp
    kd          = 0.09
    tax_rate    = 0.25
    equity_val  = mktcap_cr
    total_cap   = equity_val + max(debt_cr - cash_cr, 0)
    w_e         = equity_val / total_cap if total_cap > 0 else 0.7
    w_d         = 1 - w_e
    wacc        = w_e * ke + w_d * kd * (1 - tax_rate)

    roic        = nopat_cr / ic_cr if ic_cr > 0 else 0
    equity_chg  = wacc * ic_cr
    eva0        = nopat_cr - equity_chg

    g1          = max(rev_growth, 0.08)
    g_terminal  = 0.05
    net_debt_cr = debt_cr - cash_cr

    # ── Cover ──────────────────────────────────────────────────────────────────
    build_cover(
        wb, THEME, symbol, company_name,
        "Economic Value Added (EVA) Model",
        "Firm Value = Invested Capital + PV(EVA streams). "
        "Measures value creation above the cost of capital.",
        SHEETS_INDEX,
        {"price": price, "mktcap_cr": mktcap_cr, "sector": sector},
    )

    # ── Inputs ─────────────────────────────────────────────────────────────────
    params_def = [
        ("Risk-Free Rate (rf)",         rf,         "%",   "India 10Y G-Sec proxy",         "0.0%", False),
        ("Equity Risk Premium (ERP)",   erp,        "%",   "Damodaran India ERP estimate",  "0.0%", False),
        ("Cost of Equity (ke)",         ke,         "%",   "CAPM: rf + β × ERP",            "0.0%", True),
        ("Pre-tax Cost of Debt (kd)",   kd,         "%",   "Company borrowing rate proxy",  "0.0%", False),
        ("Tax Rate",                    tax_rate,   "%",   "Effective corporate tax rate",  "0.0%", False),
        ("WACC",                        wacc,       "%",   "Weighted cost of capital",      "0.0%", True),
        ("Terminal Growth Rate (g)",    g_terminal, "%",   "Long-run nominal GDP growth",   "0.0%", True),
        ("Near-term Growth (g1)",       g1,         "%",   "max(rev_growth, 8%)",           "0.0%", True),
        ("Weight of Equity (we)",       w_e,        "%",   "MktCap / (MktCap + Net Debt)",  "0.0%", False),
        ("Weight of Debt (wd)",         w_d,        "%",   "Net Debt / Total Capital",      "0.0%", False),
    ]
    build_inputs(wb, THEME, company_name, fin, params_def)

    # ── EVA Model sheet ────────────────────────────────────────────────────────
    ws = wb.create_sheet("EVA Model")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 16
    for ci in range(4, 11):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    # Title
    t = ws.cell(row=2, column=2, value=f"{company_name} — EVA Model")
    t.font = font(bold=True, size=13)
    ws.merge_cells("B2:J2")

    row = 4

    # ── SECTION 1: WACC & CAPITAL BASE ────────────────────────────────────────
    _sh(ws, row, "▌ WACC & CAPITAL BASE")
    row += 1

    rows_s1 = [
        ("NOPAT (₹ Cr)",               nopat_cr,       "#,##0.0", False),
        ("Invested Capital (IC) (₹ Cr)",ic_cr,          "#,##0.0", False),
        ("WACC (%)",                    wacc,           "0.00%",   False),
        ("ROIC = NOPAT / IC",           roic,           "0.00%",   False),
        ("ROIC vs WACC Spread",         roic - wacc,    "0.00%",   False),
        ("Equity Charge = WACC × IC",   equity_chg,     "#,##0.0", False),
        ("★ EVA (Year 0) = NOPAT − Equity Charge", eva0, "#,##0.0", True),
    ]
    for lbl_txt, val, fmt, is_key in rows_s1:
        if is_key:
            _key_row(ws, row, lbl_txt, val, fmt)
        else:
            _lv(ws, row, 2, lbl_txt)
            _vc(ws, row, 3, val, fmt=fmt)
        row += 1

    row += 1

    # ── SECTION 2: EVA PROJECTIONS ────────────────────────────────────────────
    _sh(ws, row, "▌ EVA PROJECTIONS (5 YEARS + TERMINAL)", span=9)
    row += 1

    # Year column headers
    yr_cols = list(range(4, 10))  # cols 4–9 = Year 1–5 + Terminal
    yr_labels = [f"Year {i}" for i in range(1, 6)] + ["Terminal"]
    for ci, lbl_txt in zip(yr_cols, yr_labels):
        c = ws.cell(row=row, column=ci, value=lbl_txt)
        c.font = font(bold=True, color="FFFFFF", size=9)
        c.fill = fill(THEME["sub"])
        c.alignment = align(h="center")
        c.border = thin_border()
    _lv(ws, row, 2, "Metric")
    row += 1

    # Build projections
    g_schedule = []
    for t in range(1, 6):
        g_t = g1 + (g_terminal - g1) * (t - 1) / 4
        g_schedule.append(g_t)

    proj_nopat = []
    proj_ic    = []
    proj_roic  = []
    proj_wacc  = []
    proj_spread= []
    proj_eva   = []
    proj_df    = []
    proj_pv    = []

    for t_idx, g_t in enumerate(g_schedule):
        n_prev = proj_nopat[-1] if proj_nopat else nopat_cr
        i_prev = proj_ic[-1]   if proj_ic    else ic_cr
        n = n_prev * (1 + g_t)
        i = i_prev * (1 + g_t * 0.8)
        r = n / i if i > 0 else 0
        sp = r - wacc
        ev = n - wacc * i
        df = 1 / (1 + wacc) ** (t_idx + 0.5)
        pv = ev * df
        proj_nopat.append(n)
        proj_ic.append(i)
        proj_roic.append(r)
        proj_wacc.append(wacc)
        proj_spread.append(sp)
        proj_eva.append(ev)
        proj_df.append(df)
        proj_pv.append(pv)

    # Terminal values
    term_eva = proj_eva[-1] * (1 + g_terminal) / (wacc - g_terminal) if wacc > g_terminal else 0
    term_df  = 1 / (1 + wacc) ** 5
    pv_term  = term_eva * term_df

    metric_rows = [
        ("NOPAT (₹ Cr)",            proj_nopat,  "#,##0.0",  False),
        ("Invested Capital (₹ Cr)", proj_ic,     "#,##0.0",  False),
        ("ROIC (%)",                proj_roic,   "0.00%",    False),
        ("WACC (%)",                proj_wacc,   "0.00%",    False),
        ("Spread (ROIC − WACC)",    proj_spread, "0.00%",    False),
        ("EVA (₹ Cr)",              proj_eva,    "#,##0.0",  False),
        ("Discount Factor (mid-yr)",proj_df,     "0.0000",   False),
        ("PV of EVA (₹ Cr)",        proj_pv,     "#,##0.0",  True),
    ]

    terminal_extras = [None, None, None, None, None,
                       proj_eva[-1] * (1 + g_terminal),
                       term_df, pv_term]

    for idx, (lbl_txt, vals, fmt, is_key) in enumerate(metric_rows):
        _lv(ws, row, 2, lbl_txt, bold=is_key)
        if is_key:
            ws.cell(row=row, column=2).fill = fill(THEME["positive_fill"])
        for ci, v in zip(yr_cols[:5], vals):
            _vc(ws, row, ci, v, fmt=fmt,
                bg=THEME["positive_fill"] if is_key else None,
                color=THEME["positive_text"] if is_key else "000000")
        # Terminal column
        tv = terminal_extras[idx]
        if tv is not None:
            _vc(ws, row, yr_cols[5], tv, fmt=fmt,
                bg=THEME["key_fill"] if is_key else None)
        row += 1

    row += 1

    # ── SECTION 3: TERMINAL VALUE ─────────────────────────────────────────────
    _sh(ws, row, "▌ TERMINAL VALUE")
    row += 1

    pv_eva_sum = sum(proj_pv)
    tv_rows = [
        ("Terminal EVA = EVA_Y5 × (1+g) / (WACC−g)",  proj_eva[-1]*(1+g_terminal)/(wacc-g_terminal) if wacc>g_terminal else 0, "#,##0.0", False),
        ("Terminal Discount Factor",                    term_df,  "0.0000",   False),
        ("★ PV of Terminal EVA (₹ Cr)",                 pv_term,  "#,##0.0",  True),
    ]
    for lbl_txt, val, fmt, is_key in tv_rows:
        if is_key:
            _key_row(ws, row, lbl_txt, val, fmt)
        else:
            _lv(ws, row, 2, lbl_txt)
            _vc(ws, row, 3, val, fmt=fmt)
        row += 1

    row += 1

    # ── SECTION 4: FIRM VALUE BRIDGE ──────────────────────────────────────────
    _sh(ws, row, "▌ FIRM VALUE BRIDGE")
    row += 1

    firm_ev    = ic_cr + pv_eva_sum + pv_term
    eq_val_cr  = firm_ev - net_debt_cr
    implied_px = eq_val_cr * 1e7 / shares_raw if shares_raw > 0 else 0
    upside     = (implied_px / price - 1) if price > 0 else 0

    bridge_rows = [
        ("Current Invested Capital (₹ Cr)",    ic_cr,       "#,##0.0", False),
        ("Add: Sum PV(EVA, Y1–Y5) (₹ Cr)",    pv_eva_sum,  "#,##0.0", False),
        ("Add: PV(Terminal EVA) (₹ Cr)",       pv_term,     "#,##0.0", False),
        ("Enterprise Value (₹ Cr)",            firm_ev,     "#,##0.0", False),
        ("Less: Net Debt (Debt − Cash) (₹ Cr)", net_debt_cr, "#,##0.0", False),
        ("Equity Value (₹ Cr)",                eq_val_cr,   "#,##0.0", False),
        ("Shares Outstanding (Cr)",            shares_cr,   "#,##0.00",False),
        ("★ Implied Share Price (₹)",          implied_px,  "#,##0.00",True),
        ("Current Market Price (₹)",           price,       "#,##0.00",False),
        ("★ Upside/(Downside)",                upside,      "0.0%",    True),
    ]
    for lbl_txt, val, fmt, is_key in bridge_rows:
        if is_key:
            _key_row(ws, row, lbl_txt, val, fmt)
        else:
            _lv(ws, row, 2, lbl_txt)
            _vc(ws, row, 3, val, fmt=fmt)
        row += 1

    row += 1

    # ── SECTION 5: VALUE CREATION METRICS ─────────────────────────────────────
    _sh(ws, row, "▌ VALUE CREATION METRICS")
    row += 1

    mva          = mktcap_cr - ic_cr
    mva_per_sh   = mva * 1e7 / shares_raw if shares_raw > 0 else 0
    spread       = roic - wacc
    val_creation = [eva * 1e7 for eva in proj_eva]

    _lv(ws, row, 2, "Market Value Added (MVA) = MktCap − IC (₹ Cr)")
    _vc(ws, row, 3, mva, "#,##0.0")
    row += 1

    _lv(ws, row, 2, "MVA per Share (₹)")
    _vc(ws, row, 3, mva_per_sh, "#,##0.00")
    row += 1

    _lv(ws, row, 2, "ROIC − WACC Spread")
    _pc(ws, row, 3, spread)
    row += 1

    _lv(ws, row, 2, "Value Creation per Year — Year headers →")
    for ci, lbl_txt in zip(yr_cols[:5], yr_labels[:5]):
        c = ws.cell(row=row, column=ci, value=lbl_txt)
        c.font = font(bold=True, color="FFFFFF", size=8)
        c.fill = fill(THEME["sub"])
        c.alignment = align(h="center")
    row += 1

    _lv(ws, row, 2, "EVA (₹ Cr)")
    cum = 0
    for ci, ev in zip(yr_cols[:5], proj_eva):
        _vc(ws, row, ci, ev, "#,##0.0",
            bg=THEME["positive_fill"] if ev > 0 else "FFC7CE",
            color=THEME["positive_text"] if ev > 0 else "9C0006")
    row += 1

    _lv(ws, row, 2, "Cumulative EVA (₹ Cr)")
    for ci, ev in zip(yr_cols[:5], proj_eva):
        cum += ev
        _vc(ws, row, ci, cum, "#,##0.0", bg=THEME["subtotal_fill"])
    row += 1

    # ── Results summary ────────────────────────────────────────────────────────
    row += 1
    results = [
        ("WACC",                   wacc,       "0.00%", False),
        ("ROIC",                   roic,       "0.00%", False),
        ("ROIC − WACC Spread",     spread,     "0.00%", False),
        ("EVA Year 0 (₹ Cr)",      eva0,       "#,##0.0", False),
        ("PV(EVA 5Y) (₹ Cr)",      pv_eva_sum, "#,##0.0", False),
        ("PV(Terminal EVA) (₹ Cr)", pv_term,   "#,##0.0", False),
        ("Enterprise Value (₹ Cr)", firm_ev,   "#,##0.0", False),
        ("Equity Value (₹ Cr)",    eq_val_cr,  "#,##0.0", False),
        ("★ Implied Share Price (₹)", implied_px, "#,##0.00", True),
        ("★ Upside/(Downside)",    upside,     "0.0%",    True),
    ]
    build_results(ws, THEME, company_name, "EVA Model", results, start_row=row)

    ws.freeze_panes = "C5"

    # ── Sensitivity sheet ──────────────────────────────────────────────────────
    ws_s = wb.create_sheet("Sensitivity")
    ws_s.sheet_view.showGridLines = False
    ws_s.column_dimensions["A"].width = 2
    ws_s.column_dimensions["B"].width = 18

    t2 = ws_s.cell(row=2, column=2, value=f"{company_name} — EVA Sensitivity Analysis")
    t2.font = font(bold=True, size=13)
    ws_s.merge_cells("B2:L2")

    wacc_vals  = [w/100 for w in range(9, 16)]   # 9%–15%
    g_vals     = [g/100 for g in range(4, 17, 2)] # 4%–16%

    base_w_idx = min(range(len(wacc_vals)), key=lambda i: abs(wacc_vals[i] - wacc))
    base_g_idx = min(range(len(g_vals)),    key=lambda i: abs(g_vals[i]   - g1))

    def _implied_price(w, g_rate):
        pv_sum = 0
        n = nopat_cr
        i = ic_cr
        for t_idx in range(5):
            g_step = g_rate + (g_terminal - g_rate) * t_idx / 4
            n = n * (1 + g_step)
            i = i * (1 + g_step * 0.8)
            ev_t = n - w * i
            df_t = 1 / (1 + w) ** (t_idx + 0.5)
            pv_sum += ev_t * df_t
        term_ev_t = (n * (1 + g_terminal) - w * i * (1 + g_terminal))
        if w > g_terminal:
            pv_t = term_ev_t / (w - g_terminal) * (1 / (1 + w) ** 5)
        else:
            pv_t = 0
        fv = ic_cr + pv_sum + pv_t
        eq = fv - net_debt_cr
        return eq * 1e7 / shares_raw if shares_raw > 0 else 0

    matrix = [[_implied_price(w, g) for g in g_vals] for w in wacc_vals]

    build_sensitivity_table(
        ws_s, THEME,
        "WACC", "NOPAT Growth",
        wacc_vals, g_vals, matrix,
        price, base_w_idx, base_g_idx,
        start_row=4,
        row_fmt="0.0%", col_fmt="0.0%",
    )

    # ── Write to bytes ─────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
