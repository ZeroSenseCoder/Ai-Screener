"""
Gordon Growth Model (GGM) — Single-Stage Dividend Discount Model
=================================================================
Core formula:  P = D₁ / (ke − g)

Damodaran Stable-Growth GGM Approach
-------------------------------------
Step 1 : g  = min(b × ROE, rf)            perpetual growth ≤ India nominal GDP (rf proxy)
Step 2 : payout_stable = 1 − g / ROE      payout consistent with stable g
Step 3 : DPS_norm = EPS × payout_stable   potential dividend in stable state
Step 4 : D₁  = DPS_norm × (1 + g)
Step 5 : ke  = rf + β × ERP               CAPM cost of equity
Step 6 : IV  = D₁ / (ke − g)

Why normalize DPS?
  Low-payout growth stocks (e.g., IT/FMCG paying 1% dividend yield) give IV << Price
  when actual DPS is used. Normalization anchors the model to what the company would
  pay in a stable, mature state — which is what a perpetuity model requires.

Implied growth rate from market price (Damodaran):
  Solve  P = D₀_norm × (1+g) / (ke − g)  for g
  → g_implied = (ke × P − D₀_norm) / (P + D₀_norm)
"""
import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.services.excel_utils import (
    font, fill, thin_border, align, safe,
    section_header, build_cover, build_inputs, build_results, build_sensitivity_table,
)

# ── Deep Purple / Gold theme ──────────────────────────────────────────────────
THEME = {
    "cover_bg":      "1A0533",
    "primary":       "2D1B69",
    "sub":           "4A2F9E",
    "accent":        "F4D03F",
    "input_color":   "4A235A",
    "positive_fill": "FEF9E7",
    "positive_text": "7D6608",
    "subtotal_fill": "FDEBD0",
    "key_fill":      "FFF2CC",
    "note_fill":     "E8F4F8",
}

# ── Internal helpers ──────────────────────────────────────────────────────────

def _col(c: int) -> str:
    return get_column_letter(c)


def _hdr(ws, row, cols, theme):
    for i, txt in enumerate(cols):
        c = ws.cell(row=row, column=2 + i, value=txt)
        c.font      = font(bold=True, color="FFFFFF", size=9)
        c.fill      = fill(theme["primary"])
        c.alignment = align(h="center")
        c.border    = thin_border()


def _lbl(ws, row, text, indent=0, bold=False):
    c = ws.cell(row=row, column=2, value=("    " * indent) + text)
    c.font      = font(bold=bold, size=9)
    c.border    = thin_border()
    c.alignment = align()
    return c


def _val(ws, row, col, val, fmt="#,##0.00", bold=False, bg=None, color="000000"):
    c = ws.cell(row=row, column=col, value=val)
    c.font          = font(bold=bold, size=9, color=color)
    c.number_format = fmt
    c.alignment     = align(h="right")
    c.border        = thin_border()
    if bg:
        c.fill = fill(bg)
    return c


# ── Core computation ──────────────────────────────────────────────────────────

def _compute(fin: dict) -> dict:
    price  = safe(fin.get("price"),  100.0)
    dps    = safe(fin.get("dps"),    0.0)
    eps    = safe(fin.get("eps"),    1.0)
    roe    = safe(fin.get("roe"),    0.12)
    beta   = safe(fin.get("beta"),   1.0)
    shares = safe(fin.get("shares"), 1e8)

    # ── CAPM ────────────────────────────────────────────────────────────────
    rf  = 0.071    # India 10Y G-Sec  (also used as perpetual growth cap)
    erp = 0.055    # Damodaran India ERP
    ke  = rf + beta * erp

    # ── Actual payout & retention ────────────────────────────────────────────
    if dps <= 0:
        dps = max(eps * 0.25, 0.01)

    if eps > 0:
        payout_actual = min(dps / eps, 0.95)
    else:
        payout_actual = 0.30
    retention_actual = 1.0 - payout_actual

    # ── Step 1: Stable growth rate ───────────────────────────────────────────
    #   g = min(b × ROE, rf)
    #   Perpetual growth cannot exceed India nominal GDP growth.
    #   rf (7.1%) is used as a proxy for long-run India nominal GDP growth.
    sustainable_g = retention_actual * roe if roe > 0 else 0.0
    g = min(sustainable_g, rf)        # cap at rf  (Damodaran stable-GGM rule)
    g = max(g, 0.02)                  # floor: at minimum 2% for any going concern

    # ── Step 2: Stable payout consistent with g ──────────────────────────────
    #   payout_stable = 1 − g / ROE  (from g = b × ROE)
    if roe > g and roe > 0:
        payout_stable = 1.0 - g / roe
    else:
        # ROE ≤ g  or ROE ≤ 0: company not growing sustainably; use actual payout
        payout_stable = payout_actual

    # Stable payout should not be less than actual (don't reduce dividends)
    payout_stable = max(payout_stable, payout_actual)
    payout_stable = min(payout_stable, 0.95)

    # ── Step 3: Normalized DPS ────────────────────────────────────────────────
    #   What the company should pay in stable state given g and ROE.
    if eps > 0:
        dps_norm = eps * payout_stable
    else:
        dps_norm = dps
    dps_norm = max(dps_norm, dps, 0.01)   # never below actual DPS

    # ── Step 4-6: Core GGM formula ───────────────────────────────────────────
    d1       = dps_norm * (1.0 + g)
    spread   = ke - g
    intrinsic_value = d1 / spread if spread > 0 else 0.0
    upside   = (intrinsic_value - price) / price if price else 0.0

    # ── For reference: GGM using ACTUAL DPS (not normalized) ─────────────────
    d1_actual      = dps * (1.0 + g)
    iv_actual_dps  = d1_actual / spread if spread > 0 else 0.0

    # ── Required D₁ to justify current market price ─────────────────────────
    req_d1 = price * spread

    # ── Implied growth rate from price (Damodaran) ───────────────────────────
    #   Solve P = DPS_norm×(1+g)/(ke−g) for g:
    #   P·ke − DPS_norm = g·(P + DPS_norm)
    #   g_implied = (ke×P − DPS_norm) / (P + DPS_norm)
    denom = price + dps_norm
    implied_g = (ke * price - dps_norm) / denom if denom > 0 else 0.0

    # ── 10-year dividend schedule (using normalized DPS, constant g) ─────────
    divs_10yr = []
    cumulative_pv = 0.0
    d_t = dps_norm
    for t in range(1, 11):
        d_t = d_t * (1.0 + g)
        pv  = d_t / (1.0 + ke) ** t
        cumulative_pv += pv
        divs_10yr.append((t, d_t, g, pv, cumulative_pv))

    pv_beyond_10  = intrinsic_value - cumulative_pv if intrinsic_value > 0 else 0.0
    pv_beyond_pct = pv_beyond_10 / intrinsic_value if intrinsic_value > 0 else 0.0

    # ── Yield & ratio metrics ────────────────────────────────────────────────
    div_yield_actual    = dps      / price if price else 0.0
    div_yield_norm      = dps_norm / price if price else 0.0
    fwd_yield_norm      = d1       / price if price else 0.0
    div_yield_intrinsic = dps_norm / intrinsic_value if intrinsic_value > 0 else 0.0
    implied_pe          = intrinsic_value / eps if eps > 0 else 0.0
    current_pe          = price / eps if eps > 0 else 0.0
    mktcap_cr           = (price * shares) / 1e7

    # ── Stable P/E justified by GGM ─────────────────────────────────────────
    #   IV = EPS × payout_stable × (1+g) / (ke−g)
    #   stable_pe = payout_stable × (1+g) / (ke−g)
    stable_pe = payout_stable * (1.0 + g) / spread if spread > 0 else 0.0

    # ── Model suitability ────────────────────────────────────────────────────
    #   HIGH  : actual div yield > 3%   (utility / mature company)
    #   MEDIUM: actual div yield > 1.5%
    #   LOW   : actual div yield ≤ 1.5% (growth stock — GGM understates value)
    if div_yield_actual >= 0.03:
        suitability = "HIGH — GGM appropriate (stable dividend payer)"
    elif div_yield_actual >= 0.015:
        suitability = "MEDIUM — GGM gives approximate guidance"
    else:
        suitability = "LOW — GGM likely understates value (growth stock with low yield)"

    # How close is implied_g to ke? If > 80% of ke, market expects near-ke growth
    implied_g_pct_ke = implied_g / ke if ke > 0 else 0.0

    # Required div yield for GGM to equal market price (= ke-g)
    req_div_yield = spread  # D1/P = ke-g at fair value

    return dict(
        # inputs
        price=price, dps=dps, eps=eps, roe=roe, beta=beta,
        rf=rf, erp=erp, ke=ke,
        # growth
        sustainable_g=sustainable_g, g=g,
        # payouts
        payout_actual=payout_actual, retention_actual=retention_actual,
        payout_stable=payout_stable,
        # normalized dividends
        dps_norm=dps_norm, d1=d1,
        # GGM output
        spread=spread, intrinsic_value=intrinsic_value, upside=upside,
        # reference: actual-DPS IV
        d1_actual=d1_actual, iv_actual_dps=iv_actual_dps,
        # diagnostics
        req_d1=req_d1, implied_g=implied_g,
        implied_g_pct_ke=implied_g_pct_ke, req_div_yield=req_div_yield,
        # suitability
        suitability=suitability, stable_pe=stable_pe, current_pe=current_pe,
        # schedule
        divs_10yr=divs_10yr, cumulative_pv=cumulative_pv,
        pv_beyond_10=pv_beyond_10, pv_beyond_pct=pv_beyond_pct,
        # yield
        div_yield_actual=div_yield_actual, div_yield_norm=div_yield_norm,
        fwd_yield_norm=fwd_yield_norm,
        div_yield_intrinsic=div_yield_intrinsic,
        implied_pe=implied_pe, mktcap_cr=mktcap_cr,
    )


# ── Sheet 3: Analysis ─────────────────────────────────────────────────────────

def _build_analysis(wb: Workbook, theme: dict, d: dict):
    ws = wb.create_sheet("Gordon Growth Model")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 54
    ws.column_dimensions["C"].width = 16
    for i in range(1, 11):
        ws.column_dimensions[_col(3 + i)].width = 11

    row = 2
    t = ws.cell(row=row, column=2,
                value="Gordon Growth Model (GGM)  —  P = D₁ / (ke − g)  [Damodaran Stable-Growth DDM]")
    t.font = font(bold=True, size=13)
    ws.merge_cells(f"B{row}:{_col(13)}{row}")
    row += 2

    # ── SECTION A: CAPM ───────────────────────────────────────────────────────
    section_header(ws, row, 2, "▌ A.  CAPM — COST OF EQUITY", theme, span=3)
    row += 1
    for lbl_txt, val, fmt_ in [
        ("Risk-Free Rate  rf  (India 10Y G-Sec)",    d["rf"],   "0.0%"),
        ("Equity Risk Premium  ERP  (Damodaran)",     d["erp"],  "0.0%"),
        ("Beta  β",                                   d["beta"], "0.00"),
        ("ke  =  rf + β × ERP",                      d["ke"],   "0.0%"),
    ]:
        _lbl(ws, row, lbl_txt, indent=1)
        _val(ws, row, 3, val, fmt_, bg=theme["subtotal_fill"])
        row += 1
    row += 1

    # ── SECTION B: Growth derivation ──────────────────────────────────────────
    section_header(ws, row, 2, "▌ B.  STABLE GROWTH DERIVATION  (Damodaran Method)", theme, span=3)
    row += 1
    for lbl_txt, val, fmt_ in [
        ("EPS  (₹)",                                          d["eps"],            "#,##0.00"),
        ("Actual DPS  D₀  (₹)",                              d["dps"],            "#,##0.00"),
        ("Actual Payout Ratio  =  DPS / EPS",                d["payout_actual"],  "0.0%"),
        ("Actual Retention Ratio  b  =  1 − Payout",         d["retention_actual"],"0.0%"),
        ("Return on Equity  ROE",                             d["roe"],            "0.0%"),
        ("Current Sustainable Growth  =  b × ROE",           d["sustainable_g"],  "0.0%"),
        ("Perpetual Growth Cap  =  rf  (India nominal GDP proxy)", d["rf"],        "0.0%"),
        ("Applied Stable Growth  g  =  min(b×ROE, rf)",      d["g"],              "0.0%"),
        ("Stable Payout  =  1 − g / ROE",                    d["payout_stable"],  "0.0%"),
        ("Normalized DPS  =  EPS × Stable Payout  (₹)",      d["dps_norm"],       "#,##0.00"),
    ]:
        _lbl(ws, row, lbl_txt, indent=1)
        _val(ws, row, 3, val, fmt_, bg=theme["positive_fill"], color=theme["positive_text"])
        row += 1
    row += 1

    # ── SECTION C: GGM Formula ────────────────────────────────────────────────
    section_header(ws, row, 2, "▌ C.  GORDON GROWTH FORMULA  —  IV = D₁ / (ke − g)", theme, span=3)
    row += 1
    formula_rows = [
        ("Normalized DPS  D₀_norm  (₹)",                        d["dps_norm"],         "#,##0.00",    False),
        ("g   Stable Dividend Growth",                           d["g"],                "0.0%",        False),
        ("D₁  =  D₀_norm × (1 + g)  (₹)",                      d["d1"],               "#,##0.00",    False),
        ("ke  =  rf + β × ERP",                                  d["ke"],               "0.0%",        False),
        ("(ke − g)  Spread",                                     d["spread"],           "0.0%",        False),
        ("★ Intrinsic Value  =  D₁ / (ke − g)  (₹)",            d["intrinsic_value"],  "#,##0.00",    True),
        ("Current Market Price  (₹)",                            d["price"],            "#,##0.00",    False),
        ("Upside / (Downside)  %",                               d["upside"],           "+0.0%;-0.0%", True),
    ]
    for lbl_txt, val, fmt_, is_key in formula_rows:
        _lbl(ws, row, lbl_txt, bold=is_key)
        _val(ws, row, 3, val, fmt_, bold=is_key,
             bg=theme["key_fill"] if is_key else None,
             color=theme["accent"] if is_key else "000000")
        row += 1
    row += 1

    # ── SECTION D: GGM Suitability Check ────────────────────────────────────
    section_header(ws, row, 2, "▌ D.  GGM SUITABILITY CHECK  —  Is this model appropriate?", theme, span=3)
    row += 1
    # Suitability label with conditional color
    suit_cell = ws.cell(row=row, column=2, value=f"    Model Suitability  :  {d['suitability']}")
    suit_cell.font = font(bold=True, size=10,
                          color="1A6B1A" if "HIGH" in d["suitability"]
                          else ("B8860B" if "MEDIUM" in d["suitability"] else "B22222"))
    suit_cell.border = thin_border()
    suit_cell.alignment = align()
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    row += 1
    suitability_rows = [
        ("GGM Stable P/E  =  payout_stable × (1+g) / (ke−g)",   d["stable_pe"],         "0.0x"),
        ("Current Market P/E  =  Price / EPS",                    d["current_pe"],        "0.0x"),
        ("GGM P/E Premium / (Discount)  vs Market",              d["current_pe"] - d["stable_pe"], "0.0x"),
        ("Actual Dividend Yield  D₀ / Price",                     d["div_yield_actual"],  "0.0%"),
        ("Required Div Yield for GGM fair value  =  (ke−g)",     d["req_div_yield"],     "0.0%"),
        ("Div Yield Gap  (Required − Actual)",                    d["req_div_yield"] - d["div_yield_actual"], "0.0%"),
        ("Implied g from Market Price",                            d["implied_g"],         "0.0%"),
        ("Implied g as % of ke  (>80% = market prices near-ke growth)", d["implied_g_pct_ke"], "0.0%"),
    ]
    for lbl_txt, val, fmt_ in suitability_rows:
        _lbl(ws, row, lbl_txt, indent=1)
        _val(ws, row, 3, val, fmt_, bg=theme["note_fill"])
        row += 1
    row += 1

    # ── SECTION E: Actual DPS reference (comparison) ─────────────────────────
    section_header(ws, row, 2, "▌ E.  REFERENCE — GGM WITH ACTUAL DPS (for comparison)", theme, span=3)
    row += 1
    for lbl_txt, val, fmt_ in [
        ("Actual DPS  D₀  (₹)",                         d["dps"],          "#,##0.00"),
        ("D₁_actual  =  D₀ × (1+g)  (₹)",               d["d1_actual"],    "#,##0.00"),
        ("IV (Actual DPS)  =  D₁_actual / (ke−g)  (₹)", d["iv_actual_dps"],"#,##0.00"),
    ]:
        _lbl(ws, row, lbl_txt, indent=1)
        _val(ws, row, 3, val, fmt_, bg=theme["positive_fill"])
        row += 1
    row += 1

    # ── SECTION F: Market Diagnostics ────────────────────────────────────────
    section_header(ws, row, 2, "▌ F.  MARKET PRICE DIAGNOSTICS", theme, span=3)
    row += 1
    for lbl_txt, val, fmt_ in [
        ("Required D₁ to justify market price  =  P × (ke−g)  (₹)", d["req_d1"],          "#,##0.00"),
        ("Normalized Dividend Yield  D₀_norm / P",                   d["div_yield_norm"],  "0.0%"),
        ("Forward Dividend Yield  D₁ / P",                           d["fwd_yield_norm"],  "0.0%"),
        ("Implied P/E at Intrinsic Value",                            d["implied_pe"],      "0.0x"),
    ]:
        _lbl(ws, row, lbl_txt, indent=1)
        _val(ws, row, 3, val, fmt_)
        row += 1
    row += 1

    # ── SECTION G: 10-Year Dividend Schedule ─────────────────────────────────
    YEAR_COLS = [f"Yr {y}" for y in range(1, 11)]
    section_header(ws, row, 2, "▌ G.  10-YEAR DIVIDEND SCHEDULE  (normalized DPS, constant g)", theme, span=11)
    row += 1
    _hdr(ws, row, ["Metric"] + YEAR_COLS, theme)
    row += 1

    divs    = [r[1] for r in d["divs_10yr"]]
    pvs     = [r[3] for r in d["divs_10yr"]]
    cum_pvs = [r[4] for r in d["divs_10yr"]]

    for lbl_txt, vals, fmt_, bg_, clr_ in [
        ("DPS  Dₜ  =  D₀_norm × (1+g)ᵗ  (₹)",   divs,    "#,##0.00", theme["positive_fill"], theme["positive_text"]),
        ("PV of Dividend  Dₜ / (1+ke)ᵗ  (₹)",    pvs,     "#,##0.00", theme["subtotal_fill"], theme["positive_text"]),
        ("Cumulative PV  Σ PV(D₁..Dₜ)  (₹)",     cum_pvs, "#,##0.00", theme["key_fill"],      theme["input_color"]),
    ]:
        _lbl(ws, row, lbl_txt, indent=1)
        for j, v in enumerate(vals):
            _val(ws, row, 3 + j, v, fmt_, bg=bg_, color=clr_ or "000000")
        row += 1
    row += 1

    # ── SECTION H: Value Decomposition ───────────────────────────────────────
    section_header(ws, row, 2, "▌ H.  PERPETUITY VALUE DECOMPOSITION", theme, span=3)
    row += 1
    for lbl_txt, val, fmt_, is_total in [
        ("PV of Dividends Yr 1–10  (₹)",          d["cumulative_pv"],  "#,##0.00", False),
        ("PV of Dividends Beyond Yr 10  (₹)",      d["pv_beyond_10"],   "#,##0.00", False),
        ("Value Beyond Yr 10  (% of IV)",           d["pv_beyond_pct"],  "0.0%",     False),
        ("Total Intrinsic Value  (₹)",              d["intrinsic_value"],"#,##0.00", True),
    ]:
        _lbl(ws, row, lbl_txt, indent=1, bold=is_total)
        _val(ws, row, 3, val, fmt_, bold=is_total,
             bg=theme["key_fill"] if is_total else None)
        row += 1

    ws.freeze_panes = "C4"


# ── Sheet 4: Results & Sensitivity ───────────────────────────────────────────

def _build_results_sheet(wb: Workbook, theme: dict, company_name: str, d: dict):
    ws = wb.create_sheet("Results & Sensitivity")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 16

    results_rows = [
        ("EPS  (₹)",                                   d["eps"],              "#,##0.00",    False),
        ("Actual DPS  D₀  (₹)",                        d["dps"],              "#,##0.00",    False),
        ("Actual Payout  =  DPS / EPS",                d["payout_actual"],    "0.0%",        False),
        ("ROE",                                         d["roe"],              "0.0%",        False),
        ("Current Sustainable Growth  b × ROE",        d["sustainable_g"],    "0.0%",        False),
        ("Applied Stable Growth  g  =  min(b×ROE, rf)",d["g"],                "0.0%",        True),
        ("Stable Payout  =  1 − g / ROE",              d["payout_stable"],    "0.0%",        False),
        ("Normalized DPS  D₀_norm  (₹)",               d["dps_norm"],         "#,##0.00",    False),
        ("ke  (Cost of Equity)",                        d["ke"],               "0.0%",        False),
        ("D₁  =  D₀_norm × (1+g)  (₹)",               d["d1"],               "#,##0.00",    False),
        ("★ Intrinsic Value  =  D₁/(ke−g)  (₹)",      d["intrinsic_value"],  "#,##0.00",    True),
        ("Current Market Price  (₹)",                  d["price"],            "#,##0.00",    False),
        ("Upside / (Downside)",                         d["upside"],           "+0.0%;-0.0%", True),
        ("Implied g from Market Price",                 d["implied_g"],        "0.0%",        False),
        ("Actual Dividend Yield",                       d["div_yield_actual"], "0.0%",        False),
        ("Implied P/E at Intrinsic Value",              d["implied_pe"],       "0.0x",        False),
    ]
    next_row = build_results(ws, theme, company_name, "Gordon Growth Model",
                             results_rows, start_row=2)

    # ── Sensitivity: ke × g using normalized DPS ─────────────────────────────
    ke_vals = [0.09, 0.10, 0.11, 0.12, 0.13, 0.14, 0.15, 0.16]
    g_vals  = [0.02, 0.03, 0.04, 0.05, 0.06, 0.07, 0.08, 0.09]
    base_ki = min(range(len(ke_vals)), key=lambda i: abs(ke_vals[i] - d["ke"]))
    base_gi = min(range(len(g_vals)),  key=lambda i: abs(g_vals[i]  - d["g"]))

    dps_norm = d["dps_norm"]
    matrix = []
    for ke_ in ke_vals:
        row_data = []
        for g_ in g_vals:
            spread_ = ke_ - g_
            if spread_ <= 0:
                row_data.append(0.0)
            else:
                row_data.append(dps_norm * (1.0 + g_) / spread_)
        matrix.append(row_data)

    build_sensitivity_table(
        ws, theme,
        row_label="ke (Cost of Equity)",
        col_label="Stable Dividend Growth  g",
        row_vals=ke_vals,
        col_vals=g_vals,
        matrix=matrix,
        current_price=d["price"],
        base_row_idx=base_ki,
        base_col_idx=base_gi,
        start_row=next_row + 2,
        start_col=2,
        row_fmt="0.0%",
        col_fmt="0.0%",
    )


# ── Public entry point ────────────────────────────────────────────────────────

def generate_excel(fin: dict, symbol: str, company_name: str) -> bytes:
    d  = _compute(fin)
    wb = Workbook()

    build_cover(
        wb, THEME, symbol, company_name,
        model_label="Gordon Growth Model  (Single-Stage DDM)",
        model_desc=(
            "P = D₁ / (ke − g).  "
            "g = min(b×ROE, rf).  "
            "DPS normalized to stable payout = 1 − g/ROE.  "
            "Damodaran stable-growth DDM formulation."
        ),
        sheets_index=[
            (1, "Cover",                "Model overview & navigation index"),
            (2, "Inputs & Assumptions", "Financial data & CAPM parameters"),
            (3, "Gordon Growth Model",  "Growth derivation, GGM formula, diagnostics & 10-yr schedule"),
            (4, "Results & Sensitivity","Summary + ke × g sensitivity heat-map"),
        ],
        meta_extra={
            "price":     d["price"],
            "mktcap_cr": d["mktcap_cr"],
            "sector":    safe(fin.get("sector"), "—"),
        },
    )

    beta = safe(fin.get("beta"), 1.0)
    params_def = [
        ("Risk-Free Rate  rf",                d["rf"],           "%", "India 10Y G-Sec / GDP proxy",  "0.0%", False),
        ("Equity Risk Premium  ERP",          d["erp"],          "%", "Damodaran India ERP",           "0.0%", False),
        ("Beta  β",                           beta,              "x", "Market beta",                   "0.00", True),
        ("Cost of Equity  ke = rf+β×ERP",     d["ke"],           "%", "CAPM required return",          "0.0%", True),
        ("ROE",                               d["roe"],          "%", "Return on Equity (trailing)",   "0.0%", True),
        ("Actual Payout  DPS/EPS",            d["payout_actual"],"%", "Trailing dividend payout",      "0.0%", False),
        ("Applied Stable Growth  g",          d["g"],            "%", "min(b×ROE, rf)",                "0.0%", True),
        ("Stable Payout  = 1−g/ROE",          d["payout_stable"],"%","Consistent with stable g",      "0.0%", True),
        ("Normalized DPS  (₹)",               d["dps_norm"],     "₹", "EPS × Stable Payout",          "#,##0.00", True),
    ]

    build_inputs(wb, THEME, company_name, fin, params_def)
    _build_analysis(wb, THEME, d)
    _build_results_sheet(wb, THEME, company_name, d)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
