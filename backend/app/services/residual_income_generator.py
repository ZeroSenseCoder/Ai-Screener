"""
Residual Income Model (RIM)
Value = BVPS + PV(Excess Returns over 10 years) + PV(Terminal RI)
Excess Return = (ROE − ke) × BVPS_start
Powerful when dividends are low or erratic.
"""
import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.services.excel_utils import (
    font, fill, thin_border, align, safe, cr,
    section_header, sub_header, label, value_cell, pct_cell,
    build_cover, build_inputs, build_results, build_sensitivity_table,
)

THEME = {
    "cover_bg":       "0D2D1A",
    "primary":        "145A32",
    "sub":            "1E8449",
    "accent":         "1ABC9C",
    "input_color":    "0E6655",
    "positive_fill":  "D1F2EB",
    "positive_text":  "0E6655",
    "subtotal_fill":  "A2D9CE",
    "key_fill":       "FDFEFE",
}

# ── helpers ───────────────────────────────────────────────────────────────────

def _col(c: int) -> str:
    return get_column_letter(c)


def _hdr(ws, row: int, cols: list, theme: dict):
    for i, txt in enumerate(cols):
        c = ws.cell(row=row, column=2 + i, value=txt)
        c.font = font(bold=True, color="FFFFFF", size=9)
        c.fill = fill(theme["primary"])
        c.alignment = align(h="center")
        c.border = thin_border()


def _lbl(ws, row, text, indent=0, bold=False):
    c = ws.cell(row=row, column=2, value=("    " * indent) + text)
    c.font = font(bold=bold, size=9)
    c.border = thin_border()
    c.alignment = align()
    return c


def _val(ws, row, col, val, fmt="#,##0.00", bold=False, bg=None, color="000000"):
    c = ws.cell(row=row, column=col, value=val)
    c.font = font(bold=bold, size=9, color=color)
    c.number_format = fmt
    c.alignment = align(h="right")
    c.border = thin_border()
    if bg:
        c.fill = fill(bg)
    return c


# ── computation ───────────────────────────────────────────────────────────────

def _compute(fin: dict):
    price    = safe(fin.get("price"),                100.0)
    bvps     = safe(fin.get("book_value_per_share"),  50.0)
    roe      = safe(fin.get("roe"),                   0.15)
    eps      = safe(fin.get("eps"),                   5.0)
    dps      = safe(fin.get("dps"),                   0.0)
    beta     = safe(fin.get("beta"),                  1.0)
    shares   = safe(fin.get("shares"),                1e8)

    # Payout ratio
    if eps > 0:
        payout = min(dps / eps, 0.90) if dps > 0 else 0.30
    else:
        payout = 0.30

    ke     = 0.071 + beta * 0.055
    g_term = 0.055

    # 10-year BVPS, NI, DPS, RI schedule
    bvps_start = []
    ni_per_shr = []
    dps_per_shr = []
    bvps_end   = []
    eq_charge  = []
    ri_list    = []
    disc_factors = []
    pv_ri_list  = []

    bvps_t = bvps
    for t in range(1, 11):
        bs = bvps_t
        ni = bs * roe
        dp = ni * payout
        be = bs + ni - dp
        ec = ke * bs
        ri = ni - ec

        df  = 1 / (1 + ke) ** t
        pv_ = ri * df

        bvps_start.append(bs)
        ni_per_shr.append(ni)
        dps_per_shr.append(dp)
        bvps_end.append(be)
        eq_charge.append(ec)
        ri_list.append(ri)
        disc_factors.append(df)
        pv_ri_list.append(pv_)

        bvps_t = be

    # Terminal RI value
    ri10      = ri_list[-1]
    spread_tv = ke - g_term
    tv_ri     = ri10 / spread_tv if spread_tv > 0 else 0
    pv_tv_ri  = tv_ri / (1 + ke) ** 10

    sum_pv_ri  = sum(pv_ri_list)
    intrinsic  = bvps + sum_pv_ri + pv_tv_ri
    upside     = (intrinsic - price) / price if price else 0

    # Key metrics
    roe_ke_spread   = roe - ke
    ptb_intrinsic   = intrinsic / bvps if bvps > 0 else 0
    ptb_market      = price / bvps if bvps > 0 else 0
    implied_pe      = intrinsic / eps if eps > 0 else 0
    mktcap_cr       = (price * shares) / 1e7

    return dict(
        price=price, bvps=bvps, roe=roe, eps=eps, dps=dps,
        payout=payout, ke=ke, g_term=g_term,
        bvps_start=bvps_start, ni_per_shr=ni_per_shr,
        dps_per_shr=dps_per_shr, bvps_end=bvps_end,
        eq_charge=eq_charge, ri_list=ri_list,
        disc_factors=disc_factors, pv_ri_list=pv_ri_list,
        ri10=ri10, tv_ri=tv_ri, pv_tv_ri=pv_tv_ri,
        sum_pv_ri=sum_pv_ri, intrinsic=intrinsic, upside=upside,
        roe_ke_spread=roe_ke_spread, ptb_intrinsic=ptb_intrinsic,
        ptb_market=ptb_market, implied_pe=implied_pe,
        mktcap_cr=mktcap_cr,
    )


# ── sheet builders ────────────────────────────────────────────────────────────

def _build_analysis(wb: Workbook, theme: dict, fin: dict, d: dict):
    ws = wb.create_sheet("Residual Income Model")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 12   # "Current" column
    for i in range(10):
        ws.column_dimensions[_col(4 + i)].width = 12

    YEAR_COLS = ["Current"] + [f"Year {y}" for y in range(1, 11)]
    row = 2
    t_title = ws.cell(row=row, column=2,
                      value="Residual Income Model (RIM)  —  Value = BVPS + PV(Excess Returns) + PV(Terminal RI)")
    t_title.font = font(bold=True, size=13)
    ws.merge_cells(f"B{row}:{_col(13)}{row}")
    row += 2

    _hdr(ws, row, ["Metric"] + YEAR_COLS, theme)
    row += 1

    # ── BOOK VALUE PER SHARE BUILD ──
    section_header(ws, row, 2, "▌ BOOK VALUE PER SHARE BUILD", theme, span=12)
    row += 1

    # "Current" column = col 3; Year 1 = col 4 … Year 10 = col 13
    _lbl(ws, row, "BVPS at Start of Period (₹)", indent=1)
    _val(ws, row, 3, d["bvps"], "#,##0.00",
         bg=theme["positive_fill"], color=theme["positive_text"])
    for j, v in enumerate(d["bvps_start"]):
        _val(ws, row, 4 + j, v, "#,##0.00",
             bg=theme["positive_fill"], color=theme["positive_text"])
    row += 1

    _lbl(ws, row, "Net Income per Share (₹)  [= BVPS × ROE]", indent=1)
    _val(ws, row, 3, d["bvps"] * d["roe"], "#,##0.00")
    for j, v in enumerate(d["ni_per_shr"]):
        _val(ws, row, 4 + j, v, "#,##0.00")
    row += 1

    _lbl(ws, row, "Dividends per Share (₹)  [= NI × payout]", indent=1)
    _val(ws, row, 3, d["bvps"] * d["roe"] * d["payout"], "#,##0.00")
    for j, v in enumerate(d["dps_per_shr"]):
        _val(ws, row, 4 + j, v, "#,##0.00")
    row += 1

    _lbl(ws, row, "BVPS at End of Period (₹)", bold=True)
    _val(ws, row, 3, d["bvps"], "#,##0.00",
         bg=theme["subtotal_fill"], color=theme["positive_text"], bold=True)
    for j, v in enumerate(d["bvps_end"]):
        _val(ws, row, 4 + j, v, "#,##0.00",
             bg=theme["subtotal_fill"], color=theme["positive_text"], bold=True)
    row += 2

    # ── RESIDUAL INCOME ──
    section_header(ws, row, 2, "▌ RESIDUAL INCOME CALCULATION", theme, span=12)
    row += 1

    _lbl(ws, row, f"Equity Charge (₹)  [= ke × BVPS_start]  ke={d['ke']:.1%}", indent=1)
    _val(ws, row, 3, d["ke"] * d["bvps"], "#,##0.00")
    for j, v in enumerate(d["eq_charge"]):
        _val(ws, row, 4 + j, v, "#,##0.00")
    row += 1

    _lbl(ws, row, "Net Income per Share (₹)", indent=1)
    _val(ws, row, 3, d["bvps"] * d["roe"], "#,##0.00")
    for j, v in enumerate(d["ni_per_shr"]):
        _val(ws, row, 4 + j, v, "#,##0.00")
    row += 1

    _lbl(ws, row, "Residual Income (₹)  [= NI − Equity Charge]", bold=True)
    _val(ws, row, 3, d["bvps"] * d["roe"] - d["ke"] * d["bvps"], "#,##0.00",
         bg=theme["key_fill"], bold=True)
    for j, v in enumerate(d["ri_list"]):
        bg_ = theme["positive_fill"] if v >= 0 else "FFC7CE"
        tc_ = theme["positive_text"] if v >= 0 else "9C0006"
        _val(ws, row, 4 + j, v, "#,##0.00", bold=True, bg=bg_, color=tc_)
    row += 1

    _lbl(ws, row, "Discount Factor  1/(1+ke)^t", indent=1)
    _val(ws, row, 3, 1.0, "0.0000")
    for j, df in enumerate(d["disc_factors"]):
        _val(ws, row, 4 + j, df, "0.0000")
    row += 1

    _lbl(ws, row, "PV of Residual Income (₹)", bold=True)
    _val(ws, row, 3, 0.0, "#,##0.00")
    for j, v in enumerate(d["pv_ri_list"]):
        bg_ = theme["subtotal_fill"] if v >= 0 else "FFC7CE"
        _val(ws, row, 4 + j, v, "#,##0.00", bold=True, bg=bg_)
    row += 2

    # ── TERMINAL VALUE ──
    section_header(ws, row, 2, "▌ TERMINAL VALUE", theme, span=12)
    row += 1
    tv_rows = [
        ("Final Year RI (Year 10)  (₹)",                    d["ri10"],    "#,##0.00"),
        (f"Terminal RI Value = RI₁₀/(ke−g)  [g={d['g_term']:.1%}]  (₹)",
                                                             d["tv_ri"],   "#,##0.00"),
        ("PV of Terminal RI Value  (₹)",                    d["pv_tv_ri"], "#,##0.00"),
    ]
    for lbl_txt, val, fmt_ in tv_rows:
        _lbl(ws, row, lbl_txt, indent=1)
        _val(ws, row, 3, val, fmt_, bg=theme["key_fill"], bold=True)
        row += 1
    row += 1

    # ── INTRINSIC VALUE BRIDGE ──
    section_header(ws, row, 2, "▌ INTRINSIC VALUE BRIDGE", theme, span=12)
    row += 1
    bridge = [
        ("Current BVPS  (₹)",                     d["bvps"],        "#,##0.00",    False),
        ("Add: Sum PV(Residual Income)  (₹)",      d["sum_pv_ri"],   "#,##0.00",    False),
        ("Add: PV(Terminal RI Value)  (₹)",         d["pv_tv_ri"],    "#,##0.00",    False),
        ("Intrinsic Value per Share  (₹)",          d["intrinsic"],   "#,##0.00",    True),
        ("Current Price  (₹)",                      d["price"],       "#,##0.00",    False),
        ("★ Upside / (Downside) %",                 d["upside"],      "+0.0%;-0.0%", True),
    ]
    for lbl_txt, val, fmt_, is_key in bridge:
        _lbl(ws, row, lbl_txt, bold=is_key)
        _val(ws, row, 3, val, fmt_, bold=is_key,
             bg=theme["key_fill"] if is_key else None,
             color=theme["accent"] if is_key else "000000")
        row += 1
    row += 1

    # ── KEY METRICS ──
    section_header(ws, row, 2, "▌ KEY METRICS", theme, span=12)
    row += 1
    metrics = [
        ("ROE vs ke Spread  (ROE − ke)",    d["roe_ke_spread"], "+0.0%;-0.0%"),
        ("Price-to-Book  (Intrinsic)",       d["ptb_intrinsic"], "0.00x"),
        ("Price-to-Book  (Current Market)", d["ptb_market"],    "0.00x"),
        ("Implied P/E  (Intrinsic / EPS)",  d["implied_pe"],    "0.0x"),
    ]
    for lbl_txt, val, fmt_ in metrics:
        _lbl(ws, row, lbl_txt, indent=1)
        _val(ws, row, 3, val, fmt_)
        row += 1

    ws.freeze_panes = "C4"


def _build_results_sheet(wb: Workbook, theme: dict, company_name: str, d: dict):
    ws = wb.create_sheet("Results & Sensitivity")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 16

    results_rows = [
        ("Current BVPS  (₹)",               d["bvps"],          "#,##0.00",    False),
        ("Sum PV(Residual Income)  (₹)",     d["sum_pv_ri"],     "#,##0.00",    False),
        ("PV of Terminal RI  (₹)",           d["pv_tv_ri"],      "#,##0.00",    False),
        ("★ Intrinsic Value per Share  (₹)", d["intrinsic"],     "#,##0.00",    True),
        ("Current Price  (₹)",               d["price"],         "#,##0.00",    False),
        ("★ Upside / (Downside)",            d["upside"],        "+0.0%;-0.0%", True),
        ("ke  (Cost of Equity)",             d["ke"],            "0.0%",        False),
        ("ROE (assumed constant)",           d["roe"],           "0.0%",        False),
        ("ROE − ke Spread",                  d["roe_ke_spread"], "+0.0%;-0.0%", False),
        ("P/B (Intrinsic)",                  d["ptb_intrinsic"], "0.00x",       False),
        ("P/B (Market)",                     d["ptb_market"],    "0.00x",       False),
        ("Implied P/E",                      d["implied_pe"],    "0.0x",        False),
    ]
    next_row = build_results(ws, theme, company_name, "Residual Income Model",
                             results_rows, start_row=2)

    # Sensitivity: ke × ROE
    ke_vals  = [0.10, 0.11, 0.12, 0.13, 0.14, 0.15, 0.16]
    roe_vals = [0.12, 0.14, 0.16, 0.18, 0.20, 0.22, 0.25]
    base_ki  = min(range(len(ke_vals)),  key=lambda i: abs(ke_vals[i]  - d["ke"]))
    base_ri  = min(range(len(roe_vals)), key=lambda i: abs(roe_vals[i] - d["roe"]))

    bvps     = d["bvps"]
    payout   = d["payout"]
    g_term   = d["g_term"]

    matrix = []
    for ke_ in ke_vals:
        row_data = []
        for roe_ in roe_vals:
            # Recompute intrinsic value
            sum_pv_ = 0.0
            bvps_t  = bvps
            ri_last = 0.0
            for t in range(1, 11):
                bs  = bvps_t
                ni  = bs * roe_
                dp  = ni * payout
                be  = bs + ni - dp
                ec  = ke_ * bs
                ri  = ni - ec
                pv_ = ri / (1 + ke_) ** t
                sum_pv_ += pv_
                bvps_t   = be
                ri_last  = ri

            spread_tv = ke_ - g_term
            if spread_tv <= 0:
                row_data.append(0.0)
                continue
            tv_   = ri_last / spread_tv
            pvtv_ = tv_ / (1 + ke_) ** 10
            iv    = bvps + sum_pv_ + pvtv_
            row_data.append(iv)
        matrix.append(row_data)

    build_sensitivity_table(
        ws, theme,
        row_label="ke (Cost of Equity)",
        col_label="ROE",
        row_vals=ke_vals,
        col_vals=roe_vals,
        matrix=matrix,
        current_price=d["price"],
        base_row_idx=base_ki,
        base_col_idx=base_ri,
        start_row=next_row + 1,
        start_col=2,
        row_fmt="0.0%",
        col_fmt="0.0%",
    )


# ── public entry point ────────────────────────────────────────────────────────

def generate_excel(fin: dict, symbol: str, company_name: str) -> bytes:
    d  = _compute(fin)
    wb = Workbook()

    build_cover(
        wb, THEME, symbol, company_name,
        model_label="Residual Income Model (RIM / EBO)",
        model_desc="Value = BVPS + PV(Excess Returns) + PV(Terminal RI). "
                   "Powerful when dividends are low or erratic.",
        sheets_index=[
            (1, "Cover",                "Model overview & index"),
            (2, "Inputs & Assumptions", "Financial data & model parameters"),
            (3, "Residual Income Model","10-yr RI projections, terminal value & intrinsic value bridge"),
            (4, "Results & Sensitivity","Summary + ke × ROE sensitivity table"),
        ],
        meta_extra={"price": d["price"], "mktcap_cr": d["mktcap_cr"],
                    "sector": safe(fin.get("sector"), "—")},
    )

    beta = safe(fin.get("beta"), 1.0)
    params_def = [
        ("Risk-Free Rate (rf)",         0.071,        "%",  "India 10Y GSec",                 "0.0%", False),
        ("Equity Risk Premium",         0.055,        "%",  "Damodaran India ERP",             "0.0%", False),
        ("Beta",                        beta,         "x",  "Market beta",                     "0.00", True),
        ("Cost of Equity (ke)",         d["ke"],      "%",  "rf + β × ERP",                    "0.0%", True),
        ("ROE (assumed constant)",      d["roe"],     "%",  "Return on equity — key driver",   "0.0%", True),
        ("Terminal Growth (g_term)",    d["g_term"],  "%",  "Long-run India nominal GDP",      "0.0%", True),
        ("Payout Ratio",                d["payout"],  "%",  "DPS / EPS (trailing)",            "0.0%", False),
        ("Current BVPS  (₹)",          d["bvps"],    "₹",  "From balance sheet",              "#,##0.00", True),
    ]

    build_inputs(wb, THEME, company_name, fin, params_def)
    _build_analysis(wb, THEME, fin, d)
    _build_results_sheet(wb, THEME, company_name, d)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
