"""
Trading Comparables Excel Generator
=====================================
Relative valuation using sector peer multiples: P/E, EV/EBITDA, P/B, P/S

Implied Price Formulas
----------------------
P/E     : Price_implied = Peer_Median_PE × EPS
EV/EBITDA: Equity_implied = (Peer_Median_EV × EBITDA − Net_Debt) / Shares
P/B     : Price_implied = Peer_Median_PB × BVPS
P/S     : Price_implied = Peer_Median_PS × Revenue / Shares   [Market Cap multiple — NO debt adj]

NOTE: P/S is a PRICE multiple (market cap / revenue), NOT an EV multiple.
      It does NOT get a debt/cash adjustment. Only EV multiples (EV/EBITDA) get that.
"""

import io
import math
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.services.excel_utils import (
    font, fill, thin_border, align, safe, cr,
    section_header, build_cover, build_inputs, build_sensitivity_table,
)

THEME = {
    "cover_bg":      "0D1B00",
    "primary":       "1B3A00",
    "sub":           "2E6400",
    "accent":        "F39C12",
    "input_color":   "7D4600",
    "positive_fill": "FEF9E7",
    "positive_text": "7D6608",
    "subtotal_fill": "FAD7A0",
    "key_fill":      "FFF2CC",
}

# ── Sector peer multiple reference tables (Indian market) ────────────────────
# Columns: (Name, P/E, EV/EBITDA, P/B, P/S)
# P/S here is Price/Sales (Market Cap / Revenue), a pure price multiple.
_TECH = [
    ("IT Services A",    28.5, 17.2, 6.1, 4.8),
    ("IT Services B",    32.0, 19.5, 7.0, 5.5),
    ("IT Services C",    26.0, 16.0, 5.5, 4.2),
    ("IT Services D",    30.5, 18.8, 6.8, 5.1),
    ("IT Services E",    35.0, 22.0, 8.0, 6.0),
    ("Software F",       27.5, 17.5, 5.9, 4.6),
    ("Software G",       33.0, 20.5, 7.5, 5.8),
    ("Tech Congl H",     29.0, 18.0, 6.3, 4.9),
]
_FMCG = [
    ("FMCG Peer A",      52.0, 32.0, 12.0, 6.0),
    ("FMCG Peer B",      48.0, 29.0, 10.5, 5.5),
    ("FMCG Peer C",      58.0, 36.0, 14.0, 7.0),
    ("FMCG Peer D",      44.0, 27.0,  9.5, 5.0),
    ("FMCG Peer E",      55.0, 33.0, 12.5, 6.5),
    ("FMCG Peer F",      60.0, 38.0, 15.0, 7.5),
    ("FMCG Peer G",      46.0, 28.5, 10.0, 5.2),
    ("FMCG Peer H",      50.0, 31.0, 11.0, 5.8),
]
_BANKING = [
    ("Bank Peer A",      12.0,  8.5, 1.8, 2.5),
    ("Bank Peer B",      14.5, 10.0, 2.2, 3.0),
    ("Bank Peer C",      10.5,  7.5, 1.5, 2.0),
    ("Bank Peer D",      16.0, 11.5, 2.5, 3.5),
    ("Bank Peer E",      11.0,  8.0, 1.6, 2.2),
    ("Bank Peer F",      13.5,  9.5, 2.0, 2.8),
    ("Bank Peer G",      15.0, 10.5, 2.3, 3.2),
    ("Bank Peer H",      12.5,  9.0, 1.9, 2.6),
]
_PHARMA = [
    ("Pharma Peer A",    22.0, 14.5, 3.5, 3.8),
    ("Pharma Peer B",    26.0, 17.0, 4.2, 4.5),
    ("Pharma Peer C",    20.0, 13.0, 3.0, 3.3),
    ("Pharma Peer D",    28.0, 18.5, 4.8, 5.0),
    ("Pharma Peer E",    24.0, 15.5, 3.8, 4.0),
    ("Pharma Peer F",    30.0, 20.0, 5.0, 5.5),
    ("Pharma Peer G",    21.0, 14.0, 3.3, 3.5),
    ("Pharma Peer H",    25.0, 16.5, 4.0, 4.2),
]
_ENERGY = [
    ("Energy Peer A",     9.0,  6.0, 1.2, 1.0),
    ("Energy Peer B",    11.0,  7.5, 1.5, 1.3),
    ("Energy Peer C",     8.0,  5.5, 1.0, 0.9),
    ("Energy Peer D",    13.0,  9.0, 1.8, 1.5),
    ("Energy Peer E",    10.0,  6.8, 1.3, 1.1),
    ("Energy Peer F",    12.0,  8.0, 1.6, 1.4),
    ("Energy Peer G",     9.5,  6.5, 1.2, 1.0),
    ("Energy Peer H",    11.5,  7.8, 1.5, 1.3),
]
_AUTO = [
    ("Auto Peer A",      18.0, 10.5, 3.0, 1.2),
    ("Auto Peer B",      22.0, 12.5, 3.5, 1.5),
    ("Auto Peer C",      15.0,  9.0, 2.5, 1.0),
    ("Auto Peer D",      25.0, 14.0, 4.0, 1.8),
    ("Auto Peer E",      20.0, 11.5, 3.2, 1.3),
    ("Auto Peer F",      17.0,  9.8, 2.8, 1.1),
    ("Auto Peer G",      23.0, 13.0, 3.7, 1.6),
    ("Auto Peer H",      19.0, 11.0, 3.0, 1.2),
]
_INFRA = [
    ("Infra Peer A",     15.0,  9.0, 2.2, 1.0),
    ("Infra Peer B",     18.0, 11.0, 2.8, 1.3),
    ("Infra Peer C",     13.0,  8.0, 1.8, 0.8),
    ("Infra Peer D",     20.0, 12.5, 3.0, 1.5),
    ("Infra Peer E",     16.0,  9.5, 2.4, 1.1),
    ("Infra Peer F",     14.0,  8.5, 2.0, 0.9),
    ("Infra Peer G",     17.0, 10.5, 2.6, 1.2),
    ("Infra Peer H",     19.0, 11.5, 2.9, 1.4),
]
_METALS = [
    ("Metals Peer A",     8.0,  5.0, 1.2, 0.8),
    ("Metals Peer B",    10.0,  6.5, 1.5, 1.0),
    ("Metals Peer C",     7.0,  4.5, 1.0, 0.7),
    ("Metals Peer D",    12.0,  7.5, 1.8, 1.2),
    ("Metals Peer E",     9.0,  5.8, 1.3, 0.9),
    ("Metals Peer F",    11.0,  6.8, 1.6, 1.1),
    ("Metals Peer G",     8.5,  5.5, 1.2, 0.8),
    ("Metals Peer H",    10.5,  6.5, 1.5, 1.0),
]
_REALTY = [
    ("Realty Peer A",    25.0, 15.0, 3.5, 5.0),
    ("Realty Peer B",    30.0, 18.0, 4.0, 6.0),
    ("Realty Peer C",    22.0, 13.0, 3.0, 4.5),
    ("Realty Peer D",    35.0, 21.0, 5.0, 7.0),
    ("Realty Peer E",    28.0, 16.5, 3.8, 5.5),
    ("Realty Peer F",    20.0, 12.0, 2.8, 4.0),
    ("Realty Peer G",    32.0, 19.0, 4.5, 6.5),
    ("Realty Peer H",    26.0, 15.5, 3.5, 5.2),
]
_TELECOM = [
    ("Telecom Peer A",   30.0, 10.0, 3.5, 3.0),
    ("Telecom Peer B",   25.0,  8.5, 3.0, 2.5),
    ("Telecom Peer C",   35.0, 12.0, 4.0, 3.5),
    ("Telecom Peer D",   28.0,  9.5, 3.2, 2.8),
    ("Telecom Peer E",   22.0,  7.5, 2.5, 2.2),
    ("Telecom Peer F",   32.0, 11.0, 3.8, 3.2),
    ("Telecom Peer G",   27.0,  9.0, 3.0, 2.6),
    ("Telecom Peer H",   33.0, 11.5, 3.7, 3.0),
]
DEFAULT_PEERS = [
    ("Sector Peer A",    20.0, 12.0, 3.0, 2.5),
    ("Sector Peer B",    22.0, 13.5, 3.3, 2.8),
    ("Sector Peer C",    18.0, 11.0, 2.7, 2.2),
    ("Sector Peer D",    25.0, 15.0, 3.8, 3.2),
    ("Sector Peer E",    21.0, 12.8, 3.1, 2.6),
    ("Sector Peer F",    23.0, 14.0, 3.5, 3.0),
    ("Sector Peer G",    19.0, 11.5, 2.8, 2.4),
    ("Sector Peer H",    24.0, 14.5, 3.7, 3.1),
]

# Maps yfinance sector names → peer table
# yfinance uses these strings for Indian (NSE/BSE) stocks
SECTOR_PEERS = {
    # Technology / IT
    "Technology":               _TECH,
    "Information Technology":   _TECH,
    "Software":                 _TECH,
    # FMCG / Consumer Staples
    "FMCG":                     _FMCG,
    "Consumer Defensive":       _FMCG,
    "Consumer Staples":         _FMCG,
    # Banking / Finance
    "Banking":                  _BANKING,
    "Financial Services":       _BANKING,
    "Financials":               _BANKING,
    "Banks":                    _BANKING,
    # Pharma / Healthcare
    "Pharmaceuticals":          _PHARMA,
    "Healthcare":               _PHARMA,
    "Health Care":              _PHARMA,
    "Biotechnology":            _PHARMA,
    # Energy / Oil & Gas
    "Energy":                   _ENERGY,
    "Oil & Gas":                _ENERGY,
    "Oil Gas & Consumable Fuels":_ENERGY,
    "Utilities":                _ENERGY,
    # Auto / Consumer Cyclical
    "Auto":                     _AUTO,
    "Automobiles":              _AUTO,
    "Consumer Cyclical":        _AUTO,
    "Consumer Discretionary":   _AUTO,
    # Infrastructure / Industrials
    "Infrastructure":           _INFRA,
    "Industrials":              _INFRA,
    "Construction":             _INFRA,
    "Capital Goods":            _INFRA,
    # Metals / Mining
    "Metals":                   _METALS,
    "Materials":                _METALS,
    "Basic Materials":          _METALS,
    "Mining":                   _METALS,
    # Real Estate
    "Real Estate":              _REALTY,
    "Realty":                   _REALTY,
    # Telecom
    "Telecom":                  _TELECOM,
    "Communication Services":   _TELECOM,
    "Telecommunications":       _TELECOM,
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _col(c):
    return get_column_letter(c)


def _median(vals):
    s = sorted(v for v in vals if v is not None and v > 0)
    if not s:
        return 0.0
    n = len(s)
    return s[n // 2] if n % 2 else (s[n // 2 - 1] + s[n // 2]) / 2.0


def _mean(vals):
    v = [x for x in vals if x is not None and x > 0]
    return sum(v) / len(v) if v else 0.0


def _lbl(ws, row, text, bold=False, bg=None):
    c = ws.cell(row=row, column=2, value=text)
    c.font   = font(bold=bold, size=9)
    c.border = thin_border()
    c.alignment = align()
    if bg:
        c.fill = fill(bg)
    return c


def _val(ws, row, col, val, fmt="#,##0.00", bold=False, bg=None, color="000000"):
    c = ws.cell(row=row, column=col, value=val if val is not None else "N/M")
    c.font          = font(bold=bold, size=9, color=color)
    c.number_format = fmt if val is not None else "@"
    c.alignment     = align(h="right")
    c.border        = thin_border()
    if bg:
        c.fill = fill(bg)
    return c


def _hdr_row(ws, row, cols, theme):
    for i, txt in enumerate(cols):
        c = ws.cell(row=row, column=2 + i, value=txt)
        c.font      = font(bold=True, color="FFFFFF", size=9)
        c.fill      = fill(theme["sub"])
        c.border    = thin_border()
        c.alignment = align(h="center" if i > 0 else "left")


def _stat_row(ws, row, lbl_txt, vals, theme):
    c = ws.cell(row=row, column=2, value=lbl_txt)
    c.font   = font(bold=True, size=9)
    c.fill   = fill(theme["subtotal_fill"])
    c.border = thin_border()
    for j, v in enumerate(vals):
        vc = ws.cell(row=row, column=3 + j, value=round(v, 2) if v else "—")
        vc.font         = font(bold=True, size=9)
        vc.number_format = "0.0x"
        vc.alignment    = align(h="right")
        vc.border       = thin_border()
        vc.fill         = fill(theme["subtotal_fill"])


def _resolve_sector(raw_sector: str):
    """Case-insensitive sector lookup with yfinance alias support."""
    if not raw_sector:
        return DEFAULT_PEERS, "General Market"
    # Direct match
    if raw_sector in SECTOR_PEERS:
        return SECTOR_PEERS[raw_sector], raw_sector
    # Case-insensitive match
    lower = raw_sector.lower()
    for key, peers in SECTOR_PEERS.items():
        if key.lower() == lower:
            return peers, key
    # Partial match
    for key, peers in SECTOR_PEERS.items():
        if key.lower() in lower or lower in key.lower():
            return peers, key
    return DEFAULT_PEERS, "General Market"


# ── Implied price calculations ────────────────────────────────────────────────

def _implied_prices(fin, med_pe, med_ev, med_pb, med_ps):
    """
    Returns (pe_iv, ev_iv, pb_iv, ps_iv) — implied price per share for each method.
    Returns 0.0 for any method where inputs are missing.
    """
    eps    = safe(fin.get("eps"), 0.0)
    bvps   = safe(fin.get("book_value_per_share"), 0.0)
    shares = safe(fin.get("shares"), 1.0)

    ebitda_cr = cr(fin.get("ebitda"))
    rev_cr    = cr(fin.get("revenue"))
    debt_cr   = cr(fin.get("total_debt"))
    cash_cr   = cr(fin.get("cash"))
    shares_cr = shares / 1e7   # shares in crore units (matches ₹Cr)

    # P/E × EPS
    pe_iv = med_pe * eps if (med_pe > 0 and eps > 0) else 0.0

    # EV/EBITDA × EBITDA → equity value → per share
    # Equity = EV − Net Debt = (EV/EBITDA × EBITDA) − Total Debt + Cash
    if med_ev > 0 and ebitda_cr > 0 and shares_cr > 0:
        equity_cr = med_ev * ebitda_cr - debt_cr + cash_cr
        ev_iv = equity_cr / shares_cr
    else:
        ev_iv = 0.0

    # P/B × BVPS
    pb_iv = med_pb * bvps if (med_pb > 0 and bvps > 0) else 0.0

    # P/S × (Revenue / Shares)  — Price multiple, NO debt/cash adjustment
    # Market Cap = P/S × Revenue → Price = Market Cap / Shares = P/S × Revenue / Shares
    if med_ps > 0 and rev_cr > 0 and shares_cr > 0:
        ps_iv = med_ps * rev_cr / shares_cr
    else:
        ps_iv = 0.0

    return pe_iv, ev_iv, pb_iv, ps_iv


# ── Sheet 3: Peer Comparables ─────────────────────────────────────────────────

def _build_analysis(wb, fin, theme):
    ws = wb.create_sheet("Peer Comparables")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 16
    ws.column_dimensions["I"].width = 16

    price     = safe(fin.get("price"), 0.0)
    shares    = safe(fin.get("shares"), 1.0)
    eps       = safe(fin.get("eps"), 0.0)
    bvps      = safe(fin.get("book_value_per_share"), 0.0)
    mktcap_cr = price * shares / 1e7
    rev_cr    = cr(fin.get("revenue"))
    ebitda_cr = cr(fin.get("ebitda"))
    ni_cr     = cr(fin.get("net_income"))
    debt_cr   = cr(fin.get("total_debt"))
    cash_cr   = cr(fin.get("cash"))

    # EV: prefer reported, fallback to mktcap + net debt
    ev_reported = cr(fin.get("enterprise_value"))
    ev_cr = ev_reported if ev_reported > 0 else (mktcap_cr + debt_cr - cash_cr)

    trailing_pe = price / eps       if eps       > 0 else None
    ev_ebitda_s = ev_cr / ebitda_cr if ebitda_cr > 0 else None
    pb_ratio    = price / bvps      if bvps      > 0 else None
    ps_ratio    = mktcap_cr / rev_cr if rev_cr   > 0 else None  # Price / Sales (market cap multiple)

    raw_sector = fin.get("sector", "")
    peers, sector_label = _resolve_sector(raw_sector)

    row = 2
    t = ws.cell(row=row, column=2,
                value="Trading Comparables  —  Relative Valuation using Sector Peer Multiples")
    t.font = font(bold=True, size=13)
    ws.merge_cells(f"B{row}:I{row}")
    row += 2

    # ── SECTION A: Subject Company Metrics ───────────────────────────────────
    section_header(ws, row, 2, "▌ A.  SUBJECT COMPANY METRICS", theme, span=8)
    row += 1
    _hdr_row(ws, row, ["Metric", "Value", "Unit", ""], theme)
    row += 1
    for lbl_txt, val, unit, fmt in [
        ("Revenue  (₹ Cr)",          rev_cr,       "₹ Cr",  "#,##0"),
        ("EBITDA  (₹ Cr)",           ebitda_cr,    "₹ Cr",  "#,##0"),
        ("Net Income  (₹ Cr)",       ni_cr,        "₹ Cr",  "#,##0"),
        ("EPS  (₹)",                 eps,          "₹",     "#,##0.00"),
        ("Book Value / Share  (₹)",  bvps,         "₹",     "#,##0.00"),
        ("Market Cap  (₹ Cr)",       mktcap_cr,    "₹ Cr",  "#,##0"),
        ("Enterprise Value  (₹ Cr)", ev_cr,        "₹ Cr",  "#,##0"),
        ("Trailing P/E  (x)",        trailing_pe,  "x",     "0.0x"),
        ("EV / EBITDA  (x)",         ev_ebitda_s,  "x",     "0.0x"),
        ("P / B  (x)",               pb_ratio,     "x",     "0.0x"),
        ("P / S  (x)  [Mkt Cap / Rev]", ps_ratio,  "x",     "0.0x"),
    ]:
        _lbl(ws, row, "    " + lbl_txt)
        _val(ws, row, 3, val, fmt)
        c = ws.cell(row=row, column=4, value=unit)
        c.font = font(size=9); c.border = thin_border()
        row += 1
    row += 1

    # ── SECTION B: Peer Multiple Benchmarks ──────────────────────────────────
    section_header(ws, row, 2,
                   f"▌ B.  PEER MULTIPLE BENCHMARKS  —  Sector: {sector_label}", theme, span=8)
    row += 1
    _hdr_row(ws, row, ["Company", "P/E (x)", "EV/EBITDA (x)", "P/B (x)", "P/S (x)"], theme)
    row += 1

    pe_vals, ev_vals, pb_vals, ps_vals = [], [], [], []
    for name, pe, ev, pb, ps in peers:
        _lbl(ws, row, "    " + name)
        for col_off, (v, lst) in enumerate([(pe, pe_vals), (ev, ev_vals), (pb, pb_vals), (ps, ps_vals)]):
            lst.append(v)
            vc = ws.cell(row=row, column=3 + col_off, value=v)
            vc.number_format = "0.0x"
            vc.alignment = align(h="right")
            vc.border    = thin_border()
            vc.font      = font(size=9)
        row += 1

    med_pe = _median(pe_vals)
    med_ev = _median(ev_vals)
    med_pb = _median(pb_vals)
    med_ps = _median(ps_vals)

    for stat_lbl, fn in [("Median", _median), ("Mean", _mean),
                          ("Min",    lambda v: min(v)),
                          ("Max",    lambda v: max(v))]:
        _stat_row(ws, row, stat_lbl,
                  [fn(pe_vals), fn(ev_vals), fn(pb_vals), fn(ps_vals)], theme)
        row += 1
    row += 1

    # ── SECTION C: Implied Valuation Matrix ──────────────────────────────────
    section_header(ws, row, 2, "▌ C.  IMPLIED VALUATION MATRIX", theme, span=8)
    row += 1
    _hdr_row(ws, row,
             ["Method", "Sector Median", "Applied To", "Implied Price (₹)", "Current (₹)", "Upside / (Down)"],
             theme)
    row += 1

    pe_iv, ev_iv, pb_iv, ps_iv = _implied_prices(fin, med_pe, med_ev, med_pb, med_ps)

    methods = [
        ("P/E Method",
         med_pe,  f"EPS = ₹{eps:.2f}",                pe_iv),
        ("EV/EBITDA Method",
         med_ev,  f"EBITDA = ₹{ebitda_cr:,.0f} Cr",   ev_iv),
        ("P/B Method",
         med_pb,  f"BVPS = ₹{bvps:.2f}",              pb_iv),
        ("P/S Method  [Mkt Cap / Rev]",
         med_ps,  f"Rev = ₹{rev_cr:,.0f} Cr",         ps_iv),
    ]
    for m_lbl, m_med, m_app, m_iv in methods:
        upside = (m_iv / price - 1) if price > 0 and m_iv > 0 else None
        _lbl(ws, row, "    " + m_lbl)
        _val(ws, row, 3, m_med,  "0.0x")
        c = ws.cell(row=row, column=4, value=m_app)
        c.font = font(size=9); c.border = thin_border()
        _val(ws, row, 5, m_iv if m_iv > 0 else None, "#,##0.00",
             bg=theme["positive_fill"], color=theme["positive_text"])
        _val(ws, row, 6, price, "#,##0.00")
        # upside cell
        if upside is not None:
            uc = ws.cell(row=row, column=7, value=upside)
            uc.number_format = "+0.0%;-0.0%"
            uc.font      = font(size=9, color="375623" if upside >= 0 else "9C0006")
            uc.alignment = align(h="right")
            uc.border    = thin_border()
        else:
            uc = ws.cell(row=row, column=7, value="N/M")
            uc.font = font(size=9); uc.border = thin_border()
        row += 1

    # Blended — only average valid (>0) methods
    valid_ivs = [v for v in [pe_iv, ev_iv, pb_iv, ps_iv] if v > 0]
    blended   = sum(valid_ivs) / len(valid_ivs) if valid_ivs else 0.0
    blended_upside = blended / price - 1 if price > 0 and blended > 0 else None

    for col in range(2, 8):
        c = ws.cell(row=row, column=col)
        c.fill   = fill(theme["key_fill"])
        c.border = thin_border()
    ws.cell(row=row, column=2,
            value=f"★ Blended Average  ({len(valid_ivs)} valid methods, equal weight)").font = font(bold=True, size=9)
    ws.cell(row=row, column=2).fill   = fill(theme["key_fill"])
    ws.cell(row=row, column=2).border = thin_border()
    _val(ws, row, 5, blended if blended > 0 else None, "#,##0.00",
         bold=True, bg=theme["key_fill"], color=theme["accent"])
    _val(ws, row, 6, price, "#,##0.00", bold=True, bg=theme["key_fill"])
    if blended_upside is not None:
        uc = ws.cell(row=row, column=7, value=blended_upside)
        uc.number_format = "+0.0%;-0.0%"
        uc.font      = font(bold=True, size=9,
                            color="375623" if blended_upside >= 0 else "9C0006")
        uc.alignment = align(h="right")
        uc.border    = thin_border()
        uc.fill      = fill(theme["key_fill"])
    row += 2

    # ── SECTION D: Football Field Summary ────────────────────────────────────
    section_header(ws, row, 2, "▌ D.  FOOTBALL FIELD  —  Valuation Range Summary", theme, span=8)
    row += 1
    _hdr_row(ws, row, ["Method", "Bear (-15%)", "Base Case", "Bull (+15%)"], theme)
    row += 1
    for f_lbl, f_base in [
        ("P/E Range",        pe_iv),
        ("EV/EBITDA Range",  ev_iv),
        ("P/B Range",        pb_iv),
        ("P/S Range",        ps_iv),
        ("Blended Range",    blended),
    ]:
        if f_base <= 0:
            continue
        _lbl(ws, row, "    " + f_lbl)
        _val(ws, row, 3, f_base * 0.85, "#,##0.00", bg="FFC7CE", color="9C0006")
        _val(ws, row, 4, f_base,        "#,##0.00", bg=theme["positive_fill"], color=theme["positive_text"])
        _val(ws, row, 5, f_base * 1.15, "#,##0.00", bg="C6EFCE", color="375623")
        # Mark current price
        cp = ws.cell(row=row, column=6, value=price)
        cp.number_format = "#,##0.00"; cp.border = thin_border()
        cp.font = font(size=9, bold=True, color=theme["accent"])
        row += 1

    ws.freeze_panes = "C4"
    return med_pe, med_ev, pe_iv, ev_iv, pb_iv, ps_iv, blended, ebitda_cr, rev_cr


# ── Sheet 4: Results & Sensitivity ───────────────────────────────────────────

def _build_sensitivity(wb, fin, theme, med_pe, med_ev, ebitda_cr, rev_cr):
    ws = wb.create_sheet("Results & Sensitivity")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 22

    price     = safe(fin.get("price"), 0.0)
    eps       = safe(fin.get("eps"), 0.0)
    debt_cr   = cr(fin.get("total_debt"))
    cash_cr   = cr(fin.get("cash"))
    shares_cr = safe(fin.get("shares"), 1.0) / 1e7

    t = ws.cell(row=2, column=2,
                value="Results & Sensitivity Analysis — Trading Comparables")
    t.font = font(bold=True, size=13, color=theme["primary"])
    ws.merge_cells("B2:L2")

    # ── Table 1: P/E × EPS Growth ────────────────────────────────────────────
    base_pe = max(round(med_pe), 5) if med_pe > 0 else 20
    pe_row_vals = [base_pe - 8, base_pe - 6, base_pe - 4, base_pe - 2, base_pe,
                   base_pe + 2, base_pe + 4, base_pe + 6, base_pe + 8]
    eps_growth_cols = [-0.10, -0.05, 0.0, 0.05, 0.10, 0.15, 0.20]
    base_row_idx = 4
    base_col_idx = eps_growth_cols.index(0.0)

    matrix1 = []
    for pe_v in pe_row_vals:
        matrix1.append([round(pe_v * eps * (1.0 + g), 2) for g in eps_growth_cols])

    next_row = build_sensitivity_table(
        ws, theme,
        row_label="P/E Multiple",
        col_label="EPS Growth",
        row_vals=pe_row_vals,
        col_vals=eps_growth_cols,
        matrix=matrix1,
        current_price=price,
        base_row_idx=base_row_idx,
        base_col_idx=base_col_idx,
        start_row=4,
        start_col=2,
        row_fmt="0.0x",
        col_fmt="0.0%",
    )

    # ── Table 2: EV/EBITDA × EBITDA Margin Δ ─────────────────────────────────
    base_ev = max(round(med_ev, 1), 3.0) if med_ev > 0 else 10.0
    ev_row_vals  = [round(base_ev - 3 + i * 1.0, 1) for i in range(7)]
    margin_deltas = [-0.03, -0.02, -0.01, 0.0, 0.01, 0.02, 0.03]
    base_ebitda_margin = (ebitda_cr / rev_cr) if rev_cr > 0 else 0.20
    ev_base_row = 3
    ev_base_col = margin_deltas.index(0.0)

    matrix2 = []
    for ev_v in ev_row_vals:
        r_row = []
        for dm in margin_deltas:
            adj_ebitda = rev_cr * (base_ebitda_margin + dm)
            if shares_cr > 0 and adj_ebitda > 0:
                eq = ev_v * adj_ebitda - debt_cr + cash_cr
                r_row.append(round(eq / shares_cr, 2))
            else:
                r_row.append(0.0)
        matrix2.append(r_row)

    build_sensitivity_table(
        ws, theme,
        row_label="EV/EBITDA Multiple",
        col_label="EBITDA Margin Δ",
        row_vals=ev_row_vals,
        col_vals=margin_deltas,
        matrix=matrix2,
        current_price=price,
        base_row_idx=ev_base_row,
        base_col_idx=ev_base_col,
        start_row=next_row + 1,
        start_col=2,
        row_fmt="0.0x",
        col_fmt="+0.0%;-0.0%;0.0%",
    )

    ws.freeze_panes = "C4"


# ── Public entry point ────────────────────────────────────────────────────────

def generate_excel(fin: dict, symbol: str, company_name: str) -> bytes:
    wb = Workbook()

    price     = safe(fin.get("price"), 0.0)
    shares    = safe(fin.get("shares"), 1.0)
    mktcap_cr = price * shares / 1e7
    sector    = fin.get("sector", "—")
    eps       = safe(fin.get("eps"), 0.0)
    ev_cr_val = cr(fin.get("enterprise_value")) or (mktcap_cr + cr(fin.get("total_debt")) - cr(fin.get("cash")))

    peers, sector_label = _resolve_sector(fin.get("sector", ""))
    med_pe  = _median([p[1] for p in peers])
    med_ev  = _median([p[2] for p in peers])
    med_pb  = _median([p[3] for p in peers])
    med_ps  = _median([p[4] for p in peers])

    build_cover(
        wb, THEME, symbol, company_name,
        model_label="Trading Comparables (Relative Valuation)",
        model_desc=(
            "P/E, EV/EBITDA, P/B, and P/S peer multiples applied to subject company fundamentals. "
            f"Sector: {sector_label}."
        ),
        sheets_index=[
            (1, "Cover",                "Model overview & metadata"),
            (2, "Inputs & Assumptions", "Financial data & model parameters"),
            (3, "Peer Comparables",     "Peer multiples, implied valuation & football field"),
            (4, "Results & Sensitivity","Sensitivity: P/E × EPS growth & EV/EBITDA × margin"),
        ],
        meta_extra={"price": price, "mktcap_cr": mktcap_cr, "sector": sector},
    )

    params_def = [
        ("Sector  (resolved to)",   sector_label, "—",  "yfinance → internal mapping",     "@",    False),
        ("Raw Sector from API",     sector,        "—",  "yfinance classification",          "@",    False),
        ("Peer P/E Median",         med_pe,        "x",  "Sector peer reference table",     "0.0x", True),
        ("Peer EV/EBITDA Median",   med_ev,        "x",  "Sector peer reference table",     "0.0x", True),
        ("Peer P/B Median",         med_pb,        "x",  "Sector peer reference table",     "0.0x", True),
        ("Peer P/S Median",         med_ps,        "x",  "Sector peer reference table",     "0.0x", True),
        ("EPS  (₹)",                eps,           "₹",  "Trailing 12 months",              "#,##0.00", True),
        ("Enterprise Value  (₹ Cr)",ev_cr_val,     "₹ Cr","Reported or computed",           "#,##0", False),
    ]
    build_inputs(wb, THEME, company_name, fin, params_def)

    med_pe_v, med_ev_v, pe_iv, ev_iv, pb_iv, ps_iv, blended, ebitda_cr_v, rev_cr_v = \
        _build_analysis(wb, fin, THEME)
    _build_sensitivity(wb, fin, THEME, med_pe_v, med_ev_v, ebitda_cr_v, rev_cr_v)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
