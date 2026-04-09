"""
DCF Free Cash Flow to Equity (FCFE) Generator
Discounts projected FCFE at cost of equity (ke) — equity-only view.
Appropriate when the firm's debt is stable or for equity-only valuation.
"""
import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.services.excel_utils import (
    font, fill, thin_border, align, safe, cr,
    section_header, sub_header, label, value_cell, pct_cell,
    build_cover, build_inputs, build_results, build_sensitivity_table,
)

THEME = {
    "cover_bg":       "061A2E",
    "primary":        "0D2B44",
    "sub":            "1A4A6E",
    "accent":         "00B4D8",
    "input_color":    "1F497D",
    "positive_fill":  "D0F0F8",
    "positive_text":  "006080",
    "subtotal_fill":  "C5E8F5",
    "key_fill":       "FFF2CC",
}

# ── helpers ──────────────────────────────────────────────────────────────────

def _col(c: int) -> str:
    return get_column_letter(c)


def _hdr(ws, row: int, cols: list, theme: dict):
    """Write year header row."""
    for i, txt in enumerate(cols):
        c = ws.cell(row=row, column=2 + i, value=txt)
        c.font = font(bold=True, color="FFFFFF", size=9)
        c.fill = fill(theme["primary"])
        c.alignment = align(h="center")
        c.border = thin_border()


def _lbl(ws, row, text, indent=0, bold=False):
    c = ws.cell(row=row, column=2, value=("    " * indent) + text)
    c.font = font(bold=bold, size=9)
    c.border = thin_border()
    c.alignment = align()
    return c


def _val(ws, row, col, val, fmt="#,##0.00", bold=False, bg=None, color="000000"):
    c = ws.cell(row=row, column=col, value=val)
    c.font = font(bold=bold, size=9, color=color)
    c.number_format = fmt
    c.alignment = align(h="right")
    c.border = thin_border()
    if bg:
        c.fill = fill(bg)
    return c


# ── core computation ──────────────────────────────────────────────────────────

def _compute(fin: dict):
    price   = safe(fin.get("price"),          100.0)
    shares  = safe(fin.get("shares"),         1e8)
    revenue = safe(fin.get("revenue"),        1e10)
    ni      = safe(fin.get("net_income"),     0.0)
    beta    = safe(fin.get("beta"),           1.0)
    rev_g   = safe(fin.get("revenue_growth"), 0.10)
    debt    = safe(fin.get("total_debt"),     0.0)
    cash_   = safe(fin.get("cash"),           0.0)

    ke   = 0.071 + beta * 0.055
    g    = 0.055
    shares_cr = shares / 1e7

    # Revenue growth schedule
    g_sched = [
        max(rev_g, 0.08),
        max(rev_g * 0.75, 0.06),
        max(rev_g * 0.55, 0.055),
        max(rev_g * 0.40, 0.055),
        0.055,
    ]

    fcfe_margin = max(0.05, min(0.40, ni / revenue if revenue else 0.10))

    revs, fcfes, pv_fcfes = [], [], []
    rev = revenue
    for t_idx, g_t in enumerate(g_sched):
        rev = rev * (1 + g_t)
        revs.append(rev)
        fcfe = rev * fcfe_margin
        fcfes.append(fcfe)
        t = t_idx + 0.5
        df = 1 / (1 + ke) ** t
        pv_fcfes.append(fcfe * df)

    tv         = fcfes[-1] * (1 + g) / (ke - g)
    pv_tv      = tv / (1 + ke) ** 5
    sum_pv     = sum(pv_fcfes)
    eq_value   = sum_pv + pv_tv
    impl_price = (eq_value / 1e7) / shares_cr if shares_cr else 0
    upside     = (impl_price - price) / price if price else 0

    return dict(
        price=price, shares_cr=shares_cr, revenue=revenue, revs=revs,
        ni=ni, fcfe_margin=fcfe_margin, fcfes=fcfes, pv_fcfes=pv_fcfes,
        g_sched=g_sched, ke=ke, g=g,
        tv=tv, pv_tv=pv_tv, sum_pv=sum_pv,
        eq_value=eq_value, impl_price=impl_price, upside=upside,
        debt=debt, cash=cash_,
    )


# ── sheet builders ────────────────────────────────────────────────────────────

def _build_analysis(wb: Workbook, theme: dict, fin: dict, d: dict):
    ws = wb.create_sheet("FCFE Projections")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 38
    for i in range(5):
        ws.column_dimensions[_col(3 + i)].width = 14

    YEARS = [f"FY+{y}E" for y in range(1, 6)]
    row = 2
    t = ws.cell(row=row, column=2, value="DCF — Free Cash Flow to Equity (FCFE) Projections")
    t.font = font(bold=True, size=13)
    ws.merge_cells(f"B{row}:{_col(7)}{row}")
    row += 2

    # Header
    _hdr(ws, row, ["Metric"] + YEARS, theme)
    row += 1

    # ── REVENUE BUILD ──
    section_header(ws, row, 2, "▌ REVENUE BUILD", theme, span=6)
    row += 1
    _lbl(ws, row, "Revenue (₹ Cr)", indent=1)
    for j, v in enumerate(d["revs"]):
        _val(ws, row, 3 + j, cr(v), "#,##0", bg=theme["positive_fill"],
             color=theme["positive_text"])
    row += 1
    _lbl(ws, row, "YoY Revenue Growth %", indent=1)
    for j, g in enumerate(d["g_sched"]):
        _val(ws, row, 3 + j, g, "0.0%")
    row += 2

    # ── FCFE BUILD ──
    section_header(ws, row, 2, "▌ FREE CASH FLOW TO EQUITY BUILD", theme, span=6)
    row += 1

    ni_per_yr = [cr(d["revs"][j]) * d["fcfe_margin"] for j in range(5)]
    dna = cr(safe(fin.get("ebitda"), 0) - safe(fin.get("net_income"), 0)) * 0.6  # proxy
    dna = max(dna, 0)
    capex = dna * 1.3  # rough proxy
    dnwc  = [cr(d["revs"][j]) * 0.01 for j in range(5)]

    _lbl(ws, row, "Net Income (₹ Cr)", indent=1)
    for j, v in enumerate(ni_per_yr):
        _val(ws, row, 3 + j, v, "#,##0")
    row += 1

    _lbl(ws, row, "Add: D&A (₹ Cr)", indent=1)
    for j in range(5):
        _val(ws, row, 3 + j, dna, "#,##0")
    row += 1

    _lbl(ws, row, "Less: CapEx (₹ Cr)", indent=1)
    for j in range(5):
        _val(ws, row, 3 + j, -capex, "#,##0")
    row += 1

    _lbl(ws, row, "Less: ΔNWC (₹ Cr)", indent=1)
    for j in range(5):
        _val(ws, row, 3 + j, -dnwc[j], "#,##0")
    row += 1

    _lbl(ws, row, "Add: Net Borrowing (₹ Cr)", indent=1)
    for j in range(5):
        _val(ws, row, 3 + j, 0, "#,##0")
    row += 1

    fcfe_cr = [cr(v) for v in d["fcfes"]]
    _lbl(ws, row, "FCFE (₹ Cr)", bold=True)
    for j, v in enumerate(fcfe_cr):
        _val(ws, row, 3 + j, v, "#,##0", bold=True,
             bg=theme["subtotal_fill"], color=theme["positive_text"])
    row += 1

    _lbl(ws, row, "FCFE Margin %", indent=1)
    for j in range(5):
        _val(ws, row, 3 + j, d["fcfe_margin"], "0.0%")
    row += 2

    # ── DISCOUNT & PV ──
    section_header(ws, row, 2, "▌ DISCOUNT & PRESENT VALUE", theme, span=6)
    row += 1
    periods = [0.5, 1.5, 2.5, 3.5, 4.5]
    _lbl(ws, row, "Discount Period (t)", indent=1)
    for j, t_ in enumerate(periods):
        _val(ws, row, 3 + j, t_, "0.0")
    row += 1

    dfs = [1 / (1 + d["ke"]) ** t_ for t_ in periods]
    _lbl(ws, row, f"Discount Factor (1/(1+ke)^t)  [ke={d['ke']:.1%}]", indent=1)
    for j, df_ in enumerate(dfs):
        _val(ws, row, 3 + j, df_, "0.0000")
    row += 1

    pv_cr = [cr(pv) for pv in d["pv_fcfes"]]
    _lbl(ws, row, "PV of FCFE (₹ Cr)", bold=True)
    for j, v in enumerate(pv_cr):
        _val(ws, row, 3 + j, v, "#,##0", bold=True,
             bg=theme["subtotal_fill"], color=theme["positive_text"])
    row += 2

    # ── TERMINAL VALUE ──
    section_header(ws, row, 2, "▌ TERMINAL VALUE", theme, span=6)
    row += 1
    _lbl(ws, row, f"Terminal Value  [FCFE₅×(1+g)/(ke−g)]  g={d['g']:.1%}", indent=1)
    _val(ws, row, 3, cr(d["tv"]), "#,##0",
         bg=theme["key_fill"], bold=True)
    row += 1
    _lbl(ws, row, "PV of Terminal Value (₹ Cr)", indent=1)
    _val(ws, row, 3, cr(d["pv_tv"]), "#,##0",
         bg=theme["key_fill"], bold=True)
    row += 2

    # ── EQUITY VALUE BRIDGE ──
    section_header(ws, row, 2, "▌ EQUITY VALUE BRIDGE", theme, span=6)
    row += 1
    rows_bridge = [
        ("Sum PV(FCFE)  (₹ Cr)",          cr(d["sum_pv"]),     "#,##0",    False),
        ("Add: PV(Terminal Value)  (₹ Cr)", cr(d["pv_tv"]),     "#,##0",    False),
        ("Equity Value  (₹ Cr)",           cr(d["eq_value"]),   "#,##0",    True),
        ("Shares Outstanding (Cr)",         d["shares_cr"],      "#,##0.00", False),
        ("★ Implied Share Price (₹)",       d["impl_price"],     "#,##0.00", True),
        ("Upside/(Downside) %",             d["upside"],         "+0.0%;-0.0%", False),
        ("Current Price (ref)  (₹)",        d["price"],          "#,##0.00", False),
    ]
    for lbl_txt, val, fmt_, is_key in rows_bridge:
        _lbl(ws, row, lbl_txt, bold=is_key)
        _val(ws, row, 3, val, fmt_, bold=is_key,
             bg=theme["key_fill"] if is_key else None,
             color=theme["accent"] if is_key else "000000")
        row += 1
    row += 1

    # ── DIAGNOSTICS ──
    section_header(ws, row, 2, "▌ DIAGNOSTICS", theme, span=6)
    row += 1
    tv_pct = cr(d["pv_tv"]) / cr(d["eq_value"]) if d["eq_value"] else 0
    diag = [
        ("TV as % of Equity Value",  tv_pct,   "0.0%"),
        ("ke Used",                  d["ke"],  "0.0%"),
        ("g (Terminal Growth) Used", d["g"],   "0.0%"),
    ]
    for lbl_txt, val, fmt_ in diag:
        _lbl(ws, row, lbl_txt, indent=1)
        _val(ws, row, 3, val, fmt_)
        row += 1

    ws.freeze_panes = "C4"


def _build_results_sheet(wb: Workbook, theme: dict, company_name: str, d: dict, fin: dict):
    ws = wb.create_sheet("Results & Sensitivity")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 16

    results_rows = [
        ("Equity Value (₹ Cr)",          cr(d["eq_value"]),   "#,##0",    False),
        ("Shares Outstanding (Cr)",       d["shares_cr"],      "#,##0.00", False),
        ("★ Implied Share Price (₹)",     d["impl_price"],     "#,##0.00", True),
        ("Current Price (₹)",            d["price"],           "#,##0.00", False),
        ("Upside / (Downside)",          d["upside"],          "+0.0%;-0.0%", True),
        ("TV as % of Equity Value",       cr(d["pv_tv"]) / cr(d["eq_value"]) if d["eq_value"] else 0,
                                                               "0.0%",     False),
        ("ke (Cost of Equity)",           d["ke"],             "0.0%",     False),
        ("Terminal Growth Rate (g)",      d["g"],              "0.0%",     False),
    ]
    next_row = build_results(ws, theme, company_name, "DCF–FCFE", results_rows, start_row=2)

    # Sensitivity: ke (rows) × g (cols)
    ke_vals = [0.10, 0.11, 0.12, 0.13, 0.14, 0.15]
    g_vals  = [0.030, 0.035, 0.040, 0.045, 0.050, 0.055]
    base_ke = d["ke"]
    base_g  = d["g"]
    base_ke_idx = min(range(len(ke_vals)), key=lambda i: abs(ke_vals[i] - base_ke))
    base_g_idx  = min(range(len(g_vals)),  key=lambda i: abs(g_vals[i]  - base_g))

    fcfe5  = d["fcfes"][-1]
    shares = d["shares_cr"]

    matrix = []
    for ke_ in ke_vals:
        row_data = []
        for g_ in g_vals:
            if ke_ <= g_:
                row_data.append(0.0)
                continue
            # recompute PV(FCFE) with new ke
            sum_pv_ = sum(
                d["fcfes"][t_] / (1 + ke_) ** (t_ + 0.5)
                for t_ in range(5)
            )
            tv_      = fcfe5 * (1 + g_) / (ke_ - g_)
            pv_tv_   = tv_ / (1 + ke_) ** 5
            eq_      = sum_pv_ + pv_tv_
            ip_      = (eq_ / 1e7) / shares if shares else 0
            row_data.append(ip_)
        matrix.append(row_data)

    build_sensitivity_table(
        ws, theme,
        row_label="ke (Cost of Equity)",
        col_label="Terminal Growth g",
        row_vals=ke_vals,
        col_vals=g_vals,
        matrix=matrix,
        current_price=d["price"],
        base_row_idx=base_ke_idx,
        base_col_idx=base_g_idx,
        start_row=next_row + 1,
        start_col=2,
        row_fmt="0.0%",
        col_fmt="0.0%",
    )


# ── public entry point ────────────────────────────────────────────────────────

def generate_excel(fin: dict, symbol: str, company_name: str) -> bytes:
    d  = _compute(fin)
    wb = Workbook()

    mktcap_cr = (d["price"] * safe(fin.get("shares"), 1e8)) / 1e7

    build_cover(
        wb, THEME, symbol, company_name,
        model_label="DCF – Free Cash Flow to Equity (FCFE)",
        model_desc="Discounts projected FCFE at cost of equity (ke). "
                   "Equity-only view; suitable when debt is stable.",
        sheets_index=[
            (1, "Cover",                "Model overview & index"),
            (2, "Inputs & Assumptions", "Financial data & model parameters"),
            (3, "FCFE Projections",     "5-year FCFE build, PV, terminal value & equity bridge"),
            (4, "Results & Sensitivity","Summary + ke × g sensitivity table"),
        ],
        meta_extra={"price": d["price"], "mktcap_cr": mktcap_cr,
                    "sector": safe(fin.get("sector"), "—")},
    )

    rev_g   = safe(fin.get("revenue_growth"), 0.10)
    beta    = safe(fin.get("beta"), 1.0)
    params_def = [
        ("Risk-Free Rate (rf)",          0.071,            "%",   "India 10Y GSec yield",         "0.0%",  False),
        ("Equity Risk Premium (ERP)",    0.055,            "%",   "Damodaran India ERP",          "0.0%",  False),
        ("Beta",                         beta,             "x",   "CAPM input",                   "0.00",  True),
        ("Cost of Equity (ke)",          d["ke"],          "%",   "rf + β × ERP",                 "0.0%",  True),
        ("Terminal Growth Rate (g)",     d["g"],           "%",   "Long-run India nominal GDP",   "0.0%",  True),
        ("FCFE Margin (base)",           d["fcfe_margin"], "%",   "NI/Revenue, capped 5%–40%",    "0.0%",  True),
        ("Revenue Growth Yr 1",          max(rev_g, 0.08), "%",  "Base historical YoY",          "0.0%",  False),
        ("Revenue Growth Yr 5",          0.055,            "%",  "Converges to g",               "0.0%",  False),
    ]

    build_inputs(wb, THEME, company_name, fin, params_def)
    _build_analysis(wb, THEME, fin, d)
    _build_results_sheet(wb, THEME, company_name, d, fin)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
