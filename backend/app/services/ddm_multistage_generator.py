"""
Multi-Stage Dividend Discount Model (DDM)  — Three-Stage Damodaran / CFA Formulation
======================================================================================
Stage 1 (Yrs 1-5) : Constant high growth    g1  = Retention Ratio × ROE
Stage 2 (Yrs 6-10): Transition              g linearly declines from g1 → g_term
Terminal           : Gordon Growth perpetuity TV = D₁₁ / (ke − g_term)

Core formula:
  V₀ = Σ [ Dₜ / (1+ke)ᵗ ]  for t = 1..10   +   TV / (1+ke)¹⁰

Growth derivation (Damodaran):
  payout    = DPS / EPS
  b         = 1 − payout          (retention ratio)
  g₁        = b × ROE             (sustainable growth rate)
  ke        = rf + β × ERP        (CAPM cost of equity)
"""
import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.services.excel_utils import (
    font, fill, thin_border, align, safe, cr,
    section_header, build_cover, build_inputs, build_results, build_sensitivity_table,
)

# ── Dark Violet / Purple theme ────────────────────────────────────────────────
THEME = {
    "cover_bg":      "0D0519",
    "primary":       "1A0A30",
    "sub":           "2E155C",
    "accent":        "9B59B6",
    "input_color":   "4A235A",
    "positive_fill": "F5EEF8",
    "positive_text": "6C3483",
    "subtotal_fill": "E8DAEF",
    "key_fill":      "FFF9C4",
    "stage2_fill":   "EDE7F6",
}

# ── Internal helpers ──────────────────────────────────────────────────────────

def _col(c: int) -> str:
    return get_column_letter(c)


def _hdr(ws, row: int, cols: list, theme: dict):
    for i, txt in enumerate(cols):
        c = ws.cell(row=row, column=2 + i, value=txt)
        c.font      = font(bold=True, color="FFFFFF", size=9)
        c.fill      = fill(theme["primary"])
        c.alignment = align(h="center")
        c.border    = thin_border()


def _lbl(ws, row, text, indent=0, bold=False):
    c = ws.cell(row=row, column=2, value=("    " * indent) + text)
    c.font      = font(bold=bold, size=9)
    c.border    = thin_border()
    c.alignment = align()
    return c


def _val(ws, row, col, val, fmt="#,##0.00", bold=False, bg=None, color="000000"):
    c = ws.cell(row=row, column=col, value=val)
    c.font         = font(bold=bold, size=9, color=color)
    c.number_format = fmt
    c.alignment    = align(h="right")
    c.border       = thin_border()
    if bg:
        c.fill = fill(bg)
    return c


# ── Core computation (Damodaran three-stage DDM) ──────────────────────────────

def _compute(fin: dict) -> dict:
    price  = safe(fin.get("price"),  100.0)
    dps    = safe(fin.get("dps"),    0.0)
    eps    = safe(fin.get("eps"),    1.0)
    beta   = safe(fin.get("beta"),   1.0)
    roe    = safe(fin.get("roe"),    0.14)   # Return on Equity (decimal)
    shares = safe(fin.get("shares"), 1e8)

    # ── CAPM cost of equity ──────────────────────────────────────────────────
    rf    = 0.071   # India 10Y G-Sec
    erp   = 0.055   # Damodaran India ERP
    ke    = rf + beta * erp

    # ── Ensure positive DPS (floor at 30% payout if unreported) ─────────────
    if dps <= 0:
        dps = max(eps * 0.30, 0.01)

    # ── Damodaran sustainable growth: g = b × ROE ───────────────────────────
    #   payout  = DPS / EPS
    #   b       = 1 − payout   (retention ratio)
    #   g₁      = b × ROE
    if eps > 0:
        payout    = min(dps / eps, 0.95)
    else:
        payout    = 0.30
    b             = 1.0 - payout                           # retention ratio
    g1_raw        = b * roe                                 # Damodaran formula
    g1            = min(max(g1_raw, 0.04), 0.30)           # clamp 4%–30%

    g_term        = 0.055                                   # India long-run nominal GDP

    # ── Project dividends over 10 years ─────────────────────────────────────
    #   Stage 1 (t = 1..5) : constant g1
    #   Stage 2 (t = 6..10): linear interpolation g1 → g_term
    dividends    = []
    growth_rates = []
    disc_factors = []
    pv_divs      = []
    cum_pv       = []

    d_t        = dps
    running_pv = 0.0

    for t in range(1, 11):
        if t <= 5:
            g_t = g1
        else:
            # linear decline: alpha=0 at t=6 means g1; alpha=1 at t=10 means g_term
            alpha = (t - 5) / 5.0
            g_t   = g1 * (1.0 - alpha) + g_term * alpha

        d_t        = d_t * (1.0 + g_t)
        df         = 1.0 / (1.0 + ke) ** t
        pv         = d_t * df
        running_pv += pv

        dividends.append(d_t)
        growth_rates.append(g_t)
        disc_factors.append(df)
        pv_divs.append(pv)
        cum_pv.append(running_pv)

    # ── Terminal Value (Gordon Growth at end of Year 10) ─────────────────────
    #   D₁₁ = D₁₀ × (1 + g_term)
    #   TV  = D₁₁ / (ke − g_term)
    #   PV(TV) = TV / (1 + ke)¹⁰
    d11    = dividends[-1] * (1.0 + g_term)
    spread = ke - g_term
    tv     = d11 / spread if spread > 0 else 0.0
    pv_tv  = tv / (1.0 + ke) ** 10

    sum_pv_divs = sum(pv_divs)
    intrinsic   = sum_pv_divs + pv_tv
    upside      = (intrinsic - price) / price if price else 0.0
    mktcap_cr   = (price * shares) / 1e7

    # EPS projections — grow at same rates as dividends (keep payout constant)
    eps_proj = []
    e_t = eps if eps > 0 else 1.0
    for t_idx in range(10):
        e_t = e_t * (1.0 + growth_rates[t_idx])
        eps_proj.append(e_t)

    # Projected payout ratios (DPS_t / EPS_t)
    payout_proj = []
    for t_idx in range(10):
        e = eps_proj[t_idx]
        p = dividends[t_idx] / e if e > 0 else payout
        payout_proj.append(min(p, 1.0))

    return dict(
        price=price, dps=dps, eps=eps, ke=ke,
        rf=rf, erp=erp, beta=beta,
        roe=roe, payout=payout, b=b,
        g1=g1, g1_raw=g1_raw, g_term=g_term,
        dividends=dividends, growth_rates=growth_rates,
        disc_factors=disc_factors, pv_divs=pv_divs, cum_pv=cum_pv,
        eps_proj=eps_proj, payout_proj=payout_proj,
        d11=d11, tv=tv, pv_tv=pv_tv,
        sum_pv_divs=sum_pv_divs, intrinsic=intrinsic,
        upside=upside, mktcap_cr=mktcap_cr,
    )


# ── Sheet 3: Analysis ─────────────────────────────────────────────────────────

def _build_analysis(wb: Workbook, theme: dict, d: dict):
    ws = wb.create_sheet("Multi-Stage DDM")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 46
    for i in range(10):
        ws.column_dimensions[_col(3 + i)].width = 11

    YEARS = [f"Yr {y}" for y in range(1, 11)]
    row = 2

    # Title
    t = ws.cell(row=row, column=2,
                value="Multi-Stage DDM  —  Three-Stage Dividend Discount Model (Yrs 1–10 + Terminal)")
    t.font = font(bold=True, size=13)
    ws.merge_cells(f"B{row}:{_col(12)}{row}")
    row += 2

    # ── SECTION A: Fundamental Growth Analysis ────────────────────────────────
    section_header(ws, row, 2, "▌ A.  FUNDAMENTAL GROWTH ANALYSIS  (Damodaran Method)", theme, span=11)
    row += 1

    growth_meta = [
        ("Trailing EPS  (₹)",                       d["eps"],     "#,##0.00"),
        ("Trailing DPS  (₹)",                       d["dps"],     "#,##0.00"),
        ("Payout Ratio  =  DPS / EPS",              d["payout"],  "0.0%"),
        ("Retention Ratio  b  =  1 − Payout",       d["b"],       "0.0%"),
        ("Return on Equity  (ROE)",                  d["roe"],     "0.0%"),
        ("Sustainable Growth  g₁  =  b × ROE",      d["g1_raw"],  "0.0%"),
        ("Clamped g₁  (applied, 4%–30%)",           d["g1"],      "0.0%"),
    ]
    for lbl_txt, val, fmt_ in growth_meta:
        _lbl(ws, row, lbl_txt, indent=1)
        _val(ws, row, 3, val, fmt_, bg=theme["positive_fill"], color=theme["positive_text"])
        row += 1
    row += 1

    # ── SECTION B: Stage Parameter Summary ───────────────────────────────────
    section_header(ws, row, 2, "▌ B.  STAGE PARAMETER SUMMARY", theme, span=11)
    row += 1
    stage_meta = [
        ("Stage 1  (Yr 1–5)   Constant High Growth",          d["g1"],     "0.0%"),
        ("Stage 2  (Yr 6–10)  Linear Decline  g₁ → g_term",  None,        None),
        ("Stage 2  Start  (Yr 6 growth rate)",                d["growth_rates"][5], "0.0%"),
        ("Stage 2  End    (Yr 10 growth rate)",               d["growth_rates"][9], "0.0%"),
        ("Terminal Growth  g_term  (India Nominal GDP)",      d["g_term"], "0.0%"),
        ("Cost of Equity  ke  =  rf + β × ERP",              d["ke"],     "0.0%"),
    ]
    for lbl_txt, val, fmt_ in stage_meta:
        _lbl(ws, row, lbl_txt, indent=1)
        if val is not None:
            _val(ws, row, 3, val, fmt_, bg=theme["subtotal_fill"], color=theme["positive_text"])
        row += 1
    row += 1

    # ── Stage label banners above columns ────────────────────────────────────
    s1 = ws.cell(row=row, column=3, value="◄── Stage 1: High Growth  g₁ = {:.1%} ──►".format(d["g1"]))
    s1.font = font(bold=True, color=theme["accent"], size=9)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
    s2 = ws.cell(row=row, column=8, value="◄── Stage 2: Transition  g₁ → g_term ──►")
    s2.font = font(bold=True, color=theme["input_color"], size=9)
    ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=12)
    row += 1

    _hdr(ws, row, ["Metric"] + YEARS, theme)
    row += 1

    # ── SECTION C: Dividend Projections ──────────────────────────────────────
    section_header(ws, row, 2, "▌ C.  DIVIDEND PROJECTIONS", theme, span=11)
    row += 1

    # Growth rate
    _lbl(ws, row, "Growth Rate  g_t  (%)", indent=1)
    for j, g in enumerate(d["growth_rates"]):
        bg = theme["positive_fill"] if j < 5 else theme["stage2_fill"]
        _val(ws, row, 3 + j, g, "0.0%", bg=bg)
    row += 1

    # DPS
    _lbl(ws, row, "DPS  Dₜ  (₹)  =  D_{t-1} × (1 + g_t)", bold=True)
    for j, v in enumerate(d["dividends"]):
        bg = theme["positive_fill"] if j < 5 else theme["stage2_fill"]
        _val(ws, row, 3 + j, v, "#,##0.00", bold=True,
             bg=bg, color=theme["positive_text"])
    row += 1

    # EPS projected
    _lbl(ws, row, "EPS  (₹)  projected at same g", indent=1)
    for j, v in enumerate(d["eps_proj"]):
        bg = theme["positive_fill"] if j < 5 else theme["stage2_fill"]
        _val(ws, row, 3 + j, v, "#,##0.00", bg=bg)
    row += 1

    # Payout projected
    _lbl(ws, row, "Payout Ratio  DPS/EPS  (%)", indent=1)
    for j, v in enumerate(d["payout_proj"]):
        _val(ws, row, 3 + j, v, "0.0%")
    row += 2

    # ── SECTION D: Present Value of Dividends ────────────────────────────────
    section_header(ws, row, 2, "▌ D.  PRESENT VALUE OF DIVIDENDS", theme, span=11)
    row += 1

    _lbl(ws, row, "Period  t", indent=1)
    for j in range(10):
        _val(ws, row, 3 + j, float(j + 1), "0")
    row += 1

    _lbl(ws, row, f"Discount Factor  1/(1+ke)ᵗ  [ke = {d['ke']:.2%}]", indent=1)
    for j, df in enumerate(d["disc_factors"]):
        _val(ws, row, 3 + j, df, "0.0000")
    row += 1

    _lbl(ws, row, "PV of Dividend  Dₜ/(1+ke)ᵗ  (₹)", bold=True)
    for j, pv in enumerate(d["pv_divs"]):
        _val(ws, row, 3 + j, pv, "#,##0.00", bold=True,
             bg=theme["subtotal_fill"], color=theme["positive_text"])
    row += 1

    _lbl(ws, row, "Cumulative PV  Σ PV(D₁..Dₜ)  (₹)", indent=1)
    for j, cpv in enumerate(d["cum_pv"]):
        _val(ws, row, 3 + j, cpv, "#,##0.00",
             bg=theme["positive_fill"], color=theme["positive_text"])
    row += 2

    # ── SECTION E: Terminal Value ─────────────────────────────────────────────
    section_header(ws, row, 2, "▌ E.  TERMINAL VALUE  (Gordon Growth at end of Year 10)", theme, span=11)
    row += 1
    tv_items = [
        ("D₁₁  =  D₁₀ × (1 + g_term)  (₹)",
             d["d11"], "#,##0.00"),
        (f"TV  =  D₁₁ / (ke − g_term)  "
         f"[ke={d['ke']:.2%}, g={d['g_term']:.2%}, spread={d['ke']-d['g_term']:.2%}]  (₹)",
             d["tv"], "#,##0.00"),
        ("PV(TV)  =  TV / (1+ke)¹⁰  (₹)",
             d["pv_tv"], "#,##0.00"),
        ("TV as % of Total Intrinsic Value",
             d["pv_tv"] / d["intrinsic"] if d["intrinsic"] else 0, "0.0%"),
    ]
    for lbl_txt, val, fmt_ in tv_items:
        _lbl(ws, row, lbl_txt, indent=1)
        _val(ws, row, 3, val, fmt_, bg=theme["key_fill"], bold=True)
        row += 1
    row += 1

    # ── SECTION F: Equity Value Bridge ───────────────────────────────────────
    section_header(ws, row, 2, "▌ F.  EQUITY VALUE BRIDGE  —  INTRINSIC VALUE", theme, span=11)
    row += 1
    bridge = [
        ("( + ) Sum PV(Dividends Yr 1–10)  (₹)",      d["sum_pv_divs"], "#,##0.00",    False),
        ("( + ) PV(Terminal Value)  (₹)",              d["pv_tv"],       "#,##0.00",    False),
        ("★  Intrinsic Value per Share  V₀  (₹)",      d["intrinsic"],   "#,##0.00",    True),
        ("     Current Market Price  (₹)",             d["price"],       "#,##0.00",    False),
        ("     Upside / (Downside)  %",                d["upside"],      "+0.0%;-0.0%", True),
    ]
    for lbl_txt, val, fmt_, is_key in bridge:
        _lbl(ws, row, lbl_txt, bold=is_key)
        _val(ws, row, 3, val, fmt_, bold=is_key,
             bg=theme["key_fill"] if is_key else None,
             color=theme["accent"] if is_key else "000000")
        row += 1

    ws.freeze_panes = "C7"


# ── Sheet 4: Results & Sensitivity ───────────────────────────────────────────

def _build_results_sheet(wb: Workbook, theme: dict, company_name: str, d: dict):
    ws = wb.create_sheet("Results & Sensitivity")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 16

    results_rows = [
        ("Trailing EPS  (₹)",                        d["eps"],         "#,##0.00",    False),
        ("Trailing DPS  (₹)",                        d["dps"],         "#,##0.00",    False),
        ("Payout Ratio  =  DPS / EPS",               d["payout"],      "0.0%",        False),
        ("Retention Ratio  b  =  1 − Payout",        d["b"],           "0.0%",        False),
        ("ROE",                                       d["roe"],         "0.0%",        False),
        ("Sustainable Growth  g₁  =  b × ROE",       d["g1_raw"],      "0.0%",        False),
        ("Applied g₁  (clamped 4%–30%)",             d["g1"],          "0.0%",        True),
        ("Terminal Growth  g_term",                   d["g_term"],      "0.0%",        False),
        ("Cost of Equity  ke  =  rf + β × ERP",      d["ke"],          "0.0%",        False),
        ("Sum PV(Dividends Yr 1–10)  (₹)",           d["sum_pv_divs"], "#,##0.00",    False),
        ("PV(Terminal Value)  (₹)",                  d["pv_tv"],       "#,##0.00",    False),
        ("★ Intrinsic Value per Share  V₀  (₹)",     d["intrinsic"],   "#,##0.00",    True),
        ("Current Market Price  (₹)",                d["price"],       "#,##0.00",    False),
        ("Upside / (Downside)",                       d["upside"],      "+0.0%;-0.0%", True),
    ]
    next_row = build_results(ws, theme, company_name, "Multi-Stage DDM",
                             results_rows, start_row=2)

    # ── Sensitivity Table: ke (rows) × g1 (cols) ─────────────────────────────
    ke_vals  = [0.09, 0.10, 0.11, 0.12, 0.13, 0.14, 0.15, 0.16]
    g1_vals  = [0.04, 0.06, 0.08, 0.10, 0.12, 0.15, 0.18, 0.22]
    base_ki  = min(range(len(ke_vals)),  key=lambda i: abs(ke_vals[i]  - d["ke"]))
    base_gi  = min(range(len(g1_vals)),  key=lambda i: abs(g1_vals[i]  - d["g1"]))

    dps    = d["dps"]
    g_term = d["g_term"]

    matrix = []
    for ke_ in ke_vals:
        row_data = []
        for g1_ in g1_vals:
            sum_pv_ = 0.0
            d_t = dps
            for t in range(1, 11):
                if t <= 5:
                    g_t = g1_
                else:
                    alpha = (t - 5) / 5.0
                    g_t   = g1_ * (1.0 - alpha) + g_term * alpha
                d_t      = d_t * (1.0 + g_t)
                sum_pv_ += d_t / (1.0 + ke_) ** t

            d11_    = d_t * (1.0 + g_term)
            spread_ = ke_ - g_term
            if spread_ <= 0:
                row_data.append(0.0)
                continue
            tv_     = d11_ / spread_
            pv_tv_  = tv_ / (1.0 + ke_) ** 10
            row_data.append(sum_pv_ + pv_tv_)
        matrix.append(row_data)

    build_sensitivity_table(
        ws, theme,
        row_label="ke (Cost of Equity)",
        col_label="Stage 1 Growth  g₁  (b × ROE)",
        row_vals=ke_vals,
        col_vals=g1_vals,
        matrix=matrix,
        current_price=d["price"],
        base_row_idx=base_ki,
        base_col_idx=base_gi,
        start_row=next_row + 2,
        start_col=2,
        row_fmt="0.0%",
        col_fmt="0.0%",
    )


# ── Public entry point ────────────────────────────────────────────────────────

def generate_excel(fin: dict, symbol: str, company_name: str) -> bytes:
    d  = _compute(fin)
    wb = Workbook()

    build_cover(
        wb, THEME, symbol, company_name,
        model_label="Multi-Stage DDM  (Three-Stage Dividend Discount Model)",
        model_desc=(
            "Stage 1 (Yr 1-5): constant g₁ = b×ROE.  "
            "Stage 2 (Yr 6-10): linear transition g₁ → g_term.  "
            "Terminal: Gordon Growth perpetuity.  "
            "Damodaran / CFA Level II formulation."
        ),
        sheets_index=[
            (1, "Cover",                "Model overview & navigation index"),
            (2, "Inputs & Assumptions", "Financial data & CAPM parameters"),
            (3, "Multi-Stage DDM",      "Growth analysis, 10-yr projections, terminal value & equity bridge"),
            (4, "Results & Sensitivity","Summary + ke × g₁ sensitivity heat-map"),
        ],
        meta_extra={
            "price":     d["price"],
            "mktcap_cr": d["mktcap_cr"],
            "sector":    safe(fin.get("sector"), "—"),
        },
    )

    beta = safe(fin.get("beta"), 1.0)
    params_def = [
        ("Risk-Free Rate  rf",          d["rf"],     "%", "India 10Y G-Sec",           "0.0%", False),
        ("Equity Risk Premium  ERP",    d["erp"],    "%", "Damodaran India ERP",        "0.0%", False),
        ("Beta  β",                     beta,        "x", "Market beta",                "0.00", True),
        ("Cost of Equity  ke = rf+β×ERP", d["ke"],  "%", "CAPM required return",       "0.0%", True),
        ("Return on Equity  ROE",       d["roe"],    "%", "Trailing annual ROE",        "0.0%", True),
        ("Payout Ratio  DPS/EPS",       d["payout"], "%", "Trailing dividend payout",  "0.0%", False),
        ("Retention Ratio  b = 1−payout", d["b"],   "%", "Earnings reinvested",        "0.0%", True),
        ("Stage 1 Growth  g₁ = b×ROE", d["g1"],     "%", "Sustainable growth (clamped)","0.0%", True),
        ("Terminal Growth  g_term",     d["g_term"], "%", "India long-run nominal GDP", "0.0%", True),
    ]

    build_inputs(wb, THEME, company_name, fin, params_def)
    _build_analysis(wb, THEME, d)
    _build_results_sheet(wb, THEME, company_name, d)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
