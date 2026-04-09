"""
Real Options Valuation
Values managerial flexibility (invest / expand / abandon) using Black-Scholes.
norm_cdf via math.erf — no scipy dependency.
"""
import io
import math
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.services.excel_utils import (
    font, fill, thin_border, align, safe, cr,
    section_header, build_cover, build_inputs, build_sensitivity_table,
)

THEME = {
    "cover_bg":      "160040",
    "primary":       "2D0080",
    "sub":           "4D00CC",
    "accent":        "E91E63",
    "input_color":   "880E4F",
    "positive_fill": "FCE4EC",
    "positive_text": "880E4F",
    "subtotal_fill": "F48FB1",
    "key_fill":      "FFF2CC",
}


def _col(c: int) -> str:
    return get_column_letter(c)


def _lbl(ws, row, text, indent=0, bold=False):
    c = ws.cell(row=row, column=2, value=("    " * indent) + text)
    c.font = font(bold=bold, size=9)
    c.border = thin_border()
    c.alignment = align()
    return c


def _val(ws, row, col, val, fmt="#,##0.00", bold=False, bg=None, color="000000"):
    c = ws.cell(row=row, column=col, value=val)
    c.font = font(bold=bold, size=9, color=color)
    c.number_format = fmt
    c.alignment = align(h="right")
    c.border = thin_border()
    if bg:
        c.fill = fill(bg)
    return c


def _hdr(ws, row, cols, theme):
    for i, txt in enumerate(cols):
        c = ws.cell(row=row, column=2 + i, value=txt)
        c.font = font(bold=True, color="FFFFFF", size=9)
        c.fill = fill(theme["primary"])
        c.alignment = align(h="center")
        c.border = thin_border()


def _norm_cdf(x: float) -> float:
    return 0.5 * (1.0 + math.erf(x / math.sqrt(2)))


def _bs_option(V, X, T, r, sigma):
    """Returns (call_val, put_val, d1, d2)."""
    if T <= 0 or sigma <= 0 or V <= 0 or X <= 0:
        call = max(V - X, 0.0)
        put  = max(X - V, 0.0)
        return call, put, 0.0, 0.0
    d1 = (math.log(V / X) + (r + sigma ** 2 / 2) * T) / (sigma * math.sqrt(T))
    d2 = d1 - sigma * math.sqrt(T)
    call = V * _norm_cdf(d1) - X * math.exp(-r * T) * _norm_cdf(d2)
    put  = X * math.exp(-r * T) * _norm_cdf(-d2) - V * _norm_cdf(-d1)
    return call, put, d1, d2


def _compute(fin: dict):
    price      = safe(fin.get("price"), 100.0)
    shares     = safe(fin.get("shares"), 1e8)
    beta       = safe(fin.get("beta"), 1.0)
    op_cf      = cr(fin.get("operating_cf"))
    ta         = cr(fin.get("total_assets"))
    sigma      = safe(fin.get("volatility_annual"), 0.35)
    sigma      = max(sigma, 0.10)

    V  = op_cf * 10        # PV proxy
    X  = ta * 0.20         # capex proxy
    T  = 3.0
    r  = 0.071

    call_val, put_val, d1, d2 = _bs_option(V, X, T, r, sigma)
    npv     = V - X
    premium = call_val - npv
    decision = "Invest now" if npv > call_val else "Wait — option value exceeds NPV"
    abandon_premium = put_val

    # Binomial tree (3 periods)
    dt    = T / 3
    u_b   = math.exp(sigma * math.sqrt(dt))
    d_b   = 1.0 / u_b
    p_b   = (math.exp(r * dt) - d_b) / (u_b - d_b) if (u_b - d_b) else 0.5
    disc  = math.exp(-r * dt)

    # Terminal nodes (4): VU3, VU2D, VUD2, VD3
    V_u3  = V * u_b ** 3
    V_u2d = V * u_b ** 2 * d_b
    V_ud2 = V * u_b * d_b ** 2
    V_d3  = V * d_b ** 3

    # Option payoffs at terminal
    ov_u3  = max(V_u3  - X, 0)
    ov_u2d = max(V_u2d - X, 0)
    ov_ud2 = max(V_ud2 - X, 0)
    ov_d3  = max(V_d3  - X, 0)

    # Step back
    ov_u2  = disc * (p_b * ov_u3  + (1 - p_b) * ov_u2d)
    ov_ud  = disc * (p_b * ov_u2d + (1 - p_b) * ov_ud2)
    ov_d2  = disc * (p_b * ov_ud2 + (1 - p_b) * ov_d3)
    V_u2   = V * u_b ** 2
    V_ud_  = V * u_b * d_b
    V_d2   = V * d_b ** 2

    ov_u   = disc * (p_b * ov_u2 + (1 - p_b) * ov_ud)
    ov_d   = disc * (p_b * ov_ud + (1 - p_b) * ov_d2)
    V_u_   = V * u_b
    V_d_   = V * d_b

    ov_0   = disc * (p_b * ov_u + (1 - p_b) * ov_d)

    strategic_val = npv + premium
    mktcap_cr = price * shares / 1e7

    return dict(
        price=price, shares=shares, beta=beta,
        V=V, X=X, T=T, r=r, sigma=sigma,
        call_val=call_val, put_val=put_val, d1=d1, d2=d2,
        npv=npv, premium=premium, decision=decision,
        abandon_premium=abandon_premium,
        u_b=u_b, d_b=d_b, p_b=p_b,
        tree_nodes={
            "V":    V,    "V_u":  V_u_,  "V_d":  V_d_,
            "V_u2": V_u2, "V_ud": V_ud_, "V_d2": V_d2,
            "V_u3": V_u3, "V_u2d":V_u2d,"V_ud2":V_ud2,"V_d3":V_d3,
        },
        option_nodes={
            "ov0":  ov_0,  "ov_u": ov_u,  "ov_d": ov_d,
            "ov_u2":ov_u2, "ov_ud":ov_ud, "ov_d2":ov_d2,
            "ov_u3":ov_u3, "ov_u2d":ov_u2d,"ov_ud2":ov_ud2,"ov_d3":ov_d3,
        },
        strategic_val=strategic_val, mktcap_cr=mktcap_cr,
    )


def _build_analysis(wb, theme, fin, d):
    ws = wb.create_sheet("Real Options Analysis")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 16
    for i in range(4, 8):
        ws.column_dimensions[_col(i)].width = 14

    row = 2
    t = ws.cell(row=row, column=2,
                value="Real Options Valuation  —  Managerial Flexibility via Black-Scholes")
    t.font = font(bold=True, size=13)
    ws.merge_cells(f"B{row}:{_col(8)}{row}")
    row += 2

    # ── PARAMETERS ──
    section_header(ws, row, 2, "▌ REAL OPTION PARAMETERS", theme, span=3)
    row += 1
    params = [
        ("Underlying Value  V = OpCF × 10  (₹ Cr)", d["V"],     "#,##0"),
        ("Investment Cost   X = Total Assets × 20%  (₹ Cr)", d["X"], "#,##0"),
        ("Time Horizon  T  (years)",                 d["T"],     "0.0"),
        ("Volatility  σ (annual)",                   d["sigma"], "0.0%"),
        ("Risk-Free Rate  r",                        d["r"],     "0.0%"),
    ]
    for lbl_txt, val, fmt_ in params:
        _lbl(ws, row, lbl_txt, indent=1)
        _val(ws, row, 3, val, fmt_)
        row += 1
    row += 1

    # ── OPTION TO INVEST (CALL) ──
    section_header(ws, row, 2, "▌ OPTION TO INVEST  (CALL)", theme, span=3)
    row += 1
    Nd1 = _norm_cdf(d["d1"])
    Nd2 = _norm_cdf(d["d2"])
    invest_rows = [
        ("d₁",                                              d["d1"],       "0.0000", False),
        ("d₂  =  d₁ − σ√T",                               d["d2"],       "0.0000", False),
        ("N(d₁)",                                           Nd1,            "0.0000", False),
        ("N(d₂)",                                           Nd2,            "0.0000", False),
        ("★ Call Value = V·N(d₁) − X·e^(−rT)·N(d₂)",      d["call_val"], "#,##0",   True),
        ("NPV = V − X",                                     d["npv"],      "#,##0",   False),
        ("Option Premium over NPV = Call − NPV",            d["premium"],  "#,##0",   False),
        ("Decision",                                        d["decision"], "@",       False),
    ]
    for lbl_txt, val, fmt_, is_key in invest_rows:
        _lbl(ws, row, lbl_txt, bold=is_key)
        _val(ws, row, 3, val, fmt_, bold=is_key,
             bg=theme["key_fill"] if is_key else None,
             color=theme["accent"] if is_key else "000000")
        row += 1
    row += 1

    # ── OPTION TO ABANDON (PUT) ──
    section_header(ws, row, 2, "▌ OPTION TO ABANDON  (PUT)", theme, span=3)
    row += 1
    Nnd2 = _norm_cdf(-d["d2"])
    Nnd1 = _norm_cdf(-d["d1"])
    abandon_rows = [
        ("★ Put Value = X·e^(−rT)·N(−d₂) − V·N(−d₁)",   d["put_val"],          "#,##0",  True),
        ("Abandonment Value Premium  (₹ Cr)",              d["abandon_premium"],  "#,##0",  False),
        ("Interpretation",                                 "Insurance against downside", "@", False),
    ]
    for lbl_txt, val, fmt_, is_key in abandon_rows:
        _lbl(ws, row, lbl_txt, bold=is_key)
        _val(ws, row, 3, val, fmt_, bold=is_key,
             bg=theme["key_fill"] if is_key else None,
             color=theme["accent"] if is_key else "000000")
        row += 1
    row += 1

    # ── BINOMIAL TREE ──
    section_header(ws, row, 2, "▌ BINOMIAL TREE  (3 PERIODS)", theme, span=5)
    row += 1
    tn = d["tree_nodes"]
    on = d["option_nodes"]
    _hdr(ws, row, ["Node", "Underlying V", "Option Value"], theme)
    row += 1
    tree_data = [
        ("t=0:  V",       tn["V"],    on["ov0"]),
        ("t=1:  V·u",     tn["V_u"],  on["ov_u"]),
        ("t=1:  V·d",     tn["V_d"],  on["ov_d"]),
        ("t=2:  V·u²",    tn["V_u2"], on["ov_u2"]),
        ("t=2:  V·ud",    tn["V_ud"], on["ov_ud"]),
        ("t=2:  V·d²",    tn["V_d2"], on["ov_d2"]),
        ("t=3:  V·u³",    tn["V_u3"], on["ov_u3"]),
        ("t=3:  V·u²d",   tn["V_u2d"],on["ov_u2d"]),
        ("t=3:  V·ud²",   tn["V_ud2"],on["ov_ud2"]),
        ("t=3:  V·d³",    tn["V_d3"], on["ov_d3"]),
    ]
    for node_lbl, v_node, ov_node in tree_data:
        _lbl(ws, row, node_lbl, indent=1)
        _val(ws, row, 3, v_node,  "#,##0")
        _val(ws, row, 4, ov_node, "#,##0")
        row += 1
    row += 1

    # ── STRATEGIC VALUE ──
    section_header(ws, row, 2, "▌ STRATEGIC VALUE SUMMARY", theme, span=3)
    row += 1
    strat = [
        ("DCF NPV = V − X  (₹ Cr)",                          d["npv"],           "#,##0",  False),
        ("Black-Scholes Option Value  (₹ Cr)",                d["call_val"],      "#,##0",  False),
        ("Real Option Premium = BS − NPV  (₹ Cr)",            d["premium"],       "#,##0",  False),
        ("★ Total Strategic Value = NPV + Premium  (₹ Cr)",   d["strategic_val"], "#,##0",  True),
    ]
    for lbl_txt, val, fmt_, is_key in strat:
        _lbl(ws, row, lbl_txt, bold=is_key)
        _val(ws, row, 3, val, fmt_, bold=is_key,
             bg=theme["key_fill"] if is_key else None,
             color=theme["accent"] if is_key else "000000")
        row += 1

    ws.freeze_panes = "C4"


def _build_results_sheet(wb, theme, company_name, d):
    ws = wb.create_sheet("Results & Sensitivity")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 16

    row = 2
    section_header(ws, row, 2, "▌ RESULTS SUMMARY", theme, span=4)
    row += 1
    results = [
        ("Underlying Value V  (₹ Cr)",               d["V"],            "#,##0",  False),
        ("Investment Cost X  (₹ Cr)",                d["X"],            "#,##0",  False),
        ("Volatility σ",                              d["sigma"],        "0.0%",   False),
        ("★ Call (Option to Invest)  (₹ Cr)",         d["call_val"],     "#,##0",  True),
        ("★ Put  (Option to Abandon) (₹ Cr)",         d["put_val"],      "#,##0",  True),
        ("NPV  (₹ Cr)",                              d["npv"],           "#,##0",  False),
        ("★ Total Strategic Value  (₹ Cr)",           d["strategic_val"],"#,##0",  True),
    ]
    for lbl_txt, val, fmt_, is_key in results:
        lc = ws.cell(row=row, column=2, value=lbl_txt)
        lc.font = font(bold=is_key, size=9)
        lc.border = thin_border()
        if is_key:
            lc.fill = fill(theme["key_fill"])
        vc = ws.cell(row=row, column=3, value=val)
        vc.font = font(bold=is_key, size=9,
                       color=theme["accent"] if is_key else "000000")
        vc.number_format = fmt_
        vc.alignment = align(h="right")
        vc.border = thin_border()
        if is_key:
            vc.fill = fill(theme["key_fill"])
        row += 1

    # Sensitivity: Volatility σ × Time Horizon T
    sigma_vals = [0.20, 0.30, 0.40, 0.50, 0.60, 0.70, 0.80]
    t_vals     = [1.0, 2.0, 3.0, 4.0, 5.0]

    V = d["V"]
    X = d["X"]
    r = d["r"]

    base_si = min(range(len(sigma_vals)), key=lambda i: abs(sigma_vals[i] - d["sigma"]))
    base_ti = 2  # T=3 is index 2

    matrix = []
    for sv in sigma_vals:
        row_data = []
        for tv in t_vals:
            call_, _, _, _ = _bs_option(V, X, tv, r, sv)
            row_data.append(round(call_, 2))
        matrix.append(row_data)

    build_sensitivity_table(
        ws, theme,
        row_label="Volatility σ",
        col_label="Time Horizon T (years)",
        row_vals=sigma_vals,
        col_vals=t_vals,
        matrix=matrix,
        current_price=d["call_val"],
        base_row_idx=base_si,
        base_col_idx=base_ti,
        start_row=row + 2,
        start_col=2,
        row_fmt="0.0%",
        col_fmt="0.0",
    )


def generate_excel(fin: dict, symbol: str, company_name: str) -> bytes:
    d  = _compute(fin)
    wb = Workbook()

    build_cover(
        wb, THEME, symbol, company_name,
        model_label="Real Options Valuation",
        model_desc="Values managerial flexibility (invest/expand/abandon) using Black-Scholes. "
                   "Real options exceed traditional NPV when uncertainty is high.",
        sheets_index=[
            (1, "Cover",                "Model overview & index"),
            (2, "Inputs & Assumptions", "Financial data & model parameters"),
            (3, "Real Options Analysis","BS pricing, binomial tree & strategic value"),
            (4, "Results & Sensitivity","Summary + σ × T sensitivity"),
        ],
        meta_extra={
            "price":     d["price"],
            "mktcap_cr": d["mktcap_cr"],
            "sector":    safe(fin.get("sector"), "—"),
        },
    )

    params_def = [
        ("Underlying Value V (₹ Cr)",   d["V"],     "₹ Cr","OpCF × 10 PV proxy",        "#,##0",  True),
        ("Investment Cost X (₹ Cr)",    d["X"],     "₹ Cr","Assets × 20% capex proxy",   "#,##0",  True),
        ("Time Horizon T (years)",       d["T"],     "yr",  "Option expiry",               "0.0",   False),
        ("Volatility σ",                d["sigma"], "%",   "Annual asset volatility",     "0.0%",  True),
        ("Risk-Free Rate r",            d["r"],     "%",   "India 10Y GSec",              "0.0%",  False),
        ("Risk-Free Rate (rf)",         0.071,      "%",   "India 10Y GSec",              "0.0%",  False),
        ("Equity Risk Premium",         0.055,      "%",   "Damodaran India ERP",          "0.0%",  False),
        ("Beta",                        d["beta"],  "x",   "Market beta",                 "0.00",  False),
    ]

    build_inputs(wb, THEME, company_name, fin, params_def)
    _build_analysis(wb, THEME, fin, d)
    _build_results_sheet(wb, THEME, company_name, d)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
