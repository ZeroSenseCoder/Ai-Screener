"""
Sum of Parts (SOTP) Valuation Model
Values each business segment at its own multiple, sums them,
and bridges to equity value. Used for conglomerates / diversified companies.
"""
import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.services.excel_utils import (
    font, fill, thin_border, align, safe, cr,
    section_header, sub_header, label, value_cell, pct_cell,
    build_cover, build_inputs, build_results, build_sensitivity_table,
)

THEME = {
    "cover_bg":       "001A33",
    "primary":        "003366",
    "sub":            "004D99",
    "accent":         "3498DB",
    "input_color":    "1A5276",
    "positive_fill":  "D6EAF8",
    "positive_text":  "1A5276",
    "subtotal_fill":  "AED6F1",
    "key_fill":       "FFF2CC",
}


def _col(c: int) -> str:
    return get_column_letter(c)


def _lbl(ws, row, text, indent=0, bold=False):
    c = ws.cell(row=row, column=2, value=("    " * indent) + text)
    c.font = font(bold=bold, size=9)
    c.border = thin_border()
    c.alignment = align()
    return c


def _val(ws, row, col, val, fmt="#,##0.00", bold=False, bg=None, color="000000"):
    c = ws.cell(row=row, column=col, value=val)
    c.font = font(bold=bold, size=9, color=color)
    c.number_format = fmt
    c.alignment = align(h="right")
    c.border = thin_border()
    if bg:
        c.fill = fill(bg)
    return c


def _hdr(ws, row, cols, theme, start_col=2):
    for i, txt in enumerate(cols):
        c = ws.cell(row=row, column=start_col + i, value=txt)
        c.font = font(bold=True, color="FFFFFF", size=9)
        c.fill = fill(theme["primary"])
        c.alignment = align(h="center")
        c.border = thin_border()


def _compute(fin: dict):
    price       = safe(fin.get("price"),          100.0)
    shares_raw  = safe(fin.get("shares"),          1e8)
    revenue_raw = safe(fin.get("revenue"),         0.0)
    ebitda_raw  = safe(fin.get("ebitda"),          0.0)
    debt_raw    = safe(fin.get("total_debt"),      0.0)
    cash_raw    = safe(fin.get("cash"),            0.0)
    rev_growth  = safe(fin.get("revenue_growth"),  0.10)
    beta        = safe(fin.get("beta"),            1.0)
    sector      = fin.get("sector", "General")

    shares_cr   = shares_raw / 1e7
    revenue_cr  = revenue_raw / 1e7
    ebitda_cr   = ebitda_raw  / 1e7
    debt_cr     = debt_raw    / 1e7
    cash_cr     = cash_raw    / 1e7
    mktcap_cr   = (price * shares_raw) / 1e7

    # Segment splits
    core_ebitda    = ebitda_cr * 0.60
    growth_ebitda  = ebitda_cr * 0.30
    other_ebitda   = ebitda_cr * 0.10

    core_mult   = 8.0
    growth_mult = 12.0
    other_mult  = 6.0

    core_ev    = core_ebitda   * core_mult
    growth_ev  = growth_ebitda * growth_mult
    other_ev   = other_ebitda  * other_mult

    gross_ev   = core_ev + growth_ev + other_ev

    # Peer notes by sector
    sector_lower = sector.lower()
    if "tech" in sector_lower or "software" in sector_lower:
        core_peer   = "Peers: IT Services 20-30x"
        growth_peer = "Peers: SaaS/Cloud 30-50x"
        other_peer  = "Peers: IT Infra 10-15x"
    elif "bank" in sector_lower or "financ" in sector_lower or "nbfc" in sector_lower:
        core_peer   = "Peers: Large Banks 1.5-3x P/B"
        growth_peer = "Peers: Small Finance Banks 2-4x P/B"
        other_peer  = "Peers: Insurance 2-3x P/B"
    elif "fmcg" in sector_lower or "consumer" in sector_lower:
        core_peer   = "Peers: FMCG 35-50x P/E"
        growth_peer = "Peers: Premium FMCG 50-70x"
        other_peer  = "Peers: Agri/Rural 15-25x"
    elif "pharma" in sector_lower or "health" in sector_lower:
        core_peer   = "Peers: Pharma Domestic 25-35x"
        growth_peer = "Peers: Pharma Export 20-30x"
        other_peer  = "Peers: Diagnostics 30-45x"
    elif "energy" in sector_lower or "oil" in sector_lower:
        core_peer   = "Peers: Upstream 5-8x EV/EBITDA"
        growth_peer = "Peers: Renewables 12-18x"
        other_peer  = "Peers: Pipelines 8-12x"
    else:
        core_peer   = "Peers: Listed conglomerates 6-10x"
        growth_peer = "Peers: Growth divisions 10-15x"
        other_peer  = "Peers: Non-core 4-8x"

    # EBITDA margin estimate
    ebitda_margin = ebitda_cr / revenue_cr if revenue_cr > 0 else 0.20

    # Corporate adjustments
    corp_overhead      = revenue_cr * 0.03
    holdco_discount    = gross_ev * 0.10
    minority_interest  = gross_ev * 0.05

    adj_ev     = gross_ev - corp_overhead - holdco_discount + cash_cr - debt_cr - minority_interest
    equity_val = adj_ev
    imp_price  = equity_val / shares_cr if shares_cr > 0 else 0.0
    upside     = (imp_price - price) / price if price else 0.0

    ke         = 0.071 + beta * 0.055

    return dict(
        price=price, shares_cr=shares_cr, revenue_cr=revenue_cr,
        ebitda_cr=ebitda_cr, debt_cr=debt_cr, cash_cr=cash_cr,
        mktcap_cr=mktcap_cr, rev_growth=rev_growth, ebitda_margin=ebitda_margin,
        core_ebitda=core_ebitda, growth_ebitda=growth_ebitda, other_ebitda=other_ebitda,
        core_mult=core_mult, growth_mult=growth_mult, other_mult=other_mult,
        core_ev=core_ev, growth_ev=growth_ev, other_ev=other_ev,
        gross_ev=gross_ev, corp_overhead=corp_overhead,
        holdco_discount=holdco_discount, minority_interest=minority_interest,
        adj_ev=adj_ev, equity_val=equity_val, imp_price=imp_price, upside=upside,
        ke=ke,
        core_peer=core_peer, growth_peer=growth_peer, other_peer=other_peer,
        sector=sector,
    )


def _build_analysis(wb: Workbook, theme: dict, fin: dict, d: dict):
    ws = wb.create_sheet("Sum of Parts")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 32

    row = 2
    t = ws.cell(row=row, column=2,
                value="Sum of Parts (SOTP) Valuation  —  Segment-by-Segment Value Build")
    t.font = font(bold=True, size=13)
    ws.merge_cells(f"B{row}:G{row}")
    row += 2

    # ── SEGMENT BREAKDOWN ──
    section_header(ws, row, 2, "▌ SEGMENT BREAKDOWN", theme, span=6)
    row += 1
    _hdr(ws, row, ["Segment", "EBITDA (₹Cr)", "EV/EBITDA Multiple",
                   "Segment EV (₹Cr)", "% of Total EV", "Notes"], theme)
    row += 1

    segs = [
        ("Core Business",       d["core_ebitda"],   d["core_mult"],   d["core_ev"],
         d["core_ev"] / d["gross_ev"] if d["gross_ev"] else 0,
         "Mature, stable – lower multiple"),
        ("Growth Division",     d["growth_ebitda"], d["growth_mult"], d["growth_ev"],
         d["growth_ev"] / d["gross_ev"] if d["gross_ev"] else 0,
         "High growth – premium multiple"),
        ("Financial / Other",   d["other_ebitda"],  d["other_mult"],  d["other_ev"],
         d["other_ev"] / d["gross_ev"] if d["gross_ev"] else 0,
         "Non-core / financial assets"),
    ]
    for seg_name, ebitda, mult, ev, pct, note in segs:
        _lbl(ws, row, seg_name, indent=1)
        _val(ws, row, 3, ebitda, "#,##0.0")
        _val(ws, row, 4, mult,   "0.0x")
        _val(ws, row, 5, ev,     "#,##0.0",  bg=theme["positive_fill"], color=theme["positive_text"])
        _val(ws, row, 6, pct,    "0.0%")
        c = ws.cell(row=row, column=7, value=note)
        c.font = font(size=9, italic=True, color="606060")
        c.border = thin_border()
        row += 1

    # TOTAL row
    _lbl(ws, row, "TOTAL GROSS EV", bold=True)
    _val(ws, row, 3, d["ebitda_cr"],  "#,##0.0", bold=True)
    _val(ws, row, 4, None, "")
    _val(ws, row, 5, d["gross_ev"],   "#,##0.0", bold=True,
         bg=theme["subtotal_fill"], color=theme["input_color"])
    _val(ws, row, 6, 1.0,             "0.0%",    bold=True)
    row += 2

    # ── SEGMENT VALUATION DETAIL ──
    section_header(ws, row, 2, "▌ SEGMENT VALUATION DETAIL", theme, span=6)
    row += 1
    detail_headers = ["Segment", "Revenue Contrib (₹Cr)",
                      "EBITDA Margin", "Growth Rate", "Multiple Rationale"]
    _hdr(ws, row, detail_headers, theme)
    row += 1

    rev_cr = d["revenue_cr"]
    em     = d["ebitda_margin"]
    rg     = d["rev_growth"]
    segs_detail = [
        ("Core Business",
         rev_cr * 0.60 if em > 0 else 0,
         em, rg * 0.70,
         f"8x – Mature segment. {d['core_peer']}"),
        ("Growth Division",
         rev_cr * 0.30 if em > 0 else 0,
         em * 1.10, rg * 1.30,
         f"12x – Growth premium. {d['growth_peer']}"),
        ("Financial / Other",
         rev_cr * 0.10 if em > 0 else 0,
         em * 0.80, rg * 0.50,
         f"6x – Below average quality. {d['other_peer']}"),
    ]
    for seg_name, rev_contrib, ebit_mg, grow, rationale in segs_detail:
        _lbl(ws, row, seg_name, indent=1)
        _val(ws, row, 3, rev_contrib, "#,##0.0")
        _val(ws, row, 4, ebit_mg,    "0.0%")
        _val(ws, row, 5, grow,       "0.0%")
        c = ws.cell(row=row, column=6, value=rationale)
        c.font = font(size=8, italic=True, color="505050")
        c.border = thin_border()
        c.alignment = align(wrap=True)
        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)
        row += 1
    row += 1

    # ── CORPORATE ADJUSTMENTS ──
    section_header(ws, row, 2, "▌ CORPORATE ADJUSTMENTS  →  EQUITY VALUE BRIDGE", theme, span=6)
    row += 1
    adj_rows = [
        ("Gross Segment EV (sum of all segments)",   d["gross_ev"],          "#,##0.0", False),
        ("Less: Corporate Overhead (-3% of revenue)",-d["corp_overhead"],    "#,##0.0", False),
        ("Less: HoldCo Discount (-10% of gross EV)", -d["holdco_discount"],  "#,##0.0", False),
        ("Add: Cash & Investments (₹ Cr)",            d["cash_cr"],          "#,##0.0", False),
        ("Less: Total Debt (₹ Cr)",                  -d["debt_cr"],          "#,##0.0", False),
        ("Less: Minority Interest (est. 5% of EV)",  -d["minority_interest"],"#,##0.0", False),
        ("★ Adjusted Enterprise / Equity Value",      d["adj_ev"],           "#,##0.0", True),
        ("Shares Outstanding (Cr)",                   d["shares_cr"],        "#,##0.00",False),
        ("★ SOTP Implied Share Price (₹)",            d["imp_price"],        "#,##0.00",True),
        ("Current Market Price (₹)",                  d["price"],            "#,##0.00",False),
        ("★ Upside / (Downside) %",                   d["upside"],           "+0.0%;-0.0%", True),
    ]
    for lbl_txt, val, fmt_, is_key in adj_rows:
        _lbl(ws, row, lbl_txt, bold=is_key)
        _val(ws, row, 3, val, fmt_, bold=is_key,
             bg=theme["key_fill"] if is_key else None,
             color=theme["accent"] if is_key else "000000")
        row += 1
    row += 1

    # ── SENSITIVITY — MULTIPLE CHANGES ──
    section_header(ws, row, 2, "▌ SENSITIVITY — MULTIPLE CHANGES", theme, span=4)
    row += 1
    _hdr(ws, row, ["Scenario", "Core Multiple", "SOTP Price (₹)", "Price Δ (₹)"], theme)
    row += 1
    s_cr = d["shares_cr"] if d["shares_cr"] > 0 else 1
    base_price = d["imp_price"]
    core_scenarios = [
        ("Core -2x", d["core_mult"] - 2),
        ("Core Base", d["core_mult"]),
        ("Core +2x", d["core_mult"] + 2),
    ]
    for sc_name, cm in core_scenarios:
        sc_ev = (d["core_ebitda"] * cm + d["growth_ebitda"] * d["growth_mult"]
                 + d["other_ebitda"] * d["other_mult"]
                 - d["corp_overhead"] - d["holdco_discount"]
                 - d["minority_interest"] - d["debt_cr"] + d["cash_cr"])
        sc_price = sc_ev / s_cr
        delta    = sc_price - base_price
        is_base  = (sc_name == "Core Base")
        _lbl(ws, row, sc_name, indent=1, bold=is_base)
        _val(ws, row, 3, cm,       "0.0x",    bold=is_base)
        _val(ws, row, 4, sc_price, "#,##0.00", bold=is_base,
             bg=theme["key_fill"] if is_base else None,
             color=theme["accent"] if is_base else "000000")
        _val(ws, row, 5, delta,    "+#,##0.00;-#,##0.00")
        row += 1
    row += 1

    growth_scenarios = [
        ("Growth -3x", d["growth_mult"] - 3),
        ("Growth Base", d["growth_mult"]),
        ("Growth +3x", d["growth_mult"] + 3),
    ]
    for sc_name, gm in growth_scenarios:
        sc_ev = (d["core_ebitda"] * d["core_mult"] + d["growth_ebitda"] * gm
                 + d["other_ebitda"] * d["other_mult"]
                 - d["corp_overhead"] - d["holdco_discount"]
                 - d["minority_interest"] - d["debt_cr"] + d["cash_cr"])
        sc_price = sc_ev / s_cr
        delta    = sc_price - base_price
        is_base  = (sc_name == "Growth Base")
        _lbl(ws, row, sc_name, indent=1, bold=is_base)
        _val(ws, row, 3, gm,       "0.0x",    bold=is_base)
        _val(ws, row, 4, sc_price, "#,##0.00", bold=is_base,
             bg=theme["key_fill"] if is_base else None,
             color=theme["accent"] if is_base else "000000")
        _val(ws, row, 5, delta,    "+#,##0.00;-#,##0.00")
        row += 1

    ws.freeze_panes = "C4"


def _build_results_sheet(wb: Workbook, theme: dict, company_name: str, d: dict):
    ws = wb.create_sheet("Results & Sensitivity")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 16

    results_rows = [
        ("Core Business EBITDA (₹ Cr)",      d["core_ebitda"],   "#,##0.0",    False),
        ("Growth Division EBITDA (₹ Cr)",    d["growth_ebitda"], "#,##0.0",    False),
        ("Financial / Other EBITDA (₹ Cr)",  d["other_ebitda"],  "#,##0.0",    False),
        ("Gross Segment EV (₹ Cr)",          d["gross_ev"],      "#,##0.0",    False),
        ("Net Corporate Adjustments (₹ Cr)", d["adj_ev"] - d["gross_ev"], "#,##0.0", False),
        ("★ Equity Value (₹ Cr)",            d["adj_ev"],        "#,##0.0",    True),
        ("Shares Outstanding (Cr)",          d["shares_cr"],     "#,##0.00",   False),
        ("★ SOTP Implied Price (₹)",         d["imp_price"],     "#,##0.00",   True),
        ("Current Price (₹)",               d["price"],          "#,##0.00",   False),
        ("★ Upside / (Downside)",            d["upside"],         "+0.0%;-0.0%",True),
    ]
    next_row = build_results(ws, theme, company_name, "SOTP", results_rows, start_row=2)

    # Sensitivity: Core Multiple × Growth Multiple
    core_mults   = [5.0, 6.0, 7.0, 8.0, 9.0, 10.0, 11.0, 12.0]
    growth_mults = [8.0, 10.0, 12.0, 14.0, 16.0, 18.0]

    base_ri = min(range(len(core_mults)),   key=lambda i: abs(core_mults[i]   - d["core_mult"]))
    base_ci = min(range(len(growth_mults)), key=lambda i: abs(growth_mults[i] - d["growth_mult"]))

    s_cr = d["shares_cr"] if d["shares_cr"] > 0 else 1
    matrix = []
    for cm in core_mults:
        row_data = []
        for gm in growth_mults:
            ev = (d["core_ebitda"] * cm + d["growth_ebitda"] * gm
                  + d["other_ebitda"] * 6.0
                  - d["corp_overhead"] - d["holdco_discount"]
                  - d["minority_interest"] - d["debt_cr"] + d["cash_cr"])
            row_data.append(ev / s_cr)
        matrix.append(row_data)

    build_sensitivity_table(
        ws, theme,
        row_label="Core Business Multiple",
        col_label="Growth Div. Multiple",
        row_vals=core_mults,
        col_vals=growth_mults,
        matrix=matrix,
        current_price=d["price"],
        base_row_idx=base_ri,
        base_col_idx=base_ci,
        start_row=next_row + 1,
        start_col=2,
        row_fmt="0.0x",
        col_fmt="0.0x",
    )


def generate_excel(fin: dict, symbol: str, company_name: str) -> bytes:
    d  = _compute(fin)
    wb = Workbook()

    build_cover(
        wb, THEME, symbol, company_name,
        model_label="Sum of Parts (SOTP) Valuation",
        model_desc="Values each business segment at its own EV/EBITDA multiple, "
                   "then bridges to equity. Best for conglomerates and diversified businesses.",
        sheets_index=[
            (1, "Cover",                "Model overview & index"),
            (2, "Inputs & Assumptions", "Financial data & model parameters"),
            (3, "Sum of Parts",         "Segment breakdown, corporate bridge & sensitivity"),
            (4, "Results & Sensitivity","Summary + Core × Growth multiple sensitivity table"),
        ],
        meta_extra={
            "price":     d["price"],
            "mktcap_cr": d["mktcap_cr"],
            "sector":    safe(fin.get("sector"), "—"),
        },
    )

    beta = safe(fin.get("beta"), 1.0)
    params_def = [
        ("Core Business EBITDA Split",   0.60,             "%",   "60% of total EBITDA",             "0.0%", True),
        ("Growth Division EBITDA Split", 0.30,             "%",   "30% of total EBITDA",             "0.0%", True),
        ("Financial / Other Split",      0.10,             "%",   "10% of total EBITDA",             "0.0%", False),
        ("Core Business EV/EBITDA",      d["core_mult"],   "x",   "8x – mature segment",             "0.0x", True),
        ("Growth Division EV/EBITDA",    d["growth_mult"], "x",   "12x – growth premium",            "0.0x", True),
        ("Financial / Other EV/EBITDA",  d["other_mult"],  "x",   "6x – lower quality / non-core",   "0.0x", False),
        ("Corporate Overhead %",         0.03,             "%",   "3% of revenue haircut",           "0.0%", False),
        ("HoldCo Discount",              0.10,             "%",   "10% conglomerate discount",       "0.0%", False),
        ("Minority Interest Estimate",   0.05,             "%",   "5% of gross EV",                  "0.0%", False),
        ("Beta",                         beta,             "x",   "Market beta",                     "0.00", False),
        ("Cost of Equity (ke)",          d["ke"],          "%",   "rf 7.1% + β × ERP 5.5%",         "0.0%", False),
    ]

    build_inputs(wb, THEME, company_name, fin, params_def)
    _build_analysis(wb, THEME, fin, d)
    _build_results_sheet(wb, THEME, company_name, d)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
