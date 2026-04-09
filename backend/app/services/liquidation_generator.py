"""
Liquidation Value Excel Generator
Floor valuation — recoverable value after applying haircuts to each asset class
and paying off all liabilities.
"""
import io
from openpyxl import Workbook

from app.services.excel_utils import (
    font, fill, thin_border, align, safe, cr,
    section_header, sub_header, label, value_cell, pct_cell,
    build_cover, build_inputs, build_results, build_sensitivity_table,
)

THEME = {
    "cover_bg":      "1A0000",
    "primary":       "4A0000",
    "sub":           "7A0000",
    "accent":        "C0392B",
    "input_color":   "922B21",
    "positive_fill": "FDEDEC",
    "positive_text": "922B21",
    "subtotal_fill": "F1948A",
    "key_fill":      "FDFEFE",
}

LIQUIDATION_COST_PCT = 0.05


# ─────────────────────────────────────────────────────────────────────────────
def _liq_per_share(total_assets_cr, total_liab_cr, shares_cr,
                   cash_rec=1.00, recv_rec=0.70, inv_rec=0.45,
                   ppe_rec=0.55, invest_rec=0.80, other_rec=0.20,
                   liq_cost_pct=LIQUIDATION_COST_PCT):
    cash_bv    = total_assets_cr * 0.05
    recv_bv    = total_assets_cr * 0.15
    inv_bv     = total_assets_cr * 0.10
    ppe_bv     = total_assets_cr * 0.40
    invest_bv  = total_assets_cr * 0.08
    other_bv   = total_assets_cr - cash_bv - recv_bv - inv_bv - ppe_bv - invest_bv
    if other_bv < 0:
        other_bv = 0.0

    total_recovery = (cash_bv   * cash_rec +
                      recv_bv   * recv_rec +
                      inv_bv    * inv_rec +
                      ppe_bv    * ppe_rec +
                      invest_bv * invest_rec +
                      other_bv  * other_rec)
    liq_costs   = total_recovery * liq_cost_pct
    surplus     = total_recovery - liq_costs - total_liab_cr
    if shares_cr <= 0:
        return 0.0
    return surplus / shares_cr * 1e7   # ₹ per share


# ─────────────────────────────────────────────────────────────────────────────
def _build_liq_analysis(wb, theme, fin):
    ws = wb.create_sheet("Liquidation Analysis")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 28

    total_assets_cr = cr(fin.get("total_assets"))
    total_liab_cr   = cr(fin.get("total_liabilities"))
    cash_cr         = cr(fin.get("cash"))
    shares          = safe(fin.get("shares"), 1e8)
    shares_cr       = shares / 1e7
    price           = safe(fin.get("price"), 0.0)

    # Asset estimates
    cash_bv    = cash_cr if cash_cr > 0 else total_assets_cr * 0.05
    recv_bv    = total_assets_cr * 0.15
    inv_bv     = total_assets_cr * 0.10
    ppe_bv     = total_assets_cr * 0.40
    invest_bv  = total_assets_cr * 0.08
    other_bv   = total_assets_cr - cash_bv - recv_bv - inv_bv - ppe_bv - invest_bv
    if other_bv < 0:
        other_bv = 0.0

    BASE_REC  = [1.00, 0.70, 0.45, 0.55, 0.80, 0.20]
    NOTES     = [
        "Full recovery assumed",
        "Bad debt and collection costs",
        "Distressed sale of stock",
        "Book value typically overstates realizable",
        "Market discount in forced sale",
        "Intangibles largely unrecoverable",
    ]
    asset_rows = [
        ("Cash & Short-term",        cash_bv,   BASE_REC[0], NOTES[0]),
        ("Accounts Receivable",      recv_bv,   BASE_REC[1], NOTES[1]),
        ("Inventory",                inv_bv,    BASE_REC[2], NOTES[2]),
        ("PP&E / Fixed Assets",      ppe_bv,    BASE_REC[3], NOTES[3]),
        ("Investments & Securities", invest_bv, BASE_REC[4], NOTES[4]),
        ("Other Assets",             other_bv,  BASE_REC[5], NOTES[5]),
    ]
    total_book   = sum(r[1] for r in asset_rows)
    total_recov  = sum(r[1] * r[2] for r in asset_rows)
    liq_costs    = total_recov * LIQUIDATION_COST_PCT
    surplus      = total_recov - liq_costs - total_liab_cr
    liq_ps       = surplus / shares_cr * 1e7 if shares_cr > 0 else 0
    premium_liq  = (price - liq_ps) / liq_ps if liq_ps else 0
    mos          = max(0, (liq_ps - price) / price) if price else 0

    row = 2
    t = ws.cell(row=row, column=2, value="Liquidation Value Analysis")
    t.font = font(bold=True, size=13)
    ws.merge_cells(f"B{row}:F{row}")

    # ── Section 1: Asset Base & Recovery ──────────────────────────────────────
    row += 2
    section_header(ws, row, 2, "▌ ASSET BASE & RECOVERY RATES", theme, span=5)
    row += 1

    hdr = ["Asset Class", "Est. Book Value (₹ Cr)", "Recovery Rate",
           "Recovery Value (₹ Cr)", "Notes"]
    for j, h in enumerate(hdr):
        c = ws.cell(row=row, column=2 + j, value=h)
        c.font = font(bold=True, color="FFFFFF", size=9)
        c.fill = fill(theme["sub"])
        c.border = thin_border()
        c.alignment = align()
    row += 1

    for name, bv, rec, note in asset_rows:
        ws.cell(row=row, column=2, value=name).font = font(size=9)
        ws.cell(row=row, column=2).border = thin_border()
        value_cell(ws, row, 3, bv,        "#,##0.0")
        pct_cell(ws,   row, 4, rec)
        value_cell(ws, row, 5, bv * rec,  "#,##0.0")
        ws.cell(row=row, column=6, value=note).font = font(size=8, color="606060", italic=True)
        ws.cell(row=row, column=6).border = thin_border()
        row += 1

    # Totals
    for col, val, fmt in [(3, total_book, "#,##0.0"), (5, total_recov, "#,##0.0")]:
        value_cell(ws, row, col, val, fmt, bold=True, bg=theme["subtotal_fill"])
    ws.cell(row=row, column=2, value="TOTAL ASSETS (Book) / TOTAL RECOVERY VALUE")
    ws.cell(row=row, column=2).font = font(bold=True, size=9)
    ws.cell(row=row, column=2).border = thin_border()
    ws.cell(row=row, column=2).fill = fill(theme["subtotal_fill"])
    row += 1

    # ── Section 2: Liquidation Value Calculation ───────────────────────────────
    row += 1
    section_header(ws, row, 2, "▌ LIQUIDATION VALUE CALCULATION", theme, span=5)
    row += 1

    calc_rows = [
        ("Total Recovery Value (₹ Cr)",         total_recov,  "#,##0.0", False),
        ("Less: Total Liabilities (₹ Cr)",       total_liab_cr,"#,##0.0", False),
        ("Less: Liquidation Costs (5% of recov)","",           "",        False),
    ]
    for lbl_txt, val, fmt, is_key in [
        ("Total Recovery Value (₹ Cr)",              total_recov,   "#,##0.0", False),
        ("Less: Total Liabilities (₹ Cr)",           total_liab_cr, "#,##0.0", False),
        ("Less: Liquidation Costs (est. 5%)",         liq_costs,     "#,##0.0", False),
        ("Liquidation Surplus / (Deficit) (₹ Cr)",   surplus,       "#,##0.0", True),
        ("Shares Outstanding (Cr)",                  shares_cr,     "#,##0.00", False),
        ("★ Liquidation Value Per Share (₹)",        liq_ps,        "#,##0.00", True),
        ("Current Market Price (₹)",                 price,         "#,##0.00", False),
        ("Premium to Liquidation",                   premium_liq,   "0.0%",    True),
        ("Margin of Safety (if liq > price)",        mos,           "0.0%",    False),
    ]:
        c = ws.cell(row=row, column=2, value=lbl_txt)
        c.font = font(bold=is_key, size=9)
        c.border = thin_border()
        if is_key:
            c.fill = fill(theme["key_fill"])
        v = ws.cell(row=row, column=3, value=val)
        v.number_format = fmt
        v.font = font(bold=is_key, size=9,
                      color=theme["accent"] if is_key else "000000")
        v.alignment = align(h="right")
        v.border = thin_border()
        if is_key:
            v.fill = fill(theme["key_fill"])
        row += 1

    # ── Section 3: Orderly vs Forced vs Fire Sale ──────────────────────────────
    row += 1
    section_header(ws, row, 2, "▌ ORDERLY vs FORCED SALE COMPARISON", theme, span=5)
    row += 1

    hdr2 = ["Scenario", "Cash Rec%", "PP&E Rec%", "Implied Value/Share (₹)", "Notes"]
    for j, h in enumerate(hdr2):
        c = ws.cell(row=row, column=2 + j, value=h)
        c.font = font(bold=True, color="FFFFFF", size=9)
        c.fill = fill(theme["sub"])
        c.border = thin_border()
        c.alignment = align()
    row += 1

    SCENARIOS = [
        ("Orderly Sale (3+ years)",   +0.15, "Higher recovery, time to find buyers"),
        ("Forced Sale (base case)",   +0.00, "Current base case assumptions"),
        ("Fire Sale (<30 days)",      -0.20, "Deep discounts, minimal buyer pool"),
    ]
    for scen_name, delta, note in SCENARIOS:
        ps = _liq_per_share(
            total_assets_cr, total_liab_cr, shares_cr,
            cash_rec   = min(1.0, BASE_REC[0] + delta),
            recv_rec   = max(0, BASE_REC[1] + delta),
            inv_rec    = max(0, BASE_REC[2] + delta),
            ppe_rec    = max(0, BASE_REC[3] + delta),
            invest_rec = max(0, BASE_REC[4] + delta),
            other_rec  = max(0, BASE_REC[5] + delta),
        )
        is_base = delta == 0.0
        ws.cell(row=row, column=2, value=scen_name).font = font(bold=is_base, size=9)
        ws.cell(row=row, column=2).border = thin_border()
        if is_base:
            ws.cell(row=row, column=2).fill = fill(theme["positive_fill"])
        pct_cell(ws, row, 3, min(1.0, BASE_REC[0] + delta))
        pct_cell(ws, row, 4, max(0, BASE_REC[3] + delta))
        value_cell(ws, row, 5, ps, "#,##0.00", bold=is_base,
                   bg=theme["positive_fill"] if is_base else None)
        ws.cell(row=row, column=6, value=note).font = font(size=8, color="606060", italic=True)
        ws.cell(row=row, column=6).border = thin_border()
        row += 1

    ws.freeze_panes = "C4"
    return ws


def _build_liq_sensitivity(wb, theme, fin):
    ws = wb.create_sheet("Sensitivity")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 22

    total_assets_cr = cr(fin.get("total_assets"))
    total_liab_cr   = cr(fin.get("total_liabilities"))
    shares          = safe(fin.get("shares"), 1e8)
    shares_cr       = shares / 1e7
    price           = safe(fin.get("price"), 0.0)

    row_vals = [0.80, 0.85, 0.90, 0.95, 1.00]   # Cash recovery %
    col_vals = [0.30, 0.40, 0.50, 0.55, 0.60, 0.65, 0.70]  # PP&E recovery %

    base_row_idx = 4   # 1.00 cash
    base_col_idx = 3   # 0.55 PP&E

    matrix = []
    for cash_rec in row_vals:
        row_data = []
        for ppe_rec in col_vals:
            ps = _liq_per_share(
                total_assets_cr, total_liab_cr, shares_cr,
                cash_rec=cash_rec, ppe_rec=ppe_rec,
            )
            row_data.append(round(ps, 2))
        matrix.append(row_data)

    row = 2
    t = ws.cell(row=row, column=2, value="Liquidation Sensitivity Analysis")
    t.font = font(bold=True, size=13)
    ws.merge_cells(f"B{row}:J{row}")
    row += 2

    build_sensitivity_table(
        ws, theme,
        row_label="Cash Recovery %",
        col_label="PP&E Recovery %",
        row_vals=row_vals,
        col_vals=col_vals,
        matrix=matrix,
        current_price=price,
        base_row_idx=base_row_idx,
        base_col_idx=base_col_idx,
        start_row=row,
    )
    return ws


# ─────────────────────────────────────────────────────────────────────────────
def generate_excel(fin: dict, symbol: str, company_name: str) -> bytes:
    wb = Workbook()

    price     = safe(fin.get("price"), 0.0)
    shares    = safe(fin.get("shares"), 1e8)
    mktcap_cr = price * shares / 1e7

    build_cover(
        wb, THEME, symbol, company_name,
        model_label="Liquidation Value Model",
        model_desc="Floor valuation: recovery value of assets after haircuts less all liabilities",
        sheets_index=[
            (1, "Cover",               "Company overview & model index"),
            (2, "Inputs & Assumptions","Financial data & model parameters"),
            (3, "Liquidation Analysis","Asset recovery, liq value, scenarios"),
            (4, "Sensitivity",         "Cash rec% × PP&E rec% implied liq value/share"),
        ],
        meta_extra={"price": price, "mktcap_cr": mktcap_cr, "sector": safe(fin.get("sector"), "—")},
    )

    params_def = [
        ("Liquidation Cost %",      LIQUIDATION_COST_PCT, "%", "Legal, admin, disposal costs",      "0.0%", True),
        ("Cash Recovery Rate",      1.00,  "%", "Full recovery for liquid assets",                   "0.0%", False),
        ("Receivables Recovery",    0.70,  "%", "After bad debts & collection costs",                "0.0%", True),
        ("Inventory Recovery",      0.45,  "%", "Distressed sale discount",                          "0.0%", True),
        ("PP&E Recovery",           0.55,  "%", "Realizable value in forced sale",                   "0.0%", True),
        ("Investments Recovery",    0.80,  "%", "Market discount in forced liquidation",             "0.0%", False),
        ("Other Assets Recovery",   0.20,  "%", "Largely intangibles — minimal value",               "0.0%", False),
        ("Orderly Sale Delta",      0.15,  "%", "Add to base rates for orderly wind-down",           "0.0%", False),
        ("Fire Sale Delta",        -0.20,  "%", "Subtract from base rates for forced fire sale",     "0.0%", False),
    ]
    build_inputs(wb, THEME, company_name, fin, params_def)

    _build_liq_analysis(wb, THEME, fin)
    _build_liq_sensitivity(wb, THEME, fin)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
