"""
Multi-Stage DCF Generator — 3-Stage Model
Stage 1: High growth (Years 1–5) at g1
Stage 2: Transition (Years 6–10) linearly interpolating g1 → g2
Stage 3: Terminal perpetuity (Gordon Growth at gT after Year 10)
Best for companies with changing growth profiles.
"""
import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.services.excel_utils import (
    font, fill, thin_border, align, safe, cr,
    section_header, sub_header, label, value_cell, pct_cell,
    build_cover, build_inputs, build_results, build_sensitivity_table,
)

THEME = {
    "cover_bg":       "071B15",
    "primary":        "0D3325",
    "sub":            "1A5C42",
    "accent":         "2ECC71",
    "input_color":    "145A32",
    "positive_fill":  "D5F5E3",
    "positive_text":  "1E8449",
    "subtotal_fill":  "A9DFBF",
    "key_fill":       "FFF9C4",
}

# ── helpers ───────────────────────────────────────────────────────────────────

def _col(c: int) -> str:
    return get_column_letter(c)


def _hdr(ws, row: int, cols: list, theme: dict):
    for i, txt in enumerate(cols):
        c = ws.cell(row=row, column=2 + i, value=txt)
        c.font = font(bold=True, color="FFFFFF", size=9)
        c.fill = fill(theme["primary"])
        c.alignment = align(h="center")
        c.border = thin_border()


def _lbl(ws, row, text, indent=0, bold=False):
    c = ws.cell(row=row, column=2, value=("    " * indent) + text)
    c.font = font(bold=bold, size=9)
    c.border = thin_border()
    c.alignment = align()
    return c


def _val(ws, row, col, val, fmt="#,##0.00", bold=False, bg=None, color="000000"):
    c = ws.cell(row=row, column=col, value=val)
    c.font = font(bold=bold, size=9, color=color)
    c.number_format = fmt
    c.alignment = align(h="right")
    c.border = thin_border()
    if bg:
        c.fill = fill(bg)
    return c


# ── computation ───────────────────────────────────────────────────────────────

def _compute(fin: dict):
    price    = safe(fin.get("price"),          100.0)
    shares   = safe(fin.get("shares"),         1e8)
    revenue  = safe(fin.get("revenue"),        1e10)
    fcf_base = safe(fin.get("fcf") or fin.get("operating_cf"), revenue * 0.08)
    beta     = safe(fin.get("beta"),           1.0)
    rev_g    = safe(fin.get("revenue_growth"), 0.12)
    debt     = safe(fin.get("total_debt"),     0.0)
    cash_    = safe(fin.get("cash"),           0.0)

    wacc  = 0.071 + beta * 0.055          # simplified WACC ≈ ke
    g1    = max(rev_g, 0.08)              # Stage 1 high growth
    g2    = max(rev_g * 0.40, 0.055)      # Stage 2 terminal transition target
    gT    = 0.055                         # Terminal perpetuity growth
    shares_cr = shares / 1e7

    # Stage 1: Years 1–5
    revs_s1, fcfs_s1, pv_s1 = [], [], []
    rev = revenue
    fcf = fcf_base
    for t in range(1, 6):
        rev = rev * (1 + g1)
        fcf = fcf * (1 + g1)
        revs_s1.append(rev)
        fcfs_s1.append(fcf)
        df = 1 / (1 + wacc) ** (t - 0.5)
        pv_s1.append(fcf * df)

    # Stage 2: Years 6–10 — linearly interpolate growth g1 → g2
    revs_s2, fcfs_s2, pv_s2, g_sched_s2 = [], [], [], []
    for t in range(6, 11):
        alpha = (t - 5) / 5          # 0.2 … 1.0
        g_t   = g1 * (1 - alpha) + g2 * alpha
        g_sched_s2.append(g_t)
        rev = rev * (1 + g_t)
        fcf = fcf * (1 + g_t)
        revs_s2.append(rev)
        fcfs_s2.append(fcf)
        df = 1 / (1 + wacc) ** (t - 0.5)
        pv_s2.append(fcf * df)

    # Terminal value at end of Year 10
    tv_fcf = fcfs_s2[-1] * (1 + gT)
    tv     = tv_fcf / (wacc - gT)
    pv_tv  = tv / (1 + wacc) ** 10

    sum_pv_s1  = sum(pv_s1)
    sum_pv_s2  = sum(pv_s2)
    ev         = sum_pv_s1 + sum_pv_s2 + pv_tv
    eq_value   = ev - debt + cash_
    impl_price = (eq_value / 1e7) / shares_cr if shares_cr else 0
    upside     = (impl_price - price) / price if price else 0

    return dict(
        price=price, shares_cr=shares_cr, revenue=revenue,
        fcf_base=fcf_base,
        revs_s1=revs_s1, fcfs_s1=fcfs_s1, pv_s1=pv_s1,
        revs_s2=revs_s2, fcfs_s2=fcfs_s2, pv_s2=pv_s2,
        g_sched_s2=g_sched_s2,
        g1=g1, g2=g2, gT=gT, wacc=wacc,
        tv=tv, pv_tv=pv_tv,
        sum_pv_s1=sum_pv_s1, sum_pv_s2=sum_pv_s2,
        ev=ev, eq_value=eq_value,
        impl_price=impl_price, upside=upside,
        debt=debt, cash=cash_,
    )


# ── sheet builders ────────────────────────────────────────────────────────────

def _build_analysis(wb: Workbook, theme: dict, fin: dict, d: dict):
    ws = wb.create_sheet("3-Stage DCF")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 40
    for i in range(10):
        ws.column_dimensions[_col(3 + i)].width = 13

    YEARS = [f"Yr {y}" for y in range(1, 11)]
    row = 2
    t = ws.cell(row=row, column=2,
                value="Multi-Stage DCF — 3-Stage FCFF Projection (Years 1–10 + Terminal)")
    t.font = font(bold=True, size=13)
    ws.merge_cells(f"B{row}:{_col(12)}{row}")
    row += 2

    # Stage labels
    s1_lbl = ws.cell(row=row, column=3, value="◄ Stage 1: High Growth ►")
    s1_lbl.font = font(bold=True, color=theme["accent"], size=9)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
    s2_lbl = ws.cell(row=row, column=8, value="◄ Stage 2: Transition ►")
    s2_lbl.font = font(bold=True, color=theme["input_color"], size=9)
    ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=12)
    row += 1

    _hdr(ws, row, ["Metric"] + YEARS, theme)
    row += 1

    # ── REVENUE & FCF ──
    section_header(ws, row, 2, "▌ REVENUE & FCF PROJECTIONS", theme, span=11)
    row += 1

    _lbl(ws, row, "Revenue (₹ Cr)", indent=1)
    for j, v in enumerate(d["revs_s1"] + d["revs_s2"]):
        bg = theme["positive_fill"] if j < 5 else "FFFFFF"
        _val(ws, row, 3 + j, cr(v), "#,##0", bg=bg, color=theme["positive_text"])
    row += 1

    g_all = [d["g1"]] * 5 + d["g_sched_s2"]
    _lbl(ws, row, "YoY Growth %", indent=1)
    for j, g in enumerate(g_all):
        _val(ws, row, 3 + j, g, "0.0%")
    row += 1

    _lbl(ws, row, "FCF (₹ Cr)", bold=True)
    for j, v in enumerate(d["fcfs_s1"] + d["fcfs_s2"]):
        bg = theme["subtotal_fill"] if j < 5 else theme["positive_fill"]
        _val(ws, row, 3 + j, cr(v), "#,##0", bold=True, bg=bg,
             color=theme["positive_text"])
    row += 1

    fcff_margins = [
        (d["fcfs_s1"][j] / d["revs_s1"][j]) if d["revs_s1"][j] else 0
        for j in range(5)
    ] + [
        (d["fcfs_s2"][j] / d["revs_s2"][j]) if d["revs_s2"][j] else 0
        for j in range(5)
    ]
    _lbl(ws, row, "FCFF Margin %", indent=1)
    for j, m in enumerate(fcff_margins):
        _val(ws, row, 3 + j, m, "0.0%")
    row += 2

    # ── WACC & DISCOUNT ──
    section_header(ws, row, 2, "▌ WACC & DISCOUNT FACTORS", theme, span=11)
    row += 1
    periods = [t - 0.5 for t in range(1, 11)]
    _lbl(ws, row, "Discount Period (midyear)", indent=1)
    for j, p in enumerate(periods):
        _val(ws, row, 3 + j, p, "0.0")
    row += 1

    dfs = [1 / (1 + d["wacc"]) ** p for p in periods]
    _lbl(ws, row, f"Discount Factor  [WACC={d['wacc']:.1%}]", indent=1)
    for j, df in enumerate(dfs):
        _val(ws, row, 3 + j, df, "0.0000")
    row += 1

    fcfs_all = d["fcfs_s1"] + d["fcfs_s2"]
    pv_all   = d["pv_s1"]  + d["pv_s2"]
    _lbl(ws, row, "PV of FCF (₹ Cr)", bold=True)
    for j, v in enumerate(pv_all):
        _val(ws, row, 3 + j, cr(v), "#,##0", bold=True,
             bg=theme["subtotal_fill"], color=theme["positive_text"])
    row += 2

    # ── TERMINAL VALUE ──
    section_header(ws, row, 2, "▌ TERMINAL VALUE (after Year 10)", theme, span=11)
    row += 1
    tv_data = [
        ("Terminal FCF (Year 11)  (₹ Cr)",              cr(d["fcfs_s2"][-1] * (1 + d["gT"])), "#,##0"),
        (f"Terminal Value  [FCF₁₁/(WACC−gT)]  (₹ Cr)", cr(d["tv"]),                           "#,##0"),
        ("PV of Terminal Value  (₹ Cr)",                cr(d["pv_tv"]),                         "#,##0"),
    ]
    for lbl_txt, val, fmt_ in tv_data:
        _lbl(ws, row, lbl_txt, indent=1)
        _val(ws, row, 3, val, fmt_, bg=theme["key_fill"], bold=True)
        row += 1
    row += 1

    # ── EV BRIDGE ──
    section_header(ws, row, 2, "▌ ENTERPRISE VALUE BRIDGE", theme, span=11)
    row += 1
    bridge = [
        ("Sum PV Stage 1 FCF  (₹ Cr)",    cr(d["sum_pv_s1"]),  "#,##0",       False),
        ("Sum PV Stage 2 FCF  (₹ Cr)",    cr(d["sum_pv_s2"]),  "#,##0",       False),
        ("PV Terminal Value  (₹ Cr)",      cr(d["pv_tv"]),      "#,##0",       False),
        ("Enterprise Value  (₹ Cr)",       cr(d["ev"]),         "#,##0",       True),
        ("Less: Total Debt  (₹ Cr)",      -cr(d["debt"]),       "#,##0",       False),
        ("Add: Cash & Equiv  (₹ Cr)",      cr(d["cash"]),       "#,##0",       False),
        ("Equity Value  (₹ Cr)",           cr(d["eq_value"]),   "#,##0",       True),
        ("Shares Outstanding (Cr)",        d["shares_cr"],      "#,##0.00",    False),
        ("★ Implied Share Price (₹)",      d["impl_price"],     "#,##0.00",    True),
        ("Upside / (Downside) %",          d["upside"],         "+0.0%;-0.0%", False),
        ("Current Price (ref)  (₹)",       d["price"],          "#,##0.00",    False),
    ]
    for lbl_txt, val, fmt_, is_key in bridge:
        _lbl(ws, row, lbl_txt, bold=is_key)
        _val(ws, row, 3, val, fmt_, bold=is_key,
             bg=theme["key_fill"] if is_key else None,
             color=theme["accent"] if is_key else "000000")
        row += 1

    ws.freeze_panes = "C5"


def _build_results_sheet(wb: Workbook, theme: dict, company_name: str, d: dict):
    ws = wb.create_sheet("Results & Sensitivity")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 16

    results_rows = [
        ("Enterprise Value (₹ Cr)",     cr(d["ev"]),         "#,##0",       False),
        ("Equity Value (₹ Cr)",         cr(d["eq_value"]),   "#,##0",       False),
        ("Shares Outstanding (Cr)",     d["shares_cr"],      "#,##0.00",    False),
        ("★ Implied Share Price (₹)",   d["impl_price"],     "#,##0.00",    True),
        ("Current Price (₹)",           d["price"],          "#,##0.00",    False),
        ("Upside / (Downside)",         d["upside"],         "+0.0%;-0.0%", True),
        ("WACC",                        d["wacc"],           "0.0%",        False),
        ("Stage 1 Growth (g1)",         d["g1"],             "0.0%",        False),
        ("Terminal Growth (gT)",        d["gT"],             "0.0%",        False),
    ]
    next_row = build_results(ws, theme, company_name, "Multi-Stage DCF",
                             results_rows, start_row=2)

    # Sensitivity: WACC × gT
    wacc_vals = [0.09, 0.10, 0.11, 0.12, 0.13, 0.14]
    gT_vals   = [0.030, 0.040, 0.050, 0.055, 0.060]
    base_wi   = min(range(len(wacc_vals)), key=lambda i: abs(wacc_vals[i] - d["wacc"]))
    base_gi   = min(range(len(gT_vals)),   key=lambda i: abs(gT_vals[i]  - d["gT"]))

    fcfs_s2_last = d["fcfs_s2"][-1]
    shares_cr    = d["shares_cr"]
    fcfs_s1      = d["fcfs_s1"]
    fcfs_s2      = d["fcfs_s2"]
    pv_s1_fcfs   = d["pv_s1"]
    pv_s2_fcfs   = d["pv_s2"]
    debt         = d["debt"]
    cash_        = d["cash"]

    matrix = []
    for wacc_ in wacc_vals:
        row_data = []
        for gT_ in gT_vals:
            if wacc_ <= gT_:
                row_data.append(0.0)
                continue
            # recompute PVs with new WACC
            sum1 = sum(fcfs_s1[t] / (1 + wacc_) ** (t + 0.5) for t in range(5))
            sum2 = sum(fcfs_s2[t] / (1 + wacc_) ** (t + 5.5) for t in range(5))
            tv_  = fcfs_s2_last * (1 + gT_) / (wacc_ - gT_)
            pvtv_ = tv_ / (1 + wacc_) ** 10
            ev_   = sum1 + sum2 + pvtv_
            eq_   = ev_ - debt + cash_
            ip_   = (eq_ / 1e7) / shares_cr if shares_cr else 0
            row_data.append(ip_)
        matrix.append(row_data)

    build_sensitivity_table(
        ws, theme,
        row_label="WACC",
        col_label="Terminal Growth gT",
        row_vals=wacc_vals,
        col_vals=gT_vals,
        matrix=matrix,
        current_price=d["price"],
        base_row_idx=base_wi,
        base_col_idx=base_gi,
        start_row=next_row + 1,
        start_col=2,
        row_fmt="0.0%",
        col_fmt="0.0%",
    )


# ── public entry point ────────────────────────────────────────────────────────

def generate_excel(fin: dict, symbol: str, company_name: str) -> bytes:
    d  = _compute(fin)
    wb = Workbook()

    mktcap_cr = (d["price"] * safe(fin.get("shares"), 1e8)) / 1e7

    build_cover(
        wb, THEME, symbol, company_name,
        model_label="Multi-Stage DCF (3-Stage FCFF)",
        model_desc="High growth → transition → terminal perpetuity. "
                   "Best for companies with evolving growth profiles.",
        sheets_index=[
            (1, "Cover",                "Model overview & index"),
            (2, "Inputs & Assumptions", "Financial data & model parameters"),
            (3, "3-Stage DCF",          "10-year FCF projections + terminal value + EV bridge"),
            (4, "Results & Sensitivity","Summary + WACC × gT sensitivity table"),
        ],
        meta_extra={"price": d["price"], "mktcap_cr": mktcap_cr,
                    "sector": safe(fin.get("sector"), "—")},
    )

    beta   = safe(fin.get("beta"), 1.0)
    params_def = [
        ("Risk-Free Rate (rf)",          0.071,     "%",  "India 10Y GSec",            "0.0%", False),
        ("Equity Risk Premium",          0.055,     "%",  "Damodaran India ERP",        "0.0%", False),
        ("Beta",                         beta,      "x",  "Market beta",                "0.00", True),
        ("WACC (simplified ke)",         d["wacc"], "%",  "rf + β × ERP",               "0.0%", True),
        ("Stage 1 Growth (g1)",          d["g1"],   "%",  "High-growth rate, yrs 1–5",  "0.0%", True),
        ("Stage 2 Target Growth (g2)",   d["g2"],   "%",  "Transition end, yr 10",      "0.0%", False),
        ("Terminal Growth (gT)",         d["gT"],   "%",  "Long-run India nominal GDP", "0.0%", True),
    ]

    build_inputs(wb, THEME, company_name, fin, params_def)
    _build_analysis(wb, THEME, fin, d)
    _build_results_sheet(wb, THEME, company_name, d)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
