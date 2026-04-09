"""
LBO (Leveraged Buyout) Excel Generator
PE-style acquisition model: buy with leverage, grow EBITDA,
exit at a multiple. Compute IRR and MOIC.
"""

import io
import math
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.services.excel_utils import (
    font, fill, thin_border, align, safe, cr,
    section_header, sub_header, label, value_cell, pct_cell,
    build_cover, build_inputs, build_results, build_sensitivity_table,
)

THEME = {
    "cover_bg":      "1A0000",
    "primary":       "3B0000",
    "sub":           "6B0000",
    "accent":        "DAA520",
    "input_color":   "7B241C",
    "positive_fill": "FEF9E7",
    "positive_text": "7D6608",
    "subtotal_fill": "FAD7A0",
    "key_fill":      "FFF2CC",
}

SHEETS_INDEX = [
    (1, "Cover",              "Model overview & metadata"),
    (2, "Inputs & Assumptions","Financial data & model parameters"),
    (3, "LBO Model",          "Transaction structure, projections & returns"),
    (4, "Sensitivity",        "Entry Multiple × Exit Multiple sensitivity table"),
]


def _lv(ws, row, col, text, bold=False, indent=0):
    return label(ws, row, col, text, indent=indent, bold=bold, theme=THEME)


def _vc(ws, row, col, val, fmt="#,##0", bold=False, bg=None, color="000000"):
    return value_cell(ws, row, col, val, fmt=fmt, bold=bold, bg=bg, color=color)


def _pc(ws, row, col, val, bold=False, bg=None):
    return pct_cell(ws, row, col, val, bold=bold, bg=bg)


def _sh(ws, row, text, span=9):
    section_header(ws, row, 2, text, THEME, span=span)


def _key_row(ws, row, lbl_txt, val, fmt="#,##0"):
    _lv(ws, row, 2, lbl_txt, bold=True)
    ws.cell(row=row, column=2).fill = fill(THEME["key_fill"])
    _vc(ws, row, 3, val, fmt=fmt, bold=True,
        bg=THEME["key_fill"], color=THEME["input_color"])


def _approx_irr(moic, years):
    if years <= 0 or moic <= 0:
        return 0.0
    return moic ** (1.0 / years) - 1.0


def generate_excel(fin: dict, symbol: str, company_name: str) -> bytes:
    wb = Workbook()

    price       = safe(fin.get("price"), 0.0)
    shares_raw  = safe(fin.get("shares"), 1e7)
    shares_cr   = shares_raw / 1e7
    mktcap_cr   = price * shares_cr
    sector      = fin.get("sector", "—")

    ebitda_cr   = cr(fin.get("ebitda"))
    revenue_cr  = cr(fin.get("revenue"))
    debt_cr     = cr(fin.get("total_debt"))
    cash_cr     = cr(fin.get("cash"))
    fcf_cr      = cr(fin.get("fcf") or fin.get("operating_cf"))
    earn_growth = safe(fin.get("earnings_growth"), safe(fin.get("revenue_growth"), 0.12))

    # LBO parameters
    entry_multiple  = 8.0
    exit_multiple   = 10.0
    debt_pct        = 0.60
    interest_rate   = 0.09
    hold_years      = 5
    g_ebitda        = min(earn_growth, 0.20)
    debt_sweep_pct  = 0.50   # 50% of (EBITDA - Interest) used to repay debt
    tax_rate        = 0.25

    entry_ev_cr     = entry_multiple * ebitda_cr
    debt_raised_cr  = debt_pct * entry_ev_cr
    equity_cr       = entry_ev_cr - debt_raised_cr
    equity_pct      = equity_cr / entry_ev_cr if entry_ev_cr > 0 else 0
    int_exp_yr1     = debt_raised_cr * interest_rate

    # ── Cover ──────────────────────────────────────────────────────────────────
    build_cover(
        wb, THEME, symbol, company_name,
        "Leveraged Buyout (LBO) Model",
        "PE acquisition model: debt-funded buyout → EBITDA growth → exit. "
        "Returns measured by MOIC and IRR.",
        SHEETS_INDEX,
        {"price": price, "mktcap_cr": mktcap_cr, "sector": sector},
    )

    # ── Inputs ─────────────────────────────────────────────────────────────────
    params_def = [
        ("Entry EV/EBITDA Multiple",    entry_multiple, "x",  "Acquisition entry multiple",      "#,##0.0", True),
        ("Exit EV/EBITDA Multiple",     exit_multiple,  "x",  "Exit (sale) multiple assumption", "#,##0.0", True),
        ("Debt % of Entry EV",          debt_pct,       "%",  "Senior debt financing ratio",     "0.0%",    True),
        ("Interest Rate on Debt",       interest_rate,  "%",  "Blended senior debt rate",        "0.0%",    False),
        ("Hold Period (years)",         hold_years,     "yrs","Investment hold period",           "#,##0",   True),
        ("EBITDA Growth Rate",          g_ebitda,       "%",  "Annual EBITDA growth assumption", "0.0%",    True),
        ("Debt Repayment Sweep %",      debt_sweep_pct, "%",  "% of free cash used for debt paydown", "0.0%", False),
        ("Tax Rate",                    tax_rate,       "%",  "Effective corporate tax rate",    "0.0%",    False),
    ]
    build_inputs(wb, THEME, company_name, fin, params_def)

    # ── LBO Model sheet ────────────────────────────────────────────────────────
    ws = wb.create_sheet("LBO Model")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 16
    for ci in range(4, 10):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    t = ws.cell(row=2, column=2, value=f"{company_name} — LBO Model")
    t.font = font(bold=True, size=13)
    ws.merge_cells("B2:J2")

    row = 4

    # ── SECTION 1: TRANSACTION STRUCTURE ──────────────────────────────────────
    _sh(ws, row, "▌ TRANSACTION STRUCTURE")
    row += 1

    s1_rows = [
        ("EBITDA — Entry Year (₹ Cr)",         ebitda_cr,      "#,##0.0", False),
        ("Entry EV/EBITDA Multiple",            entry_multiple, "#,##0.0x",False),
        ("Entry Enterprise Value (₹ Cr)",       entry_ev_cr,    "#,##0.0", False),
        ("Debt Raised (60% of EV) (₹ Cr)",     debt_raised_cr, "#,##0.0", False),
        ("Equity Contributed (₹ Cr)",           equity_cr,      "#,##0.0", False),
        ("Equity as % of EV",                   equity_pct,     "0.0%",    False),
        ("Total Sources = Uses (₹ Cr)",         entry_ev_cr,    "#,##0.0", False),
        ("Interest Rate on Debt",               interest_rate,  "0.0%",    False),
        ("Interest Expense Year 1 (₹ Cr)",      int_exp_yr1,    "#,##0.0", False),
    ]
    for lbl_txt, val, fmt, is_key in s1_rows:
        _lv(ws, row, 2, lbl_txt)
        _vc(ws, row, 3, val, fmt=fmt)
        row += 1

    row += 1

    # ── SECTION 2: EBITDA PROJECTIONS & DEBT SCHEDULE ─────────────────────────
    _sh(ws, row, "▌ EBITDA PROJECTIONS & DEBT SCHEDULE (5 YEARS)", span=9)
    row += 1

    yr_cols   = list(range(4, 9))
    yr_labels = [f"Year {i}" for i in range(1, 6)]

    _lv(ws, row, 2, "Metric")
    for ci, lbl_txt in zip(yr_cols, yr_labels):
        c = ws.cell(row=row, column=ci, value=lbl_txt)
        c.font = font(bold=True, color="FFFFFF", size=9)
        c.fill = fill(THEME["sub"])
        c.alignment = align(h="center")
        c.border = thin_border()
    row += 1

    proj_revenue   = []
    proj_ebitda    = []
    proj_margin    = []
    proj_int_exp   = []
    proj_repayment = []
    proj_open_debt = []
    proj_close_debt= []
    proj_leverage  = []

    ebitda_margin_base = ebitda_cr / revenue_cr if revenue_cr > 0 else 0.20

    open_debt = debt_raised_cr
    for t_idx in range(5):
        rev   = revenue_cr * (1 + g_ebitda) ** (t_idx + 1)
        ebit  = ebitda_cr  * (1 + g_ebitda) ** (t_idx + 1)
        marg  = ebit / rev if rev > 0 else 0
        int_e = open_debt * interest_rate
        repay = max(0, (ebit - int_e) * debt_sweep_pct)
        close = max(0, open_debt - repay)
        lev   = close / ebit if ebit > 0 else 0

        proj_revenue.append(rev)
        proj_ebitda.append(ebit)
        proj_margin.append(marg)
        proj_int_exp.append(int_e)
        proj_repayment.append(repay)
        proj_open_debt.append(open_debt)
        proj_close_debt.append(close)
        proj_leverage.append(lev)
        open_debt = close

    metric_rows2 = [
        ("Revenue (₹ Cr)",              proj_revenue,    "#,##0.0", False),
        ("EBITDA (₹ Cr)",               proj_ebitda,     "#,##0.0", False),
        ("EBITDA Margin %",             proj_margin,     "0.0%",    False),
        ("Interest Expense (₹ Cr)",     proj_int_exp,    "#,##0.0", False),
        ("Mandatory Debt Repayment (₹ Cr)", proj_repayment, "#,##0.0", False),
        ("Opening Debt (₹ Cr)",         proj_open_debt,  "#,##0.0", False),
        ("Closing Debt (₹ Cr)",         proj_close_debt, "#,##0.0", False),
        ("Leverage (Debt/EBITDA)",       proj_leverage,   "#,##0.0x",False),
    ]

    for lbl_txt, vals, fmt, is_key in metric_rows2:
        _lv(ws, row, 2, lbl_txt, bold=is_key)
        for ci, v in zip(yr_cols, vals):
            _vc(ws, row, ci, v, fmt=fmt,
                bg=THEME["subtotal_fill"] if is_key else None)
        row += 1

    row += 1

    # ── SECTION 3: EXIT ANALYSIS ───────────────────────────────────────────────
    _sh(ws, row, "▌ EXIT ANALYSIS")
    row += 1

    exit_ebitda   = proj_ebitda[hold_years - 1] if hold_years <= 5 else proj_ebitda[-1]
    exit_ev_cr    = exit_multiple * exit_ebitda
    exit_debt_cr  = proj_close_debt[hold_years - 1] if hold_years <= 5 else proj_close_debt[-1]
    equity_proc   = exit_ev_cr - exit_debt_cr
    moic          = equity_proc / equity_cr if equity_cr > 0 else 0
    irr           = _approx_irr(moic, hold_years)
    debt_paydown  = debt_raised_cr - exit_debt_cr

    exit_rows = [
        ("Exit Year (Hold Period)",             hold_years,     "#,##0",   False),
        ("Exit EBITDA (₹ Cr)",                  exit_ebitda,    "#,##0.0", False),
        ("Exit EV/EBITDA Multiple",             exit_multiple,  "#,##0.0x",False),
        ("Exit Enterprise Value (₹ Cr)",        exit_ev_cr,     "#,##0.0", False),
        ("Less: Closing Debt at Exit (₹ Cr)",   exit_debt_cr,   "#,##0.0", False),
        ("Equity Proceeds (₹ Cr)",              equity_proc,    "#,##0.0", False),
        ("Initial Equity Investment (₹ Cr)",    equity_cr,      "#,##0.0", False),
        ("★ MOIC = Exit Equity / Entry Equity", moic,           "#,##0.0x",True),
        ("★ IRR (approx) = MOIC^(1/t) − 1",    irr,            "0.0%",    True),
    ]
    for lbl_txt, val, fmt, is_key in exit_rows:
        if is_key:
            _key_row(ws, row, lbl_txt, val, fmt)
        else:
            _lv(ws, row, 2, lbl_txt)
            _vc(ws, row, 3, val, fmt=fmt)
        row += 1

    row += 1

    # ── SECTION 4: RETURNS WATERFALL ──────────────────────────────────────────
    _sh(ws, row, "▌ RETURNS WATERFALL")
    row += 1

    ebitda_growth_value = (exit_ebitda - ebitda_cr) * exit_multiple
    multiple_expansion  = (exit_multiple - entry_multiple) * ebitda_cr
    debt_paydown_value  = debt_paydown

    waterfall = [
        ("Entry Equity Invested (₹ Cr)",        equity_cr,            "#,##0.0"),
        ("Add: EBITDA Growth Value (₹ Cr)",      ebitda_growth_value,  "#,##0.0"),
        ("Add: Multiple Expansion Value (₹ Cr)", multiple_expansion,   "#,##0.0"),
        ("Add: Debt Paydown Value (₹ Cr)",       debt_paydown_value,   "#,##0.0"),
        ("Total Exit Equity (₹ Cr)",             equity_proc,          "#,##0.0"),
        ("MOIC",                                 moic,                 "#,##0.0x"),
        ("IRR (approx)",                         irr,                  "0.0%"),
    ]
    for lbl_txt, val, fmt in waterfall:
        _lv(ws, row, 2, lbl_txt)
        _vc(ws, row, 3, val, fmt=fmt,
            bg=THEME["subtotal_fill"] if lbl_txt.startswith("Total") else None)
        row += 1

    row += 1

    # ── SECTION 5: RETURNS SCENARIOS ──────────────────────────────────────────
    _sh(ws, row, "▌ RETURNS SCENARIOS")
    row += 1

    scenario_hdr = ws.cell(row=row, column=2, value="Scenario")
    scenario_hdr.font = font(bold=True, color="FFFFFF", size=9)
    scenario_hdr.fill = fill(THEME["primary"])
    scenario_hdr.border = thin_border()
    for ci, lbl_txt in [(3, "Entry Multiple"), (4, "Exit Multiple"),
                         (5, "EBITDA Growth"), (6, "MOIC"), (7, "IRR")]:
        c = ws.cell(row=row, column=ci, value=lbl_txt)
        c.font = font(bold=True, color="FFFFFF", size=9)
        c.fill = fill(THEME["primary"])
        c.alignment = align(h="center")
        c.border = thin_border()
    row += 1

    scenarios = [
        ("Bear",  entry_multiple + 2, exit_multiple - 1, max(g_ebitda - 0.03, 0.02)),
        ("Base",  entry_multiple,     exit_multiple,     g_ebitda),
        ("Bull",  entry_multiple - 1, exit_multiple + 2, g_ebitda + 0.03),
    ]
    for sc_name, en_m, ex_m, g_sc in scenarios:
        en_ev  = en_m * ebitda_cr
        d_sc   = debt_pct * en_ev
        eq_sc  = en_ev - d_sc
        od     = d_sc
        ex_ebi = ebitda_cr
        for t_idx in range(hold_years):
            ex_ebi = ex_ebi * (1 + g_sc)
            int_sc  = od * interest_rate
            rep_sc  = max(0, (ex_ebi - int_sc) * debt_sweep_pct)
            od      = max(0, od - rep_sc)
        ex_ev_sc = ex_m * ex_ebi
        eq_pr    = ex_ev_sc - od
        m_sc     = eq_pr / eq_sc if eq_sc > 0 else 0
        irr_sc   = _approx_irr(m_sc, hold_years)
        bg       = THEME["positive_fill"] if m_sc >= 2 else ("FFC7CE" if m_sc < 1.5 else None)
        _lv(ws, row, 2, sc_name)
        _vc(ws, row, 3, en_m,  "#,##0.0x", bg=bg)
        _vc(ws, row, 4, ex_m,  "#,##0.0x", bg=bg)
        _pc(ws, row, 5, g_sc,  bg=bg)
        _vc(ws, row, 6, m_sc,  "#,##0.0x", bold=True, bg=THEME["key_fill"])
        _pc(ws, row, 7, irr_sc, bold=True, bg=THEME["key_fill"])
        row += 1

    ws.freeze_panes = "C5"

    # ── Sensitivity: Entry Multiple × Exit Multiple ────────────────────────────
    ws_s = wb.create_sheet("Sensitivity")
    ws_s.sheet_view.showGridLines = False
    ws_s.column_dimensions["A"].width = 2
    ws_s.column_dimensions["B"].width = 18

    t2 = ws_s.cell(row=2, column=2, value=f"{company_name} — LBO Sensitivity: MOIC")
    t2.font = font(bold=True, size=13)
    ws_s.merge_cells("B2:L2")

    entry_mults = [6.0, 7.0, 8.0, 9.0, 10.0, 11.0, 12.0]
    exit_mults  = [7.0, 8.0, 9.0, 10.0, 11.0, 12.0, 13.0, 14.0]

    base_en_idx = min(range(len(entry_mults)), key=lambda i: abs(entry_mults[i] - entry_multiple))
    base_ex_idx = min(range(len(exit_mults)),  key=lambda i: abs(exit_mults[i]  - exit_multiple))

    def _moic_matrix(en_m, ex_m):
        en_ev_m  = en_m * ebitda_cr
        d_m      = debt_pct * en_ev_m
        eq_m     = en_ev_m - d_m
        od_m     = d_m
        e_ebi    = ebitda_cr
        for t_idx in range(hold_years):
            e_ebi  = e_ebi * (1 + g_ebitda)
            int_m  = od_m * interest_rate
            rep_m  = max(0, (e_ebi - int_m) * debt_sweep_pct)
            od_m   = max(0, od_m - rep_m)
        ex_ev_m  = ex_m * e_ebi
        eq_pr_m  = ex_ev_m - od_m
        m_val    = eq_pr_m / eq_m if eq_m > 0 else 0
        return round(m_val, 2)

    # Build MOIC matrix — use current_price = base MOIC for colouring
    base_moic = _moic_matrix(entry_multiple, exit_multiple)
    moic_matrix = [[_moic_matrix(en_m, ex_m) for ex_m in exit_mults] for en_m in entry_mults]

    # Override build_sensitivity_table title note by writing custom header
    note = ws_s.cell(row=3, column=2,
        value="Sensitivity: Entry Multiple (rows) × Exit Multiple (cols) — Implied MOIC (x)")
    note.font = font(bold=True, size=10, color=THEME["input_color"])
    ws_s.merge_cells("B3:K3")

    # Custom table (MOIC not share price)
    start_row = 4
    start_col = 2

    # Corner
    ws_s.cell(row=start_row, column=start_col,
              value="Entry \\ Exit").font = font(bold=True, size=9)

    for j, ex_m in enumerate(exit_mults):
        c = ws_s.cell(row=start_row, column=start_col + 1 + j, value=f"{ex_m:.0f}x")
        c.font = font(bold=True, color="FFFFFF", size=9)
        c.fill = fill(THEME["primary"])
        c.alignment = align(h="center")
        c.border = thin_border()
        ws_s.column_dimensions[get_column_letter(start_col + 1 + j)].width = 10

    for i, en_m in enumerate(entry_mults):
        r = start_row + 1 + i
        rh = ws_s.cell(row=r, column=start_col, value=f"{en_m:.0f}x")
        rh.font = font(bold=True, color="FFFFFF", size=9)
        rh.fill = fill(THEME["primary"])
        rh.alignment = align(h="center")
        rh.border = thin_border()

        for j, ex_m in enumerate(exit_mults):
            mv = moic_matrix[i][j]
            c  = ws_s.cell(row=r, column=start_col + 1 + j, value=mv)
            c.number_format = "#,##0.0x"
            c.alignment = align(h="right")
            c.border = thin_border()
            is_base = (i == base_en_idx and j == base_ex_idx)
            if is_base:
                c.fill = fill("FFD700")
                c.font = font(bold=True, size=9)
            elif mv >= 3.0:
                c.fill = fill("C6EFCE"); c.font = font(size=9, color="375623")
            elif mv < 1.5:
                c.fill = fill("FFC7CE"); c.font = font(size=9, color="9C0006")
            else:
                c.font = font(size=9)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
