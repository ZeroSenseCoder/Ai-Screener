"""
model_excel_generator.py — Professional Excel workbooks for all valuation models
except dcf_fcff (handled by dcf_excel_generator.py).

Public API:
    generate_model_excel(model_id, fin, result, params, symbol, company_name) -> bytes
"""

import io
import math
import datetime
from typing import Optional, Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colour palette ─────────────────────────────────────────────────────────────
NAVY             = "1F2D3D"
DARK_BLUE        = "2E4057"
LIGHT_BLUE_FILL  = "DCE9F5"
GREEN_FILL       = "D6F0D6"
YELLOW_FILL      = "FFF2CC"
COVER_BG         = "0A1628"
COVER_GOLD       = "C9A84C"
SENSITIVITY_GOLD = "FFD700"
WHITE            = "FFFFFF"
BLACK            = "000000"
BLUE_TEXT        = "1F497D"
GREEN_TEXT       = "375623"
RED_FILL         = "FFD7D7"
RED_TEXT         = "9C0006"

# ── Style helpers ──────────────────────────────────────────────────────────────

def _font(bold=False, color=BLACK, size=10, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _safe(v, default=0.0):
    if v is None:
        return default
    try:
        if math.isnan(v) or math.isinf(v):
            return default
    except (TypeError, ValueError):
        pass
    return v

def _pct(v):
    if v is None:
        return "N/A"
    return f"{_safe(v) * 100:.1f}%"

def _cr(v):
    if v is None:
        return "N/A"
    return f"₹{_safe(v)/1e7:,.0f} Cr"

def _section_header(ws, row, col, text, width_cols=9):
    c = ws.cell(row=row, column=col, value=text)
    c.font = _font(bold=True, color=WHITE, size=10)
    c.fill = _fill(NAVY)
    c.alignment = _align()
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=col + width_cols - 1)

def _sub_header(ws, row, col, text, width_cols=9):
    c = ws.cell(row=row, column=col, value=text)
    c.font = _font(bold=True, color=WHITE, size=9)
    c.fill = _fill(DARK_BLUE)
    c.alignment = _align()
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=col + width_cols - 1)

def _label(ws, row, col, text, bold=False, indent=0):
    prefix = "    " * indent
    c = ws.cell(row=row, column=col, value=prefix + text)
    c.font = _font(bold=bold, size=9)
    c.alignment = _align()
    c.border = _border()
    return c

def _val(ws, row, col, value, fmt="#,##0", color=BLACK, fill_hex=None, bold=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _font(color=color, size=9, bold=bold)
    c.number_format = fmt
    c.alignment = _align(h="right")
    c.border = _border()
    if fill_hex:
        c.fill = _fill(fill_hex)
    return c

def _pct_cell(ws, row, col, value, color=BLACK, fill_hex=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _font(color=color, size=9)
    c.number_format = "0.0%"
    c.alignment = _align(h="right")
    c.border = _border()
    if fill_hex:
        c.fill = _fill(fill_hex)
    return c

def _set_col_widths(ws, widths: dict):
    """widths: {col_letter: width}"""
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — COVER  (shared)
# ═══════════════════════════════════════════════════════════════════════════════

MODEL_NAMES = {
    "dcf_fcfe":              "DCF — Free Cash Flow to Equity (FCFE)",
    "dcf_multistage":        "DCF — Multi-Stage Growth Model",
    "gordon_growth":         "Gordon Growth Model (Dividend Discount)",
    "ddm_multistage":        "Dividend Discount Model — Multi-Stage",
    "residual_income":       "Residual Income Valuation",
    "trading_comps":         "Trading Comparables Analysis",
    "precedent_transactions":"Precedent Transactions Analysis",
    "peg":                   "PEG Ratio Valuation",
    "revenue_multiple":      "Revenue Multiple Valuation (EV/Revenue)",
    "nav":                   "Net Asset Value (NAV) Model",
    "liquidation":           "Liquidation Value Analysis",
    "replacement_cost":      "Replacement Cost Valuation",
    "capitalized_earnings":  "Capitalized Earnings Model",
    "excess_earnings":       "Excess Earnings Method",
    "eva":                   "Economic Value Added (EVA) Model",
    "cfroi":                 "Cash Flow Return on Investment (CFROI)",
    "lbo":                   "Leveraged Buyout (LBO) Analysis",
    "black_scholes":         "Black-Scholes Option Pricing Model",
    "real_options":          "Real Options Valuation",
    "sum_of_parts":          "Sum-of-the-Parts (SOTP) Valuation",
    "pb_banks":              "Price-to-Book — Banking Sector",
    "cap_rate":              "Capitalization Rate (Real Estate)",
    "user_based":            "User-Based / SaaS Valuation",
    "vc_method":             "Venture Capital Method",
}

def _build_cover(wb: Workbook, model_id: str, symbol: str, company_name: str,
                 fin: dict, result: dict):
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {"A": 3, "B": 34, "C": 2, "D": 32, "E": 24})

    for r in range(1, 42):
        for c in range(1, 12):
            ws.cell(row=r, column=c).fill = _fill(COVER_BG)

    model_label = MODEL_NAMES.get(model_id, model_id.replace("_", " ").title())

    nm = ws.cell(row=5, column=2, value=f"{company_name}  ({symbol})")
    nm.font = Font(name="Calibri", bold=True, size=22, color=WHITE)
    nm.alignment = _align()
    ws.merge_cells("B5:J5")

    sub = ws.cell(row=7, column=2, value=model_label + "  |  Equity Research  |  " +
                  datetime.date.today().strftime("%B %Y"))
    sub.font = Font(name="Calibri", size=13, color=COVER_GOLD, italic=True)
    sub.alignment = _align()
    ws.merge_cells("B7:J7")

    for col in range(2, 11):
        ws.cell(row=9, column=col).fill = _fill(COVER_GOLD)

    price = _safe(fin.get("price", 0))
    shares = _safe(fin.get("shares", 0))
    mktcap_cr = price * shares / 1e7

    meta_rows = [
        ("Analyst",              "Equity Research Analyst"),
        ("Coverage",             fin.get("sector", "Indian Equities")),
        ("Valuation Model",      model_label),
        ("Fiscal Year End",      "March 31"),
        ("Base Currency",        "INR (₹)"),
        ("Share Price (ref)",    f"₹{price:,.2f}"),
        ("Shares Outstanding",   f"{shares/1e7:.1f} Cr"),
        ("Market Cap (ref)",     f"₹{mktcap_cr:,.0f} Cr"),
        ("Intrinsic Value",      f"₹{_safe(result.get('intrinsic_value', 0)):,.2f}"),
        ("Upside / (Downside)",  f"{_safe(result.get('upside_pct', 0)):.1f}%"),
        ("Model Date",           datetime.date.today().strftime("%B %d, %Y").replace(" 0", " ")),
    ]

    for i, (label, value) in enumerate(meta_rows):
        r = 11 + i
        lc = ws.cell(row=r, column=2, value=label)
        lc.font = Font(name="Calibri", size=10, color="A0A0A0")
        vc = ws.cell(row=r, column=4, value=value)
        vc.font = Font(name="Calibri", size=10, color=WHITE, bold=True)

    for col in range(2, 11):
        ws.cell(row=24, column=col).fill = _fill(COVER_GOLD)

    idx = ws.cell(row=26, column=2, value="MODEL INDEX")
    idx.font = Font(name="Calibri", bold=True, size=11, color=COVER_GOLD)

    sheets = [
        ("1", "Cover",                   "Company overview and model summary"),
        ("2", "Inputs & Assumptions",    "Key model parameters and financial data"),
        ("3", "Analysis",                "Model-specific calculations and projections"),
        ("4", "Results & Sensitivity",   "Intrinsic value, upside, and sensitivity analysis"),
    ]
    for i, (num, sheet, desc) in enumerate(sheets):
        r = 28 + i
        ws.cell(row=r, column=2, value=num).font = Font(name="Calibri", size=10, color=COVER_GOLD, bold=True)
        ws.cell(row=r, column=3, value=sheet).font = Font(name="Calibri", size=10, color=WHITE, bold=True)
        ws.cell(row=r, column=5, value=desc).font = Font(name="Calibri", size=10, color="808080")

    disc = ws.cell(row=39, column=2,
        value="DISCLAIMER: This model is for educational and illustrative purposes only. "
              "It does not constitute investment advice. Financial data sourced from NSE/BSE "
              "filings and public sources. Assumptions are model estimates only. "
              "Past performance is not indicative of future results.")
    disc.font = Font(name="Calibri", size=8, color="606060", italic=True)
    disc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.merge_cells("B39:J40")
    ws.row_dimensions[5].height = 30
    ws.row_dimensions[7].height = 22


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — INPUTS & ASSUMPTIONS  (shared framework)
# ═══════════════════════════════════════════════════════════════════════════════

def _build_inputs(wb: Workbook, model_id: str, company_name: str,
                  fin: dict, params: dict):
    ws = wb.create_sheet("Inputs & Assumptions")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {"A": 2, "B": 40, "C": 18, "D": 18, "E": 14, "F": 46})

    t = ws.cell(row=2, column=2, value=f"{company_name} — Inputs & Assumptions")
    t.font = _font(bold=True, size=13)
    ws.merge_cells("B2:F2")

    sub = ws.cell(row=3, column=2,
        value="Blue text = hardcoded inputs  |  Yellow background = key driver  |  All monetary values in ₹")
    sub.font = _font(size=8, color="808080", italic=True)
    ws.merge_cells("B3:F3")

    # Header row
    row = 5
    for col, hdr in [(2, "Parameter"), (3, "Value"), (4, "Unit"), (5, "Source"), (6, "Description")]:
        c = ws.cell(row=row, column=col, value=hdr)
        c.font = _font(bold=True, color=WHITE, size=9)
        c.fill = _fill(NAVY)
        c.alignment = _align(h="center")
        c.border = _border()

    # ── Financial Data section ──────────────────────────────────────────────
    row = 6
    _section_header(ws, row, 2, "FINANCIAL DATA  (from screener)", width_cols=5)
    row += 1

    fin_rows = [
        ("Share Price",         _safe(fin.get("price")),          "₹",      "#,##0.00"),
        ("Shares Outstanding",  _safe(fin.get("shares"))/1e7,     "Cr",     "#,##0.00"),
        ("Revenue (TTM)",       _safe(fin.get("revenue"))/1e7,    "₹ Cr",   "#,##0"),
        ("EBITDA (TTM)",        _safe(fin.get("ebitda"))/1e7,     "₹ Cr",   "#,##0"),
        ("Net Income (TTM)",    _safe(fin.get("net_income"))/1e7, "₹ Cr",   "#,##0"),
        ("Free Cash Flow",      _safe(fin.get("fcf"))/1e7,        "₹ Cr",   "#,##0"),
        ("Operating CF",        _safe(fin.get("operating_cf"))/1e7,"₹ Cr",  "#,##0"),
        ("Total Debt",          _safe(fin.get("total_debt"))/1e7, "₹ Cr",   "#,##0"),
        ("Cash & Equivalents",  _safe(fin.get("cash"))/1e7,       "₹ Cr",   "#,##0"),
        ("Total Assets",        _safe(fin.get("total_assets"))/1e7,"₹ Cr",  "#,##0"),
        ("Total Liabilities",   _safe(fin.get("total_liabilities"))/1e7,"₹ Cr","#,##0"),
        ("Book Value/Share",    _safe(fin.get("book_value_per_share")),"₹",  "#,##0.00"),
        ("EPS (TTM)",           _safe(fin.get("eps")),             "₹",      "#,##0.00"),
        ("DPS (Annual)",        _safe(fin.get("dps")),             "₹",      "#,##0.00"),
        ("Beta",                _safe(fin.get("beta", 1.0)),       "x",      "0.00"),
        ("Revenue Growth",      _safe(fin.get("revenue_growth")),  "%",      "0.0%"),
        ("Earnings Growth",     _safe(fin.get("earnings_growth")), "%",      "0.0%"),
        ("Profit Margin",       _safe(fin.get("profit_margins")),  "%",      "0.0%"),
        ("ROE",                 _safe(fin.get("roe")),             "%",      "0.0%"),
        ("ROA",                 _safe(fin.get("roa")),             "%",      "0.0%"),
        ("NOPAT",               _safe(fin.get("nopat"))/1e7,      "₹ Cr",   "#,##0"),
        ("Invested Capital",    _safe(fin.get("invested_capital"))/1e7,"₹ Cr","#,##0"),
        ("Annual Volatility",   _safe(fin.get("volatility_annual",0.3)),"σ", "0.0%"),
    ]

    for label, value, unit, fmt in fin_rows:
        _label(ws, row, 2, label)
        _val(ws, row, 3, value, fmt=fmt, color=BLUE_TEXT)
        c = ws.cell(row=row, column=4, value=unit)
        c.font = _font(size=9, color="808080")
        c.alignment = _align(h="center")
        c.border = _border()
        c = ws.cell(row=row, column=5, value="yfinance / screener.in")
        c.font = _font(size=9, color="808080")
        c.border = _border()
        row += 1

    # ── Model Parameters section ─────────────────────────────────────────────
    row += 1
    _section_header(ws, row, 2, "MODEL PARAMETERS  (user inputs)", width_cols=5)
    row += 1

    for key, value in params.items():
        label = key.replace("_", " ").title()
        _label(ws, row, 2, label)
        if isinstance(value, float) and abs(value) < 2.0 and key not in (
                "entry_multiple","exit_multiple","pe_multiple","pb_multiple",
                "ev_ebitda_multiple","ev_revenue_multiple","target_peg",
                "asset_life","hold_years","shares","ebitda","revenue"):
            _pct_cell(ws, row, 3, value, fill_hex=YELLOW_FILL)
            c = ws.cell(row=row, column=4, value="%")
        else:
            if isinstance(value, (int, float)):
                _val(ws, row, 3, value, fmt="#,##0.00", color=BLUE_TEXT, fill_hex=YELLOW_FILL)
            else:
                c2 = ws.cell(row=row, column=3, value=str(value))
                c2.font = _font(color=BLUE_TEXT, size=9)
                c2.fill = _fill(YELLOW_FILL)
                c2.border = _border()
            c = ws.cell(row=row, column=4, value="—")
        c.font = _font(size=9, color="808080")
        c.alignment = _align(h="center")
        c.border = _border()
        desc_cell = ws.cell(row=row, column=6, value=f"User-defined: {label}")
        desc_cell.font = _font(size=9, color="808080", italic=True)
        desc_cell.border = _border()
        row += 1

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — RESULTS & SENSITIVITY  (shared framework)
# ═══════════════════════════════════════════════════════════════════════════════

def _build_results_sensitivity(wb: Workbook, model_id: str, company_name: str,
                                fin: dict, result: dict, params: dict,
                                sens_fn):
    ws = wb.create_sheet("Results & Sensitivity")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {"A": 2, "B": 36, "C": 16, "D": 16,
                         "E": 16, "F": 16, "G": 16, "H": 16})

    t = ws.cell(row=2, column=2, value=f"{company_name} — Results & Sensitivity")
    t.font = _font(bold=True, size=13)
    ws.merge_cells("B2:H2")

    # Results section
    row = 4
    _section_header(ws, row, 2, "VALUATION RESULTS", width_cols=7)
    row += 1

    iv   = _safe(result.get("intrinsic_value", 0))
    cp   = _safe(result.get("current_price", _safe(fin.get("price", 0))))
    upside = _safe(result.get("upside_pct", ((iv/cp - 1)*100) if cp else 0))

    results_data = [
        ("Intrinsic Value (Model Output)", iv,     "#,##0.00", GREEN_FILL, GREEN_TEXT),
        ("Current Market Price",           cp,     "#,##0.00", None,       BLUE_TEXT),
        ("Upside / (Downside)",            upside/100, "0.0%",
         GREEN_FILL if upside >= 0 else RED_FILL,
         GREEN_TEXT if upside >= 0 else RED_TEXT),
    ]

    for label, value, fmt, fill, color in results_data:
        _label(ws, row, 2, label, bold=True)
        _val(ws, row, 3, value, fmt=fmt, color=color, fill_hex=fill, bold=True)
        for col in range(4, 9):
            c = ws.cell(row=row, column=col)
            c.border = _border()
        row += 1

    # Model-specific metrics
    row += 1
    _sub_header(ws, row, 2, "KEY METRICS", width_cols=7)
    row += 1
    skip_keys = {"intrinsic_value", "current_price", "upside_pct", "sensitivity",
                 "year_details", "dividend_schedule"}
    for key, value in result.items():
        if key in skip_keys or not isinstance(value, (int, float)):
            continue
        label = key.replace("_", " ").title()
        _label(ws, row, 2, label)
        if abs(_safe(value)) < 5 and key not in ("moic", "irr"):
            _pct_cell(ws, row, 3, _safe(value))
        else:
            _val(ws, row, 3, _safe(value), fmt="#,##0.00")
        row += 1

    # Sensitivity table
    row += 2
    _section_header(ws, row, 2, "SENSITIVITY ANALYSIS", width_cols=7)
    row += 1

    note = ws.cell(row=row, column=2,
        value="Gold = base case  |  Green = premium to current price  |  Red = discount to current price")
    note.font = _font(size=8, color="808080", italic=True)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    row += 2

    sens_fn(ws, row, fin, result, params, cp)
    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SENSITIVITY HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def _write_sens_table(ws, start_row, start_col, row_label, col_label,
                      row_vals, col_vals, value_matrix, base_row_idx, base_col_idx,
                      current_price):
    """Write a 5x5 (or NxM) sensitivity table with heat-map colouring."""
    # Column header row
    r = start_row
    lbl = ws.cell(row=r, column=start_col, value=f"{row_label} \\ {col_label}")
    lbl.font = _font(bold=True, color=WHITE, size=9)
    lbl.fill = _fill(NAVY)
    lbl.alignment = _align(h="center")
    lbl.border = _border()

    for j, cv in enumerate(col_vals):
        c = ws.cell(row=r, column=start_col + 1 + j, value=cv)
        c.font = _font(bold=True, color=WHITE, size=9)
        c.fill = _fill(DARK_BLUE)
        c.number_format = "0.0%" if abs(cv) < 2 else "0.00"
        c.alignment = _align(h="center")
        c.border = _border()

    for i, rv in enumerate(row_vals):
        r = start_row + 1 + i
        rc = ws.cell(row=r, column=start_col, value=rv)
        rc.font = _font(bold=True, size=9)
        rc.fill = _fill(LIGHT_BLUE_FILL)
        rc.number_format = "0.0%" if abs(rv) < 2 else "0.00"
        rc.alignment = _align(h="right")
        rc.border = _border()

        for j, val in enumerate(value_matrix[i]):
            cell = ws.cell(row=r, column=start_col + 1 + j, value=val)
            cell.number_format = "#,##0.00"
            cell.alignment = _align(h="center")
            cell.border = _border()
            is_base = (i == base_row_idx and j == base_col_idx)
            if is_base:
                cell.fill = _fill(SENSITIVITY_GOLD)
                cell.font = _font(bold=True, size=9)
            elif val >= current_price:
                cell.fill = _fill(GREEN_FILL)
                cell.font = _font(color=GREEN_TEXT, size=9)
            else:
                cell.fill = _fill(RED_FILL)
                cell.font = _font(color=RED_TEXT, size=9)


def _range_around(base, pct=0.30, steps=5):
    """5 values: base ±30% in equal steps."""
    lo = base * (1 - pct)
    hi = base * (1 + pct)
    step = (hi - lo) / (steps - 1)
    return [lo + step * i for i in range(steps)]


# ═══════════════════════════════════════════════════════════════════════════════
# MODEL-SPECIFIC ANALYSIS SHEETS
# ═══════════════════════════════════════════════════════════════════════════════

# ── DCF FCFE ──────────────────────────────────────────────────────────────────

def _analysis_dcf_fcfe(wb, company_name, fin, result, params):
    ws = wb.create_sheet("Analysis")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {"A": 2, "B": 36, "C": 14, "D": 14, "E": 14,
                         "F": 14, "G": 14, "H": 14})

    t = ws.cell(row=2, column=2, value=f"{company_name} — DCF FCFE Analysis")
    t.font = _font(bold=True, size=13)
    ws.merge_cells("B2:H2")

    row = 4
    _section_header(ws, row, 2, "FREE CASH FLOW TO EQUITY — YEAR-BY-YEAR PROJECTIONS", width_cols=7)
    row += 1

    years = list(range(1, 6))
    _label(ws, row, 2, "Year", bold=True)
    for i, yr in enumerate(years):
        c = ws.cell(row=row, column=3 + i, value=f"Year {yr}")
        c.font = _font(bold=True, color=WHITE, size=9)
        c.fill = _fill(DARK_BLUE)
        c.alignment = _align(h="center")
        c.border = _border()
    row += 1

    g1   = _safe(params.get("growth_stage1", fin.get("earnings_growth", 0.15)))
    g2   = _safe(params.get("terminal_growth", 0.05))
    coe  = _safe(params.get("cost_of_equity", 0.12))
    ni0  = _safe(fin.get("net_income", 1e10))
    capex = _safe(fin.get("fcf", ni0 * 0.7))

    ni_vals, fcfe_vals, disc_vals, pv_vals = [], [], [], []
    for yr in years:
        g = g1 if yr <= 3 else g1 * 0.5 + g2 * 0.5
        ni  = ni0 * ((1 + g) ** yr)
        fcfe = ni * 0.80
        df  = 1 / ((1 + coe) ** yr)
        pv  = fcfe * df
        ni_vals.append(ni); fcfe_vals.append(fcfe)
        disc_vals.append(df); pv_vals.append(pv)

    rows_data = [
        ("Net Income (₹ Cr)",        [v/1e7 for v in ni_vals],    "#,##0",  None),
        ("(−) Capex & Reinvestment", [v/1e7*0.15 for v in ni_vals],"#,##0", None),
        ("(+) D&A",                  [v/1e7*0.08 for v in ni_vals],"#,##0", None),
        ("(+/−) Working Capital Δ",  [v/1e7*0.03 for v in ni_vals],"#,##0", None),
        ("FCFE (₹ Cr)",              [v/1e7 for v in fcfe_vals],  "#,##0",  LIGHT_BLUE_FILL),
        ("Discount Factor",          disc_vals,                    "0.0000", None),
        ("PV of FCFE (₹ Cr)",        [v/1e7 for v in pv_vals],   "#,##0",  GREEN_FILL),
    ]
    for label, vals, fmt, fill in rows_data:
        _label(ws, row, 2, label, bold=(fill is not None))
        for i, v in enumerate(vals):
            _val(ws, row, 3 + i, v, fmt=fmt, fill_hex=fill)
        row += 1

    row += 1
    _section_header(ws, row, 2, "VALUATION BRIDGE", width_cols=7)
    row += 1

    pv_sum   = sum(pv_vals)
    tv_fcfe  = fcfe_vals[-1] * (1 + g2) / (coe - g2) if (coe - g2) > 0 else 0
    pv_tv    = tv_fcfe / ((1 + coe) ** len(years))
    eq_val   = pv_sum + pv_tv
    iv_share = eq_val / _safe(fin.get("shares", 1), 1)

    bridge = [
        ("PV of FCFE (5 Years, ₹ Cr)",  pv_sum/1e7,    "#,##0"),
        ("Terminal Value FCFE (₹ Cr)",   tv_fcfe/1e7,   "#,##0"),
        ("PV of Terminal Value (₹ Cr)",  pv_tv/1e7,     "#,##0"),
        ("Total Equity Value (₹ Cr)",    eq_val/1e7,    "#,##0"),
        ("Shares Outstanding (Cr)",      _safe(fin.get("shares"))/1e7, "#,##0.00"),
        ("Intrinsic Value Per Share (₹)",iv_share,      "#,##0.00"),
    ]
    for label, value, fmt in bridge:
        _label(ws, row, 2, label, bold=("Intrinsic" in label or "Total" in label))
        _val(ws, row, 3, value, fmt=fmt,
             fill_hex=GREEN_FILL if "Intrinsic" in label else None,
             color=GREEN_TEXT if "Intrinsic" in label else BLACK)
        row += 1

    _append_year_details(ws, row, result)
    return ws


def _append_year_details(ws, row, result):
    year_details = result.get("year_details", [])
    if not year_details:
        return
    row += 2
    _section_header(ws, row, 2, "DETAILED YEAR-BY-YEAR DATA (from model)", width_cols=7)
    row += 1
    if isinstance(year_details[0], dict):
        keys = list(year_details[0].keys())
        for j, k in enumerate(keys):
            c = ws.cell(row=row, column=2 + j, value=k.replace("_", " ").title())
            c.font = _font(bold=True, color=WHITE, size=9)
            c.fill = _fill(DARK_BLUE)
            c.border = _border()
        row += 1
        for entry in year_details:
            for j, k in enumerate(keys):
                v = entry.get(k, "")
                c = ws.cell(row=row, column=2 + j, value=v)
                c.font = _font(size=9)
                c.border = _border()
                if isinstance(v, float):
                    c.number_format = "0.0%" if abs(v) < 2 else "#,##0.00"
                c.alignment = _align(h="right" if isinstance(v, (int, float)) else "left")
            row += 1


# ── DCF MULTISTAGE ────────────────────────────────────────────────────────────

def _analysis_dcf_multistage(wb, company_name, fin, result, params):
    ws = wb.create_sheet("Analysis")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {"A": 2, "B": 36, "C": 14, "D": 14, "E": 14,
                         "F": 14, "G": 14, "H": 14})

    t = ws.cell(row=2, column=2, value=f"{company_name} — Multi-Stage DCF Analysis")
    t.font = _font(bold=True, size=13)
    ws.merge_cells("B2:H2")

    row = 4
    _section_header(ws, row, 2, "MULTI-STAGE CASH FLOW PROJECTIONS (10 YEARS)", width_cols=7)
    row += 1

    g1   = _safe(params.get("growth_stage1", 0.20))
    g2   = _safe(params.get("growth_stage2", 0.12))
    gT   = _safe(params.get("terminal_growth", 0.05))
    wacc = _safe(params.get("wacc", 0.12))
    fcf0 = _safe(fin.get("fcf", fin.get("nopat", 1e10)))

    _label(ws, row, 2, "Stage", bold=True)
    _label(ws, row, 3, "Year", bold=True)
    _label(ws, row, 4, "Growth Rate", bold=True)
    _label(ws, row, 5, "FCF (₹ Cr)", bold=True)
    _label(ws, row, 6, "Disc. Factor", bold=True)
    _label(ws, row, 7, "PV FCF (₹ Cr)", bold=True)
    row += 1

    pv_total = 0.0
    last_fcf = fcf0
    for yr in range(1, 11):
        stage = "High Growth" if yr <= 5 else "Transition"
        g = g1 if yr <= 5 else g2
        last_fcf = last_fcf * (1 + g)
        df = 1 / ((1 + wacc) ** yr)
        pv = last_fcf * df
        pv_total += pv
        fill = LIGHT_BLUE_FILL if yr % 2 == 0 else None
        _label(ws, row, 2, stage)
        _val(ws, row, 3, yr, fmt="0")
        _pct_cell(ws, row, 4, g, fill_hex=fill)
        _val(ws, row, 5, last_fcf/1e7, fmt="#,##0", fill_hex=fill)
        _val(ws, row, 6, df, fmt="0.0000", fill_hex=fill)
        _val(ws, row, 7, pv/1e7, fmt="#,##0", fill_hex=fill)
        row += 1

    row += 1
    _section_header(ws, row, 2, "TERMINAL VALUE & EQUITY BRIDGE", width_cols=7)
    row += 1
    tv       = last_fcf * (1 + gT) / (wacc - gT) if (wacc - gT) > 0 else 0
    pv_tv    = tv / ((1 + wacc) ** 10)
    debt     = _safe(fin.get("total_debt", 0))
    cash_val = _safe(fin.get("cash", 0))
    eq_val   = pv_total + pv_tv - debt + cash_val
    shares   = _safe(fin.get("shares", 1), 1)
    iv       = eq_val / shares

    bridge = [
        ("PV of FCF (Stages 1+2, ₹ Cr)", pv_total/1e7, "#,##0"),
        ("Terminal Value (₹ Cr)",          tv/1e7,       "#,##0"),
        ("PV of Terminal Value (₹ Cr)",    pv_tv/1e7,    "#,##0"),
        ("(−) Total Debt (₹ Cr)",          debt/1e7,     "#,##0"),
        ("(+) Cash & Equivalents (₹ Cr)",  cash_val/1e7, "#,##0"),
        ("Enterprise Value (₹ Cr)",        (pv_total+pv_tv)/1e7, "#,##0"),
        ("Equity Value (₹ Cr)",            eq_val/1e7,   "#,##0"),
        ("Intrinsic Value Per Share (₹)",  iv,           "#,##0.00"),
    ]
    for label, value, fmt in bridge:
        bold = "Intrinsic" in label or "Equity Value" in label
        _label(ws, row, 2, label, bold=bold)
        _val(ws, row, 3, value, fmt=fmt,
             fill_hex=GREEN_FILL if "Intrinsic" in label else (LIGHT_BLUE_FILL if "Enterprise" in label else None),
             color=GREEN_TEXT if "Intrinsic" in label else BLACK, bold=bold)
        row += 1

    _append_year_details(ws, row, result)
    return ws


# ── GORDON GROWTH ─────────────────────────────────────────────────────────────

def _analysis_gordon_growth(wb, company_name, fin, result, params):
    ws = wb.create_sheet("Analysis")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {"A": 2, "B": 36, "C": 18, "D": 18, "E": 18, "F": 18})

    t = ws.cell(row=2, column=2, value=f"{company_name} — Gordon Growth Model Analysis")
    t.font = _font(bold=True, size=13)
    ws.merge_cells("B2:F2")

    dps     = _safe(params.get("dps", fin.get("dps", 5)))
    coe     = _safe(params.get("cost_of_equity", 0.12))
    g       = _safe(params.get("terminal_growth", 0.05))
    iv      = dps * (1 + g) / (coe - g) if (coe - g) > 0 else 0

    row = 4
    _section_header(ws, row, 2, "GORDON GROWTH MODEL — DIVIDEND STREAM", width_cols=5)
    row += 1

    _label(ws, row, 2, "Year", bold=True)
    _label(ws, row, 3, "Dividend (₹)", bold=True)
    _label(ws, row, 4, "Disc. Factor", bold=True)
    _label(ws, row, 5, "PV of Dividend", bold=True)
    row += 1

    pv_sum = 0.0
    last_div = dps
    for yr in range(1, 11):
        last_div = last_div * (1 + g)
        df = 1 / ((1 + coe) ** yr)
        pv = last_div * df
        pv_sum += pv
        fill = LIGHT_BLUE_FILL if yr % 2 == 0 else None
        _val(ws, row, 2, yr, fmt="0")
        _val(ws, row, 3, last_div, fmt="#,##0.00", fill_hex=fill)
        _val(ws, row, 4, df, fmt="0.0000", fill_hex=fill)
        _val(ws, row, 5, pv, fmt="#,##0.00", fill_hex=fill)
        row += 1

    row += 1
    _section_header(ws, row, 2, "GORDON GROWTH — FORMULA DERIVATION", width_cols=5)
    row += 1

    data = [
        ("Current DPS (D₀)",               dps,           "#,##0.00"),
        ("Expected DPS Next Year (D₁)",    dps*(1+g),     "#,##0.00"),
        ("Dividend Growth Rate (g)",        g,             "0.0%"),
        ("Cost of Equity (Ke)",             coe,           "0.0%"),
        ("(Ke − g) Spread",                coe - g,       "0.0%"),
        ("Gordon Value = D₁ / (Ke − g)",   iv,            "#,##0.00"),
        ("PV Explicit Dividends (10yr)",    pv_sum,        "#,##0.00"),
        ("Current Market Price",           _safe(fin.get("price")), "#,##0.00"),
        ("Upside / (Downside)",            (iv/_safe(fin.get("price",1))-1), "0.0%"),
    ]
    for label, value, fmt in data:
        bold = "Gordon Value" in label or "Intrinsic" in label
        _label(ws, row, 2, label, bold=bold)
        if "%" in fmt:
            _pct_cell(ws, row, 3, value,
                      fill_hex=GREEN_FILL if "Gordon Value" in label else None)
        else:
            _val(ws, row, 3, value, fmt=fmt,
                 fill_hex=GREEN_FILL if "Gordon Value" in label else None,
                 color=GREEN_TEXT if "Gordon Value" in label else BLACK)
        row += 1

    return ws


# ── DDM MULTISTAGE ────────────────────────────────────────────────────────────

def _analysis_ddm_multistage(wb, company_name, fin, result, params):
    ws = wb.create_sheet("Analysis")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {"A": 2, "B": 36, "C": 16, "D": 16, "E": 16, "F": 16, "G": 16})

    t = ws.cell(row=2, column=2, value=f"{company_name} — Multi-Stage DDM Analysis")
    t.font = _font(bold=True, size=13)
    ws.merge_cells("B2:G2")

    dps   = _safe(params.get("dps", fin.get("dps", 5)))
    g1    = _safe(params.get("growth_stage1", 0.15))
    g2    = _safe(params.get("terminal_growth", 0.05))
    coe   = _safe(params.get("cost_of_equity", 0.12))

    row = 4
    _section_header(ws, row, 2, "DIVIDEND SCHEDULE — MULTI-STAGE", width_cols=6)
    row += 1

    headers = ["Year", "Stage", "Growth Rate", "Dividend (₹)", "Disc. Factor", "PV (₹)"]
    for j, h in enumerate(headers):
        c = ws.cell(row=row, column=2 + j, value=h)
        c.font = _font(bold=True, color=WHITE, size=9)
        c.fill = _fill(DARK_BLUE)
        c.alignment = _align(h="center")
        c.border = _border()
    row += 1

    pv_sum = 0.0
    last_div = dps
    for yr in range(1, 11):
        stage = "High Growth" if yr <= 5 else "Stable Growth"
        g = g1 if yr <= 5 else g2
        last_div = last_div * (1 + g)
        df = 1 / ((1 + coe) ** yr)
        pv = last_div * df
        pv_sum += pv
        fill = LIGHT_BLUE_FILL if yr % 2 == 0 else None
        _val(ws, row, 2, yr, fmt="0")
        c = ws.cell(row=row, column=3, value=stage)
        c.font = _font(size=9); c.border = _border(); c.alignment = _align()
        _pct_cell(ws, row, 4, g, fill_hex=fill)
        _val(ws, row, 5, last_div, fmt="#,##0.00", fill_hex=fill)
        _val(ws, row, 6, df, fmt="0.0000", fill_hex=fill)
        _val(ws, row, 7, pv, fmt="#,##0.00", fill_hex=fill)
        row += 1

    row += 1
    _section_header(ws, row, 2, "TERMINAL VALUE & INTRINSIC VALUE", width_cols=6)
    row += 1
    tv = last_div * (1 + g2) / (coe - g2) if (coe - g2) > 0 else 0
    pv_tv = tv / ((1 + coe) ** 10)
    iv = pv_sum + pv_tv

    data = [
        ("PV of Explicit Dividends (₹)", pv_sum,   "#,##0.00"),
        ("Terminal Dividend Value (₹)",   tv,       "#,##0.00"),
        ("PV of Terminal Value (₹)",      pv_tv,    "#,##0.00"),
        ("Intrinsic Value Per Share (₹)", iv,       "#,##0.00"),
    ]
    for label, value, fmt in data:
        bold = "Intrinsic" in label
        _label(ws, row, 2, label, bold=bold)
        _val(ws, row, 3, value, fmt=fmt,
             fill_hex=GREEN_FILL if bold else None,
             color=GREEN_TEXT if bold else BLACK)
        row += 1

    return ws


# ── RESIDUAL INCOME ───────────────────────────────────────────────────────────

def _analysis_residual_income(wb, company_name, fin, result, params):
    ws = wb.create_sheet("Analysis")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {"A": 2, "B": 36, "C": 14, "D": 14, "E": 14,
                         "F": 14, "G": 14, "H": 14})

    t = ws.cell(row=2, column=2, value=f"{company_name} — Residual Income Valuation")
    t.font = _font(bold=True, size=13)
    ws.merge_cells("B2:H2")

    bvps0  = _safe(params.get("book_value_per_share", fin.get("book_value_per_share", 100)))
    roe    = _safe(params.get("roe", fin.get("roe", 0.18)))
    coe    = _safe(params.get("cost_of_equity", 0.12))
    g      = _safe(params.get("terminal_growth", 0.04))
    payout = _safe(params.get("payout_ratio", 0.30))

    row = 4
    _section_header(ws, row, 2, "RESIDUAL INCOME SCHEDULE (10 YEARS)", width_cols=7)
    row += 1

    headers = ["Year", "BV/Share", "EPS", "Req. Return", "Res. Income", "Disc. Factor", "PV of RI"]
    for j, h in enumerate(headers):
        c = ws.cell(row=row, column=2 + j, value=h)
        c.font = _font(bold=True, color=WHITE, size=9)
        c.fill = _fill(DARK_BLUE)
        c.alignment = _align(h="center")
        c.border = _border()
    row += 1

    bvps = bvps0
    pv_sum = 0.0
    last_ri = 0.0
    for yr in range(1, 11):
        eps = bvps * roe
        req = bvps * coe
        ri  = eps - req
        df  = 1 / ((1 + coe) ** yr)
        pv  = ri * df
        pv_sum += pv
        last_ri = ri
        fill = LIGHT_BLUE_FILL if yr % 2 == 0 else None
        _val(ws, row, 2, yr, fmt="0")
        _val(ws, row, 3, bvps, fmt="#,##0.00", fill_hex=fill)
        _val(ws, row, 4, eps, fmt="#,##0.00", fill_hex=fill)
        _val(ws, row, 5, req, fmt="#,##0.00", fill_hex=fill)
        _val(ws, row, 6, ri, fmt="#,##0.00", fill_hex=GREEN_FILL if ri > 0 else None)
        _val(ws, row, 7, df, fmt="0.0000")
        _val(ws, row, 8, pv, fmt="#,##0.00", fill_hex=fill)
        bvps = bvps + eps * (1 - payout)
        row += 1

    row += 1
    _section_header(ws, row, 2, "TERMINAL VALUE & INTRINSIC VALUE", width_cols=7)
    row += 1
    tv = last_ri * (1 + g) / (coe - g) if (coe - g) > 0 else 0
    pv_tv = tv / ((1 + coe) ** 10)
    iv = bvps0 + pv_sum + pv_tv

    data = [
        ("Book Value Per Share (₹)",      bvps0,    "#,##0.00"),
        ("PV of Residual Income (₹)",     pv_sum,   "#,##0.00"),
        ("Terminal RI Value (₹)",         tv,       "#,##0.00"),
        ("PV of Terminal RI (₹)",         pv_tv,    "#,##0.00"),
        ("Intrinsic Value Per Share (₹)", iv,       "#,##0.00"),
    ]
    for label, value, fmt in data:
        bold = "Intrinsic" in label
        _label(ws, row, 2, label, bold=bold)
        _val(ws, row, 3, value, fmt=fmt,
             fill_hex=GREEN_FILL if bold else None,
             color=GREEN_TEXT if bold else BLACK)
        row += 1

    return ws


# ── TRADING COMPS ─────────────────────────────────────────────────────────────

def _analysis_trading_comps(wb, company_name, fin, result, params):
    ws = wb.create_sheet("Analysis")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {"A": 2, "B": 26, "C": 14, "D": 14, "E": 14,
                         "F": 14, "G": 14, "H": 14})

    t = ws.cell(row=2, column=2, value=f"{company_name} — Trading Comparables Analysis")
    t.font = _font(bold=True, size=13)
    ws.merge_cells("B2:H2")

    pe   = _safe(params.get("pe_multiple", 25))
    pb   = _safe(params.get("pb_multiple", 4))
    ev_e = _safe(params.get("ev_ebitda_multiple", 15))
    eps  = _safe(fin.get("eps", 50))
    bvps = _safe(fin.get("book_value_per_share", 300))
    ebitda = _safe(fin.get("ebitda", 1e10))
    shares = _safe(fin.get("shares", 1), 1)
    debt  = _safe(fin.get("total_debt", 0))
    cash  = _safe(fin.get("cash", 0))

    row = 4
    _section_header(ws, row, 2, "TRADING COMPS — MULTIPLE ANALYSIS", width_cols=7)
    row += 1

    headers = ["Multiple", "Base Value", "Applied Multiple", "Implied Value (₹)",
               "Current Price", "Premium/(Discount)"]
    for j, h in enumerate(headers):
        c = ws.cell(row=row, column=2 + j, value=h)
        c.font = _font(bold=True, color=WHITE, size=9)
        c.fill = _fill(DARK_BLUE)
        c.alignment = _align(h="center")
        c.border = _border()
    row += 1

    price = _safe(fin.get("price", 800))
    eq_from_ev = lambda ev: (ev - debt + cash) / shares

    comp_data = [
        ("P/E Multiple",         eps,                      pe,   eps * pe),
        ("P/B Multiple",         bvps,                     pb,   bvps * pb),
        ("EV/EBITDA",            ebitda/1e7,               ev_e, eq_from_ev(ebitda * ev_e)),
    ]

    for label, base, mult, implied in comp_data:
        prem = (implied / price - 1) if price else 0
        fill = GREEN_FILL if prem >= 0 else RED_FILL
        _label(ws, row, 2, label)
        _val(ws, row, 3, base, fmt="#,##0.00")
        _val(ws, row, 4, mult, fmt="0.0x")
        _val(ws, row, 5, implied, fmt="#,##0.00", fill_hex=fill)
        _val(ws, row, 6, price, fmt="#,##0.00")
        _pct_cell(ws, row, 7, prem, fill_hex=fill,
                  color=GREEN_TEXT if prem >= 0 else RED_TEXT)
        row += 1

    row += 2
    _section_header(ws, row, 2, "SENSITIVITY — IMPLIED VALUE AT DIFFERENT MULTIPLES", width_cols=7)
    row += 1

    note = ws.cell(row=row, column=2, value="Table shows Implied Price at P/E multiples (rows) × P/B multiples (cols)")
    note.font = _font(size=9, color="808080", italic=True)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    row += 2

    pe_range = [pe * (0.7 + 0.15*i) for i in range(5)]
    pb_range = [pb * (0.7 + 0.15*i) for i in range(5)]

    lbl = ws.cell(row=row, column=2, value="P/E \\ P/B")
    lbl.font = _font(bold=True, color=WHITE, size=9)
    lbl.fill = _fill(NAVY)
    lbl.border = _border()
    for j, pv_ in enumerate(pb_range):
        c = ws.cell(row=row, column=3 + j, value=round(pv_, 1))
        c.font = _font(bold=True, color=WHITE, size=9)
        c.fill = _fill(DARK_BLUE)
        c.number_format = "0.0"
        c.alignment = _align(h="center")
        c.border = _border()
    row += 1

    for i, pe_ in enumerate(pe_range):
        rc = ws.cell(row=row, column=2, value=round(pe_, 1))
        rc.font = _font(bold=True, size=9)
        rc.fill = _fill(LIGHT_BLUE_FILL)
        rc.number_format = "0.0"
        rc.alignment = _align(h="right")
        rc.border = _border()
        for j, pb_ in enumerate(pb_range):
            avg = (eps * pe_ + bvps * pb_) / 2
            cell = ws.cell(row=row, column=3 + j, value=avg)
            cell.number_format = "#,##0.00"
            cell.alignment = _align(h="center")
            cell.border = _border()
            is_base = (abs(pe_ - pe) < 1 and abs(pb_ - pb) < 1)
            if is_base:
                cell.fill = _fill(SENSITIVITY_GOLD)
                cell.font = _font(bold=True, size=9)
            elif avg >= price:
                cell.fill = _fill(GREEN_FILL); cell.font = _font(color=GREEN_TEXT, size=9)
            else:
                cell.fill = _fill(RED_FILL); cell.font = _font(color=RED_TEXT, size=9)
        row += 1

    return ws


# ── PRECEDENT TRANSACTIONS ────────────────────────────────────────────────────

def _analysis_precedent_transactions(wb, company_name, fin, result, params):
    ws = wb.create_sheet("Analysis")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {"A": 2, "B": 30, "C": 18, "D": 18, "E": 18, "F": 18, "G": 18})

    t = ws.cell(row=2, column=2, value=f"{company_name} — Precedent Transactions")
    t.font = _font(bold=True, size=13)
    ws.merge_cells("B2:G2")

    ev_mult = _safe(params.get("ev_ebitda_multiple", 14))
    premium = _safe(params.get("deal_premium", 0.25))
    ebitda  = _safe(fin.get("ebitda", 1e10))
    shares  = _safe(fin.get("shares", 1), 1)
    debt    = _safe(fin.get("total_debt", 0))
    cash    = _safe(fin.get("cash", 0))
    price   = _safe(fin.get("price", 800))

    row = 4
    _section_header(ws, row, 2, "PRECEDENT TRANSACTIONS — DEAL METRICS", width_cols=6)
    row += 1

    ev    = ebitda * ev_mult
    eq_v  = ev - debt + cash
    iv    = eq_v / shares
    iv_p  = iv * (1 + premium)

    data = [
        ("EBITDA (₹ Cr)",                ebitda/1e7,   "#,##0"),
        ("EV/EBITDA Multiple Applied",   ev_mult,      "0.0x"),
        ("Enterprise Value (₹ Cr)",      ev/1e7,       "#,##0"),
        ("(−) Net Debt (₹ Cr)",          (debt-cash)/1e7, "#,##0"),
        ("Equity Value (₹ Cr)",          eq_v/1e7,     "#,##0"),
        ("Equity Value Per Share (₹)",   iv,           "#,##0.00"),
        ("Control Premium Applied",      premium,      "0.0%"),
        ("Implied Price incl. Premium (₹)", iv_p,      "#,##0.00"),
        ("Current Market Price (₹)",     price,        "#,##0.00"),
        ("Premium to Market Price",      (iv_p/price-1), "0.0%"),
    ]
    for label, value, fmt in data:
        bold = "Implied Price" in label
        _label(ws, row, 2, label, bold=bold)
        if "%" in fmt:
            _pct_cell(ws, row, 3, value,
                      fill_hex=GREEN_FILL if bold else None)
        else:
            _val(ws, row, 3, value, fmt=fmt,
                 fill_hex=GREEN_FILL if bold else None,
                 color=GREEN_TEXT if bold else BLACK)
        row += 1

    row += 2
    _section_header(ws, row, 2, "TRANSACTION COMPS SUMMARY", width_cols=6)
    row += 1
    headers = ["Transaction", "EV/EBITDA", "Deal Premium", "Implied EV (₹ Cr)", "Implied Price (₹)"]
    for j, h in enumerate(headers):
        c = ws.cell(row=row, column=2 + j, value=h)
        c.font = _font(bold=True, color=WHITE, size=9)
        c.fill = _fill(DARK_BLUE); c.alignment = _align(h="center"); c.border = _border()
    row += 1

    scenarios = [
        ("Low Case",    ev_mult * 0.85, premium * 0.7),
        ("Base Case",   ev_mult,        premium),
        ("High Case",   ev_mult * 1.15, premium * 1.3),
        ("Bull Case",   ev_mult * 1.25, premium * 1.5),
        ("Bear Case",   ev_mult * 0.75, premium * 0.5),
    ]
    for name, em, pr in scenarios:
        ev_s  = ebitda * em
        iv_s  = (ev_s - debt + cash) / shares * (1 + pr)
        prem_ = iv_s / price - 1
        fill  = GREEN_FILL if prem_ >= 0 else RED_FILL
        _label(ws, row, 2, name)
        _val(ws, row, 3, em, fmt="0.0x")
        _pct_cell(ws, row, 4, pr)
        _val(ws, row, 5, ev_s/1e7, fmt="#,##0")
        _val(ws, row, 6, iv_s, fmt="#,##0.00", fill_hex=fill,
             color=GREEN_TEXT if prem_ >= 0 else RED_TEXT)
        row += 1

    return ws


# ── PEG ───────────────────────────────────────────────────────────────────────

def _analysis_peg(wb, company_name, fin, result, params):
    ws = wb.create_sheet("Analysis")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {"A": 2, "B": 36, "C": 18, "D": 18, "E": 18, "F": 18})

    t = ws.cell(row=2, column=2, value=f"{company_name} — PEG Ratio Valuation")
    t.font = _font(bold=True, size=13)
    ws.merge_cells("B2:F2")

    eps  = _safe(fin.get("eps", 50))
    g    = _safe(params.get("earnings_growth_pct", fin.get("earnings_growth", 0.15)))
    peg  = _safe(params.get("target_peg", 1.0))
    price = _safe(fin.get("price", 800))

    g_pct = g * 100 if g < 2 else g
    pe_implied = peg * g_pct
    iv = eps * pe_implied

    row = 4
    _section_header(ws, row, 2, "PEG RATIO — FORMULA BREAKDOWN", width_cols=5)
    row += 1

    data = [
        ("EPS (TTM, ₹)",                eps,       "#,##0.00"),
        ("Earnings Growth Rate (g)",    g if g < 2 else g/100, "0.0%"),
        ("Growth Rate for PEG (× 100)", g_pct,     "0.0"),
        ("Target PEG Ratio",            peg,       "0.00"),
        ("Implied P/E = PEG × g%",      pe_implied,"0.0x"),
        ("Implied Price = EPS × P/E",   iv,        "#,##0.00"),
        ("Current Market Price (₹)",    price,     "#,##0.00"),
        ("Actual P/E",                  price/eps if eps else 0, "0.0x"),
        ("Actual PEG = P/E ÷ g",        (price/eps/g_pct) if (eps and g_pct) else 0, "0.00"),
        ("Upside / (Downside)",         (iv/price - 1) if price else 0, "0.0%"),
    ]
    for label, value, fmt in data:
        bold = "Implied Price" in label
        _label(ws, row, 2, label, bold=bold)
        if "%" in fmt:
            _pct_cell(ws, row, 3, value, fill_hex=GREEN_FILL if bold else None)
        else:
            _val(ws, row, 3, value, fmt=fmt, fill_hex=GREEN_FILL if bold else None,
                 color=GREEN_TEXT if bold else BLACK)
        row += 1

    row += 2
    _section_header(ws, row, 2, "PEG SENSITIVITY — IMPLIED PRICE MATRIX", width_cols=5)
    row += 1
    note = ws.cell(row=row, column=2, value="Rows = Earnings Growth %  |  Cols = Target PEG ratio")
    note.font = _font(size=9, color="808080", italic=True)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    row += 2

    g_vals  = [g * (0.7 + 0.15*i) for i in range(5)]
    peg_vals = [peg * (0.7 + 0.15*i) for i in range(5)]

    lbl = ws.cell(row=row, column=2, value="g% \\ PEG")
    lbl.font = _font(bold=True, color=WHITE, size=9); lbl.fill = _fill(NAVY); lbl.border = _border()
    for j, pv in enumerate(peg_vals):
        c = ws.cell(row=row, column=3 + j, value=round(pv, 2))
        c.font = _font(bold=True, color=WHITE, size=9)
        c.fill = _fill(DARK_BLUE); c.number_format = "0.00"
        c.alignment = _align(h="center"); c.border = _border()
    row += 1

    for i, gv in enumerate(g_vals):
        gv_pct = gv * 100 if gv < 2 else gv
        rc = ws.cell(row=row, column=2, value=round(gv, 4))
        rc.font = _font(bold=True, size=9); rc.fill = _fill(LIGHT_BLUE_FILL)
        rc.number_format = "0.0%"; rc.alignment = _align(h="right"); rc.border = _border()
        for j, pv in enumerate(peg_vals):
            implied = eps * pv * gv_pct
            cell = ws.cell(row=row, column=3 + j, value=implied)
            cell.number_format = "#,##0.00"; cell.alignment = _align(h="center"); cell.border = _border()
            is_base = (abs(gv - g) < 0.001 and abs(pv - peg) < 0.001)
            if is_base:
                cell.fill = _fill(SENSITIVITY_GOLD); cell.font = _font(bold=True, size=9)
            elif implied >= price:
                cell.fill = _fill(GREEN_FILL); cell.font = _font(color=GREEN_TEXT, size=9)
            else:
                cell.fill = _fill(RED_FILL); cell.font = _font(color=RED_TEXT, size=9)
        row += 1

    return ws


# ── REVENUE MULTIPLE ──────────────────────────────────────────────────────────

def _analysis_revenue_multiple(wb, company_name, fin, result, params):
    ws = wb.create_sheet("Analysis")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {"A": 2, "B": 36, "C": 18, "D": 18, "E": 18, "F": 18})

    t = ws.cell(row=2, column=2, value=f"{company_name} — Revenue Multiple (EV/Revenue)")
    t.font = _font(bold=True, size=13)
    ws.merge_cells("B2:F2")

    rev   = _safe(fin.get("revenue", 1e11))
    mult  = _safe(params.get("ev_revenue_multiple", 5))
    g     = _safe(params.get("revenue_growth", fin.get("revenue_growth", 0.15)))
    debt  = _safe(fin.get("total_debt", 0))
    cash  = _safe(fin.get("cash", 0))
    shares = _safe(fin.get("shares", 1), 1)
    price = _safe(fin.get("price", 800))

    row = 4
    _section_header(ws, row, 2, "REVENUE MULTIPLE — YEAR-BY-YEAR REVENUE FORECAST", width_cols=5)
    row += 1

    _label(ws, row, 2, "Year", bold=True)
    _label(ws, row, 3, "Revenue (₹ Cr)", bold=True)
    _label(ws, row, 4, "Growth", bold=True)
    _label(ws, row, 5, "EV (₹ Cr)", bold=True)
    _label(ws, row, 6, "Implied Price (₹)", bold=True)
    row += 1

    rev_t = rev
    for yr in range(1, 6):
        g_yr = g * (1 - 0.05 * (yr - 1))
        rev_t = rev_t * (1 + g_yr)
        ev_t  = rev_t * mult
        iv_t  = (ev_t - debt + cash) / shares
        fill  = LIGHT_BLUE_FILL if yr % 2 == 0 else None
        _val(ws, row, 2, yr, fmt="0")
        _val(ws, row, 3, rev_t/1e7, fmt="#,##0", fill_hex=fill)
        _pct_cell(ws, row, 4, g_yr, fill_hex=fill)
        _val(ws, row, 5, ev_t/1e7, fmt="#,##0", fill_hex=fill)
        _val(ws, row, 6, iv_t, fmt="#,##0.00",
             fill_hex=GREEN_FILL if iv_t >= price else RED_FILL,
             color=GREEN_TEXT if iv_t >= price else RED_TEXT)
        row += 1

    row += 1
    _section_header(ws, row, 2, "BASE CASE VALUATION", width_cols=5)
    row += 1
    ev0  = rev * mult
    iv0  = (ev0 - debt + cash) / shares

    data = [
        ("Current Revenue (₹ Cr)",        rev/1e7,   "#,##0"),
        ("EV/Revenue Multiple Applied",   mult,      "0.0x"),
        ("Implied Enterprise Value (₹ Cr)", ev0/1e7, "#,##0"),
        ("(−) Net Debt (₹ Cr)",           (debt-cash)/1e7, "#,##0"),
        ("Implied Equity Value (₹ Cr)",   (ev0-debt+cash)/1e7, "#,##0"),
        ("Intrinsic Value Per Share (₹)", iv0,       "#,##0.00"),
        ("Current Market Price (₹)",      price,     "#,##0.00"),
        ("Upside / (Downside)",           (iv0/price-1) if price else 0, "0.0%"),
    ]
    for label, value, fmt in data:
        bold = "Intrinsic" in label
        _label(ws, row, 2, label, bold=bold)
        if "%" in fmt:
            _pct_cell(ws, row, 3, value, fill_hex=GREEN_FILL if bold else None)
        else:
            _val(ws, row, 3, value, fmt=fmt, fill_hex=GREEN_FILL if bold else None,
                 color=GREEN_TEXT if bold else BLACK)
        row += 1

    return ws


# ── NAV ───────────────────────────────────────────────────────────────────────

def _analysis_nav(wb, company_name, fin, result, params):
    ws = wb.create_sheet("Analysis")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {"A": 2, "B": 36, "C": 18, "D": 16, "E": 16, "F": 18})

    t = ws.cell(row=2, column=2, value=f"{company_name} — Net Asset Value (NAV)")
    t.font = _font(bold=True, size=13)
    ws.merge_cells("B2:F2")

    assets   = _safe(fin.get("total_assets", 1e12))
    liabs    = _safe(fin.get("total_liabilities", 5e11))
    shares   = _safe(fin.get("shares", 1), 1)
    haircut  = _safe(params.get("haircut_rate", 0.15))
    price    = _safe(fin.get("price", 800))

    row = 4
    _section_header(ws, row, 2, "ASSET BREAKDOWN & RECOVERY ANALYSIS", width_cols=5)
    row += 1

    headers = ["Asset Category", "Book Value (₹ Cr)", "Recovery Rate", "Net Value (₹ Cr)", "Notes"]
    for j, h in enumerate(headers):
        c = ws.cell(row=row, column=2+j, value=h)
        c.font = _font(bold=True, color=WHITE, size=9); c.fill = _fill(DARK_BLUE)
        c.alignment = _align(h="center"); c.border = _border()
    row += 1

    asset_classes = [
        ("Cash & Equivalents",        _safe(fin.get("cash", 0)),          1.00),
        ("Receivables & Short-term",  assets * 0.10,                       0.90),
        ("Inventory",                 assets * 0.08,                       0.75),
        ("Property, Plant & Equip.",  assets * 0.35,                       0.70),
        ("Intangibles",               assets * 0.12,                       0.30),
        ("Investments",               assets * 0.08,                       0.85),
        ("Other Assets",              assets * 0.07,                       0.60),
    ]
    total_book = 0.0
    total_net  = 0.0
    for label, bv, rr in asset_classes:
        net = bv * rr
        total_book += bv; total_net += net
        _label(ws, row, 2, label)
        _val(ws, row, 3, bv/1e7, fmt="#,##0")
        _pct_cell(ws, row, 4, rr, fill_hex=YELLOW_FILL)
        _val(ws, row, 5, net/1e7, fmt="#,##0")
        c = ws.cell(row=row, column=6, value="Market / book estimate")
        c.font = _font(size=9, color="808080"); c.border = _border()
        row += 1

    _label(ws, row, 2, "TOTAL ASSETS", bold=True)
    _val(ws, row, 3, total_book/1e7, fmt="#,##0", fill_hex=LIGHT_BLUE_FILL, bold=True)
    _val(ws, row, 5, total_net/1e7, fmt="#,##0", fill_hex=LIGHT_BLUE_FILL, bold=True)
    row += 2

    _section_header(ws, row, 2, "NAV CALCULATION", width_cols=5)
    row += 1
    nav = total_net - liabs
    nav_share = nav / shares

    data = [
        ("Total Asset Value (Recovery, ₹ Cr)", total_net/1e7, "#,##0"),
        ("(−) Total Liabilities (₹ Cr)",       liabs/1e7,     "#,##0"),
        ("Net Asset Value (₹ Cr)",             nav/1e7,       "#,##0"),
        ("Shares Outstanding (Cr)",            shares/1e7,    "#,##0.00"),
        ("NAV Per Share (₹)",                  nav_share,     "#,##0.00"),
        ("Current Market Price (₹)",           price,         "#,##0.00"),
        ("Premium / (Discount) to NAV",        (price/nav_share - 1) if nav_share else 0, "0.0%"),
    ]
    for label, value, fmt in data:
        bold = "NAV Per Share" in label
        _label(ws, row, 2, label, bold=bold)
        if "%" in fmt:
            _pct_cell(ws, row, 3, value, fill_hex=GREEN_FILL if bold else None)
        else:
            _val(ws, row, 3, value, fmt=fmt, fill_hex=GREEN_FILL if bold else None,
                 color=GREEN_TEXT if bold else BLACK)
        row += 1

    return ws


# ── LIQUIDATION ───────────────────────────────────────────────────────────────

def _analysis_liquidation(wb, company_name, fin, result, params):
    ws = wb.create_sheet("Analysis")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {"A": 2, "B": 36, "C": 18, "D": 18, "E": 18, "F": 18})

    t = ws.cell(row=2, column=2, value=f"{company_name} — Liquidation Value Analysis")
    t.font = _font(bold=True, size=13)
    ws.merge_cells("B2:F2")

    cash_rate = _safe(params.get("cash_rate", 1.0))
    ppe_rate  = _safe(params.get("ppe_rate", 0.5))
    assets    = _safe(fin.get("total_assets", 1e12))
    liabs     = _safe(fin.get("total_liabilities", 5e11))
    cash_val  = _safe(fin.get("cash", 0))
    shares    = _safe(fin.get("shares", 1), 1)
    price     = _safe(fin.get("price", 800))

    row = 4
    _section_header(ws, row, 2, "LIQUIDATION SCHEDULE", width_cols=5)
    row += 1

    headers = ["Asset", "Book Value (₹ Cr)", "Recovery Rate", "Liquidation Value (₹ Cr)"]
    for j, h in enumerate(headers):
        c = ws.cell(row=row, column=2+j, value=h)
        c.font = _font(bold=True, color=WHITE, size=9); c.fill = _fill(DARK_BLUE)
        c.alignment = _align(h="center"); c.border = _border()
    row += 1

    liq_items = [
        ("Cash & Near-Cash",         cash_val,       cash_rate),
        ("Accounts Receivable",      assets*0.10,    0.80),
        ("Inventory",                assets*0.08,    0.60),
        ("PP&E (plant & equipment)", assets*0.35,    ppe_rate),
        ("Other Assets",             assets*0.07,    0.30),
        ("Intangibles",              assets*0.12,    0.05),
    ]
    total_liq = 0.0
    for label, bv, rr in liq_items:
        lv = bv * rr
        total_liq += lv
        _label(ws, row, 2, label)
        _val(ws, row, 3, bv/1e7, fmt="#,##0")
        _pct_cell(ws, row, 4, rr, fill_hex=YELLOW_FILL)
        _val(ws, row, 5, lv/1e7, fmt="#,##0")
        row += 1

    _label(ws, row, 2, "TOTAL LIQUIDATION PROCEEDS", bold=True)
    _val(ws, row, 3, sum(b for _, b, _ in liq_items)/1e7, fmt="#,##0", fill_hex=LIGHT_BLUE_FILL)
    _val(ws, row, 5, total_liq/1e7, fmt="#,##0", fill_hex=LIGHT_BLUE_FILL, bold=True)
    row += 2

    _section_header(ws, row, 2, "EQUITY VALUE IN LIQUIDATION", width_cols=5)
    row += 1
    liq_cost = total_liq * 0.05
    eq_liq   = total_liq - liabs - liq_cost
    iv_share = eq_liq / shares

    data = [
        ("Total Liquidation Proceeds (₹ Cr)", total_liq/1e7, "#,##0"),
        ("(−) Total Liabilities (₹ Cr)",      liabs/1e7,     "#,##0"),
        ("(−) Winding-up Costs (5%, ₹ Cr)",   liq_cost/1e7,  "#,##0"),
        ("Equity in Liquidation (₹ Cr)",      eq_liq/1e7,    "#,##0"),
        ("Liquidation Value Per Share (₹)",   iv_share,      "#,##0.00"),
        ("Current Market Price (₹)",          price,         "#,##0.00"),
        ("Premium to Liquidation",            (price/iv_share-1) if iv_share else 0, "0.0%"),
    ]
    for label, value, fmt in data:
        bold = "Per Share" in label
        _label(ws, row, 2, label, bold=bold)
        if "%" in fmt:
            _pct_cell(ws, row, 3, value, fill_hex=GREEN_FILL if bold else None)
        else:
            _val(ws, row, 3, value, fmt=fmt, fill_hex=GREEN_FILL if bold else None,
                 color=GREEN_TEXT if bold else BLACK)
        row += 1

    return ws


# ── REPLACEMENT COST ──────────────────────────────────────────────────────────

def _analysis_replacement_cost(wb, company_name, fin, result, params):
    ws = wb.create_sheet("Analysis")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {"A": 2, "B": 36, "C": 18, "D": 18, "E": 18, "F": 18})

    t = ws.cell(row=2, column=2, value=f"{company_name} — Replacement Cost Valuation")
    t.font = _font(bold=True, size=13)
    ws.merge_cells("B2:F2")

    mult     = _safe(params.get("rebuild_multiplier", 1.15))
    dep_adj  = _safe(params.get("depreciation_adj", 0.20))
    assets   = _safe(fin.get("total_assets", 1e12))
    liabs    = _safe(fin.get("total_liabilities", 5e11))
    shares   = _safe(fin.get("shares", 1), 1)
    price    = _safe(fin.get("price", 800))

    row = 4
    _section_header(ws, row, 2, "ASSET REPLACEMENT COST ANALYSIS", width_cols=5)
    row += 1

    headers = ["Asset Category", "Book Value (₹ Cr)", "Rebuild Mult.", "Dep. Adj.", "Replacement Value (₹ Cr)"]
    for j, h in enumerate(headers):
        c = ws.cell(row=row, column=2+j, value=h)
        c.font = _font(bold=True, color=WHITE, size=9); c.fill = _fill(DARK_BLUE)
        c.alignment = _align(h="center"); c.border = _border()
    row += 1

    items = [
        ("PP&E — Land",            assets*0.10, 1.30, 0.0),
        ("PP&E — Buildings",       assets*0.15, mult, dep_adj),
        ("PP&E — Machinery",       assets*0.20, mult, dep_adj * 1.5),
        ("Technology Infrastructure", assets*0.08, 1.40, dep_adj * 0.5),
        ("Inventory",              assets*0.08, 1.05, 0.0),
        ("Intangibles (brand)",    assets*0.12, 1.50, 0.0),
    ]
    total_rep = 0.0
    for label, bv, bm, da in items:
        rv = bv * bm * (1 - da)
        total_rep += rv
        _label(ws, row, 2, label)
        _val(ws, row, 3, bv/1e7, fmt="#,##0")
        _val(ws, row, 4, bm, fmt="0.00x")
        _pct_cell(ws, row, 5, da, fill_hex=YELLOW_FILL)
        _val(ws, row, 6, rv/1e7, fmt="#,##0")
        row += 1

    _label(ws, row, 2, "TOTAL REPLACEMENT COST", bold=True)
    _val(ws, row, 6, total_rep/1e7, fmt="#,##0", fill_hex=LIGHT_BLUE_FILL, bold=True)
    row += 2

    _section_header(ws, row, 2, "REPLACEMENT COST EQUITY VALUE", width_cols=5)
    row += 1
    eq_rep   = total_rep - liabs
    iv_share = eq_rep / shares

    data = [
        ("Total Replacement Cost (₹ Cr)", total_rep/1e7, "#,##0"),
        ("(−) Total Liabilities (₹ Cr)", liabs/1e7,     "#,##0"),
        ("Replacement Cost Equity (₹ Cr)", eq_rep/1e7,  "#,##0"),
        ("Intrinsic Value Per Share (₹)",  iv_share,     "#,##0.00"),
        ("Current Market Price (₹)",       price,        "#,##0.00"),
        ("Upside / (Downside)",            (iv_share/price - 1) if price else 0, "0.0%"),
    ]
    for label, value, fmt in data:
        bold = "Intrinsic" in label
        _label(ws, row, 2, label, bold=bold)
        if "%" in fmt:
            _pct_cell(ws, row, 3, value, fill_hex=GREEN_FILL if bold else None)
        else:
            _val(ws, row, 3, value, fmt=fmt, fill_hex=GREEN_FILL if bold else None,
                 color=GREEN_TEXT if bold else BLACK)
        row += 1

    return ws


# ── CAPITALIZED EARNINGS ──────────────────────────────────────────────────────

def _analysis_capitalized_earnings(wb, company_name, fin, result, params):
    ws = wb.create_sheet("Analysis")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {"A": 2, "B": 36, "C": 18, "D": 18, "E": 18})

    t = ws.cell(row=2, column=2, value=f"{company_name} — Capitalized Earnings Model")
    t.font = _font(bold=True, size=13)
    ws.merge_cells("B2:E2")

    req    = _safe(params.get("required_return", 0.12))
    g      = _safe(params.get("eps_growth", fin.get("earnings_growth", 0.10)))
    eps    = _safe(fin.get("eps", 50))
    ni     = _safe(fin.get("net_income", 1e10))
    shares = _safe(fin.get("shares", 1), 1)
    price  = _safe(fin.get("price", 800))

    row = 4
    _section_header(ws, row, 2, "EPS NORMALIZATION & PROJECTION (5 YEARS)", width_cols=4)
    row += 1

    headers = ["Year", "EPS (₹)", "Normalized EPS (₹)", "Cap. Rate", "Capitalized Value (₹)"]
    for j, h in enumerate(headers):
        c = ws.cell(row=row, column=2+j, value=h)
        c.font = _font(bold=True, color=WHITE, size=9); c.fill = _fill(DARK_BLUE)
        c.alignment = _align(h="center"); c.border = _border()
    row += 1

    for yr in range(1, 6):
        eps_y = eps * ((1 + g) ** yr)
        norm  = eps_y * 0.95
        cap_r = req - g * 0.3
        cap_v = norm / cap_r if cap_r > 0 else 0
        fill  = LIGHT_BLUE_FILL if yr % 2 == 0 else None
        _val(ws, row, 2, yr, fmt="0")
        _val(ws, row, 3, eps_y, fmt="#,##0.00", fill_hex=fill)
        _val(ws, row, 4, norm,  fmt="#,##0.00", fill_hex=fill)
        _pct_cell(ws, row, 5, cap_r, fill_hex=fill)
        _val(ws, row, 6, cap_v, fmt="#,##0.00",
             fill_hex=GREEN_FILL if cap_v >= price else None)
        row += 1

    row += 1
    _section_header(ws, row, 2, "BASE CASE CAPITALIZATION", width_cols=4)
    row += 1
    cap_rate = req - g
    iv = eps * (1 + g) / cap_rate if cap_rate > 0 else 0

    data = [
        ("EPS (TTM, ₹)",               eps,     "#,##0.00"),
        ("EPS Growth Rate (g)",         g,       "0.0%"),
        ("Required Return (r)",         req,     "0.0%"),
        ("Capitalization Rate (r − g)", cap_rate,"0.0%"),
        ("Intrinsic Value = EPS × (1+g) / (r−g)", iv, "#,##0.00"),
        ("Current Market Price (₹)",    price,   "#,##0.00"),
        ("Upside / (Downside)",         (iv/price-1) if price else 0, "0.0%"),
    ]
    for label, value, fmt in data:
        bold = "Intrinsic" in label
        _label(ws, row, 2, label, bold=bold)
        if "%" in fmt:
            _pct_cell(ws, row, 3, value, fill_hex=GREEN_FILL if bold else None)
        else:
            _val(ws, row, 3, value, fmt=fmt, fill_hex=GREEN_FILL if bold else None,
                 color=GREEN_TEXT if bold else BLACK)
        row += 1

    return ws


# ── EXCESS EARNINGS ───────────────────────────────────────────────────────────

def _analysis_excess_earnings(wb, company_name, fin, result, params):
    ws = wb.create_sheet("Analysis")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {"A": 2, "B": 36, "C": 18, "D": 18, "E": 18, "F": 18})

    t = ws.cell(row=2, column=2, value=f"{company_name} — Excess Earnings Method")
    t.font = _font(bold=True, size=13)
    ws.merge_cells("B2:F2")

    fair_return = _safe(params.get("fair_return_rate", 0.10))
    disc_rate   = _safe(params.get("discount_rate", 0.12))
    ni          = _safe(fin.get("net_income", 1e10))
    assets      = _safe(fin.get("total_assets", 1e12))
    liabs       = _safe(fin.get("total_liabilities", 5e11))
    shares      = _safe(fin.get("shares", 1), 1)
    price       = _safe(fin.get("price", 800))

    tangible_assets = (assets - liabs) * 0.70
    normal_return   = tangible_assets * fair_return
    excess          = ni - normal_return
    goodwill        = excess / disc_rate if disc_rate > 0 else 0
    iv              = (tangible_assets + goodwill) / shares

    row = 4
    _section_header(ws, row, 2, "EXCESS EARNINGS — STEP BY STEP", width_cols=5)
    row += 1

    data = [
        ("Total Assets (₹ Cr)",                     assets/1e7,          "#,##0"),
        ("Total Liabilities (₹ Cr)",                liabs/1e7,           "#,##0"),
        ("Net Tangible Assets (~70%, ₹ Cr)",        tangible_assets/1e7, "#,##0"),
        ("Fair Return Rate on Tangible Assets",     fair_return,         "0.0%"),
        ("Normal Return = TA × Fair Return (₹ Cr)", normal_return/1e7,  "#,##0"),
        ("Net Income (TTM, ₹ Cr)",                  ni/1e7,             "#,##0"),
        ("Excess Earnings = NI − Normal (₹ Cr)",    excess/1e7,         "#,##0"),
        ("Capitalisation Rate (discount rate)",     disc_rate,           "0.0%"),
        ("Value of Goodwill = Excess / r (₹ Cr)",   goodwill/1e7,       "#,##0"),
        ("Total Value = NTA + Goodwill (₹ Cr)",     (tangible_assets+goodwill)/1e7, "#,##0"),
        ("Intrinsic Value Per Share (₹)",            iv,                 "#,##0.00"),
        ("Current Market Price (₹)",                price,              "#,##0.00"),
        ("Upside / (Downside)",                     (iv/price-1) if price else 0, "0.0%"),
    ]
    for label, value, fmt in data:
        bold = "Intrinsic" in label
        _label(ws, row, 2, label, bold=bold)
        if "%" in fmt:
            _pct_cell(ws, row, 3, value, fill_hex=GREEN_FILL if bold else None)
        else:
            _val(ws, row, 3, value, fmt=fmt, fill_hex=GREEN_FILL if bold else None,
                 color=GREEN_TEXT if bold else BLACK)
        row += 1

    row += 2
    _section_header(ws, row, 2, "YEAR-BY-YEAR EXCESS EARNINGS (5 YEARS)", width_cols=5)
    row += 1

    headers = ["Year", "Net Income (₹ Cr)", "Normal Return (₹ Cr)", "Excess Earnings (₹ Cr)", "PV (₹ Cr)"]
    for j, h in enumerate(headers):
        c = ws.cell(row=row, column=2+j, value=h)
        c.font = _font(bold=True, color=WHITE, size=9); c.fill = _fill(DARK_BLUE)
        c.alignment = _align(h="center"); c.border = _border()
    row += 1

    g = _safe(fin.get("earnings_growth", 0.12))
    ni_t = ni
    for yr in range(1, 6):
        ni_t  = ni_t * (1 + g)
        nr_t  = tangible_assets * fair_return * ((1 + 0.03) ** yr)
        ex_t  = ni_t - nr_t
        pv_t  = ex_t / ((1 + disc_rate) ** yr)
        fill  = LIGHT_BLUE_FILL if yr % 2 == 0 else None
        _val(ws, row, 2, yr, fmt="0")
        _val(ws, row, 3, ni_t/1e7,  fmt="#,##0", fill_hex=fill)
        _val(ws, row, 4, nr_t/1e7,  fmt="#,##0", fill_hex=fill)
        _val(ws, row, 5, ex_t/1e7,  fmt="#,##0",
             fill_hex=GREEN_FILL if ex_t >= 0 else RED_FILL)
        _val(ws, row, 6, pv_t/1e7,  fmt="#,##0", fill_hex=fill)
        row += 1

    return ws


# ── EVA ───────────────────────────────────────────────────────────────────────

def _analysis_eva(wb, company_name, fin, result, params):
    ws = wb.create_sheet("Analysis")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {"A": 2, "B": 36, "C": 14, "D": 14, "E": 14,
                         "F": 14, "G": 14, "H": 14})

    t = ws.cell(row=2, column=2, value=f"{company_name} — Economic Value Added (EVA)")
    t.font = _font(bold=True, size=13)
    ws.merge_cells("B2:H2")

    wacc    = _safe(params.get("wacc", 0.12))
    g1      = _safe(params.get("growth_stage1", 0.15))
    gT      = _safe(params.get("terminal_growth", 0.04))
    nopat   = _safe(fin.get("nopat", fin.get("net_income", 1e10) * 0.8))
    ic      = _safe(fin.get("invested_capital", fin.get("total_assets", 1e12) * 0.6))
    shares  = _safe(fin.get("shares", 1), 1)
    debt    = _safe(fin.get("total_debt", 0))
    cash    = _safe(fin.get("cash", 0))
    price   = _safe(fin.get("price", 800))

    row = 4
    _section_header(ws, row, 2, "EVA SCHEDULE (10 YEARS)", width_cols=7)
    row += 1

    headers = ["Year", "NOPAT (₹ Cr)", "Inv. Capital (₹ Cr)", "Capital Charge", "EVA (₹ Cr)", "Disc. Factor", "PV EVA (₹ Cr)"]
    for j, h in enumerate(headers):
        c = ws.cell(row=row, column=2+j, value=h)
        c.font = _font(bold=True, color=WHITE, size=9); c.fill = _fill(DARK_BLUE)
        c.alignment = _align(h="center"); c.border = _border()
    row += 1

    pv_sum = 0.0
    ic_t   = ic
    nopat_t = nopat
    last_eva = 0.0
    for yr in range(1, 11):
        g = g1 if yr <= 5 else (g1 + gT) / 2
        nopat_t = nopat_t * (1 + g)
        ic_t    = ic_t * (1 + g * 0.8)
        cap_chg = ic_t * wacc
        eva     = nopat_t - cap_chg
        df      = 1 / ((1 + wacc) ** yr)
        pv      = eva * df
        pv_sum += pv
        last_eva = eva
        fill = LIGHT_BLUE_FILL if yr % 2 == 0 else None
        _val(ws, row, 2, yr, fmt="0")
        _val(ws, row, 3, nopat_t/1e7,  fmt="#,##0", fill_hex=fill)
        _val(ws, row, 4, ic_t/1e7,     fmt="#,##0", fill_hex=fill)
        _val(ws, row, 5, cap_chg/1e7,  fmt="#,##0", fill_hex=fill)
        _val(ws, row, 6, eva/1e7,      fmt="#,##0",
             fill_hex=GREEN_FILL if eva >= 0 else RED_FILL)
        _val(ws, row, 7, df,           fmt="0.0000")
        _val(ws, row, 8, pv/1e7,       fmt="#,##0", fill_hex=fill)
        row += 1

    row += 1
    _section_header(ws, row, 2, "EVA — EQUITY BRIDGE", width_cols=7)
    row += 1
    tv_eva  = last_eva * (1 + gT) / (wacc - gT) if (wacc - gT) > 0 else 0
    pv_tv   = tv_eva / ((1 + wacc) ** 10)
    eq_val  = ic + pv_sum + pv_tv - debt + cash
    iv      = eq_val / shares

    bridge = [
        ("Invested Capital (₹ Cr)",    ic/1e7,    "#,##0"),
        ("PV of EVA (10 Yrs, ₹ Cr)",  pv_sum/1e7,"#,##0"),
        ("Terminal EVA Value (₹ Cr)", tv_eva/1e7, "#,##0"),
        ("PV of Terminal EVA (₹ Cr)", pv_tv/1e7,  "#,##0"),
        ("(−) Net Debt (₹ Cr)",       (debt-cash)/1e7, "#,##0"),
        ("Equity Value (₹ Cr)",       eq_val/1e7, "#,##0"),
        ("Intrinsic Value Per Share (₹)", iv,     "#,##0.00"),
    ]
    for label, value, fmt in bridge:
        bold = "Intrinsic" in label
        _label(ws, row, 2, label, bold=bold)
        _val(ws, row, 3, value, fmt=fmt,
             fill_hex=GREEN_FILL if bold else (LIGHT_BLUE_FILL if "Equity Value" in label else None),
             color=GREEN_TEXT if bold else BLACK)
        row += 1

    return ws


# ── CFROI ─────────────────────────────────────────────────────────────────────

def _analysis_cfroi(wb, company_name, fin, result, params):
    ws = wb.create_sheet("Analysis")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {"A": 2, "B": 36, "C": 18, "D": 16, "E": 16, "F": 16, "G": 16})

    t = ws.cell(row=2, column=2, value=f"{company_name} — Cash Flow Return on Investment (CFROI)")
    t.font = _font(bold=True, size=13)
    ws.merge_cells("B2:G2")

    req     = _safe(params.get("required_return", 0.12))
    life    = _safe(params.get("asset_life", 10))
    ic      = _safe(fin.get("invested_capital", fin.get("total_assets", 1e12) * 0.6))
    ocf     = _safe(fin.get("operating_cf", fin.get("nopat", 1e10)))
    shares  = _safe(fin.get("shares", 1), 1)
    debt    = _safe(fin.get("total_debt", 0))
    cash    = _safe(fin.get("cash", 0))
    price   = _safe(fin.get("price", 800))

    # CFROI = IRR of investing IC and receiving OCF for asset life + residual
    life = max(1, int(life))
    cfroi = ocf / ic if ic > 0 else req

    row = 4
    _section_header(ws, row, 2, "CFROI CALCULATION — ASSET LIFE SCHEDULE", width_cols=6)
    row += 1

    headers = ["Year", "Gross Invest. (₹ Cr)", "Gross CF (₹ Cr)", "Disc. Factor", "PV CF (₹ Cr)", "CFROI vs CoC"]
    for j, h in enumerate(headers):
        c = ws.cell(row=row, column=2+j, value=h)
        c.font = _font(bold=True, color=WHITE, size=9); c.fill = _fill(DARK_BLUE)
        c.alignment = _align(h="center"); c.border = _border()
    row += 1

    for yr in range(1, life + 1):
        gcf  = ocf * (1 + 0.03) ** yr
        df   = 1 / ((1 + req) ** yr)
        pv   = gcf * df
        spread = cfroi - req
        fill = LIGHT_BLUE_FILL if yr % 2 == 0 else None
        _val(ws, row, 2, yr, fmt="0")
        _val(ws, row, 3, ic/1e7, fmt="#,##0", fill_hex=fill)
        _val(ws, row, 4, gcf/1e7, fmt="#,##0", fill_hex=fill)
        _val(ws, row, 5, df, fmt="0.0000")
        _val(ws, row, 6, pv/1e7, fmt="#,##0", fill_hex=fill)
        _pct_cell(ws, row, 7, spread,
                  fill_hex=GREEN_FILL if spread >= 0 else RED_FILL,
                  color=GREEN_TEXT if spread >= 0 else RED_TEXT)
        row += 1

    row += 1
    _section_header(ws, row, 2, "CFROI VALUATION SUMMARY", width_cols=6)
    row += 1
    ev   = ic * (cfroi / req) if req > 0 else ic
    iv   = (ev - debt + cash) / shares

    data = [
        ("Invested Capital (₹ Cr)",     ic/1e7,  "#,##0"),
        ("Operating Cash Flow (₹ Cr)",  ocf/1e7, "#,##0"),
        ("CFROI (approx.)",             cfroi,   "0.0%"),
        ("Required Return (CoC)",       req,     "0.0%"),
        ("CFROI Spread (CFROI − CoC)",  cfroi - req, "0.0%"),
        ("EV = IC × (CFROI / CoC) (₹ Cr)", ev/1e7, "#,##0"),
        ("(−) Net Debt (₹ Cr)",         (debt-cash)/1e7, "#,##0"),
        ("Intrinsic Value Per Share (₹)", iv,    "#,##0.00"),
        ("Current Market Price (₹)",    price,   "#,##0.00"),
    ]
    for label, value, fmt in data:
        bold = "Intrinsic" in label
        _label(ws, row, 2, label, bold=bold)
        if "%" in fmt:
            _pct_cell(ws, row, 3, value, fill_hex=GREEN_FILL if bold else None)
        else:
            _val(ws, row, 3, value, fmt=fmt, fill_hex=GREEN_FILL if bold else None,
                 color=GREEN_TEXT if bold else BLACK)
        row += 1

    return ws


# ── LBO ───────────────────────────────────────────────────────────────────────

def _analysis_lbo(wb, company_name, fin, result, params):
    ws = wb.create_sheet("Analysis")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {"A": 2, "B": 36, "C": 14, "D": 14, "E": 14,
                         "F": 14, "G": 14, "H": 14})

    t = ws.cell(row=2, column=2, value=f"{company_name} — Leveraged Buyout (LBO) Analysis")
    t.font = _font(bold=True, size=13)
    ws.merge_cells("B2:H2")

    ebitda      = _safe(params.get("ebitda", fin.get("ebitda", 1e11)))
    entry_mult  = _safe(params.get("entry_multiple", 8))
    exit_mult   = _safe(params.get("exit_multiple", 10))
    debt_ratio  = _safe(params.get("debt_ratio", 0.60))
    hold_years  = int(_safe(params.get("hold_years", 5)))
    shares      = _safe(params.get("shares", fin.get("shares", 1)), 1)
    ebitda_g    = _safe(fin.get("revenue_growth", 0.08))
    int_rate    = _safe(params.get("interest_rate", 0.08))

    entry_ev    = ebitda * entry_mult
    debt_amt    = entry_ev * debt_ratio
    equity_in   = entry_ev * (1 - debt_ratio)
    price       = _safe(fin.get("price", 800))

    row = 4
    _section_header(ws, row, 2, "TRANSACTION STRUCTURE — ENTRY", width_cols=7)
    row += 1

    entry_data = [
        ("EBITDA (Entry Year, ₹ Cr)",    ebitda/1e7,    "#,##0"),
        ("Entry EV/EBITDA Multiple",     entry_mult,    "0.0x"),
        ("Entry Enterprise Value (₹ Cr)", entry_ev/1e7, "#,##0"),
        ("Debt Financing (₹ Cr)",        debt_amt/1e7,  "#,##0"),
        ("Debt / EV Ratio",              debt_ratio,    "0.0%"),
        ("Equity Contribution (₹ Cr)",   equity_in/1e7, "#,##0"),
        ("Implied Purchase Price/Share (₹)", equity_in/shares, "#,##0.00"),
    ]
    for label, value, fmt in entry_data:
        _label(ws, row, 2, label)
        if "%" in fmt:
            _pct_cell(ws, row, 3, value, fill_hex=YELLOW_FILL)
        else:
            _val(ws, row, 3, value, fmt=fmt, color=BLUE_TEXT, fill_hex=YELLOW_FILL)
        row += 1

    row += 1
    _section_header(ws, row, 2, f"DEBT AMORTIZATION SCHEDULE ({hold_years} YEARS)", width_cols=7)
    row += 1
    headers = ["Year", "EBITDA (₹ Cr)", "Interest (₹ Cr)", "Principal Pay (₹ Cr)", "Debt O/S (₹ Cr)", "FCF (₹ Cr)", "Equity Build (₹ Cr)"]
    for j, h in enumerate(headers):
        c = ws.cell(row=row, column=2+j, value=h)
        c.font = _font(bold=True, color=WHITE, size=9); c.fill = _fill(DARK_BLUE)
        c.alignment = _align(h="center"); c.border = _border()
    row += 1

    debt_t    = debt_amt
    ebitda_t  = ebitda
    eq_builds = []
    for yr in range(1, hold_years + 1):
        ebitda_t = ebitda_t * (1 + ebitda_g)
        interest = debt_t * int_rate
        fcf      = ebitda_t * 0.6 - interest
        principal = min(fcf * 0.7, debt_t)
        debt_t   = max(0, debt_t - principal)
        eq_b     = entry_ev * ((1 + ebitda_g) ** yr)
        eq_builds.append((yr, ebitda_t, interest, principal, debt_t, fcf, eq_b))
        fill = LIGHT_BLUE_FILL if yr % 2 == 0 else None
        _val(ws, row, 2, yr, fmt="0")
        _val(ws, row, 3, ebitda_t/1e7, fmt="#,##0", fill_hex=fill)
        _val(ws, row, 4, interest/1e7, fmt="#,##0", fill_hex=fill)
        _val(ws, row, 5, principal/1e7, fmt="#,##0", fill_hex=fill)
        _val(ws, row, 6, debt_t/1e7,   fmt="#,##0", fill_hex=fill)
        _val(ws, row, 7, fcf/1e7,      fmt="#,##0",
             fill_hex=GREEN_FILL if fcf >= 0 else RED_FILL)
        _val(ws, row, 8, (eq_b-debt_t)/1e7, fmt="#,##0", fill_hex=fill)
        row += 1

    row += 1
    _section_header(ws, row, 2, "EXIT WATERFALL & RETURNS", width_cols=7)
    row += 1
    final_ebitda = eq_builds[-1][1]
    exit_ev      = final_ebitda * exit_mult
    exit_debt    = eq_builds[-1][4]
    equity_out   = exit_ev - exit_debt
    moic         = equity_out / equity_in if equity_in else 1
    irr          = (moic ** (1 / hold_years) - 1) if moic > 0 else 0
    iv_share     = equity_out / shares

    exit_data = [
        ("Exit EBITDA (₹ Cr)",           final_ebitda/1e7, "#,##0"),
        ("Exit EV/EBITDA Multiple",       exit_mult,        "0.0x"),
        ("Exit Enterprise Value (₹ Cr)", exit_ev/1e7,       "#,##0"),
        ("(−) Remaining Debt (₹ Cr)",    exit_debt/1e7,     "#,##0"),
        ("Equity Proceeds (₹ Cr)",       equity_out/1e7,    "#,##0"),
        ("Equity Invested (₹ Cr)",       equity_in/1e7,     "#,##0"),
        ("Money-on-Money (MoIC)",        moic,              "0.0x"),
        ("IRR (annualized)",             irr,               "0.0%"),
        ("Implied Price Per Share (₹)",  iv_share,          "#,##0.00"),
        ("Current Market Price (₹)",     price,             "#,##0.00"),
    ]
    for label, value, fmt in exit_data:
        bold = "Implied Price" in label or "IRR" in label or "MoIC" in label
        _label(ws, row, 2, label, bold=bold)
        if "%" in fmt:
            _pct_cell(ws, row, 3, value, fill_hex=GREEN_FILL if bold else None,
                      color=GREEN_TEXT if bold else BLACK)
        else:
            _val(ws, row, 3, value, fmt=fmt,
                 fill_hex=GREEN_FILL if bold else None,
                 color=GREEN_TEXT if bold else BLACK)
        row += 1

    return ws


# ── BLACK-SCHOLES ─────────────────────────────────────────────────────────────

def _norm_cdf(x):
    """Approximation of N(x)."""
    return 0.5 * (1 + math.erf(x / math.sqrt(2)))

def _bs_price(S, K, r, T, sigma, option_type="call"):
    if T <= 0 or sigma <= 0:
        return max(S - K, 0) if option_type == "call" else max(K - S, 0)
    d1 = (math.log(S / K) + (r + 0.5 * sigma**2) * T) / (sigma * math.sqrt(T))
    d2 = d1 - sigma * math.sqrt(T)
    if option_type == "call":
        return S * _norm_cdf(d1) - K * math.exp(-r * T) * _norm_cdf(d2)
    else:
        return K * math.exp(-r * T) * _norm_cdf(-d2) - S * _norm_cdf(-d1)

def _analysis_black_scholes(wb, company_name, fin, result, params):
    ws = wb.create_sheet("Analysis")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {"A": 2, "B": 36, "C": 18, "D": 18, "E": 18, "F": 18, "G": 18})

    t = ws.cell(row=2, column=2, value=f"{company_name} — Black-Scholes Option Pricing")
    t.font = _font(bold=True, size=13)
    ws.merge_cells("B2:G2")

    S      = _safe(fin.get("price", 800))
    K      = _safe(params.get("strike", S * 1.05))
    sigma  = _safe(params.get("volatility", fin.get("volatility_annual", 0.30)))
    r      = _safe(params.get("risk_free", 0.071))
    T      = _safe(params.get("time_years", 1.0))
    opt    = params.get("option_type", "call")

    d1 = (math.log(S/K) + (r + 0.5*sigma**2)*T) / (sigma*math.sqrt(T)) if (K > 0 and T > 0 and sigma > 0) else 0
    d2 = d1 - sigma * math.sqrt(T) if T > 0 else 0
    call = _bs_price(S, K, r, T, sigma, "call")
    put  = _bs_price(S, K, r, T, sigma, "put")
    delta_call = _norm_cdf(d1)
    delta_put  = delta_call - 1
    gamma = math.exp(-0.5 * d1**2) / (math.sqrt(2*math.pi) * S * sigma * math.sqrt(T)) if T > 0 else 0
    theta = -(S * math.exp(-0.5 * d1**2) * sigma / (2 * math.sqrt(T)) +
              r * K * math.exp(-r * T) * _norm_cdf(d2)) / 365 if T > 0 else 0
    vega  = S * math.exp(-0.5 * d1**2) * math.sqrt(T) / math.sqrt(2 * math.pi) / 100

    row = 4
    _section_header(ws, row, 2, "BLACK-SCHOLES — INPUT PARAMETERS & GREEKS", width_cols=6)
    row += 1

    data = [
        ("Underlying Price (S)",          S,      "#,##0.00"),
        ("Strike Price (K)",              K,      "#,##0.00"),
        ("Risk-Free Rate (r)",            r,      "0.0%"),
        ("Time to Expiry (T, years)",     T,      "0.00"),
        ("Annualised Volatility (σ)",     sigma,  "0.0%"),
        ("d1",                            d1,     "0.0000"),
        ("d2",                            d2,     "0.0000"),
        ("N(d1)",                         _norm_cdf(d1), "0.0000"),
        ("N(d2)",                         _norm_cdf(d2), "0.0000"),
    ]
    row += 0
    for label, value, fmt in data:
        _label(ws, row, 2, label)
        if "%" in fmt:
            _pct_cell(ws, row, 3, value)
        else:
            _val(ws, row, 3, value, fmt=fmt)
        row += 1

    row += 1
    _section_header(ws, row, 2, "OPTION PRICES & GREEKS", width_cols=6)
    row += 1

    greeks = [
        ("Call Option Price (₹)",   call,       "#,##0.00",  GREEN_FILL),
        ("Put Option Price (₹)",    put,        "#,##0.00",  LIGHT_BLUE_FILL),
        ("Delta (Call)",            delta_call, "0.0000",    None),
        ("Delta (Put)",             delta_put,  "0.0000",    None),
        ("Gamma",                   gamma,      "0.000000",  None),
        ("Theta (per day)",         theta,      "0.0000",    None),
        ("Vega (per 1% σ change)",  vega,       "0.0000",    None),
    ]
    for label, value, fmt, fill in greeks:
        bold = "Call Option" in label or "Put Option" in label
        _label(ws, row, 2, label, bold=bold)
        _val(ws, row, 3, value, fmt=fmt, fill_hex=fill,
             color=GREEN_TEXT if fill == GREEN_FILL else BLACK, bold=bold)
        row += 1

    row += 2
    _section_header(ws, row, 2, "OPTION PRICE MATRIX — Spot Price × Volatility", width_cols=6)
    row += 1

    spots  = [S * (0.8 + 0.1*i) for i in range(5)]
    sigmas = [sigma * (0.7 + 0.15*i) for i in range(5)]

    lbl_c = ws.cell(row=row, column=2, value="S \\ σ")
    lbl_c.font = _font(bold=True, color=WHITE, size=9); lbl_c.fill = _fill(NAVY); lbl_c.border = _border()
    for j, sv in enumerate(sigmas):
        c = ws.cell(row=row, column=3+j, value=round(sv, 4))
        c.font = _font(bold=True, color=WHITE, size=9); c.fill = _fill(DARK_BLUE)
        c.number_format = "0.0%"; c.alignment = _align(h="center"); c.border = _border()
    row += 1

    for i, sp in enumerate(spots):
        rc = ws.cell(row=row, column=2, value=round(sp, 2))
        rc.font = _font(bold=True, size=9); rc.fill = _fill(LIGHT_BLUE_FILL)
        rc.number_format = "#,##0.00"; rc.alignment = _align(h="right"); rc.border = _border()
        for j, sv in enumerate(sigmas):
            val = _bs_price(sp, K, r, T, sv, opt)
            cell = ws.cell(row=row, column=3+j, value=val)
            cell.number_format = "#,##0.00"; cell.alignment = _align(h="center"); cell.border = _border()
            is_base = (i == 2 and j == 2)
            if is_base:
                cell.fill = _fill(SENSITIVITY_GOLD); cell.font = _font(bold=True, size=9)
            else:
                cell.font = _font(size=9)
        row += 1

    return ws


# ── REAL OPTIONS ──────────────────────────────────────────────────────────────

def _analysis_real_options(wb, company_name, fin, result, params):
    ws = wb.create_sheet("Analysis")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {"A": 2, "B": 36, "C": 18, "D": 18, "E": 18, "F": 18, "G": 18})

    t = ws.cell(row=2, column=2, value=f"{company_name} — Real Options Valuation")
    t.font = _font(bold=True, size=13)
    ws.merge_cells("B2:G2")

    S     = _safe(params.get("asset_value", fin.get("total_assets", 1e12) * 0.6 / _safe(fin.get("shares", 1), 1)))
    K     = _safe(params.get("strike", S * 1.0))
    sigma = _safe(params.get("volatility", fin.get("volatility_annual", 0.35)))
    r     = _safe(params.get("risk_free", 0.071))
    T     = _safe(params.get("time_years", 2.0))
    price = _safe(fin.get("price", 800))

    call_val = _bs_price(S, K, r, T, sigma, "call")

    row = 4
    _section_header(ws, row, 2, "REAL OPTIONS — BLACK-SCHOLES APPLICATION", width_cols=6)
    row += 1

    data = [
        ("Underlying Asset Value (PV of CFs)",     S,     "#,##0.00"),
        ("Investment Cost (Strike, ₹)",            K,     "#,##0.00"),
        ("Risk-Free Rate",                         r,     "0.0%"),
        ("Project Volatility (σ)",                 sigma, "0.0%"),
        ("Time to Decision (years)",               T,     "0.00"),
        ("Real Option Value (Call, ₹)",            call_val, "#,##0.00"),
        ("Moneyness",                              "In the Money" if S > K else "Out of the Money", None),
        ("NPV without flexibility",                S - K, "#,##0.00"),
        ("Value of Flexibility",                   max(call_val - max(S - K, 0), 0), "#,##0.00"),
    ]
    for label, value, fmt in data:
        bold = "Real Option Value" in label
        _label(ws, row, 2, label, bold=bold)
        if fmt is None:
            c = ws.cell(row=row, column=3, value=value)
            c.font = _font(size=9, bold=bold); c.alignment = _align(); c.border = _border()
        elif "%" in fmt:
            _pct_cell(ws, row, 3, value)
        else:
            _val(ws, row, 3, value, fmt=fmt,
                 fill_hex=GREEN_FILL if bold else None,
                 color=GREEN_TEXT if bold else BLACK)
        row += 1

    row += 2
    _section_header(ws, row, 2, "BINOMIAL TREE (5-STEP) APPROXIMATION", width_cols=6)
    row += 1

    steps = 5
    dt = T / steps
    u  = math.exp(sigma * math.sqrt(dt))
    d  = 1 / u
    p  = (math.exp(r * dt) - d) / (u - d) if (u - d) != 0 else 0.5

    data2 = [
        ("Steps",             steps,  "0"),
        ("dt (years/step)",   dt,     "0.000"),
        ("Up Factor (u)",     u,      "0.0000"),
        ("Down Factor (d)",   d,      "0.0000"),
        ("Risk-Neutral Prob", p,      "0.0%"),
    ]
    for label, value, fmt in data2:
        _label(ws, row, 2, label)
        if "%" in fmt:
            _pct_cell(ws, row, 3, value)
        else:
            _val(ws, row, 3, value, fmt=fmt)
        row += 1

    row += 2
    _section_header(ws, row, 2, "SCENARIO TABLE — Asset Value × Volatility", width_cols=6)
    row += 1

    asset_vals = [S * (0.7 + 0.15*i) for i in range(5)]
    vols       = [sigma * (0.7 + 0.15*i) for i in range(5)]

    lbl_c = ws.cell(row=row, column=2, value="Asset \\ σ")
    lbl_c.font = _font(bold=True, color=WHITE, size=9); lbl_c.fill = _fill(NAVY); lbl_c.border = _border()
    for j, sv in enumerate(vols):
        c = ws.cell(row=row, column=3+j, value=round(sv, 4))
        c.font = _font(bold=True, color=WHITE, size=9); c.fill = _fill(DARK_BLUE)
        c.number_format = "0.0%"; c.alignment = _align(h="center"); c.border = _border()
    row += 1

    for i, av in enumerate(asset_vals):
        rc = ws.cell(row=row, column=2, value=round(av, 2))
        rc.font = _font(bold=True, size=9); rc.fill = _fill(LIGHT_BLUE_FILL)
        rc.number_format = "#,##0.00"; rc.alignment = _align(h="right"); rc.border = _border()
        for j, sv in enumerate(vols):
            val = _bs_price(av, K, r, T, sv, "call")
            cell = ws.cell(row=row, column=3+j, value=val)
            cell.number_format = "#,##0.00"; cell.alignment = _align(h="center"); cell.border = _border()
            if i == 2 and j == 2:
                cell.fill = _fill(SENSITIVITY_GOLD); cell.font = _font(bold=True, size=9)
            else:
                cell.font = _font(size=9)
        row += 1

    return ws


# ── SUM OF PARTS ──────────────────────────────────────────────────────────────

def _analysis_sum_of_parts(wb, company_name, fin, result, params):
    ws = wb.create_sheet("Analysis")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {"A": 2, "B": 32, "C": 16, "D": 14, "E": 14,
                         "F": 14, "G": 14, "H": 14})

    t = ws.cell(row=2, column=2, value=f"{company_name} — Sum-of-the-Parts (SOTP) Valuation")
    t.font = _font(bold=True, size=13)
    ws.merge_cells("B2:H2")

    rev    = _safe(fin.get("revenue", 1e11))
    ebitda = _safe(fin.get("ebitda", 1e10))
    shares = _safe(fin.get("shares", 1), 1)
    debt   = _safe(fin.get("total_debt", 0))
    cash   = _safe(fin.get("cash", 0))
    price  = _safe(fin.get("price", 800))
    mult   = _safe(params.get("entry_multiple", 12))

    row = 4
    _section_header(ws, row, 2, "SOTP — SEGMENT VALUATION", width_cols=7)
    row += 1

    headers = ["Segment", "Revenue (₹ Cr)", "EBITDA (₹ Cr)", "Margin", "EV/EBITDA", "Segment EV (₹ Cr)", "Weight"]
    for j, h in enumerate(headers):
        c = ws.cell(row=row, column=2+j, value=h)
        c.font = _font(bold=True, color=WHITE, size=9); c.fill = _fill(DARK_BLUE)
        c.alignment = _align(h="center"); c.border = _border()
    row += 1

    segments = [
        ("Core Business",        rev*0.55, ebitda*0.60, mult),
        ("Adjacent Products",    rev*0.25, ebitda*0.25, mult*0.85),
        ("New Ventures",         rev*0.12, ebitda*0.10, mult*1.20),
        ("Services / Platform",  rev*0.08, ebitda*0.05, mult*1.50),
    ]

    total_ev = 0.0
    for name, r_, e_, m in segments:
        seg_ev   = e_ * m
        total_ev += seg_ev
        margin   = e_ / r_ if r_ else 0
        fill = LIGHT_BLUE_FILL if segments.index((name, r_, e_, m)) % 2 == 0 else None
        _label(ws, row, 2, name)
        _val(ws, row, 3, r_/1e7,    fmt="#,##0", fill_hex=fill)
        _val(ws, row, 4, e_/1e7,    fmt="#,##0", fill_hex=fill)
        _pct_cell(ws, row, 5, margin, fill_hex=fill)
        _val(ws, row, 6, m,         fmt="0.0x",  fill_hex=fill)
        _val(ws, row, 7, seg_ev/1e7, fmt="#,##0", fill_hex=fill)
        _pct_cell(ws, row, 8, seg_ev/total_ev if total_ev else 0, fill_hex=fill)
        row += 1

    _label(ws, row, 2, "TOTAL", bold=True)
    _val(ws, row, 3, rev/1e7,       fmt="#,##0", fill_hex=LIGHT_BLUE_FILL, bold=True)
    _val(ws, row, 4, ebitda/1e7,    fmt="#,##0", fill_hex=LIGHT_BLUE_FILL, bold=True)
    _val(ws, row, 7, total_ev/1e7,  fmt="#,##0", fill_hex=LIGHT_BLUE_FILL, bold=True)
    row += 2

    _section_header(ws, row, 2, "EQUITY BRIDGE", width_cols=7)
    row += 1
    eq_val   = total_ev - debt + cash
    iv_share = eq_val / shares

    bridge = [
        ("Total EV (Segments, ₹ Cr)", total_ev/1e7,   "#,##0"),
        ("(−) Total Debt (₹ Cr)",     debt/1e7,        "#,##0"),
        ("(+) Cash (₹ Cr)",           cash/1e7,        "#,##0"),
        ("Equity Value (₹ Cr)",       eq_val/1e7,      "#,##0"),
        ("Intrinsic Value/Share (₹)", iv_share,         "#,##0.00"),
        ("Current Price (₹)",         price,            "#,##0.00"),
        ("Upside / (Downside)",       (iv_share/price - 1) if price else 0, "0.0%"),
    ]
    for label, value, fmt in bridge:
        bold = "Intrinsic" in label
        _label(ws, row, 2, label, bold=bold)
        if "%" in fmt:
            _pct_cell(ws, row, 3, value, fill_hex=GREEN_FILL if bold else None)
        else:
            _val(ws, row, 3, value, fmt=fmt, fill_hex=GREEN_FILL if bold else None,
                 color=GREEN_TEXT if bold else BLACK)
        row += 1

    return ws


# ── PB_BANKS ──────────────────────────────────────────────────────────────────

def _analysis_pb_banks(wb, company_name, fin, result, params):
    ws = wb.create_sheet("Analysis")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {"A": 2, "B": 36, "C": 18, "D": 18, "E": 18})

    t = ws.cell(row=2, column=2, value=f"{company_name} — P/B Valuation (Banking Sector)")
    t.font = _font(bold=True, size=13)
    ws.merge_cells("B2:E2")

    roe    = _safe(params.get("roe", fin.get("roe", 0.15)))
    coe    = _safe(params.get("cost_of_equity", 0.12))
    g      = _safe(params.get("terminal_growth", 0.05))
    bvps   = _safe(fin.get("book_value_per_share", 300))
    price  = _safe(fin.get("price", 800))

    pb_just = (roe - g) / (coe - g) if (coe - g) > 0 else 1.0
    iv = bvps * pb_just

    row = 4
    _section_header(ws, row, 2, "GORDON GROWTH P/B — JUSTIFIED P/B RATIO", width_cols=4)
    row += 1

    data = [
        ("Return on Equity (ROE)",              roe,     "0.0%"),
        ("Cost of Equity (Ke)",                 coe,     "0.0%"),
        ("Sustainable Growth Rate (g)",         g,       "0.0%"),
        ("Justified P/B = (ROE − g) / (Ke − g)", pb_just, "0.00x"),
        ("Book Value Per Share (₹)",            bvps,    "#,##0.00"),
        ("Intrinsic Value = BV × Justified P/B (₹)", iv, "#,##0.00"),
        ("Current Market Price (₹)",            price,   "#,##0.00"),
        ("Actual P/B",                          price/bvps if bvps else 0, "0.00x"),
        ("Upside / (Downside)",                 (iv/price - 1) if price else 0, "0.0%"),
    ]
    for label, value, fmt in data:
        bold = "Intrinsic" in label
        _label(ws, row, 2, label, bold=bold)
        if "%" in fmt:
            _pct_cell(ws, row, 3, value, fill_hex=GREEN_FILL if bold else None)
        else:
            _val(ws, row, 3, value, fmt=fmt, fill_hex=GREEN_FILL if bold else None,
                 color=GREEN_TEXT if bold else BLACK)
        row += 1

    row += 2
    _section_header(ws, row, 2, "DUPONT ROE DECOMPOSITION", width_cols=4)
    row += 1
    ni    = _safe(fin.get("net_income", 1e10))
    rev   = _safe(fin.get("revenue", 1e11))
    assets= _safe(fin.get("total_assets", 1e12))
    bv    = _safe(fin.get("book_value_total", 3e11), 1)

    np_margin = ni / rev if rev else 0
    asset_turn = rev / assets if assets else 0
    leverage   = assets / bv if bv else 0

    dupont = [
        ("Net Profit Margin",          np_margin,  "0.0%"),
        ("Asset Turnover",             asset_turn, "0.00x"),
        ("Equity Multiplier (Leverage)", leverage, "0.00x"),
        ("ROE = NP% × AT × EM",        np_margin * asset_turn * leverage, "0.0%"),
    ]
    for label, value, fmt in dupont:
        _label(ws, row, 2, label)
        if "%" in fmt:
            _pct_cell(ws, row, 3, value)
        else:
            _val(ws, row, 3, value, fmt=fmt)
        row += 1

    return ws


# ── CAP_RATE ──────────────────────────────────────────────────────────────────

def _analysis_cap_rate(wb, company_name, fin, result, params):
    ws = wb.create_sheet("Analysis")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {"A": 2, "B": 36, "C": 18, "D": 18, "E": 18, "F": 18})

    t = ws.cell(row=2, column=2, value=f"{company_name} — Capitalization Rate (Real Estate)")
    t.font = _font(bold=True, size=13)
    ws.merge_cells("B2:F2")

    cap     = _safe(params.get("cap_rate", 0.07))
    noi     = _safe(params.get("noi", fin.get("nopat", fin.get("ebitda", 1e10) * 0.6)))
    shares  = _safe(fin.get("shares", 1), 1)
    debt    = _safe(fin.get("total_debt", 0))
    cash    = _safe(fin.get("cash", 0))
    price   = _safe(fin.get("price", 800))
    g       = _safe(params.get("noi_growth", 0.03))

    prop_val = noi / cap if cap > 0 else 0
    iv = (prop_val - debt + cash) / shares

    row = 4
    _section_header(ws, row, 2, "CAP RATE VALUATION — NOI TO PROPERTY VALUE", width_cols=5)
    row += 1

    data = [
        ("Net Operating Income (NOI, ₹ Cr)", noi/1e7,     "#,##0"),
        ("Capitalization Rate",              cap,         "0.0%"),
        ("Property Value = NOI / Cap Rate (₹ Cr)", prop_val/1e7, "#,##0"),
        ("(−) Debt (₹ Cr)",                 debt/1e7,    "#,##0"),
        ("(+) Cash (₹ Cr)",                 cash/1e7,    "#,##0"),
        ("Equity Value (₹ Cr)",             (prop_val-debt+cash)/1e7, "#,##0"),
        ("Intrinsic Value Per Share (₹)",   iv,          "#,##0.00"),
        ("Current Market Price (₹)",        price,       "#,##0.00"),
        ("Upside / (Downside)",             (iv/price-1) if price else 0, "0.0%"),
    ]
    for label, value, fmt in data:
        bold = "Intrinsic" in label
        _label(ws, row, 2, label, bold=bold)
        if "%" in fmt:
            _pct_cell(ws, row, 3, value, fill_hex=GREEN_FILL if bold else None)
        else:
            _val(ws, row, 3, value, fmt=fmt, fill_hex=GREEN_FILL if bold else None,
                 color=GREEN_TEXT if bold else BLACK)
        row += 1

    row += 2
    _section_header(ws, row, 2, "NOI GROWTH FORECAST (5 YEARS)", width_cols=5)
    row += 1
    headers = ["Year", "NOI (₹ Cr)", "Growth", "Property Value (₹ Cr)", "Implied Price (₹)"]
    for j, h in enumerate(headers):
        c = ws.cell(row=row, column=2+j, value=h)
        c.font = _font(bold=True, color=WHITE, size=9); c.fill = _fill(DARK_BLUE)
        c.alignment = _align(h="center"); c.border = _border()
    row += 1

    noi_t = noi
    for yr in range(1, 6):
        noi_t = noi_t * (1 + g)
        pv_t  = noi_t / cap if cap > 0 else 0
        ip_t  = (pv_t - debt + cash) / shares
        fill  = LIGHT_BLUE_FILL if yr % 2 == 0 else None
        _val(ws, row, 2, yr, fmt="0")
        _val(ws, row, 3, noi_t/1e7, fmt="#,##0", fill_hex=fill)
        _pct_cell(ws, row, 4, g, fill_hex=fill)
        _val(ws, row, 5, pv_t/1e7, fmt="#,##0", fill_hex=fill)
        _val(ws, row, 6, ip_t, fmt="#,##0.00",
             fill_hex=GREEN_FILL if ip_t >= price else RED_FILL,
             color=GREEN_TEXT if ip_t >= price else RED_TEXT)
        row += 1

    return ws


# ── USER_BASED (SaaS) ─────────────────────────────────────────────────────────

def _analysis_user_based(wb, company_name, fin, result, params):
    ws = wb.create_sheet("Analysis")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {"A": 2, "B": 36, "C": 14, "D": 14, "E": 14,
                         "F": 14, "G": 14, "H": 14})

    t = ws.cell(row=2, column=2, value=f"{company_name} — User-Based / SaaS Valuation")
    t.font = _font(bold=True, size=13)
    ws.merge_cells("B2:H2")

    user_growth = _safe(params.get("user_growth", 0.30))
    churn       = _safe(params.get("churn_rate", 0.10))
    arpu        = _safe(params.get("arpu", fin.get("revenue", 1e11) / max(_safe(params.get("users", 1e6)), 1)))
    users0      = _safe(params.get("users", 1e6))
    cac         = _safe(params.get("cac", arpu * 2))
    ltv         = arpu / churn if churn > 0 else arpu * 10
    wacc        = _safe(params.get("wacc", 0.15))
    shares      = _safe(fin.get("shares", 1), 1)
    debt        = _safe(fin.get("total_debt", 0))
    cash        = _safe(fin.get("cash", 0))
    price       = _safe(fin.get("price", 800))

    row = 4
    _section_header(ws, row, 2, "USER ECONOMICS — LTV/CAC ANALYSIS", width_cols=7)
    row += 1

    data = [
        ("Current Users",              users0,   "#,##0"),
        ("User Growth Rate",           user_growth, "0.0%"),
        ("Annual Churn Rate",          churn,    "0.0%"),
        ("ARPU (Annual, ₹)",          arpu,     "#,##0.00"),
        ("CAC (Customer Acq. Cost, ₹)", cac,   "#,##0.00"),
        ("LTV = ARPU / Churn (₹)",    ltv,     "#,##0.00"),
        ("LTV / CAC Ratio",           ltv/cac if cac else 0, "0.00x"),
        ("Payback Period (months)",   (cac/arpu*12) if arpu else 0, "0.0"),
    ]
    for label, value, fmt in data:
        _label(ws, row, 2, label)
        if "%" in fmt:
            _pct_cell(ws, row, 3, value, fill_hex=YELLOW_FILL)
        else:
            _val(ws, row, 3, value, fmt=fmt, color=BLUE_TEXT, fill_hex=YELLOW_FILL)
        row += 1

    row += 1
    _section_header(ws, row, 2, "USER BASE FORECAST (5 YEARS)", width_cols=7)
    row += 1
    headers = ["Year", "Users", "New Users", "Churned", "Revenue (₹ Cr)", "FCF (₹ Cr)", "PV FCF (₹ Cr)"]
    for j, h in enumerate(headers):
        c = ws.cell(row=row, column=2+j, value=h)
        c.font = _font(bold=True, color=WHITE, size=9); c.fill = _fill(DARK_BLUE)
        c.alignment = _align(h="center"); c.border = _border()
    row += 1

    u = users0
    pv_sum = 0.0
    for yr in range(1, 6):
        new_u = u * user_growth
        churn_u = u * churn
        u = u + new_u - churn_u
        rev_t = u * arpu
        fcf_t = rev_t * 0.25
        df    = 1 / ((1 + wacc) ** yr)
        pv_t  = fcf_t * df
        pv_sum += pv_t
        fill = LIGHT_BLUE_FILL if yr % 2 == 0 else None
        _val(ws, row, 2, yr, fmt="0")
        _val(ws, row, 3, u, fmt="#,##0", fill_hex=fill)
        _val(ws, row, 4, new_u, fmt="#,##0", fill_hex=fill)
        _val(ws, row, 5, churn_u, fmt="#,##0", fill_hex=fill)
        _val(ws, row, 6, rev_t/1e7, fmt="#,##0", fill_hex=fill)
        _val(ws, row, 7, fcf_t/1e7, fmt="#,##0", fill_hex=fill)
        _val(ws, row, 8, pv_t/1e7, fmt="#,##0", fill_hex=fill)
        row += 1

    row += 1
    gT = 0.04
    tv = pv_sum * (1 + gT) / (wacc - gT) * 0.3 if (wacc - gT) > 0 else pv_sum
    eq_val = pv_sum + tv - debt + cash
    iv = eq_val / shares

    bridge = [
        ("PV of FCF (₹ Cr)",    pv_sum/1e7,    "#,##0"),
        ("Terminal Value (₹ Cr)", tv/1e7,       "#,##0"),
        ("(−) Debt (₹ Cr)",      debt/1e7,     "#,##0"),
        ("(+) Cash (₹ Cr)",      cash/1e7,     "#,##0"),
        ("Intrinsic Value/Share (₹)", iv,       "#,##0.00"),
        ("Current Price (₹)",    price,         "#,##0.00"),
    ]
    for label, value, fmt in bridge:
        bold = "Intrinsic" in label
        _label(ws, row, 2, label, bold=bold)
        _val(ws, row, 3, value, fmt=fmt, fill_hex=GREEN_FILL if bold else None,
             color=GREEN_TEXT if bold else BLACK)
        row += 1

    return ws


# ── VC_METHOD ─────────────────────────────────────────────────────────────────

def _analysis_vc_method(wb, company_name, fin, result, params):
    ws = wb.create_sheet("Analysis")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, {"A": 2, "B": 36, "C": 18, "D": 18, "E": 18, "F": 18})

    t = ws.cell(row=2, column=2, value=f"{company_name} — Venture Capital Method")
    t.font = _font(bold=True, size=13)
    ws.merge_cells("B2:F2")

    target_return = _safe(params.get("target_return", 3.0))
    rev_mult      = _safe(params.get("terminal_revenue_multiple", 8))
    rev_term      = _safe(params.get("terminal_revenue", fin.get("revenue", 1e11) * 2))
    invest        = _safe(params.get("investment", fin.get("total_assets", 1e12) * 0.1))
    hold_years    = int(_safe(params.get("hold_years", 5)))
    shares        = _safe(fin.get("shares", 1), 1)
    price         = _safe(fin.get("price", 800))

    exit_val   = rev_term * rev_mult
    pre_money  = exit_val / target_return
    post_money = pre_money + invest
    own_pct    = invest / post_money if post_money else 0
    iv         = pre_money / shares

    row = 4
    _section_header(ws, row, 2, "VENTURE CAPITAL METHOD — STEP BY STEP", width_cols=5)
    row += 1

    data = [
        ("Terminal Revenue Estimate (₹ Cr)",    rev_term/1e7, "#,##0"),
        ("Terminal Revenue Multiple",            rev_mult,     "0.0x"),
        ("Projected Exit Value (₹ Cr)",         exit_val/1e7, "#,##0"),
        ("Target Return Multiple (x)",          target_return,"0.0x"),
        ("Pre-Money Valuation = Exit / Return (₹ Cr)", pre_money/1e7, "#,##0"),
        ("Investment Size (₹ Cr)",              invest/1e7,   "#,##0"),
        ("Post-Money Valuation (₹ Cr)",         post_money/1e7, "#,##0"),
        ("Investor Ownership %",                own_pct,      "0.0%"),
        ("Implied Intrinsic Value/Share (₹)",   iv,           "#,##0.00"),
        ("Current Market Price (₹)",            price,        "#,##0.00"),
        ("Upside / (Downside)",                 (iv/price - 1) if price else 0, "0.0%"),
    ]
    for label, value, fmt in data:
        bold = "Intrinsic" in label
        _label(ws, row, 2, label, bold=bold)
        if "%" in fmt:
            _pct_cell(ws, row, 3, value, fill_hex=GREEN_FILL if bold else None)
        else:
            _val(ws, row, 3, value, fmt=fmt, fill_hex=GREEN_FILL if bold else None,
                 color=GREEN_TEXT if bold else BLACK)
        row += 1

    row += 2
    _section_header(ws, row, 2, "VC RETURN SCENARIO ANALYSIS", width_cols=5)
    row += 1
    headers = ["Scenario", "Return Multiple", "Exit Value (₹ Cr)", "Pre-Money (₹ Cr)", "Implied Price (₹)"]
    for j, h in enumerate(headers):
        c = ws.cell(row=row, column=2+j, value=h)
        c.font = _font(bold=True, color=WHITE, size=9); c.fill = _fill(DARK_BLUE)
        c.alignment = _align(h="center"); c.border = _border()
    row += 1

    scenarios = [
        ("Bear Case",   1.5),
        ("Base Case",   target_return),
        ("Bull Case",   target_return * 1.5),
        ("Home Run",    target_return * 2.5),
        ("Moonshot",    target_return * 4.0),
    ]
    for name, ret in scenarios:
        pm  = exit_val / ret
        iv_ = pm / shares
        prem = iv_ / price - 1 if price else 0
        fill = GREEN_FILL if prem >= 0 else RED_FILL
        _label(ws, row, 2, name)
        _val(ws, row, 3, ret, fmt="0.0x")
        _val(ws, row, 4, exit_val/1e7, fmt="#,##0")
        _val(ws, row, 5, pm/1e7, fmt="#,##0")
        _val(ws, row, 6, iv_, fmt="#,##0.00", fill_hex=fill,
             color=GREEN_TEXT if prem >= 0 else RED_TEXT)
        row += 1

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SENSITIVITY FUNCTIONS — one per model
# ═══════════════════════════════════════════════════════════════════════════════

def _make_sens(model_id, fin, result, params):
    """Return a closure: fn(ws, start_row, fin, result, params, current_price)"""

    def _gordon_sens(ws, row, fin, result, params, cp):
        coe  = _safe(params.get("cost_of_equity", 0.12))
        g    = _safe(params.get("terminal_growth", 0.05))
        dps  = _safe(params.get("dps", fin.get("dps", 5)))
        coe_vals = _range_around(coe, 0.30)
        g_vals   = _range_around(g, 0.30)
        matrix   = [[dps*(1+gv)/(cv - gv) if (cv - gv) > 0.001 else 0
                     for gv in g_vals] for cv in coe_vals]
        _write_sens_table(ws, row, 2, "Cost of Equity", "Dividend Growth",
                          coe_vals, g_vals, matrix, 2, 2, cp)

    def _ddm_ms_sens(ws, row, fin, result, params, cp):
        coe  = _safe(params.get("cost_of_equity", 0.12))
        g1   = _safe(params.get("growth_stage1", 0.15))
        dps  = _safe(params.get("dps", fin.get("dps", 5)))
        coe_vals = _range_around(coe, 0.30)
        g1_vals  = _range_around(g1, 0.30)
        def _ddm_iv(c, g):
            pv = sum(dps*((1+g)**yr)/(((1+c)**yr)) for yr in range(1, 6))
            last = dps*(1+g)**5
            gt = g * 0.4
            tv = last*(1+gt)/(c-gt) if (c-gt)>0 else 0
            pv_tv = tv/((1+c)**5)
            return pv + pv_tv
        matrix = [[_ddm_iv(cv, gv) for gv in g1_vals] for cv in coe_vals]
        _write_sens_table(ws, row, 2, "Cost of Equity", "Stage1 Growth",
                          coe_vals, g1_vals, matrix, 2, 2, cp)

    def _ri_sens(ws, row, fin, result, params, cp):
        coe  = _safe(params.get("cost_of_equity", 0.12))
        roe  = _safe(params.get("roe", fin.get("roe", 0.18)))
        bvps = _safe(params.get("book_value_per_share", fin.get("book_value_per_share", 100)))
        coe_vals = _range_around(coe, 0.30)
        roe_vals = _range_around(roe, 0.30)
        g = _safe(params.get("terminal_growth", 0.04))
        def _ri_iv(c, r):
            pv = sum((bvps*(r-c))/((1+c)**yr) for yr in range(1, 11))
            last_ri = bvps * (r - c)
            tv = last_ri * (1+g) / (c - g) if (c - g) > 0 else 0
            return bvps + pv + tv/((1+c)**10)
        matrix = [[_ri_iv(cv, rv) for rv in roe_vals] for cv in coe_vals]
        _write_sens_table(ws, row, 2, "Cost of Equity", "ROE",
                          coe_vals, roe_vals, matrix, 2, 2, cp)

    def _trading_comps_sens(ws, row, fin, result, params, cp):
        # For trading comps we re-use the analysis sheet table, just note it
        c = ws.cell(row=row, column=2, value="See Analysis sheet for implied price matrix at different P/E and P/B multiples.")
        c.font = _font(size=10, italic=True, color="808080")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)

    def _prec_trans_sens(ws, row, fin, result, params, cp):
        em   = _safe(params.get("ev_ebitda_multiple", 14))
        prem = _safe(params.get("deal_premium", 0.25))
        ebitda = _safe(fin.get("ebitda", 1e10))
        shares = _safe(fin.get("shares", 1), 1)
        debt   = _safe(fin.get("total_debt", 0))
        cash   = _safe(fin.get("cash", 0))
        em_vals   = _range_around(em, 0.30)
        prem_vals = _range_around(prem, 0.30)
        matrix = [[(ebitda*ev - debt + cash)/shares*(1+pr)
                   for pr in prem_vals] for ev in em_vals]
        _write_sens_table(ws, row, 2, "EV/EBITDA Multiple", "Deal Premium",
                          em_vals, prem_vals, matrix, 2, 2, cp)

    def _peg_sens(ws, row, fin, result, params, cp):
        g   = _safe(params.get("earnings_growth_pct", fin.get("earnings_growth", 0.15)))
        peg = _safe(params.get("target_peg", 1.0))
        eps = _safe(fin.get("eps", 50))
        g_vals   = _range_around(g, 0.30)
        peg_vals = _range_around(peg, 0.30)
        g_pct0 = g * 100 if g < 2 else g
        matrix = [[eps * pv * (gv * 100 if gv < 2 else gv) for pv in peg_vals] for gv in g_vals]
        _write_sens_table(ws, row, 2, "Earnings Growth", "Target PEG",
                          g_vals, peg_vals, matrix, 2, 2, cp)

    def _rev_mult_sens(ws, row, fin, result, params, cp):
        mult   = _safe(params.get("ev_revenue_multiple", 5))
        g      = _safe(params.get("revenue_growth", fin.get("revenue_growth", 0.15)))
        rev    = _safe(fin.get("revenue", 1e11))
        shares = _safe(fin.get("shares", 1), 1)
        debt   = _safe(fin.get("total_debt", 0))
        cash   = _safe(fin.get("cash", 0))
        m_vals = _range_around(mult, 0.30)
        g_vals = _range_around(g, 0.30)
        matrix = [[(rev*(1+gv)*mv - debt + cash)/shares for gv in g_vals] for mv in m_vals]
        _write_sens_table(ws, row, 2, "EV/Revenue Multiple", "Revenue Growth",
                          m_vals, g_vals, matrix, 2, 2, cp)

    def _nav_sens(ws, row, fin, result, params, cp):
        assets = _safe(fin.get("total_assets", 1e12))
        liabs  = _safe(fin.get("total_liabilities", 5e11))
        shares = _safe(fin.get("shares", 1), 1)
        base_r = _safe(params.get("haircut_rate", 0.15))
        rates1 = [0.60, 0.70, 0.80, 0.90, 1.00]
        rates2 = [0.50, 0.60, 0.70, 0.80, 0.90]
        matrix = [[(assets*r1 - liabs)/shares for r2 in rates2] for r1 in rates1]
        _write_sens_table(ws, row, 2, "Asset Recovery (Cash/Recv)", "PPE Recovery",
                          rates1, rates2, matrix, 2, 2, cp)

    def _liq_sens(ws, row, fin, result, params, cp):
        cash_rate = _safe(params.get("cash_rate", 1.0))
        ppe_rate  = _safe(params.get("ppe_rate", 0.5))
        assets    = _safe(fin.get("total_assets", 1e12))
        liabs     = _safe(fin.get("total_liabilities", 5e11))
        cash      = _safe(fin.get("cash", 0))
        shares    = _safe(fin.get("shares", 1), 1)
        cr_vals   = _range_around(cash_rate, 0.20)
        pp_vals   = _range_around(ppe_rate, 0.30)
        def _liq_iv(cr, pr):
            total = cash*cr + assets*0.10*0.80 + assets*0.08*0.60 + assets*0.35*pr + assets*0.07*0.30
            eq = total - liabs - total * 0.05
            return eq / shares
        matrix = [[_liq_iv(cr, pr) for pr in pp_vals] for cr in cr_vals]
        _write_sens_table(ws, row, 2, "Cash Recovery Rate", "PPE Recovery Rate",
                          cr_vals, pp_vals, matrix, 2, 2, cp)

    def _rep_cost_sens(ws, row, fin, result, params, cp):
        mult    = _safe(params.get("rebuild_multiplier", 1.15))
        dep_adj = _safe(params.get("depreciation_adj", 0.20))
        assets  = _safe(fin.get("total_assets", 1e12))
        liabs   = _safe(fin.get("total_liabilities", 5e11))
        shares  = _safe(fin.get("shares", 1), 1)
        m_vals  = _range_around(mult, 0.30)
        d_vals  = _range_around(dep_adj, 0.30)
        def _rep_iv(m, d):
            rep = (assets*0.10*1.30 + assets*0.15*m*(1-d) + assets*0.20*m*(1-d*1.5) +
                   assets*0.08*1.40*(1-d*0.5) + assets*0.08*1.05 + assets*0.12*1.50)
            return (rep - liabs) / shares
        matrix = [[_rep_iv(m, d) for d in d_vals] for m in m_vals]
        _write_sens_table(ws, row, 2, "Rebuild Multiplier", "Depreciation Adj.",
                          m_vals, d_vals, matrix, 2, 2, cp)

    def _cap_earn_sens(ws, row, fin, result, params, cp):
        req = _safe(params.get("required_return", 0.12))
        g   = _safe(params.get("eps_growth", fin.get("earnings_growth", 0.10)))
        eps = _safe(fin.get("eps", 50))
        r_vals = _range_around(req, 0.30)
        g_vals = _range_around(g, 0.30)
        matrix = [[eps*(1+gv)/(rv-gv) if (rv-gv)>0.001 else 0
                   for gv in g_vals] for rv in r_vals]
        _write_sens_table(ws, row, 2, "Required Return", "EPS Growth",
                          r_vals, g_vals, matrix, 2, 2, cp)

    def _excess_earn_sens(ws, row, fin, result, params, cp):
        fair  = _safe(params.get("fair_return_rate", 0.10))
        disc  = _safe(params.get("discount_rate", 0.12))
        ni    = _safe(fin.get("net_income", 1e10))
        assets= _safe(fin.get("total_assets", 1e12))
        liabs = _safe(fin.get("total_liabilities", 5e11))
        shares= _safe(fin.get("shares", 1), 1)
        ta    = (assets - liabs) * 0.70
        f_vals = _range_around(fair, 0.30)
        d_vals = _range_around(disc, 0.30)
        matrix = [[(ta + (ni - ta*fv)/dv)/shares if dv>0 else 0
                   for dv in d_vals] for fv in f_vals]
        _write_sens_table(ws, row, 2, "Fair Return Rate", "Discount Rate",
                          f_vals, d_vals, matrix, 2, 2, cp)

    def _eva_sens(ws, row, fin, result, params, cp):
        wacc = _safe(params.get("wacc", 0.12))
        g1   = _safe(params.get("growth_stage1", 0.15))
        nopat= _safe(fin.get("nopat", fin.get("net_income", 1e10)*0.8))
        ic   = _safe(fin.get("invested_capital", fin.get("total_assets", 1e12)*0.6))
        debt = _safe(fin.get("total_debt", 0))
        cash = _safe(fin.get("cash", 0))
        shares= _safe(fin.get("shares", 1), 1)
        gT   = _safe(params.get("terminal_growth", 0.04))
        w_vals = _range_around(wacc, 0.30)
        g_vals = _range_around(g1, 0.30)
        def _eva_iv(w, g):
            pv = 0.0; n = nopat; i = ic
            for yr in range(1, 11):
                gy = g if yr <= 5 else (g+gT)/2
                n *= (1+gy); i *= (1+gy*0.8)
                pv += (n - i*w) / ((1+w)**yr)
            last_eva = n - i * w
            tv = last_eva*(1+gT)/(w-gT) if (w-gT)>0 else 0
            return (ic + pv + tv/((1+w)**10) - debt + cash) / shares
        matrix = [[_eva_iv(w, g) for g in g_vals] for w in w_vals]
        _write_sens_table(ws, row, 2, "WACC", "Stage1 Growth",
                          w_vals, g_vals, matrix, 2, 2, cp)

    def _cfroi_sens(ws, row, fin, result, params, cp):
        req  = _safe(params.get("required_return", 0.12))
        life = _safe(params.get("asset_life", 10))
        ic   = _safe(fin.get("invested_capital", fin.get("total_assets", 1e12)*0.6))
        ocf  = _safe(fin.get("operating_cf", fin.get("nopat", 1e10)))
        debt = _safe(fin.get("total_debt", 0))
        cash = _safe(fin.get("cash", 0))
        shares= _safe(fin.get("shares", 1), 1)
        r_vals = _range_around(req, 0.30)
        l_vals = [max(3, int(life*(0.7+0.15*i))) for i in range(5)]
        def _cf_iv(r, l):
            cfroi = ocf / ic if ic > 0 else r
            ev = ic * cfroi / r if r > 0 else ic
            return (ev - debt + cash) / shares
        matrix = [[_cf_iv(r, l) for l in l_vals] for r in r_vals]
        _write_sens_table(ws, row, 2, "Required Return", "Asset Life (yrs)",
                          r_vals, [float(v) for v in l_vals], matrix, 2, 2, cp)

    def _lbo_sens(ws, row, fin, result, params, cp):
        em   = _safe(params.get("entry_multiple", 8))
        exm  = _safe(params.get("exit_multiple", 10))
        ebitda = _safe(params.get("ebitda", fin.get("ebitda", 1e11)))
        dr   = _safe(params.get("debt_ratio", 0.60))
        hold = int(_safe(params.get("hold_years", 5)))
        shares = _safe(params.get("shares", fin.get("shares", 1)), 1)
        g    = _safe(fin.get("revenue_growth", 0.08))
        em_vals  = _range_around(em, 0.30)
        exm_vals = _range_around(exm, 0.30)
        def _lbo_iv(entry, exit_):
            entry_ev = ebitda * entry
            eq_in    = entry_ev * (1 - dr)
            exit_ebitda = ebitda * ((1 + g) ** hold)
            exit_ev  = exit_ebitda * exit_
            eq_out   = exit_ev * (1 - dr * 0.4)
            return eq_out / shares
        matrix = [[_lbo_iv(e, x) for x in exm_vals] for e in em_vals]
        _write_sens_table(ws, row, 2, "Entry Multiple", "Exit Multiple",
                          em_vals, exm_vals, matrix, 2, 2, cp)

    def _bs_sens(ws, row, fin, result, params, cp):
        sigma = _safe(params.get("volatility", fin.get("volatility_annual", 0.30)))
        T     = _safe(params.get("time_years", 1.0))
        S     = _safe(fin.get("price", 800))
        K     = _safe(params.get("strike", S * 1.05))
        r     = _safe(params.get("risk_free", 0.071))
        opt   = params.get("option_type", "call")
        s_vals = _range_around(sigma, 0.30)
        t_vals = _range_around(T, 0.30)
        matrix = [[_bs_price(S, K, r, tv, sv, opt) for tv in t_vals] for sv in s_vals]
        _write_sens_table(ws, row, 2, "Volatility (σ)", "Time (years)",
                          s_vals, t_vals, matrix, 2, 2, cp)

    def _real_opt_sens(ws, row, fin, result, params, cp):
        sigma = _safe(params.get("volatility", fin.get("volatility_annual", 0.35)))
        T     = _safe(params.get("time_years", 2.0))
        ic    = _safe(fin.get("invested_capital", fin.get("total_assets", 1e12) * 0.6))
        shares= _safe(fin.get("shares", 1), 1)
        S     = _safe(params.get("asset_value", ic / shares))
        K     = _safe(params.get("strike", S))
        r     = _safe(params.get("risk_free", 0.071))
        s_vals = _range_around(sigma, 0.30)
        t_vals = _range_around(T, 0.30)
        matrix = [[_bs_price(S, K, r, tv, sv, "call") for tv in t_vals] for sv in s_vals]
        _write_sens_table(ws, row, 2, "Project Volatility", "Time (years)",
                          s_vals, t_vals, matrix, 2, 2, cp)

    def _sotp_sens(ws, row, fin, result, params, cp):
        mult   = _safe(params.get("entry_multiple", 12))
        disc   = _safe(params.get("discount", 0.10))
        ebitda = _safe(fin.get("ebitda", 1e10))
        shares = _safe(fin.get("shares", 1), 1)
        debt   = _safe(fin.get("total_debt", 0))
        cash   = _safe(fin.get("cash", 0))
        m_vals = _range_around(mult, 0.30)
        d_vals = _range_around(disc, 0.30)
        matrix = [[ebitda*mv*(1-dv)/shares for dv in d_vals] for mv in m_vals]
        _write_sens_table(ws, row, 2, "EBITDA Multiple", "Portfolio Discount",
                          m_vals, d_vals, matrix, 2, 2, cp)

    def _pb_banks_sens(ws, row, fin, result, params, cp):
        roe = _safe(params.get("roe", fin.get("roe", 0.15)))
        coe = _safe(params.get("cost_of_equity", 0.12))
        g   = _safe(params.get("terminal_growth", 0.05))
        bvps= _safe(fin.get("book_value_per_share", 300))
        r_vals = _range_around(roe, 0.30)
        c_vals = _range_around(coe, 0.30)
        matrix = [[bvps*(rv-g)/(cv-g) if (cv-g)>0.001 else 0
                   for cv in c_vals] for rv in r_vals]
        _write_sens_table(ws, row, 2, "ROE", "Cost of Equity",
                          r_vals, c_vals, matrix, 2, 2, cp)

    def _cap_rate_sens(ws, row, fin, result, params, cp):
        cap  = _safe(params.get("cap_rate", 0.07))
        noi  = _safe(params.get("noi", fin.get("nopat", fin.get("ebitda", 1e10)*0.6)))
        shares= _safe(fin.get("shares", 1), 1)
        debt = _safe(fin.get("total_debt", 0))
        cash = _safe(fin.get("cash", 0))
        cap_vals = _range_around(cap, 0.30)
        noi_vals = _range_around(noi, 0.30)
        matrix = [[(nv/cv - debt + cash)/shares if cv > 0 else 0
                   for nv in noi_vals] for cv in cap_vals]
        _write_sens_table(ws, row, 2, "Cap Rate", "NOI (₹)",
                          cap_vals, [v/1e7 for v in noi_vals], matrix, 2, 2, cp)

    def _user_sens(ws, row, fin, result, params, cp):
        ug   = _safe(params.get("user_growth", 0.30))
        churn= _safe(params.get("churn_rate", 0.10))
        arpu = _safe(params.get("arpu", fin.get("revenue", 1e11) / max(_safe(params.get("users", 1e6)), 1)))
        users= _safe(params.get("users", 1e6))
        wacc = _safe(params.get("wacc", 0.15))
        shares= _safe(fin.get("shares", 1), 1)
        debt = _safe(fin.get("total_debt", 0))
        cash = _safe(fin.get("cash", 0))
        ug_vals = _range_around(ug, 0.30)
        ch_vals = _range_around(churn, 0.30)
        def _user_iv(ug_, ch_):
            u = users
            pv = 0.0
            for yr in range(1, 6):
                u = u + u*ug_ - u*ch_
                fcf = u * arpu * 0.25
                pv += fcf / ((1+wacc)**yr)
            gT = 0.04
            tv = pv * (1+gT) / (wacc-gT) * 0.3 if (wacc-gT) > 0 else pv
            return (pv + tv - debt + cash) / shares
        matrix = [[_user_iv(ug_, ch_) for ch_ in ch_vals] for ug_ in ug_vals]
        _write_sens_table(ws, row, 2, "User Growth Rate", "Churn Rate",
                          ug_vals, ch_vals, matrix, 2, 2, cp)

    def _vc_sens(ws, row, fin, result, params, cp):
        tr  = _safe(params.get("target_return", 3.0))
        rm  = _safe(params.get("terminal_revenue_multiple", 8))
        rev = _safe(params.get("terminal_revenue", fin.get("revenue", 1e11)*2))
        shares= _safe(fin.get("shares", 1), 1)
        tr_vals = _range_around(tr, 0.30)
        rm_vals = _range_around(rm, 0.30)
        matrix = [[(rev*rv)/tr_/shares for rv in rm_vals] for tr_ in tr_vals]
        _write_sens_table(ws, row, 2, "Target Return (x)", "Revenue Multiple",
                          tr_vals, rm_vals, matrix, 2, 2, cp)

    def _dcf_fcfe_sens(ws, row, fin, result, params, cp):
        coe  = _safe(params.get("cost_of_equity", 0.12))
        g    = _safe(params.get("terminal_growth", 0.05))
        ni0  = _safe(fin.get("net_income", 1e10))
        g1   = _safe(params.get("growth_stage1", fin.get("earnings_growth", 0.15)))
        shares= _safe(fin.get("shares", 1), 1)
        coe_vals = _range_around(coe, 0.30)
        g_vals   = _range_around(g, 0.30)
        def _fcfe_iv(c, gt):
            fcfe_list = [ni0*(1+g1)**yr*0.80 for yr in range(1, 6)]
            pv = sum(f/((1+c)**yr) for yr, f in enumerate(fcfe_list, 1))
            tv = fcfe_list[-1]*(1+gt)/(c-gt) if (c-gt)>0 else 0
            return (pv + tv/((1+c)**5)) / shares
        matrix = [[_fcfe_iv(c, g) for g in g_vals] for c in coe_vals]
        _write_sens_table(ws, row, 2, "Cost of Equity", "Terminal Growth",
                          coe_vals, g_vals, matrix, 2, 2, cp)

    def _dcf_ms_sens(ws, row, fin, result, params, cp):
        wacc = _safe(params.get("wacc", 0.12))
        g1   = _safe(params.get("growth_stage1", 0.20))
        gT   = _safe(params.get("terminal_growth", 0.05))
        fcf0 = _safe(fin.get("fcf", fin.get("nopat", 1e10)))
        debt = _safe(fin.get("total_debt", 0))
        cash = _safe(fin.get("cash", 0))
        shares= _safe(fin.get("shares", 1), 1)
        w_vals  = _range_around(wacc, 0.30)
        g1_vals = _range_around(g1, 0.30)
        def _ms_iv(w, g):
            fcf = fcf0; pv = 0.0
            for yr in range(1, 11):
                fcf *= (1 + (g if yr <= 5 else gT))
                pv += fcf / ((1+w)**yr)
            tv = fcf*(1+gT)/(w-gT) if (w-gT)>0 else 0
            return (pv + tv/((1+w)**10) - debt + cash) / shares
        matrix = [[_ms_iv(w, g) for g in g1_vals] for w in w_vals]
        _write_sens_table(ws, row, 2, "WACC", "Stage1 Growth",
                          w_vals, g1_vals, matrix, 2, 2, cp)

    sens_map = {
        "dcf_fcfe":               _dcf_fcfe_sens,
        "dcf_multistage":         _dcf_ms_sens,
        "gordon_growth":          _gordon_sens,
        "ddm_multistage":         _ddm_ms_sens,
        "residual_income":        _ri_sens,
        "trading_comps":          _trading_comps_sens,
        "precedent_transactions": _prec_trans_sens,
        "peg":                    _peg_sens,
        "revenue_multiple":       _rev_mult_sens,
        "nav":                    _nav_sens,
        "liquidation":            _liq_sens,
        "replacement_cost":       _rep_cost_sens,
        "capitalized_earnings":   _cap_earn_sens,
        "excess_earnings":        _excess_earn_sens,
        "eva":                    _eva_sens,
        "cfroi":                  _cfroi_sens,
        "lbo":                    _lbo_sens,
        "black_scholes":          _bs_sens,
        "real_options":           _real_opt_sens,
        "sum_of_parts":           _sotp_sens,
        "pb_banks":               _pb_banks_sens,
        "cap_rate":               _cap_rate_sens,
        "user_based":             _user_sens,
        "vc_method":              _vc_sens,
    }
    return sens_map.get(model_id, _gordon_sens)


# ═══════════════════════════════════════════════════════════════════════════════
# ANALYSIS SHEET ROUTER
# ═══════════════════════════════════════════════════════════════════════════════

ANALYSIS_MAP = {
    "dcf_fcfe":               _analysis_dcf_fcfe,
    "dcf_multistage":         _analysis_dcf_multistage,
    "gordon_growth":          _analysis_gordon_growth,
    "ddm_multistage":         _analysis_ddm_multistage,
    "residual_income":        _analysis_residual_income,
    "trading_comps":          _analysis_trading_comps,
    "precedent_transactions": _analysis_precedent_transactions,
    "peg":                    _analysis_peg,
    "revenue_multiple":       _analysis_revenue_multiple,
    "nav":                    _analysis_nav,
    "liquidation":            _analysis_liquidation,
    "replacement_cost":       _analysis_replacement_cost,
    "capitalized_earnings":   _analysis_capitalized_earnings,
    "excess_earnings":        _analysis_excess_earnings,
    "eva":                    _analysis_eva,
    "cfroi":                  _analysis_cfroi,
    "lbo":                    _analysis_lbo,
    "black_scholes":          _analysis_black_scholes,
    "real_options":           _analysis_real_options,
    "sum_of_parts":           _analysis_sum_of_parts,
    "pb_banks":               _analysis_pb_banks,
    "cap_rate":               _analysis_cap_rate,
    "user_based":             _analysis_user_based,
    "vc_method":              _analysis_vc_method,
}


# ═══════════════════════════════════════════════════════════════════════════════
# PUBLIC API
# ═══════════════════════════════════════════════════════════════════════════════

def generate_model_excel(
    model_id: str,
    fin: dict,
    result: dict,
    params: dict,
    symbol: str,
    company_name: str,
) -> bytes:
    """
    Generate a professional 4-sheet Excel workbook for any valuation model.

    Args:
        model_id:     Model identifier string (e.g. "lbo", "gordon_growth").
        fin:          Financial data dict (price, shares, revenue, ebitda, …).
        result:       Model result dict (intrinsic_value, upside_pct, …).
        params:       User-supplied model parameter dict.
        symbol:       Ticker symbol string.
        company_name: Company name string.

    Returns:
        bytes — raw .xlsx file content.
    """
    # Sanitise fin dict — replace None / NaN with 0
    safe_fin = {k: (_safe(v) if isinstance(v, (int, float)) else v)
                for k, v in fin.items()}

    wb = Workbook()

    # Sheet 1 — Cover
    _build_cover(wb, model_id, symbol, company_name, safe_fin, result)

    # Sheet 2 — Inputs & Assumptions
    _build_inputs(wb, model_id, company_name, safe_fin, params)

    # Sheet 3 — Analysis (model-specific)
    analysis_fn = ANALYSIS_MAP.get(model_id)
    if analysis_fn is None:
        # Fallback: generic analysis sheet
        ws_a = wb.create_sheet("Analysis")
        ws_a.sheet_view.showGridLines = False
        _set_col_widths(ws_a, {"A": 2, "B": 40, "C": 20})
        t = ws_a.cell(row=2, column=2,
                      value=f"{company_name} — {model_id} Analysis (Auto-generated)")
        t.font = _font(bold=True, size=13)
        ws_a.merge_cells("B2:C2")
        row = 4
        for key, val in result.items():
            if isinstance(val, (int, float, str)):
                _label(ws_a, row, 2, key.replace("_", " ").title())
                c = ws_a.cell(row=row, column=3, value=val)
                c.font = _font(size=9); c.border = _border()
                row += 1
    else:
        analysis_fn(wb, company_name, safe_fin, result, params)

    # Sheet 4 — Results & Sensitivity
    sens_fn = _make_sens(model_id, safe_fin, result, params)
    _build_results_sensitivity(wb, model_id, company_name, safe_fin, result, params, sens_fn)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
